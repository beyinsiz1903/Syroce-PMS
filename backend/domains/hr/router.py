"""
Domain Router: HR Operations

HR complete suite, F&B complete suite for department managers.

Türk İş Kanunu uyumlu defaultlar (2026):
  - Aylık standart saat: 195 (45 sa/hf × 4.33)
  - Saatlik brüt asgari taban: 140 TL (yaklaşık 2026 asgari ücret)
  - Fazla mesai zammı: %50 (overtime_rate = hourly_rate * 1.5)
  - Para birimi: TRY
"""
import base64
import io
import uuid
from datetime import UTC, date, datetime, timedelta, timezone
from typing import Any, Literal

from bson import ObjectId
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorGridFSBucket
from pydantic import BaseModel, Field, field_validator
from pymongo.errors import DuplicateKeyError

from core.database import _raw_db, db
from core.security import get_current_user

# GridFS bucket — personel belgeleri için (5MB üstü destek + memory verimi).
# Eski kayıtlar `data_b64` alanı üzerinden okunmaya devam eder (geriye uyum).
_hr_docs_bucket = AsyncIOMotorGridFSBucket(_raw_db, bucket_name='staff_docs')
from core.audit import log_audit_event  # v2 HR Foundation (Task #262)
from models.schemas import User
from modules.pms_core.role_permission_service import (  # v96 DW
    RolePermissionService,
    require_op,
)


# ============= HR v2 Foundation helpers (Task #262) =============

_HR_RPS = RolePermissionService()


def _user_has_hr_op(user: User, op: str) -> bool:
    """True iff `user` `op` izinine sahip (super_admin/admin bypass dahil).

    PII maskeleme + dept scope kararlarında kullanılır; require_op dependency'sini
    Depends dışında, response serializer içinde sorgulamamızı sağlar.
    """
    from core.security import _is_super_admin
    try:
        if _is_super_admin(user):
            return True
    except Exception:
        pass
    role = getattr(user, "role", None)
    granted = getattr(user, "granted_permissions", None)
    try:
        return _HR_RPS.check_permission(role, op, granted_permissions=granted)
    except Exception:
        return False


def _user_assigned_department(user: User) -> str | None:
    """Department-Manager scope alanı. User şemasında opsiyonel; dict/attr both ok."""
    val = getattr(user, "assigned_department", None)
    if not val and hasattr(user, "model_dump"):
        try:
            val = user.model_dump().get("assigned_department")
        except Exception:
            val = None
    return (val or None) if isinstance(val, str) else None


_PII_PHONE_FIELDS = ("phone", "mobile", "emergency_phone")
_PII_ID_FIELDS = ("national_id", "identity_number", "tc_kimlik", "tc")
_PII_BANK_FIELDS = ("iban", "bank_iban", "bank_account")
_PII_SALARY_FIELDS = (
    "salary", "monthly_salary", "hourly_rate", "gross_salary",
    "net_salary", "base_salary",
)


def _mask_phone(v: str | None) -> str | None:
    if not v or not isinstance(v, str):
        return v
    if v.startswith("aes256gcm:"):
        return ""
    if len(v) <= 4:
        return "***"
    return f"***{v[-4:]}"


def _mask_id(v: str | None) -> str | None:
    if not v or not isinstance(v, str):
        return v
    if v.startswith("aes256gcm:"):
        return ""
    if len(v) <= 4:
        return "***"
    return f"***-**-{v[-4:]}"


def _mask_iban(v: str | None) -> str | None:
    if not v or not isinstance(v, str):
        return v
    if v.startswith("aes256gcm:"):
        return ""
    compact = v.replace(" ", "")
    if len(compact) <= 4:
        return "****"
    return f"****{compact[-4:]}"


def _mask_hr_pii(
    record: dict | None,
    current_user: User,
    *,
    self_id: str | None = None,
    self_email: str | None = None,
) -> dict | None:
    """Rol-bazlı PII maskeleme — staff/profile/salary serializer'larda kullanılır.

    Tam görünürlük YALNIZCA `manage_hr` perm'iyle (HR Admin / HR Manager /
    super_admin). `view_hr` (Finance/Supervisor) yeterli DEĞİL — payroll
    raporları için ayrı `view_hr_payroll` perm'i var, ancak hassas PII
    (TC/IBAN/telefon) için ÖZEL olarak `manage_hr` gerekir (KVKK least-privilege).

    Diğer roller için:
      - phone/mobile/emergency_phone → son 4 hane
      - national_id/tc → `***-**-1234`
      - iban → son 4 hane
      - salary/hourly_rate → maskelenir (`None`)

    Self-service istisnası: kaydın id'i current_user.id ile eşleşir VEYA
    e-posta normalize edilmiş halde eşleşirse (users-derived staff ve
    user↔staff link table'sız ortamda determinist self-mapping) → unmask.
    """
    if not record or not isinstance(record, dict):
        return record
    # Tam görünürlük (Task #264 — strict allowlist, KVKK least-privilege):
    # yalnız HR Admin (manage_hr perm) + Finance role + super_admin role +
    # self-service. view_hr_payroll TEK BAŞINA YETMEZ — payroll-view perm'i
    # tutar görmek içindir, IBAN/TC için ayrı entitlement gerekir.
    if _user_has_hr_op(current_user, "manage_hr"):
        return record
    role_lc = (getattr(current_user, "role", "") or "").lower()
    if role_lc in ("finance", "super_admin"):
        return record
    # Self-service: id eşleşmesi.
    rec_id = record.get("id") or record.get("staff_id") or record.get("user_id")
    if self_id and rec_id and str(rec_id) == str(self_id):
        return record
    # Self-service: e-posta eşleşmesi (user↔staff link table yokken
    # deterministic mapping). Boş/None e-posta eşleşmesi YASAK.
    rec_email = (record.get("email") or "").strip().lower()
    me_email = (self_email or "").strip().lower()
    if rec_email and me_email and rec_email == me_email:
        return record
    rec = dict(record)
    for f in _PII_PHONE_FIELDS:
        if f in rec:
            rec[f] = _mask_phone(rec.get(f))
    for f in _PII_ID_FIELDS:
        if f in rec:
            rec[f] = _mask_id(rec.get(f))
    for f in _PII_BANK_FIELDS:
        if f in rec:
            rec[f] = _mask_iban(rec.get(f))
    for f in _PII_SALARY_FIELDS:
        if f in rec:
            rec[f] = None
    return rec


def _authorize_staff_access(
    staff: dict | None,
    current_user: User,
    *,
    require_manage: bool = False,
) -> None:
    """Per-record authorization gate for staff detail endpoints.

    Centralized RBAC + tenant + assigned_department + self-service kontrolü.
    Mevcut auth zinciri (Depends(get_current_user) + per-route require_op)
    üzerine eklenir; ID-known IDOR / cross-department leakage'ı engeller.

    Karar matrisi (sırayla):
      1. `staff` None ise → çağıran 404 yükseltsin (early-return).
      2. `manage_hr` perm'i varsa (HR Admin / SUPER_ADMIN / SUPERVISOR) → ALLOW.
      3. Self-service: kaydın id'i veya e-postası current_user ile eşleşir →
         ALLOW (require_manage olsa bile — kullanıcı kendi performansını/
         maaşını görebilir; profil endpoint'iyle tutarlı).
      4. `require_manage=True` ise (örn. performans notları, başka kullanıcı) → 403.
      5. `view_hr` perm'i + (assigned_department yok VEYA dept eşleşir) → ALLOW.
      6. Aksi → 403.
    """
    if not staff or not isinstance(staff, dict):
        return
    if _user_has_hr_op(current_user, "manage_hr"):
        return
    # Self-service (id veya e-posta eşleşmesi). `require_manage=True` olsa
    # bile kullanıcı KENDİ kaydına bakıyorsa erişim ALLOW — performans/maaş
    # gibi alanlarda da self bypass profil endpoint'iyle tutarlı.
    self_id = str(getattr(current_user, "id", "") or "")
    rec_id = str(staff.get("id") or "")
    self_email = (str(getattr(current_user, "email", "") or "")).strip().lower()
    rec_email = (str(staff.get("email") or "")).strip().lower()
    is_self = bool(
        (self_id and rec_id and self_id == rec_id)
        or (self_email and rec_email and self_email == rec_email)
    )
    if is_self:
        return
    if require_manage:
        raise HTTPException(status_code=403, detail="Yetkiniz yok (manage_hr gerekir).")
    # Dept scope (Department-Manager).
    if not _user_has_hr_op(current_user, "view_hr"):
        raise HTTPException(status_code=403, detail="Yetkiniz yok (view_hr gerekir).")
    assigned = _user_assigned_department(current_user)
    if assigned:
        staff_dept = (staff.get("department") or "").strip()
        if staff_dept and staff_dept.lower() != assigned.strip().lower():
            raise HTTPException(
                status_code=403,
                detail="Departman dışı personel kaydına erişim engellendi.",
            )


async def _audit(
    user: User,
    action: str,
    entity_type: str,
    entity_id: str,
    details: str,
    *,
    before: dict | None = None,
    after: dict | None = None,
    severity: str = "info",
) -> None:
    """Best-effort audit logger — exception'lar HR mutation akışını bozmamalı."""
    try:
        await log_audit_event(
            tenant_id=getattr(user, "tenant_id", None) or "",
            user_id=str(getattr(user, "id", "") or ""),
            action=action,
            entity_type=entity_type,
            entity_id=str(entity_id) if entity_id else "",
            details=details,
            before_value=before,
            after_value=after,
            db=db,
            severity=severity,
        )
    except Exception:
        # Audit yazımı best-effort; mutation akışını bloklamaz.
        pass

router = APIRouter(prefix="/api", tags=["hr-operations"])

# ============= TR Labor Defaults =============
TR_DEFAULT_HOURLY_RATE = 140.0   # 2026 asgari ücret yaklaşık saatlik brüt
TR_DEFAULT_MONTHLY_HOURS = 195   # 45 sa/hafta × 4.33 hafta
TR_DEFAULT_OVERTIME_MULTIPLIER = 1.5  # %50 zam (İş K. m.41)
TR_CURRENCY = "TRY"

# Türkiye saat dilimi (UTC+3, sabit — yaz saati uygulaması yok)
TR_TZ = timezone(timedelta(hours=3))


def _today_local() -> date:
    """Türkiye TZ'ne göre bugünün tarihi — UTC vs local skew'unu engeller."""
    return datetime.now(TR_TZ).date()


# ============= HR COMPLETE SUITE =============

async def _get_staff_map(tenant_id: str):
    staff = await db.staff_members.find({'tenant_id': tenant_id}, {'_id': 0}).to_list(500)
    return {member['id']: member for member in staff}


async def _verify_staff_in_tenant(staff_id: str, tenant_id: str) -> dict | None:
    """
    Bug DAK round-7: Cross-tenant clock-in açığı.
    staff_id'nin gerçekten bu tenant'a ait olduğunu doğrular.
    staff_members'da bulunamazsa users tablosundan (derived staff) kontrol eder.
    """
    if not staff_id:
        return None
    sm = await db.staff_members.find_one(
        {'id': staff_id, 'tenant_id': tenant_id},
        {'_id': 0}
    )
    if sm:
        return sm
    # users tablosundan türetilmiş staff (HRv2 — derived_from=users)
    user = await db.users.find_one(
        {'id': staff_id, 'tenant_id': tenant_id, 'is_active': True},
        {'_id': 0, 'id': 1, 'name': 1, 'role': 1, 'email': 1}
    )
    if user:
        return {
            'id': user['id'],
            'tenant_id': tenant_id,
            'name': user.get('name') or 'Personel',
            'department': user.get('role') or 'staff',
            'derived_from': 'users',
        }
    return None


# ============= Pydantic Models =============

class ClockInRequest(BaseModel):
    staff_id: str = Field(..., min_length=1, max_length=128)


class LeaveRequestPayload(BaseModel):
    staff_id: str = Field(..., min_length=1, max_length=128)
    staff_name: str | None = Field(None, max_length=200)
    leave_type: Literal['annual', 'sick', 'maternity', 'paternity', 'unpaid', 'bereavement', 'excused'] = 'annual'
    start_date: str = Field(..., description="ISO date YYYY-MM-DD")
    end_date: str = Field(..., description="ISO date YYYY-MM-DD")
    reason: str | None = Field(None, max_length=500)
    total_days: int | None = Field(None, ge=0, le=365)

    @field_validator('start_date', 'end_date')
    @classmethod
    def _valid_iso(cls, v: str) -> str:
        try:
            datetime.fromisoformat(v).date()
        except Exception as exc:
            raise ValueError("Geçersiz tarih formatı (YYYY-MM-DD)") from exc
        return v


class LeaveDecision(BaseModel):
    # Task #263: 2-aşamalı onay state machine.
    #   pending → dept_approve → dept_approved → approve → approved
    #   pending / dept_approved → reject → rejected (note ZORUNLU)
    decision: Literal['dept_approve', 'approve', 'reject']
    note: str | None = Field(None, max_length=500)


class PayrollFinalizePayload(BaseModel):
    month: str = Field(..., pattern=r'^\d{4}-\d{2}$')


# ============= Payroll v2 (Task #264) =============
# `payroll_runs` koleksiyonu: dry-run preview + draft + locked lifecycle.
# `payroll_revisions` koleksiyonu: locked sonrası düzeltme audit zinciri.

PAYROLL_EXTRA_KINDS = (
    'bonus',          # prim
    'meal',           # yemek
    'transport',      # yol
    'advance',        # avans (kesinti yönü)
    'deduction',      # diğer kesinti
)


class PayrollExtraLine(BaseModel):
    """Personel başına manuel ekleme/kesinti satırı.

    `kind`:
      - bonus    → brüte eklenir (prim)
      - meal     → brüte eklenir (yemek yardımı, vergi-dışı varsayım yok)
      - transport→ brüte eklenir (yol yardımı)
      - advance  → nete kesinti (avans mahsup)
      - deduction→ nete kesinti (diğer)
    `amount` her zaman pozitif TRY; yönü `kind` belirler.
    """
    staff_id: str = Field(..., min_length=1, max_length=128)
    kind: Literal['bonus', 'meal', 'transport', 'advance', 'deduction']
    amount: float = Field(..., ge=0, le=1_000_000)
    note: str | None = Field(None, max_length=200)


class PayrollSavePayload(BaseModel):
    """`POST /hr/payroll/{month}/save` — draft run upsert.

    Idempotent: aynı (tenant, period_month) için açık draft varsa üzerine yazar.
    `extras`: per-staff manuel kalemler (avans/prim/yemek/yol/kesinti).
    """
    extras: list[PayrollExtraLine] = Field(default_factory=list, max_length=2000)
    note: str | None = Field(None, max_length=500)


class PayrollRevisionPayload(BaseModel):
    """`POST /hr/payroll/{run_id}/revisions` — locked run düzeltmesi.

    Yeni bir draft run açar (parent_run_id link); locked run değişmez.
    `reason` zorunlu (KVKK + iş hukuku audit).
    """
    reason: str = Field(..., min_length=3, max_length=500)
    extras: list[PayrollExtraLine] = Field(default_factory=list, max_length=2000)


class PerformanceReviewPayload(BaseModel):
    staff_id: str = Field(..., min_length=1, max_length=128)
    reviewer_name: str | None = Field(None, max_length=200)
    period: str | None = Field(None, max_length=64)
    overall_score: float = Field(..., ge=0, le=10)
    goals: str | None = Field(None, max_length=2000)
    strengths: str | None = Field(None, max_length=2000)
    improvement_areas: str | None = Field(None, max_length=2000)
    template_id: str | None = Field(None, max_length=128)
    competency_scores: dict[str, float] | None = None  # {"Müşteri ilişkileri": 8.5, ...}


class JobPostingPayload(BaseModel):
    """Personel ihtiyaç talebi / iş ilanı formu.

    Bu model iki amaca hizmet eder:
    - `request_type='internal_request'` (varsayılan): Departman müdürü
      personel ihtiyacı bildirir, HR yöneticisi onaylayınca aday alımına açılır.
    - `request_type='public_posting'`: Direkt kabul edilmiş açık pozisyon
      (legacy davranış; ileride dış kanal entegrasyonu eklenirse kullanılır).
    """
    title: str = Field(..., min_length=1, max_length=200)
    department: str = Field(..., max_length=100)
    description: str | None = Field(None, max_length=5000)
    employment_type: Literal['full_time', 'part_time', 'seasonal', 'intern', 'contract'] = 'full_time'
    location: str | None = Field(None, max_length=200)
    salary_range: str | None = Field(None, max_length=100)
    # Personel talebi alanları
    request_type: Literal['internal_request', 'public_posting'] = 'internal_request'
    headcount_needed: int = Field(1, ge=1, le=50)
    urgency: Literal['low', 'normal', 'high', 'critical'] = 'normal'
    justification: str | None = Field(None, max_length=2000)
    needed_by: str | None = Field(None, pattern=r'^\d{4}-\d{2}-\d{2}$')


class JobDecisionPayload(BaseModel):
    note: str | None = Field(None, max_length=500)


# ============= Attendance =============

@router.post("/hr/clock-in")
async def clock_in(payload: ClockInRequest, current_user: User = Depends(get_current_user)):
    """Personel giriş kaydı — tenant scoped, TR-TZ tarih, validated staff_id."""
    staff = await _verify_staff_in_tenant(payload.staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı veya bu tenant'a ait değil")

    today_iso = _today_local().isoformat()
    # Aynı gün içinde açık clock-in varsa, çift kayıt önle
    existing = await db.attendance_records.find_one({
        'tenant_id': current_user.tenant_id,
        'staff_id': payload.staff_id,
        'date': today_iso,
        'clock_out': None,
    })
    if existing:
        return {'success': False, 'message': 'Açık giriş kaydı mevcut', 'time': existing.get('clock_in')}

    record = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'staff_id': payload.staff_id,
        'staff_name': staff.get('name'),
        'date': today_iso,
        'clock_in': datetime.now(UTC).isoformat(),
        'clock_out': None,
        'status': 'present',
        'created_at': datetime.now(UTC).isoformat()
    }
    await db.attendance_records.insert_one(record)
    return {'success': True, 'message': 'Giriş kaydedildi', 'time': record['clock_in']}


@router.post("/hr/clock-out")
async def clock_out(payload: ClockInRequest, current_user: User = Depends(get_current_user)):
    # v109 Bug DAK round-6 (T08 P1): tenant_id scoped on both find and update.
    today_iso = _today_local().isoformat()
    record = await db.attendance_records.find_one({
        'tenant_id': current_user.tenant_id,
        'staff_id': payload.staff_id,
        'date': today_iso,
        'clock_out': None,
    })
    if record:
        clock_out_time = datetime.now(UTC)
        clock_in_time = datetime.fromisoformat(record['clock_in'].replace('Z', '+00:00'))
        hours = (clock_out_time - clock_in_time).total_seconds() / 3600
        await db.attendance_records.update_one(
            {'id': record['id'], 'tenant_id': current_user.tenant_id},
            {'$set': {'clock_out': clock_out_time.isoformat(), 'total_hours': round(hours, 2)}}
        )
        return {'success': True, 'hours_worked': round(hours, 2)}
    return {'success': False, 'message': 'Açık giriş kaydı bulunamadı'}


def _parse_date_range(start: str | None, end: str | None, days: int = 7):
    today = _today_local()
    start_date = datetime.fromisoformat(start).date() if start else today - timedelta(days=days)
    end_date = datetime.fromisoformat(end).date() if end else today
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    return start_date, end_date


@router.get("/hr/attendance/records")
async def get_attendance_records(
    start_date: str | None = None,
    end_date: str | None = None,
    staff_id: str | None = None,
    limit: int = 500,
    current_user: User = Depends(get_current_user)
):
    start_dt, end_dt = _parse_date_range(start_date, end_date, days=7)
    query: dict[str, Any] = {
        'tenant_id': current_user.tenant_id,
        'date': {'$gte': start_dt.isoformat(), '$lte': end_dt.isoformat()}
    }
    if staff_id:
        query['staff_id'] = staff_id
    records = await db.attendance_records.find(query, {'_id': 0}).sort('clock_in', -1).limit(limit).to_list(limit)
    staff_map = await _get_staff_map(current_user.tenant_id)
    for record in records:
        staff = staff_map.get(record['staff_id'])
        if staff:
            record['staff_name'] = staff.get('name') or record.get('staff_name')
            record['department'] = staff.get('department')
            record['position'] = staff.get('position')
    return {
        'records': records,
        'total': len(records),
        'range': {'start': start_dt.isoformat(), 'end': end_dt.isoformat()}
    }


@router.get("/hr/attendance/summary")
async def get_attendance_summary(
    start_date: str | None = None,
    end_date: str | None = None,
    current_user: User = Depends(get_current_user)
):
    start_dt, end_dt = _parse_date_range(start_date, end_date, days=30)
    query = {
        'tenant_id': current_user.tenant_id,
        'date': {'$gte': start_dt.isoformat(), '$lte': end_dt.isoformat()}
    }
    records = await db.attendance_records.find(query, {'_id': 0}).to_list(2000)
    staff_map = await _get_staff_map(current_user.tenant_id)

    summary: dict[str, Any] = {}
    for record in records:
        staff_id = record['staff_id']
        summary.setdefault(staff_id, {
            'staff_id': staff_id,
            'total_hours': 0,
            'days_present': 0,
            'overtime_hours': 0,
            'late_count': 0
        })
        summary[staff_id]['total_hours'] += record.get('total_hours', 0)
        summary[staff_id]['days_present'] += 1

    for staff_id, data in summary.items():
        staff = staff_map.get(staff_id, {})
        data['staff_name'] = staff.get('name', staff_id)
        data['department'] = staff.get('department', 'unknown')
        data['position'] = staff.get('position', 'Staff')
        overtime_threshold = staff.get('monthly_hours', TR_DEFAULT_MONTHLY_HOURS)
        if data['total_hours'] > overtime_threshold:
            data['overtime_hours'] = round(data['total_hours'] - overtime_threshold, 2)
        data['average_daily_hours'] = round(
            data['total_hours'] / data['days_present'], 2
        ) if data['days_present'] else 0

    summary_list = sorted(summary.values(), key=lambda x: x['staff_name'])
    total_hours = round(sum(item['total_hours'] for item in summary_list), 2)

    # Aktif personel sayısı: hem staff_members'da olanlar hem users
    # tablosundan türeyenler (HRv2 derived staff). `_get_staff_map` yalnızca
    # staff_members'ı döndürdüğü için users tarafını ayrıca sayıyoruz.
    explicit_count = await db.staff_members.count_documents(
        {'tenant_id': current_user.tenant_id, 'active': True}
    )
    derived_roles = ['housekeeping', 'front_desk', 'supervisor',
                     'finance', 'sales', 'admin']
    derived_count = await db.users.count_documents({
        'tenant_id': current_user.tenant_id,
        'is_active': True,
        'role': {'$in': derived_roles},
    })
    total_active_staff = explicit_count + derived_count

    # Devam kayıtlı personel başına ortalamayı koru, bilgi amaçlı
    # genel ortalamayı (gerçek headcount'a göre) ayrı alanla sunalım.
    avg_hours = round(total_hours / len(summary_list), 2) if summary_list else 0
    avg_hours_per_active_staff = (
        round(total_hours / total_active_staff, 2) if total_active_staff else 0
    )

    return {
        'summary': summary_list,
        'range': {'start': start_dt.isoformat(), 'end': end_dt.isoformat()},
        'metrics': {
            'staff_count': len(summary_list),
            'total_active_staff': total_active_staff,
            'total_hours': total_hours,
            'avg_hours_per_staff': avg_hours,
            'avg_hours_per_active_staff': avg_hours_per_active_staff,
        }
    }


# ============= Leave =============

@router.post("/hr/leave-request")
async def create_leave_request(
    payload: LeaveRequestPayload,
    current_user: User = Depends(get_current_user),
):
    """
    İzin talebi oluştur. Bug DAK round-7: Önceden permission gate yoktu ve
    Pydantic doğrulaması yapılmıyordu — kullanıcı başkası adına talep yaratabiliyordu.
    Şimdi: staff_id mutlaka aynı tenant'ta olmalı; HR yönetici yetkisi yoksa
    kullanıcı SADECE kendi user.id'sine eşleşen staff_id ile talep oluşturabilir.
    """
    staff = await _verify_staff_in_tenant(payload.staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı veya bu tenant'a ait değil")

    # Self-leave kontrolü: HR yönetici izni olmayan kullanıcı sadece kendi adına talep edebilir
    is_self = (payload.staff_id == getattr(current_user, 'id', None))
    if not is_self:
        # HR yönetici yetkisi gereken roller: admin/supervisor/finance
        manager_roles = {'admin', 'supervisor', 'finance'}
        if (getattr(current_user, 'role', None) or '').lower() not in manager_roles:
            raise HTTPException(
                status_code=403,
                detail="Başka personel adına izin talebi oluşturma yetkiniz yok"
            )

    start = datetime.fromisoformat(payload.start_date).date()
    end = datetime.fromisoformat(payload.end_date).date()
    if end < start:
        raise HTTPException(status_code=400, detail="Bitiş tarihi başlangıçtan önce olamaz")
    total_days = payload.total_days if payload.total_days is not None else (end - start).days + 1

    leave = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'staff_id': payload.staff_id,
        'staff_name': payload.staff_name or staff.get('name') or 'Personel',
        'leave_type': payload.leave_type,
        'start_date': payload.start_date,
        'end_date': payload.end_date,
        'total_days': total_days,
        'reason': payload.reason,
        'status': 'pending',
        'requested_by': getattr(current_user, 'id', None),
        'created_at': datetime.now(UTC).isoformat(),
    }
    await db.leave_requests.insert_one(leave)

    # HR yöneticilerine bildirim — onay bekleyen talep
    await _notify_hr_managers(
        current_user.tenant_id,
        kind='leave_request',
        title='Yeni izin talebi',
        body=(
            f"{leave['staff_name']} • {payload.leave_type} • "
            f"{payload.start_date} → {payload.end_date} ({total_days} gün)"
        ),
        link=f'/hr?tab=leave&id={leave["id"]}',
        ref_id=leave['id'],
    )

    return {'success': True, 'leave_id': leave['id'], 'total_days': total_days, 'status': 'pending'}


@router.get("/hr/leave-requests")
async def list_leave_requests(
    status: str | None = None,
    staff_id: str | None = None,
    current_user: User = Depends(get_current_user),
):
    query: dict[str, Any] = {'tenant_id': current_user.tenant_id}
    if status:
        query['status'] = status
    if staff_id:
        query['staff_id'] = staff_id
    items = await db.leave_requests.find(query, {'_id': 0}).sort('created_at', -1).to_list(500)
    counts = {'pending': 0, 'approved': 0, 'rejected': 0}
    for item in items:
        counts[item.get('status', 'pending')] = counts.get(item.get('status', 'pending'), 0) + 1
    return {'items': items, 'total': len(items), 'counts': counts}


async def _apply_leave_to_shifts(
    tenant_id: str, leave: dict,
) -> int:
    """Final onaylı izin → izin gününe `shift_schedules` üzerinde
    status='on_leave' satırı upsert. Lock YARATMAZ (izin kapsayan gün
    çakışma kontratının dışındadır; overlap guard sadece aktif vardiya
    için anlamlıdır). Idempotent: aynı (staff,date,leave_id) için
    yeniden çağrılırsa duplicate açmaz."""
    start = date.fromisoformat(leave['start_date'])
    end = date.fromisoformat(leave['end_date'])
    cur = start
    written = 0
    leave_type = leave.get('leave_type')
    leave_notes = f"İzin: {leave_type}" if leave_type else "İzin"
    while cur <= end:
        d_iso = cur.isoformat()
        # Task #263: Mevcut shift'i `on_leave` olarak işaretle (eski davranış
        # yeni satır eklerdi → planner çift kayıt görüyordu). Mevcut shift
        # varsa status='on_leave' + leave_id setle; lock konvansiyonel olarak
        # release edilmez çünkü on_leave kayıt çakışma kontratının dışında.
        existing = await db.shift_schedules.find_one({
            'tenant_id': tenant_id,
            'staff_id': leave['staff_id'],
            'shift_date': d_iso,
            'status': {'$ne': 'cancelled'},
        })
        if existing is not None:
            await db.shift_schedules.update_one(
                {'id': existing['id'], 'tenant_id': tenant_id},
                {'$set': {
                    'status': 'on_leave',
                    'leave_id': leave['id'],
                    'leave_type': leave_type,
                    'notes': leave_notes,
                    'on_leave_applied_at': datetime.now(UTC).isoformat(),
                    'on_leave_applied_via': 'leave_approval',
                }},
            )
            if existing.get('status') != 'on_leave' or existing.get('leave_id') != leave['id']:
                written += 1
        else:
            # Mevcut shift yoksa yeni on_leave kayıt aç (idempotent: aynı leave_id
            # için tekrar çağrılırsa update_one match eder, insert YOK).
            res = await db.shift_schedules.update_one(
                {
                    'tenant_id': tenant_id,
                    'staff_id': leave['staff_id'],
                    'shift_date': d_iso,
                    'status': 'on_leave',
                    'leave_id': leave['id'],
                },
                {
                    '$setOnInsert': {
                        'id': str(uuid.uuid4()),
                        'tenant_id': tenant_id,
                        'staff_id': leave['staff_id'],
                        'staff_name': leave.get('staff_name'),
                        'shift_date': d_iso,
                        'shift_type': 'off',
                        'start_time': '00:00',
                        'end_time': '23:59',
                        'crosses_midnight': False,
                        'end_date': d_iso,
                        'status': 'on_leave',
                        'leave_id': leave['id'],
                        'leave_type': leave_type,
                        'notes': leave_notes,
                        'created_at': datetime.now(UTC).isoformat(),
                        'created_via': 'leave_approval',
                    },
                },
                upsert=True,
            )
            if res.upserted_id is not None:
                written += 1
        cur += timedelta(days=1)
    return written


@router.post("/hr/leave-request/{leave_id}/decision")
async def decide_leave_request(
    leave_id: str,
    payload: LeaveDecision,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),  # HR yönetici yetkisi
):
    leave = await db.leave_requests.find_one({
        'tenant_id': current_user.tenant_id, 'id': leave_id
    })
    if not leave:
        raise HTTPException(status_code=404, detail="İzin talebi bulunamadı")

    current_status = leave.get('status', 'pending')
    # Task #263: 2-aşamalı state machine.
    if payload.decision == 'reject':
        if current_status not in ('pending', 'dept_approved'):
            raise HTTPException(
                status_code=400,
                detail=f"Bu talep zaten karara bağlanmış (status={current_status})",
            )
        if not (payload.note and payload.note.strip()):
            raise HTTPException(
                status_code=400,
                detail="Red için gerekçe (note) zorunludur",
            )
        new_status = 'rejected'
    elif payload.decision == 'dept_approve':
        if current_status != 'pending':
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Departman onayı sadece pending durumdan verilebilir "
                    f"(mevcut: {current_status})"
                ),
            )
        new_status = 'dept_approved'
    else:  # 'approve' (final HR) — Task #263: strict chain enforcement
        if current_status != 'dept_approved':
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Final onay sadece departman onayından sonra verilebilir "
                    f"(mevcut: {current_status}). Önce 'dept_approve' aşaması gerekli."
                ),
            )
        new_status = 'approved'

    update_set: dict[str, Any] = {
        'status': new_status,
        'decision_note': payload.note,
        'decided_at': datetime.now(UTC).isoformat(),
    }
    decision_history = leave.get('decision_history') or []
    decision_history.append({
        'stage': payload.decision,
        'status': new_status,
        'by': getattr(current_user, 'id', None),
        'at': datetime.now(UTC).isoformat(),
        'note': payload.note,
    })
    update_set['decision_history'] = decision_history
    if new_status == 'dept_approved':
        update_set['dept_approved_by'] = getattr(current_user, 'id', None)
        update_set['dept_approved_at'] = datetime.now(UTC).isoformat()
    elif new_status == 'approved':
        update_set['decided_by'] = getattr(current_user, 'id', None)
        update_set['approved_by'] = getattr(current_user, 'id', None)
        update_set['approved_at'] = datetime.now(UTC).isoformat()
    else:
        update_set['decided_by'] = getattr(current_user, 'id', None)

    await db.leave_requests.update_one(
        {'tenant_id': current_user.tenant_id, 'id': leave_id},
        {'$set': update_set},
    )

    on_leave_written = 0
    if new_status == 'approved':
        # Final onay: shift_schedules üstüne on_leave upsert.
        on_leave_written = await _apply_leave_to_shifts(
            current_user.tenant_id, {**leave, **update_set, 'id': leave_id},
        )

    # Talep sahibine bildirim — kararı duyur
    requester_id = leave.get('requested_by') or leave.get('staff_id')
    if requester_id:
        notif_title = {
            'approved': 'İzin talebiniz onaylandı',
            'dept_approved': 'İzin talebiniz departman onayı aldı (HR onayı bekleniyor)',
            'rejected': 'İzin talebiniz reddedildi',
        }[new_status]
        await _notify_user(
            current_user.tenant_id,
            user_id=requester_id,
            kind=f'leave_{new_status}',
            title=notif_title,
            body=(
                f"{leave.get('start_date')} → {leave.get('end_date')} • "
                f"{leave.get('total_days', 0)} gün"
                + (f" • Not: {payload.note}" if payload.note else '')
            ),
            link=f'/hr?tab=leave&id={leave_id}',
            ref_id=leave_id,
        )
    return {
        'success': True,
        'status': new_status,
        'on_leave_shifts_created': on_leave_written,
    }


@router.get("/hr/leave/calendar")
async def leave_calendar(
    month: str = Query(..., pattern=r'^\d{4}-\d{2}$'),
    department: str | None = None,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """Ay görünümü: per-staff per-day matrix. Sadece status IN
    (dept_approved, approved) izinler döner. Frontend takvim renkler."""
    yyyy, mm = month.split('-')
    start = date(int(yyyy), int(mm), 1)
    next_month = start.replace(day=28) + timedelta(days=4)
    end = next_month - timedelta(days=next_month.day)

    query: dict[str, Any] = {
        'tenant_id': current_user.tenant_id,
        'status': {'$in': ['dept_approved', 'approved']},
        'start_date': {'$lte': end.isoformat()},
        'end_date': {'$gte': start.isoformat()},
    }
    leaves = await db.leave_requests.find(
        query, {'_id': 0, 'id': 1, 'staff_id': 1, 'staff_name': 1,
                'leave_type': 1, 'status': 1,
                'start_date': 1, 'end_date': 1},
    ).to_list(2000)

    # Department filter (staff lookup)
    if department:
        staff_in_dept: set[str] = set()
        async for s in db.staff_members.find(
            {'tenant_id': current_user.tenant_id, 'department': department},
            {'_id': 0, 'id': 1},
        ):
            staff_in_dept.add(s['id'])
        leaves = [l for l in leaves if l['staff_id'] in staff_in_dept]

    # Per-day occupancy: {day_iso: [staff_id, ...]}
    by_day: dict[str, list[dict]] = {}
    cur = start
    while cur <= end:
        by_day[cur.isoformat()] = []
        cur += timedelta(days=1)
    for l in leaves:
        ls = max(date.fromisoformat(l['start_date']), start)
        le = min(date.fromisoformat(l['end_date']), end)
        c = ls
        while c <= le:
            by_day[c.isoformat()].append({
                'leave_id': l['id'],
                'staff_id': l['staff_id'],
                'staff_name': l.get('staff_name'),
                'leave_type': l.get('leave_type'),
                'status': l.get('status'),
            })
            c += timedelta(days=1)
    return {
        'month': month,
        'range': {'start': start.isoformat(), 'end': end.isoformat()},
        'days': by_day,
        'leaves': leaves,
        'total': len(leaves),
    }


# ============= Notification helpers (HR akışı için) =============

def _build_notification_doc(
    tenant_id: str, *, user_id: str | None, kind: str, title: str, body: str,
    link: str | None, ref_id: str | None,
) -> dict:
    """notifications koleksiyon şemasıyla uyumlu doc üretir.

    notification_router.list_notifications {type, title, message, action_url}
    alanlarını okur — bu helper bu adlandırmayı kullanır; ek olarak
    'kind' ve 'ref_id' metadata içine konur (debugging / future filtering için).
    """
    now_iso = datetime.now(UTC).isoformat()
    return {
        'id': str(uuid.uuid4()),
        'tenant_id': tenant_id,
        'user_id': user_id,
        'type': kind,                # list endpoint 'type' okuyor
        'title': title,
        'message': body,             # list endpoint 'message' okuyor
        'body': body,                # backward-compat (bazı yerler 'body' kullanıyor)
        'priority': 'normal',
        'action_url': link,          # list endpoint 'action_url' okuyor
        'metadata': {'kind': kind, 'ref_id': ref_id} if ref_id else {'kind': kind},
        'read': False,
        'created_at': now_iso,
        'sent_at': now_iso,
    }


async def _notify_hr_managers(
    tenant_id: str, *, kind: str, title: str, body: str,
    link: str | None = None, ref_id: str | None = None,
):
    """HR yönetici rollerindeki tüm aktif kullanıcılara in-app bildirim gönder."""
    cursor = db.users.find({
        'tenant_id': tenant_id,
        'is_active': True,
        'role': {'$in': ['admin', 'supervisor', 'finance']},
    }, {'_id': 0, 'id': 1})
    user_ids = [u['id'] async for u in cursor if u.get('id')]
    if not user_ids:
        return
    docs = [
        _build_notification_doc(
            tenant_id, user_id=uid, kind=kind, title=title, body=body,
            link=link, ref_id=ref_id,
        )
        for uid in user_ids
    ]
    try:
        await db.notifications.insert_many(docs)
    except Exception as exc:  # pragma: no cover
        logger.warning("HR bildirimi yazılamadı: %s", exc)


async def _notify_user(
    tenant_id: str, *, user_id: str, kind: str, title: str, body: str,
    link: str | None = None, ref_id: str | None = None,
):
    """Tek kullanıcıya in-app bildirim gönder."""
    try:
        await db.notifications.insert_one(_build_notification_doc(
            tenant_id, user_id=user_id, kind=kind, title=title, body=body,
            link=link, ref_id=ref_id,
        ))
    except Exception as exc:  # pragma: no cover
        logger.warning("HR bildirimi yazılamadı: %s", exc)


# ============= Payroll =============

# NOTE: The dynamic `GET /hr/payroll/{month}` is intentionally declared
# AFTER the static export/finalize routes so FastAPI matches static first.

def _compute_payroll_for_month(
    records: list[dict],
    staff_map: dict[str, dict],
    month: str,
) -> list[dict]:
    payroll_rows: dict[str, float] = {}
    for record in records:
        staff_id = record['staff_id']
        payroll_rows.setdefault(staff_id, 0.0)
        payroll_rows[staff_id] += record.get('total_hours', 0)

    payroll: list[dict] = []
    for staff_id, total_hours in payroll_rows.items():
        staff = staff_map.get(staff_id, {})
        hourly_rate = float(staff.get('hourly_rate') or TR_DEFAULT_HOURLY_RATE)
        monthly_hours = float(staff.get('monthly_hours') or TR_DEFAULT_MONTHLY_HOURS)
        overtime_rate = float(staff.get('overtime_rate') or hourly_rate * TR_DEFAULT_OVERTIME_MULTIPLIER)
        overtime_hours = max(0.0, total_hours - monthly_hours)
        base_hours = total_hours - overtime_hours
        gross_pay = (base_hours * hourly_rate) + (overtime_hours * overtime_rate)

        # Basitleştirilmiş TR kesinti modeli (yaklaşık):
        # SGK işçi payı %14, işsizlik %1, gelir vergisi %15 (matrah - SGK), damga %0.759
        sgk_employee = round(gross_pay * 0.14, 2)
        unemployment = round(gross_pay * 0.01, 2)
        income_tax_base = max(0.0, gross_pay - sgk_employee - unemployment)
        income_tax = round(income_tax_base * 0.15, 2)
        stamp_tax = round(gross_pay * 0.00759, 2)
        total_deductions = round(sgk_employee + unemployment + income_tax + stamp_tax, 2)
        net_pay = round(gross_pay - total_deductions, 2)

        payroll.append({
            'staff_id': staff_id,
            'staff_name': staff.get('name', staff_id),
            'department': staff.get('department', 'unknown'),
            'period_month': month,
            'total_hours': round(total_hours, 2),
            'overtime_hours': round(overtime_hours, 2),
            'hourly_rate': round(hourly_rate, 2),
            'overtime_rate': round(overtime_rate, 2),
            'gross_pay': round(gross_pay, 2),
            'sgk_employee': sgk_employee,
            'unemployment': unemployment,
            'income_tax': income_tax,
            'stamp_tax': stamp_tax,
            'total_deductions': total_deductions,
            'net_salary': net_pay,
            'currency': TR_CURRENCY,
        })
    return payroll


async def _build_payroll(month: str | None, tenant_id: str):
    today = _today_local()
    if month:
        start_dt = datetime.fromisoformat(f"{month}-01").date()
    else:
        start_dt = date(today.year, today.month, 1)
    next_month = start_dt.replace(day=28) + timedelta(days=4)
    end_dt = next_month - timedelta(days=next_month.day)

    query = {
        'tenant_id': tenant_id,
        'date': {'$gte': start_dt.isoformat(), '$lte': end_dt.isoformat()}
    }
    records = await db.attendance_records.find(query, {'_id': 0}).to_list(5000)
    staff_map = await _get_staff_map(tenant_id)
    period_month = start_dt.strftime('%Y-%m')
    payroll = _compute_payroll_for_month(records, staff_map, period_month)
    return period_month, payroll


@router.get("/hr/payroll/export")
async def export_payroll(
    month: str | None = None,
    format: str = 'json',
    current_user: User = Depends(get_current_user),
    # Bug DAK round-7: Maaş PII'sı için yetki gate'i (KVKK + iş hukuku).
    _perm=Depends(require_op("view_hr")),
):
    period_month, payroll = await _build_payroll(month, current_user.tenant_id)

    response: dict[str, Any] = {
        'month': period_month,
        'payroll': payroll,
        'staff_count': len(payroll),
        'total_gross_pay': round(sum(row['gross_pay'] for row in payroll), 2),
        'total_net_pay': round(sum(row['net_salary'] for row in payroll), 2),
        'currency': TR_CURRENCY,
    }
    await _audit(
        current_user, 'payroll.export', 'payroll', period_month,
        f"Bordro export (format={format}, staff={len(payroll)})",
        severity='info',
    )

    # NOT: format=csv için inline base64 desteği kaldırıldı. UI artık
    # /hr/payroll/export/csv (StreamingResponse) endpoint'ini kullanıyor —
    # data: URL'in 2MB tarayıcı limitini aşan büyük tenant'larda doğru çözüm.
    return response


@router.get("/hr/payroll/export/csv")
async def export_payroll_csv_stream(
    month: str | None = Query(None),
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """Streaming CSV download — büyük dosyalar için data: URL limiti yok."""
    period_month, payroll = await _build_payroll(month, current_user.tenant_id)

    import csv

    from core.csv_safe import safe_dict_writerow
    buf = io.StringIO()
    fields = [
        'staff_id', 'staff_name', 'department', 'period_month',
        'total_hours', 'overtime_hours', 'hourly_rate', 'overtime_rate',
        'gross_pay', 'sgk_employee', 'unemployment', 'income_tax',
        'stamp_tax', 'total_deductions', 'net_salary', 'currency'
    ]
    writer = csv.DictWriter(buf, fieldnames=fields)
    writer.writeheader()
    for row in payroll:
        safe_dict_writerow(writer, row)
    csv_text = buf.getvalue()

    await _audit(
        current_user, 'payroll.export_csv', 'payroll', period_month,
        f"Bordro CSV stream (satır={len(payroll)})",
        severity='info',
    )
    return StreamingResponse(
        iter([csv_text]),
        media_type='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': f'attachment; filename="payroll_{period_month}.csv"'
        }
    )


@router.post("/hr/payroll/finalize", deprecated=True)
async def finalize_payroll(
    payload: PayrollFinalizePayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    """**DEPRECATED (Task #264).** Geri uyumluluk için korunur.

    Eski client'lar yeni akışa (`/save` → `/finalize`) geçmelidir; bu
    endpoint hâlâ önceki sözleşmeyi (`payroll_records` overwrite) verir
    ama OpenAPI'da `deprecated=true` görünür. Yeni gelişimde
    `POST /hr/payroll/{month}/save` + `POST /hr/payroll/{run_id}/finalize`
    kullanılmalıdır (immutable lifecycle + audit + revision desteği).
    """
    period_month, payroll = await _build_payroll(payload.month, current_user.tenant_id)
    if not payroll:
        return {'success': False, 'message': 'Bu ay için attendance kaydı yok', 'count': 0}

    await db.payroll_records.delete_many({
        'tenant_id': current_user.tenant_id,
        'period_month': period_month,
    })
    for row in payroll:
        row['id'] = str(uuid.uuid4())
        row['tenant_id'] = current_user.tenant_id
        row['finalized_by'] = getattr(current_user, 'id', None)
        row['finalized_at'] = datetime.now(UTC).isoformat()
    await db.payroll_records.insert_many(payroll)
    await _audit(
        current_user, 'payroll.legacy_finalize', 'payroll_records', period_month,
        f"Legacy finalize (deprecated) — ay={period_month}, satır={len(payroll)}",
        severity='warning',
    )
    return {
        'success': True,
        'period_month': period_month,
        'count': len(payroll),
        'total_gross': round(sum(r['gross_pay'] for r in payroll), 2),
        'total_net': round(sum(r['net_salary'] for r in payroll), 2),
        'currency': TR_CURRENCY,
    }


# ---------- Payroll v2 helpers (Task #264) ----------

def _payroll_apply_extras_and_overtime(
    base_rows: list[dict],
    extras: list[dict],
    overtime_by_staff: dict[str, dict],
    leaves_by_staff: dict[str, dict] | None = None,
) -> list[dict]:
    """Base payroll satırlarına (saat × ücret) ekstra kalemleri ve onaylı
    mesai (overtime_requests) saatlerini uygular; her satıra `line_items`
    listesi ve yeniden hesaplanmış brüt/net döner.

    Hesap doktrin:
      1. base satır brüt'ünü baz alır (saat × ücret + attendance-derived mesai)
      2. + onaylı mesai (overtime_requests) → ek brüt (hours × hourly × 1.5)
      3. + bonus / meal / transport → ek brüt
      4. = yeni brüt → standart TR kesinti modeli (SGK/işsizlik/vergi/damga)
      5. − advance / deduction → nete kesinti (post-tax)
    """
    extras_by_staff: dict[str, list[dict]] = {}
    for ex in extras:
        extras_by_staff.setdefault(ex['staff_id'], []).append(ex)

    out: list[dict] = []
    for row in base_rows:
        sid = row['staff_id']
        hourly_rate = float(row.get('hourly_rate') or TR_DEFAULT_HOURLY_RATE)
        line_items: list[dict] = [
            {
                'kind': 'base',
                'label': 'Esas ücret (saat × tarife)',
                'hours': row.get('total_hours', 0) - row.get('overtime_hours', 0),
                'rate': hourly_rate,
                'amount': round(
                    (row.get('total_hours', 0) - row.get('overtime_hours', 0)) * hourly_rate,
                    2,
                ),
                'direction': 'earning',
            },
        ]
        if row.get('overtime_hours', 0) > 0:
            line_items.append({
                'kind': 'overtime_attendance',
                'label': 'Mesai (devam-türetilmiş, >195h)',
                'hours': row['overtime_hours'],
                'rate': row.get('overtime_rate', hourly_rate * TR_DEFAULT_OVERTIME_MULTIPLIER),
                'amount': round(
                    row['overtime_hours']
                    * (row.get('overtime_rate') or hourly_rate * TR_DEFAULT_OVERTIME_MULTIPLIER),
                    2,
                ),
                'direction': 'earning',
            })

        ot = overtime_by_staff.get(sid)
        if ot and ot.get('hours', 0) > 0:
            ot_amount = round(
                ot['hours'] * hourly_rate * TR_DEFAULT_OVERTIME_MULTIPLIER, 2,
            )
            line_items.append({
                'kind': 'overtime_approved',
                'label': f"Onaylı mesai ({ot['requests']} talep)",
                'hours': round(ot['hours'], 2),
                'rate': round(hourly_rate * TR_DEFAULT_OVERTIME_MULTIPLIER, 2),
                'amount': ot_amount,
                'direction': 'earning',
            })

        added_earnings = 0.0
        post_tax_deductions = 0.0
        for ex in extras_by_staff.get(sid, []):
            amt = round(float(ex['amount']), 2)
            kind = ex['kind']
            if kind in ('bonus', 'meal', 'transport'):
                added_earnings += amt
                line_items.append({
                    'kind': kind,
                    'label': {'bonus': 'Prim', 'meal': 'Yemek', 'transport': 'Yol'}[kind],
                    'amount': amt,
                    'direction': 'earning',
                    'note': ex.get('note'),
                })
            else:  # advance / deduction
                post_tax_deductions += amt
                line_items.append({
                    'kind': kind,
                    'label': {'advance': 'Avans mahsubu', 'deduction': 'Kesinti'}[kind],
                    'amount': amt,
                    'direction': 'deduction',
                    'note': ex.get('note'),
                })

        # İzin etkisi (Task #264 post-review): onaylı leave_requests'ten
        # ay'a düşen gün sayıları → ücretsiz izin günleri NET'ten 8h × hourly
        # düşülür (eksik gün), ücretli izin günleri SGK günü için raporlanır.
        lv = (leaves_by_staff or {}).get(sid) or {}
        paid_leave_days = int(lv.get('paid_days') or 0)
        unpaid_leave_days = int(lv.get('unpaid_days') or 0)
        if unpaid_leave_days > 0:
            unpaid_amount = round(unpaid_leave_days * 8 * hourly_rate, 2)
            post_tax_deductions += unpaid_amount
            line_items.append({
                'kind': 'leave_unpaid_deduction',
                'label': f"Ücretsiz izin kesintisi ({unpaid_leave_days} gün)",
                'days': unpaid_leave_days,
                'rate': hourly_rate,
                'amount': unpaid_amount,
                'direction': 'deduction',
            })
        if paid_leave_days > 0:
            line_items.append({
                'kind': 'leave_paid_info',
                'label': f"Ücretli izin (bilgi, {paid_leave_days} gün)",
                'days': paid_leave_days,
                'amount': 0.0,
                'direction': 'info',
            })

        ot_added = sum(
            li['amount'] for li in line_items if li['kind'] == 'overtime_approved'
        )
        new_gross = round(
            row['gross_pay'] + ot_added + added_earnings, 2,
        )

        sgk = round(new_gross * 0.14, 2)
        unemp = round(new_gross * 0.01, 2)
        income_tax = round(max(0.0, new_gross - sgk - unemp) * 0.15, 2)
        stamp = round(new_gross * 0.00759, 2)
        tax_total = round(sgk + unemp + income_tax + stamp, 2)
        net = round(new_gross - tax_total - post_tax_deductions, 2)

        new_row = dict(row)
        # SGK günü: 30 - ücretsiz izin günleri (Türk iş hukuku doktrin —
        # ücretli izin SGK gününü etkilemez, ücretsiz izin düşer).
        sgk_days = max(0, 30 - unpaid_leave_days)
        eksik_gun = unpaid_leave_days
        new_row.update({
            'gross_pay': new_gross,
            'sgk_employee': sgk,
            'unemployment': unemp,
            'income_tax': income_tax,
            'stamp_tax': stamp,
            'total_deductions': round(tax_total + post_tax_deductions, 2),
            'tax_deductions': tax_total,
            'extra_deductions': round(post_tax_deductions, 2),
            'extra_earnings': round(ot_added + added_earnings, 2),
            'leave_days_paid': paid_leave_days,
            'leave_days_unpaid': unpaid_leave_days,
            'sgk_days': sgk_days,
            'eksik_gun': eksik_gun,
            'net_salary': net,
            'line_items': line_items,
        })
        out.append(new_row)
    return out


async def _payroll_collect_leaves(
    tenant_id: str, period_month: str,
) -> dict[str, dict]:
    """`leave_requests` status=approved + start/end overlap ile ay'a düşen
    gün sayılarını topla → per-staff {paid_days, unpaid_days}. SGK günü ve
    eksik gün hesabı için kullanılır. Read-only; finalize ASLA çağrılmaz.

    `leave_type` 'unpaid' / 'unpaid_leave' / 'ucretsiz' → unpaid; aksi paid.
    """
    yyyy, mm = period_month.split('-')
    start_iso = f'{yyyy}-{mm}-01'
    nm = date(int(yyyy), int(mm), 28) + timedelta(days=4)
    end_iso = (nm - timedelta(days=nm.day)).isoformat()
    p_start = date.fromisoformat(start_iso)
    p_end = date.fromisoformat(end_iso)
    by_staff: dict[str, dict] = {}
    cursor = db.leave_requests.find({
        'tenant_id': tenant_id,
        'status': 'approved',
        'start_date': {'$lte': end_iso},
        'end_date': {'$gte': start_iso},
    }, {'_id': 0, 'staff_id': 1, 'start_date': 1, 'end_date': 1, 'leave_type': 1})
    async for r in cursor:
        sid = r.get('staff_id')
        if not sid:
            continue
        try:
            s = date.fromisoformat(str(r.get('start_date') or '')[:10])
            e = date.fromisoformat(str(r.get('end_date') or '')[:10])
        except Exception:
            continue
        s = max(s, p_start)
        e = min(e, p_end)
        if e < s:
            continue
        days = (e - s).days + 1
        lt = (r.get('leave_type') or '').strip().lower()
        is_unpaid = lt in ('unpaid', 'unpaid_leave', 'ucretsiz', 'ücretsiz')
        b = by_staff.setdefault(sid, {'paid_days': 0, 'unpaid_days': 0, 'requests': 0})
        if is_unpaid:
            b['unpaid_days'] += days
        else:
            b['paid_days'] += days
        b['requests'] += 1
    return by_staff


async def _payroll_collect_overtime(tenant_id: str, period_month: str) -> dict[str, dict]:
    """Bordro modülünün onaylı mesai tüketimi — **GET /hr/overtime/ready-for-payroll**
    kontratının iç-süreç eşdeğeri. Aynı sorgu/şema kullanılır
    (`overtime_requests` status=approved + work_date in [month_start..end])
    böylece HTTP endpoint ile parity sağlanır; HTTP self-call yapmaktan
    kaçınılır (timeout/circular-dep önlemi). Read-only; finalize ASLA
    çağrılmaz. Şema kontrat ile bağlı — değiştirilirse her iki yer birlikte
    güncellenmeli (parity test'i CI'da yakalar).
    """
    yyyy, mm = period_month.split('-')
    start_iso = f'{yyyy}-{mm}-01'
    nm = date(int(yyyy), int(mm), 28) + timedelta(days=4)
    end_iso = (nm - timedelta(days=nm.day)).isoformat()
    by_staff: dict[str, dict] = {}
    cursor = db.overtime_requests.find({
        'tenant_id': tenant_id,
        'status': 'approved',
        'work_date': {'$gte': start_iso, '$lte': end_iso},
    }, {'_id': 0, 'staff_id': 1, 'hours': 1})
    async for r in cursor:
        sid = r['staff_id']
        b = by_staff.setdefault(sid, {'hours': 0.0, 'requests': 0})
        b['hours'] += float(r.get('hours') or 0)
        b['requests'] += 1
    return by_staff


async def _build_payroll_v2(
    tenant_id: str, month: str, extras: list[dict] | None = None,
) -> tuple[str, list[dict], dict[str, Any]]:
    """Dry-run v2 compute — base + approved overtime + extras → enriched rows
    with line_items. Pure function over DB reads; no writes."""
    period_month, base = await _build_payroll(month, tenant_id)
    ot_map = await _payroll_collect_overtime(tenant_id, period_month)
    lv_map = await _payroll_collect_leaves(tenant_id, period_month)
    enriched = _payroll_apply_extras_and_overtime(
        base, extras or [], ot_map, lv_map,
    )
    summary = {
        'staff_count': len(enriched),
        'total_gross': round(sum(r['gross_pay'] for r in enriched), 2),
        'total_net': round(sum(r['net_salary'] for r in enriched), 2),
        'total_extra_earnings': round(
            sum(r.get('extra_earnings', 0) for r in enriched), 2),
        'total_extra_deductions': round(
            sum(r.get('extra_deductions', 0) for r in enriched), 2),
        'total_leave_days_paid': sum(r.get('leave_days_paid', 0) for r in enriched),
        'total_leave_days_unpaid': sum(r.get('leave_days_unpaid', 0) for r in enriched),
        'currency': TR_CURRENCY,
    }
    return period_month, enriched, summary


def _payroll_run_to_response(
    run: dict, current_user: User,
) -> dict:
    """`payroll_runs` doc → API response; IBAN/TC PII mask per row unless
    manage_hr OR self."""
    out = dict(run)
    out.pop('_id', None)
    rows = out.get('rows') or []
    masked: list[dict] = []
    self_id = str(getattr(current_user, 'id', '') or '')
    self_email = str(getattr(current_user, 'email', '') or '')
    for r in rows:
        m = _mask_hr_pii(r, current_user, self_id=self_id, self_email=self_email)
        masked.append(m or r)
    out['rows'] = masked
    return out


# ---------- Payroll v2 RBAC gate (Task #264, post-review P1) ----------
# Contract (Task #264 — entitlement-based, KVKK least-privilege):
#   • FINALIZE (locked, immutable): manage_hr permission (HR Admin) OR
#     finance/super_admin role.
#   • DRAFT / REVISION (mutable preview): yukarıdakiler + view_hr_payroll
#     permission (HR Manager — finalize edemez ama hazırlayabilir).
#   • Diğer herkes 403.
# Ham `admin` rolü ARTIK YETMEZ — HR Admin yetkisi `manage_hr` perm'i ile
# açıkça verilmelidir (role-literal gating kaldırıldı).
_PAYROLL_PRIVILEGED_ROLES = frozenset({'finance', 'super_admin'})


def _payroll_lifecycle_gate(user: User, *, allow_hr_manager: bool) -> None:
    role = (getattr(user, 'role', '') or '').lower()
    # Full-lifecycle gate (finalize dahil).
    if _user_has_hr_op(user, 'manage_hr') or role in _PAYROLL_PRIVILEGED_ROLES:
        return
    # Draft/revision için ek olarak view_hr_payroll perm'i yeterli.
    if allow_hr_manager and _user_has_hr_op(user, 'view_hr_payroll'):
        return
    raise HTTPException(
        status_code=403,
        detail=(
            "Bordro yaşam döngüsü yetkiniz yok "
            "(yalnızca HR Admin (manage_hr) / Finance / Süper Admin"
            + (" / HR Manager (view_hr_payroll)" if allow_hr_manager else "")
            + ")."
        ),
    )


# ---------- Payroll v2 endpoints (Task #264) ----------

@router.get("/hr/payroll/runs")
async def list_payroll_runs(
    month: str | None = Query(None, pattern=r'^\d{4}-\d{2}$'),
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr_payroll")),
):
    """Tenant'a ait payroll_runs listele — opsiyonel month filtresi."""
    q: dict[str, Any] = {'tenant_id': current_user.tenant_id}
    if month:
        q['period_month'] = month
    cursor = db.payroll_runs.find(q, {'_id': 0, 'rows': 0}).sort('created_at', -1)
    items = await cursor.to_list(200)
    return {'items': items, 'count': len(items)}


@router.get("/hr/payroll/runs/{run_id}")
async def get_payroll_run(
    run_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr_payroll")),
):
    """Tek run + line_items detay (PII masked unless manage_hr/self)."""
    run = await db.payroll_runs.find_one(
        {'id': run_id, 'tenant_id': current_user.tenant_id}, {'_id': 0},
    )
    if not run:
        raise HTTPException(status_code=404, detail="Bordro çalışması bulunamadı")
    return _payroll_run_to_response(run, current_user)


@router.get("/hr/payroll/runs/{run_id}/revisions")
async def list_payroll_revisions(
    run_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr_payroll")),
):
    """Bir run için açılmış tüm revizyonlar (zaman sıralı, audit trail)."""
    run = await db.payroll_runs.find_one(
        {'id': run_id, 'tenant_id': current_user.tenant_id}, {'_id': 0, 'id': 1},
    )
    if not run:
        raise HTTPException(status_code=404, detail="Bordro çalışması bulunamadı")
    items = await db.payroll_revisions.find(
        {'tenant_id': current_user.tenant_id, 'parent_run_id': run_id}, {'_id': 0},
    ).sort('created_at', -1).to_list(200)
    return {'items': items, 'count': len(items)}


@router.post("/hr/payroll/{month}/save")
async def save_payroll_draft(
    month: str,
    payload: PayrollSavePayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """Draft bordro kaydet — idempotent upsert (aynı ay için var olan
    draft üstüne yazılır). LOCKED bir run varsa 409 → revizyon yolu açılmalı.

    Bu endpoint HİÇBİR muhasebe kaydı oluşturmaz; sadece `payroll_runs`
    koleksiyonuna draft snapshot yazar. Asıl muhasebe etkisi `finalize` ile.

    RBAC: HR Admin / Finance / Süper Admin / HR Manager / Supervisor.
    """
    _payroll_lifecycle_gate(current_user, allow_hr_manager=True)

    existing_locked = await db.payroll_runs.find_one(
        {
            'tenant_id': current_user.tenant_id,
            'period_month': month,
            'status': 'locked',
        }, {'_id': 0, 'id': 1},
    )
    if existing_locked:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Bu ay ({month}) için kilitlenmiş bir bordro var "
                f"(run_id={existing_locked['id']}). Değişiklik için revizyon açın."
            ),
        )

    extras = [ex.model_dump() for ex in payload.extras]
    period_month, rows, summary = await _build_payroll_v2(
        current_user.tenant_id, month, extras=extras,
    )
    if not rows:
        raise HTTPException(
            status_code=400,
            detail="Bu ay için devam kaydı yok — draft oluşturulamaz.",
        )

    now_iso = datetime.now(UTC).isoformat()
    uid = getattr(current_user, 'id', None)
    filter_q = {
        'tenant_id': current_user.tenant_id,
        'period_month': period_month,
        'status': 'draft',
    }
    update_q = {
        '$setOnInsert': {
            'id': str(uuid.uuid4()),
            'created_at': now_iso,
            'created_by': uid,
            'parent_run_id': None,
            'finalized_at': None,
            'finalized_by': None,
        },
        '$set': {
            'tenant_id': current_user.tenant_id,
            'period_month': period_month,
            'status': 'draft',
            'rows': rows,
            'summary': summary,
            'extras': extras,
            'note': payload.note,
            'updated_at': now_iso,
            'updated_by': uid,
        },
    }
    try:
        res = await db.payroll_runs.update_one(filter_q, update_q, upsert=True)
        is_idempotent_update = res.upserted_id is None
        upserted_id = res.upserted_id
    except DuplicateKeyError:
        # Post-review P1 (round 1): concurrent insert race; partial unique on
        # (tenant_id, period_month) where status='draft' rejected the
        # second insert. Re-apply update on the now-existing draft.
        res = await db.payroll_runs.update_one(filter_q, {'$set': update_q['$set']})
        is_idempotent_update = True
        upserted_id = None
    # Post-review P1 (round 2): save↔finalize TOCTOU. existing_locked
    # kontrolü ile upsert arasında başka bir istek draft'ı locked'a
    # çevirebilir. Bu durumda upsert filter (status='draft') eşleşmez ve
    # AYNI ay için YENİ bir draft insert edilir → "locked + new draft"
    # immutability ihlali. Çözüm: yazımdan sonra locked yeniden doğrula;
    # eğer locked varsa ve biz az önce insert ettiysek, draft'ı geri al.
    locked_after = await db.payroll_runs.find_one(
        {
            'tenant_id': current_user.tenant_id,
            'period_month': period_month,
            'status': 'locked',
        }, {'_id': 0, 'id': 1},
    )
    if locked_after and upserted_id is not None:
        # Bizim insert ettiğimiz draft satırını geri al (rollback).
        try:
            await db.payroll_runs.delete_one({
                'tenant_id': current_user.tenant_id,
                'period_month': period_month,
                'status': 'draft',
                '_id': upserted_id,
            })
        except Exception:
            pass
        raise HTTPException(
            status_code=409,
            detail=(
                f"Bu ay ({period_month}) için bordro arada kilitlendi "
                f"(run_id={locked_after['id']}). Değişiklik için revizyon açın."
            ),
        )
    doc = await db.payroll_runs.find_one(filter_q, {'_id': 0, 'id': 1})
    run_id = (doc or {}).get('id')
    await _audit(
        current_user, 'payroll.save_draft', 'payroll_run', run_id or '?',
        f"Draft bordro kaydedildi (ay={period_month}, satır={len(rows)})",
        severity='info',
    )
    return {
        'success': True,
        'run_id': run_id,
        'period_month': period_month,
        'status': 'draft',
        'summary': summary,
        'is_idempotent_update': is_idempotent_update,
    }


@router.post("/hr/payroll/{run_id}/finalize")
async def finalize_payroll_run(
    run_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """Draft → locked geçişi. Bir kere kilitlenince satırlar immutable.
    Değişiklik için `revisions` yolu kullanılmalı.

    RBAC: yalnızca HR Admin / Finance / Süper Admin (HR Manager finalize
    edemez — iş kuralı, _payroll_lifecycle_gate ile zorlanır)."""
    _payroll_lifecycle_gate(current_user, allow_hr_manager=False)

    run = await db.payroll_runs.find_one(
        {'id': run_id, 'tenant_id': current_user.tenant_id}, {'_id': 0},
    )
    if not run:
        raise HTTPException(status_code=404, detail="Bordro çalışması bulunamadı")
    if run.get('status') == 'locked':
        return {
            'success': True, 'run_id': run_id, 'status': 'locked',
            'message': 'Zaten kilitli (idempotent).',
        }
    if run.get('status') != 'draft':
        raise HTTPException(
            status_code=409,
            detail=f"Sadece draft durumdaki bordro kilitlenebilir (mevcut={run.get('status')}).",
        )

    now_iso = datetime.now(UTC).isoformat()
    result = await db.payroll_runs.update_one(
        {
            'id': run_id,
            'tenant_id': current_user.tenant_id,
            'status': 'draft',  # CAS guard
        },
        {'$set': {
            'status': 'locked',
            'finalized_at': now_iso,
            'finalized_by': getattr(current_user, 'id', None),
            'updated_at': now_iso,
            'updated_by': getattr(current_user, 'id', None),
        }},
    )
    if result.modified_count == 0:
        raise HTTPException(
            status_code=409,
            detail="Bordro kilitlenemedi (yarış durumu); tekrar deneyin.",
        )

    await _audit(
        current_user, 'payroll.finalize', 'payroll_run', run_id,
        f"Bordro kilitlendi (ay={run.get('period_month')})",
        severity='warning',
    )
    return {
        'success': True, 'run_id': run_id, 'status': 'locked',
        'period_month': run.get('period_month'),
        'finalized_at': now_iso,
    }


@router.post("/hr/payroll/{run_id}/revisions")
async def revise_payroll_run(
    run_id: str,
    payload: PayrollRevisionPayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """Locked run için revizyon: parent IMMUTABLE kalır; yeni bir DRAFT
    run açılır + `payroll_revisions` audit satırı yazılır. Reason zorunlu.

    RBAC: HR Admin / Finance / Süper Admin / HR Manager / Supervisor.
    Aynı ay için açık draft varsa partial unique index `DuplicateKeyError`
    fırlatır — revizyon açılamaz (önce mevcut draft kilitlenmeli)."""
    _payroll_lifecycle_gate(current_user, allow_hr_manager=True)
    parent = await db.payroll_runs.find_one(
        {'id': run_id, 'tenant_id': current_user.tenant_id}, {'_id': 0},
    )
    if not parent:
        raise HTTPException(status_code=404, detail="Üst bordro bulunamadı")
    if parent.get('status') != 'locked':
        raise HTTPException(
            status_code=409,
            detail="Revizyon yalnızca kilitlenmiş bordro üzerinden açılabilir.",
        )

    extras = [ex.model_dump() for ex in payload.extras]
    period_month, rows, summary = await _build_payroll_v2(
        current_user.tenant_id, parent['period_month'], extras=extras,
    )
    now_iso = datetime.now(UTC).isoformat()
    new_run_id = str(uuid.uuid4())
    new_doc = {
        'id': new_run_id,
        'tenant_id': current_user.tenant_id,
        'period_month': period_month,
        'status': 'draft',
        'rows': rows,
        'summary': summary,
        'extras': extras,
        'note': f"Revizyon: {payload.reason}",
        'created_at': now_iso,
        'created_by': getattr(current_user, 'id', None),
        'updated_at': now_iso,
        'updated_by': getattr(current_user, 'id', None),
        'finalized_at': None,
        'finalized_by': None,
        'parent_run_id': run_id,
    }
    try:
        await db.payroll_runs.insert_one(new_doc)
    except DuplicateKeyError:
        raise HTTPException(
            status_code=409,
            detail=(
                "Bu ay için açık bir taslak zaten var — yeni revizyon açabilmek "
                "için önce mevcut taslağı kilitleyin (tek-aktif-draft kuralı)."
            ),
        )

    diff = {
        'gross_before': parent.get('summary', {}).get('total_gross'),
        'gross_after': summary['total_gross'],
        'net_before': parent.get('summary', {}).get('total_net'),
        'net_after': summary['total_net'],
        'staff_before': parent.get('summary', {}).get('staff_count'),
        'staff_after': summary['staff_count'],
    }
    rev_doc = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'parent_run_id': run_id,
        'new_run_id': new_run_id,
        'period_month': period_month,
        'reason': payload.reason,
        'diff': diff,
        'created_at': now_iso,
        'created_by': getattr(current_user, 'id', None),
    }
    await db.payroll_revisions.insert_one(rev_doc)
    await _audit(
        current_user, 'payroll.revision_open', 'payroll_run', run_id,
        (
            f"Revizyon açıldı (ay={period_month}, yeni_draft={new_run_id}, "
            f"sebep={payload.reason[:80]})"
        ),
        severity='warning',
    )
    return {
        'success': True,
        'parent_run_id': run_id,
        'new_run_id': new_run_id,
        'status': 'draft',
        'diff': diff,
    }


@router.get("/hr/payroll/runs/{run_id}/export.xlsx")
async def export_payroll_run_xlsx(
    run_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr_payroll")),
):
    """Bir run'ı Excel olarak indir (line_items detay)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    run = await db.payroll_runs.find_one(
        {'id': run_id, 'tenant_id': current_user.tenant_id}, {'_id': 0},
    )
    if not run:
        raise HTTPException(status_code=404, detail="Bordro çalışması bulunamadı")
    resp_run = _payroll_run_to_response(run, current_user)
    rows = resp_run.get('rows') or []

    wb = Workbook()
    ws = wb.active
    ws.title = f"Bordro {run.get('period_month', '')}"
    headers = [
        'Personel', 'Departman', 'Toplam Saat', 'Mesai Saat',
        'Brüt', 'SGK+İşsizlik', 'Vergi+Damga',
        'Ek Kazanç', 'Ek Kesinti', 'Net', 'Para Birimi',
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    for r_idx, row in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1, value=row.get('staff_name') or row.get('staff_id'))
        ws.cell(row=r_idx, column=2, value=row.get('department'))
        ws.cell(row=r_idx, column=3, value=float(row.get('total_hours') or 0))
        ws.cell(row=r_idx, column=4, value=float(row.get('overtime_hours') or 0))
        ws.cell(row=r_idx, column=5, value=float(row.get('gross_pay') or 0))
        ws.cell(row=r_idx, column=6, value=float(
            (row.get('sgk_employee') or 0) + (row.get('unemployment') or 0)))
        ws.cell(row=r_idx, column=7, value=float(
            (row.get('income_tax') or 0) + (row.get('stamp_tax') or 0)))
        ws.cell(row=r_idx, column=8, value=float(row.get('extra_earnings') or 0))
        ws.cell(row=r_idx, column=9, value=float(row.get('extra_deductions') or 0))
        ws.cell(row=r_idx, column=10, value=float(row.get('net_salary') or 0))
        ws.cell(row=r_idx, column=11, value=row.get('currency') or TR_CURRENCY)
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 16

    # Detail sheet: line_items per staff
    ws2 = wb.create_sheet('Kalemler')
    ws2_headers = ['Personel', 'Kalem', 'Açıklama', 'Saat', 'Tarife', 'Tutar', 'Yön', 'Not']
    for col, h in enumerate(ws2_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    row_cursor = 2
    for row in rows:
        name = row.get('staff_name') or row.get('staff_id')
        for li in row.get('line_items') or []:
            ws2.cell(row=row_cursor, column=1, value=name)
            ws2.cell(row=row_cursor, column=2, value=li.get('kind'))
            ws2.cell(row=row_cursor, column=3, value=li.get('label'))
            ws2.cell(row=row_cursor, column=4, value=li.get('hours'))
            ws2.cell(row=row_cursor, column=5, value=li.get('rate'))
            ws2.cell(row=row_cursor, column=6, value=li.get('amount'))
            ws2.cell(row=row_cursor, column=7, value=li.get('direction'))
            ws2.cell(row=row_cursor, column=8, value=li.get('note'))
            row_cursor += 1
    for col_idx in range(1, len(ws2_headers) + 1):
        ws2.column_dimensions[chr(64 + col_idx)].width = 18

    resp = _xlsx_stream(wb)
    resp.headers['Content-Disposition'] = (
        f'attachment; filename="payroll_run_{run_id}.xlsx"'
    )
    await _audit(
        current_user, 'payroll.export_xlsx', 'payroll_run', run_id,
        f"Bordro XLSX indirildi (run_id={run_id})",
        severity='info',
    )
    return resp


# Task #264 round 6: explicit allow-list (fail-closed). Sadece iç staff
# rolleri kendi bordrosuna erişebilir; agency/guest/bilinmeyen rol 403.
# UserRole enum tabanlı — yeni eklenen roller buraya bilinçli olarak
# dahil edilmedikçe otomatik olarak reddedilir.
_PAYROLL_ME_ALLOWED_ROLES = frozenset({
    'super_admin', 'admin', 'supervisor', 'front_desk', 'housekeeping',
    'sales', 'finance', 'procurement', 'staff',
})


@router.get("/hr/payroll/me")
async def get_my_payroll(
    month: str | None = Query(None, pattern=r'^\d{4}-\d{2}$'),
    current_user: User = Depends(get_current_user),
):
    """Self-service: kullanıcının kendi locked bordro satırı. Yalnız `locked`
    runlardan veri döner — taslak görünmez (KVKK + iş hukuku doktrin).

    RBAC (Task #264 post-review round 6, fail-closed allow-list):
    `_PAYROLL_ME_ALLOWED_ROLES` setindeki iç staff rolleri erişebilir;
    guest / agency_admin / agency_agent / bilinmeyen rol 403 alır. Erişim
    açılsa bile object-level self-match (`staff_id`/e-posta) ile yalnız
    kendi satırı döner.
    """
    role_lc = (getattr(current_user, 'role', '') or '').lower()
    if role_lc not in _PAYROLL_ME_ALLOWED_ROLES:
        raise HTTPException(
            status_code=403,
            detail="Bordro self-service yalnız iç staff için açıktır.",
        )
    self_id = str(getattr(current_user, 'id', '') or '')
    self_email = (str(getattr(current_user, 'email', '') or '')).strip().lower()
    q: dict[str, Any] = {
        'tenant_id': current_user.tenant_id,
        'status': 'locked',
    }
    if month:
        q['period_month'] = month
    runs = await db.payroll_runs.find(q, {'_id': 0}).sort('period_month', -1).to_list(24)
    my_rows: list[dict] = []
    for run in runs:
        for r in run.get('rows') or []:
            rid = str(r.get('staff_id') or '')
            remail = (str(r.get('email') or '')).strip().lower()
            if (self_id and rid == self_id) or (
                self_email and remail and remail == self_email
            ):
                my_rows.append({
                    'period_month': run['period_month'],
                    'run_id': run['id'],
                    'finalized_at': run.get('finalized_at'),
                    **r,
                })
    return {'items': my_rows, 'count': len(my_rows)}


@router.get("/hr/payroll/{month}/dept-summary")
async def get_payroll_dept_summary(
    month: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """Departman müdürü için aggregate özet — NET ve BRÜT TUTAR DÖNMEZ
    (Task #264 post-review: no-amount doktrin); yalnızca personel sayısı,
    gün toplamı ve mesai saat toplamı.

    Tüm runlar değil, en güncel LOCKED runlardan; locked yoksa son draft'tan
    (manager planlama amaçlı). Tek bir periyot için."""
    runs = await db.payroll_runs.find(
        {'tenant_id': current_user.tenant_id, 'period_month': month},
        {'_id': 0},
    ).sort('finalized_at', -1).to_list(20)
    chosen = next((r for r in runs if r.get('status') == 'locked'), None)
    if chosen is None:
        chosen = next((r for r in runs if r.get('status') == 'draft'), None)
    if chosen is None:
        return {'period_month': month, 'departments': [], 'source': 'none'}

    assigned = _user_assigned_department(current_user)
    by_dept: dict[str, dict] = {}
    for r in chosen.get('rows') or []:
        dept = (r.get('department') or 'unknown').strip().lower() or 'unknown'
        if assigned and dept != assigned.strip().lower() and not _user_has_hr_op(
            current_user, 'view_hr_payroll',
        ):
            continue
        agg = by_dept.setdefault(dept, {
            'department': dept,
            'staff_count': 0,
            'total_hours': 0.0,
            'overtime_hours': 0.0,
            'sgk_days': 0,
        })
        agg['staff_count'] += 1
        agg['total_hours'] = round(agg['total_hours'] + (r.get('total_hours') or 0), 2)
        agg['overtime_hours'] = round(
            agg['overtime_hours'] + (r.get('overtime_hours') or 0), 2,
        )
        agg['sgk_days'] += int(r.get('sgk_days') or 0)
    return {
        'period_month': month,
        'run_id': chosen['id'],
        'source': chosen.get('status'),
        'departments': list(by_dept.values()),
    }


@router.get("/hr/payroll/{month}")
async def get_payroll(
    month: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """Per-month payroll preview — DAİMA dry-run (saf hesap, hiç yazma).

    `runs`: bu aya ait `payroll_runs` (draft+locked) özet listesi.
    `latest_locked` / `latest_draft`: en güncel ilgili run referansları.
    `payroll`: dry-run hesap (legacy `payroll_records` kayıtları DA döner —
    geri uyum; UI yeni `runs` listesini tercih etmelidir).

    Task #264: tüm muhasebe etkili veriler `payroll_runs` üzerinden;
    bu endpoint asla DB'ye yazmaz (`is_dry_run=true`).
    """
    # Dry-run compute (extras yok — pure baseline)
    period_month, rows, summary = await _build_payroll_v2(
        current_user.tenant_id, month, extras=[],
    )

    # Dept-scope (Task #264 post-review round 2): department-manager
    # (view_hr + assigned_department + view_hr_payroll YOK + manage_hr YOK)
    # için response degrade modu:
    #   • rows yalnız kendi departman + monetary alanlar strip
    #   • summary kendi scope üzerinden YENİDEN hesap, brüt/net yok
    #   • runs özetlerindeki `summary.total_gross/net` strip
    #   • legacy payroll_records hiç döndürülmez (no-amount doktrin)
    assigned = _user_assigned_department(current_user)
    is_dept_only = bool(
        assigned
        and not _user_has_hr_op(current_user, "view_hr_payroll")
        and not _user_has_hr_op(current_user, "manage_hr")
        and (getattr(current_user, "role", "") or "").lower() not in (
            "admin", "super_admin", "finance",
        )
    )

    if is_dept_only:
        a_lc = assigned.strip().lower()  # type: ignore[union-attr]
        rows = [
            r for r in rows
            if (r.get('department') or '').strip().lower() == a_lc
        ]
        # Monetary alanları strip — yalnız days/hours metrikleri kalsın.
        _AMOUNT_FIELDS = (
            'gross_pay', 'net_salary', 'sgk_employee', 'unemployment',
            'income_tax', 'stamp_tax', 'total_deductions', 'tax_deductions',
            'extra_deductions', 'extra_earnings', 'hourly_rate',
            'overtime_rate', 'base_salary', 'line_items',
        )
        scoped_rows: list[dict] = []
        for r in rows:
            r2 = {k: v for k, v in r.items() if k not in _AMOUNT_FIELDS}
            scoped_rows.append(r2)
        rows = scoped_rows
        # Summary'i scoped rows üzerinden yeniden hesapla — no-amount.
        summary = {
            'staff_count': len(rows),
            'total_hours': round(sum(r.get('total_hours', 0) for r in rows), 2),
            'overtime_hours': round(
                sum(r.get('overtime_hours', 0) for r in rows), 2,
            ),
            'total_leave_days_paid': sum(
                r.get('leave_days_paid', 0) for r in rows
            ),
            'total_leave_days_unpaid': sum(
                r.get('leave_days_unpaid', 0) for r in rows
            ),
            'total_sgk_days': sum(r.get('sgk_days', 0) for r in rows),
            'scope': 'department_only',
            'department': assigned,
        }

    # PII mask satırlar üzerinde
    self_id = str(getattr(current_user, 'id', '') or '')
    self_email = str(getattr(current_user, 'email', '') or '')
    masked_rows = [
        _mask_hr_pii(r, current_user, self_id=self_id, self_email=self_email) or r
        for r in rows
    ]

    runs_cursor = db.payroll_runs.find(
        {'tenant_id': current_user.tenant_id, 'period_month': month},
        {'_id': 0, 'rows': 0},
    ).sort('created_at', -1)
    runs = await runs_cursor.to_list(50)
    if is_dept_only:
        # Run özetlerinden tutar bilgilerini strip.
        for r in runs:
            s = r.get('summary')
            if isinstance(s, dict):
                for k in (
                    'total_gross', 'total_net', 'total_extra_earnings',
                    'total_extra_deductions',
                ):
                    s.pop(k, None)
    latest_locked = next((r for r in runs if r.get('status') == 'locked'), None)
    latest_draft = next((r for r in runs if r.get('status') == 'draft'), None)

    # Geri uyum: legacy finalize edilmiş `payroll_records` — dept-only için
    # legacy satırlar gizlenir (no-amount doktrin).
    legacy: list[dict] = []
    if not is_dept_only:
        legacy = await db.payroll_records.find(
            {'tenant_id': current_user.tenant_id, 'period_month': month},
            {'_id': 0},
        ).to_list(500)

    payload: dict[str, Any] = {
        'is_dry_run': True,
        'period_month': period_month,
        'payroll': masked_rows,
        'summary': summary,
        'staff_count': summary.get('staff_count'),
        'currency': TR_CURRENCY,
        'runs': runs,
        'latest_locked_run': latest_locked,
        'latest_draft_run': latest_draft,
        'legacy_payroll_records': legacy,
        'legacy_count': len(legacy),
    }
    if not is_dept_only:
        payload['total_gross_pay'] = summary.get('total_gross')
        payload['total_net_pay'] = summary.get('total_net')
    else:
        payload['scope'] = 'department_only'
        payload['department'] = assigned
    return payload


# ============= Performance =============

@router.post("/hr/performance")
async def create_performance_review(
    payload: PerformanceReviewPayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    staff = await _verify_staff_in_tenant(payload.staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")

    # Task #254 (F8D-v2 § 32 P1): terminal-state guard — bir personel için
    # aynı `period` (örn. "2027-Q4") tekrar review oluşturulmamalı. Olmayan
    # uniqueness gate olmadan stres testlerinde duplicate kayıt oluşuyor ve
    # raporlarda çift sayım, kazanç hesabında çift faktör riskine yol açıyor.
    # `period` boş ise ad-hoc review sayılır (rapor dönemi yok) → guard
    # uygulanmaz. Aksi halde (tenant_id, staff_id, period) tekil olmalı.
    if payload.period:
        existing = await db.performance_reviews.find_one({
            'tenant_id': current_user.tenant_id,
            'staff_id': payload.staff_id,
            'period': payload.period,
        }, {'_id': 0, 'id': 1})
        if existing:
            raise HTTPException(
                status_code=409,
                detail=f"Bu personel için {payload.period} dönemine ait performans değerlendirmesi zaten mevcut (id={existing.get('id')}).",
            )

    review = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'staff_id': payload.staff_id,
        'staff_name': staff.get('name'),
        'reviewer_id': getattr(current_user, 'id', None),
        'reviewer_name': payload.reviewer_name,
        'period': payload.period,
        'overall_score': payload.overall_score,
        'goals': payload.goals,
        'strengths': payload.strengths,
        'improvement_areas': payload.improvement_areas,
        'template_id': payload.template_id,
        'competency_scores': payload.competency_scores or {},
        'reviewed_at': datetime.now(UTC).isoformat(),
    }
    await db.performance_reviews.insert_one(review)
    return {'success': True, 'review_id': review['id']}


@router.get("/hr/performance")
async def list_performance_reviews(
    staff_id: str | None = None,
    current_user: User = Depends(get_current_user),
):
    query: dict[str, Any] = {'tenant_id': current_user.tenant_id}
    if staff_id:
        query['staff_id'] = staff_id
    items = await db.performance_reviews.find(query, {'_id': 0}).sort('reviewed_at', -1).to_list(500)
    avg = round(sum(i.get('overall_score', 0) for i in items) / len(items), 2) if items else 0
    return {'items': items, 'total': len(items), 'avg_score': avg}


# ============= Recruitment =============

@router.post("/hr/job-posting")
async def create_job_posting(
    payload: JobPostingPayload,
    current_user: User = Depends(get_current_user),
):
    """Personel ihtiyaç talebi oluştur (her departman müdürü açabilir).

    - internal_request → status=pending_approval, HR yöneticisine bildirim gider.
    - public_posting → yalnızca HR yetkisi ile, doğrudan status=active.
    """
    # public_posting için ek yetki gerekir; internal_request herkese açık
    if payload.request_type == 'public_posting':
        role = (getattr(current_user, 'role', None) or '').lower()
        if role not in {'admin', 'supervisor', 'finance'}:
            raise HTTPException(
                status_code=403,
                detail="Doğrudan ilan yayınlamak için HR yöneticisi yetkisi gerekir"
            )
        initial_status = 'active'
    else:
        initial_status = 'pending_approval'

    job = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'title': payload.title,
        'department': payload.department,
        'description': payload.description,
        'employment_type': payload.employment_type,
        'location': payload.location,
        'salary_range': payload.salary_range,
        'request_type': payload.request_type,
        'headcount_needed': payload.headcount_needed,
        'urgency': payload.urgency,
        'justification': payload.justification,
        'needed_by': payload.needed_by,
        'status': initial_status,
        'applicants_count': 0,
        'created_by': getattr(current_user, 'id', None),
        'created_by_name': getattr(current_user, 'name', None) or getattr(current_user, 'email', None),
        'created_at': datetime.now(UTC).isoformat(),
    }
    await db.job_postings.insert_one(job)

    # Yöneticiye bildirim — onay gerektiren talepler için
    if initial_status == 'pending_approval':
        await _notify_hr_managers(
            current_user.tenant_id,
            kind='hr_request',
            title='Yeni personel talebi',
            body=(
                f"{job.get('created_by_name') or 'Bir müdür'} • "
                f"{payload.department} • {payload.title} ({payload.headcount_needed} kişi)"
                + (f" • aciliyet: {payload.urgency}" if payload.urgency != 'normal' else '')
            ),
            link=f'/hr?tab=recruitment&job={job["id"]}',
            ref_id=job['id'],
        )

    return {'success': True, 'job_id': job['id'], 'status': initial_status}


@router.post("/hr/job-posting/{job_id}/approve")
async def approve_job_posting(
    job_id: str,
    payload: JobDecisionPayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    job = await db.job_postings.find_one({'tenant_id': current_user.tenant_id, 'id': job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Talep bulunamadı")
    if job.get('status') not in ('pending_approval', None):
        raise HTTPException(status_code=400, detail="Bu talep zaten karara bağlanmış")
    await db.job_postings.update_one(
        {'tenant_id': current_user.tenant_id, 'id': job_id},
        {'$set': {
            'status': 'active',
            'approved_by': getattr(current_user, 'id', None),
            'approved_by_name': getattr(current_user, 'name', None),
            'approved_at': datetime.now(UTC).isoformat(),
            'decision_note': payload.note,
        }}
    )
    # Talep sahibine bildirim
    if job.get('created_by'):
        await _notify_user(
            current_user.tenant_id,
            user_id=job['created_by'],
            kind='hr_request_approved',
            title='Personel talebiniz onaylandı',
            body=f"{job.get('title')} • aday eklemeye açıldı",
            link=f'/hr?tab=recruitment&job={job_id}',
            ref_id=job_id,
        )
    return {'success': True, 'status': 'active'}


@router.post("/hr/job-posting/{job_id}/reject")
async def reject_job_posting(
    job_id: str,
    payload: JobDecisionPayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    job = await db.job_postings.find_one({'tenant_id': current_user.tenant_id, 'id': job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Talep bulunamadı")
    await db.job_postings.update_one(
        {'tenant_id': current_user.tenant_id, 'id': job_id},
        {'$set': {
            'status': 'rejected',
            'rejected_by': getattr(current_user, 'id', None),
            'rejected_at': datetime.now(UTC).isoformat(),
            'decision_note': payload.note,
        }}
    )
    if job.get('created_by'):
        await _notify_user(
            current_user.tenant_id,
            user_id=job['created_by'],
            kind='hr_request_rejected',
            title='Personel talebiniz reddedildi',
            body=(payload.note or job.get('title') or '—'),
            link=f'/hr?tab=recruitment&job={job_id}',
            ref_id=job_id,
        )
    return {'success': True, 'status': 'rejected'}


@router.get("/hr/job-postings")
async def list_job_postings(
    status: str | None = None,
    current_user: User = Depends(get_current_user),
):
    query: dict[str, Any] = {'tenant_id': current_user.tenant_id}
    if status:
        query['status'] = status
    items = await db.job_postings.find(query, {'_id': 0}).sort('created_at', -1).to_list(500)
    return {'items': items, 'total': len(items)}


@router.post("/hr/job-posting/{job_id}/close")
async def close_job_posting(
    job_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    res = await db.job_postings.update_one(
        {'tenant_id': current_user.tenant_id, 'id': job_id},
        {'$set': {'status': 'closed', 'closed_at': datetime.now(UTC).isoformat()}}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="İş ilanı bulunamadı")
    return {'success': True}


# ============= F&B COMPLETE SUITE =============

async def _get_ingredient_map(tenant_id: str):
    ingredients = await db.ingredients.find({'tenant_id': tenant_id}, {'_id': 0}).to_list(1000)
    return {ing['id']: ing for ing in ingredients}, ingredients


def _enrich_recipe_cost(recipe: dict, ingredient_map: dict[str, dict]):
    enriched_lines = []
    total_cost = 0
    for line in recipe.get('ingredients', []):
        ingredient = ingredient_map.get(line.get('ingredient_id'))
        unit_cost = line.get('unit_cost') or (ingredient.get('unit_cost') if ingredient else 0)
        quantity = line.get('quantity', 1)
        waste_pct = line.get('waste_pct', 0)
        line_cost = unit_cost * quantity * (1 + waste_pct / 100)
        total_cost += line_cost

        enriched_lines.append({
            'ingredient_id': line.get('ingredient_id'),
            'ingredient_name': line.get('ingredient_name') or (ingredient.get('name') if ingredient else 'Unknown'),
            'unit': line.get('unit') or (ingredient.get('unit') if ingredient else ''),
            'quantity': quantity,
            'unit_cost': round(unit_cost, 2),
            'waste_pct': waste_pct,
            'line_cost': round(line_cost, 2),
            'supplier': ingredient.get('supplier') if ingredient else None
        })

    recipe['cost_breakdown'] = enriched_lines
    recipe['total_cost'] = round(total_cost, 2)
    selling_price = recipe.get('selling_price', 0)
    recipe['gp_percentage'] = round(((selling_price - total_cost) / selling_price * 100), 1) if selling_price else 0
    recipe['ingredient_count'] = len(enriched_lines)
    return recipe


@router.post("/fnb/recipes")
async def create_recipe(recipe_data: dict, current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_sales")),  # v99 DW
):
    ingredient_map, _ = await _get_ingredient_map(current_user.tenant_id)

    recipe = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'dish_name': recipe_data.get('recipe_name') or recipe_data.get('dish_name', 'Unnamed Recipe'),
        'category': recipe_data.get('category', 'general'),
        'portion_size': recipe_data.get('portion_size', '1 portion'),
        'preparation_time': recipe_data.get('preparation_time', 15),
        'ingredients': recipe_data.get('ingredients', []),
        'selling_price': recipe_data.get('selling_price', 0),
        'active': True,
        'created_at': datetime.now(UTC).isoformat()
    }

    recipe = _enrich_recipe_cost(recipe, ingredient_map)
    await db.recipes.insert_one(recipe)

    return {
        'success': True,
        'recipe_id': recipe['id'],
        'gp_percentage': recipe['gp_percentage'],
        'total_cost': recipe['total_cost']
    }


@router.get("/fnb/recipes")
async def get_recipes(current_user: User = Depends(get_current_user)):
    ingredient_map, _ = await _get_ingredient_map(current_user.tenant_id)
    recipes = await db.recipes.find(
        {'tenant_id': current_user.tenant_id, 'active': True},
        {'_id': 0}
    ).sort('dish_name', 1).to_list(200)

    enriched = [_enrich_recipe_cost(dict(recipe), ingredient_map) for recipe in recipes]
    avg_gp = sum([r.get('gp_percentage', 0) for r in enriched]) / len(enriched) if enriched else 0
    avg_food_cost = sum([r.get('total_cost', 0) for r in enriched]) / len(enriched) if enriched else 0

    return {
        'recipes': enriched,
        'total': len(enriched),
        'avg_gp': round(avg_gp, 1),
        'avg_food_cost': round(avg_food_cost, 2)
    }


@router.get("/fnb/recipes/{recipe_id}")
async def get_recipe(recipe_id: str, current_user: User = Depends(get_current_user)):
    recipe = await db.recipes.find_one({'tenant_id': current_user.tenant_id, 'id': recipe_id}, {'_id': 0})
    if not recipe:
        raise HTTPException(status_code=404, detail="Recipe not found")
    ingredient_map, _ = await _get_ingredient_map(current_user.tenant_id)
    recipe = _enrich_recipe_cost(recipe, ingredient_map)
    return {'recipe': recipe}


@router.put("/fnb/recipes/{recipe_id}")
async def update_recipe(recipe_id: str, recipe_data: dict, current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_sales")),  # v96 DW
):
    recipe = await db.recipes.find_one({'tenant_id': current_user.tenant_id, 'id': recipe_id})
    if not recipe:
        raise HTTPException(status_code=404, detail="Recipe not found")

    updated = {**recipe, **recipe_data}
    ingredient_map, _ = await _get_ingredient_map(current_user.tenant_id)
    updated = _enrich_recipe_cost(updated, ingredient_map)

    await db.recipes.update_one(
        {'tenant_id': current_user.tenant_id, 'id': recipe_id},
        {'$set': updated}
    )
    updated.pop('_id', None)
    return {'success': True, 'recipe': updated}

@router.post("/fnb/beo")
async def create_beo(beo_data: dict, current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_sales")),  # v99 DW
):
    beo = {
        'id': str(uuid.uuid4()), 'tenant_id': current_user.tenant_id,
        **beo_data, 'status': 'confirmed',
        'created_at': datetime.now(UTC).isoformat()
    }
    await db.banquet_event_orders.insert_one(beo)
    return {'success': True, 'beo_id': beo['id'], 'message': 'BEO olusturuldu'}

# NOTE: Kitchen Display System endpoints moved to pos_fnb_router.py


@router.get("/fnb/ingredients")
async def list_ingredients(current_user: User = Depends(get_current_user)):
    ingredients = await db.ingredients.find(
        {'tenant_id': current_user.tenant_id},
        {'_id': 0}
    ).sort('name', 1).to_list(500)

    low_stock = [ing for ing in ingredients if ing.get('current_stock', 0) <= ing.get('reorder_point', 0)]
    total_value = sum((ing.get('current_stock', 0) * ing.get('unit_cost', 0)) for ing in ingredients)

    categories = {}
    for ing in ingredients:
        cat = ing.get('category', 'other')
        categories.setdefault(cat, 0)
        categories[cat] += 1

    return {
        'ingredients': ingredients,
        'summary': {
            'total_items': len(ingredients),
            'low_stock': len(low_stock),
            'inventory_value': round(total_value, 2),
            'categories': categories
        }
    }


@router.post("/fnb/ingredients")
async def add_ingredient(ing_data: dict, current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_sales")),  # v99 DW
):
    ingredient = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'name': ing_data['name'],
        'category': ing_data.get('category', 'general'),
        'unit': ing_data.get('unit', 'kg'),
        'current_stock': ing_data.get('current_stock', 0),
        'par_level': ing_data.get('par_level', 0),
        'reorder_point': ing_data.get('reorder_point', 0),
        'unit_cost': ing_data.get('unit_cost', 0),
        'supplier': ing_data.get('supplier'),
        'last_order_date': ing_data.get('last_order_date'),
        'created_at': datetime.now(UTC).isoformat()
    }
    await db.ingredients.insert_one(ingredient)
    ingredient.pop('_id', None)
    return {'success': True, 'ingredient': ingredient}


@router.put("/fnb/ingredients/{ingredient_id}")
async def update_ingredient(ingredient_id: str, ing_data: dict, current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_sales")),  # v99 DW
):
    ingredient = await db.ingredients.find_one({'tenant_id': current_user.tenant_id, 'id': ingredient_id})
    if not ingredient:
        raise HTTPException(status_code=404, detail="Ingredient not found")

    updated = {**ingredient, **ing_data}
    await db.ingredients.update_one(
        {'tenant_id': current_user.tenant_id, 'id': ingredient_id},
        {'$set': updated}
    )
    updated.pop('_id', None)
    return {'success': True, 'ingredient': updated}


# ============= Departments / Positions (HR settings) =============

class DepartmentPayload(BaseModel):
    name: str = Field(..., min_length=1, max_length=80)
    code: str | None = Field(None, max_length=40)
    description: str | None = Field(None, max_length=300)


class PositionPayload(BaseModel):
    title: str = Field(..., min_length=1, max_length=120)
    # v2 HR (Task #262): department FK ZORUNLU. Pozisyon dept'siz yaratılırsa
    # RBAC scope'u (Department Manager) ve raporlama tutarsız kalır; KVKK
    # forensic için master-data referans bütünlüğü gerekli.
    department: str = Field(..., min_length=1, max_length=80)
    default_hourly_rate: float | None = Field(None, ge=0, le=100000)


@router.get("/hr/departments")
async def list_departments(
    include_inactive: bool = False,
    current_user: User = Depends(get_current_user),
):
    """Departman master data. Varsayılan: yalnızca aktif. `include_inactive=true`
    pasif kayıtları da döner (settings ekranı için)."""
    q: dict = {'tenant_id': current_user.tenant_id}
    if not include_inactive:
        # Eski kayıtlarda `active` alanı yok → varsayılan aktif say (yok ya da True).
        q['$or'] = [{'active': {'$exists': False}}, {'active': True}]
    items = await db.hr_departments.find(q, {'_id': 0}).sort('name', 1).to_list(200)
    return {'items': items, 'total': len(items)}


@router.post("/hr/departments")
async def create_department(
    payload: DepartmentPayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    code = (payload.code or payload.name.lower().replace(' ', '_'))[:40]
    if await db.hr_departments.find_one(
        {'tenant_id': current_user.tenant_id, 'code': code}
    ):
        raise HTTPException(status_code=409, detail="Bu departman kodu zaten var")
    item = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'name': payload.name,
        'code': code,
        'description': payload.description,
        'active': True,
        'created_at': datetime.now(UTC).isoformat(),
    }
    await db.hr_departments.insert_one(item)
    item.pop('_id', None)
    await _audit(
        current_user, "hr.department.create", "hr_department", item['id'],
        f"Departman oluşturuldu: {item['name']} (kod={code})",
        after=item, severity="info",
    )
    return {'success': True, 'department': item}


@router.delete("/hr/departments/{dept_id}")
async def delete_department(
    dept_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    """Soft-delete (pasifleştirme). Aktif personeli olan departman silinemez.

    v2 Foundation guard: dept_code/dept_name eşleşen aktif staff_members veya
    role→dept eşlemesi ile uyumlu aktif kullanıcı varsa 409 atılır. Pasifleşme
    sonrası kayıt listede görünmez (`include_inactive=true` ile geri gelir).
    """
    tid = current_user.tenant_id
    existing = await db.hr_departments.find_one({'tenant_id': tid, 'id': dept_id}, {'_id': 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Departman bulunamadı")
    dept_codes = {existing.get('code'), existing.get('name')}
    dept_codes.discard(None)
    active_count = await db.staff_members.count_documents({
        'tenant_id': tid,
        'department': {'$in': list(dept_codes)},
        '$or': [{'active': {'$exists': False}}, {'active': True}],
    }) if dept_codes else 0
    if active_count > 0:
        raise HTTPException(
            status_code=409,
            detail=f"Departmanda {active_count} aktif personel var. Önce personeli başka departmana taşıyın veya ayrılış işleyin.",
        )
    await db.hr_departments.update_one(
        {'tenant_id': tid, 'id': dept_id},
        {'$set': {'active': False, 'deactivated_at': datetime.now(UTC).isoformat()}},
    )
    await _audit(
        current_user, "hr.department.deactivate", "hr_department", dept_id,
        f"Departman pasifleştirildi: {existing.get('name')}",
        before=existing, severity="warning",
    )
    return {'success': True, 'soft_deleted': True}


@router.post("/hr/departments/sync-from-staff")
async def sync_departments_from_staff(
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    """Personelde geçen serbest departman string'lerinden master data oluştur.

    Idempotent: mevcut `code` ile çakışan kayıtlar atlanır. `staff_members` +
    `users` (role→dept eşlemesi) taranır. UI'daki "Personelden Senkronize Et"
    aksiyonu bunu çağırır.
    """
    tid = current_user.tenant_id
    # Mevcut master data code seti.
    existing = await db.hr_departments.find(
        {'tenant_id': tid}, {'_id': 0, 'code': 1, 'name': 1},
    ).to_list(500)
    existing_codes = {(d.get('code') or '').lower() for d in existing if d.get('code')}
    existing_codes |= {(d.get('name') or '').lower() for d in existing if d.get('name')}
    # Staff'tan benzersiz departmanlar.
    pipeline = [
        {'$match': {'tenant_id': tid, 'department': {'$type': 'string', '$ne': ''}}},
        {'$group': {'_id': '$department', 'count': {'$sum': 1}}},
    ]
    seen: dict[str, int] = {}
    async for row in db.staff_members.aggregate(pipeline):
        key = (row.get('_id') or '').strip()
        if key:
            seen[key] = (seen.get(key, 0) + int(row.get('count', 0)))
    # Users-derived role'lerden (rol→dept) ekle.
    role_to_dept = {
        'housekeeping': 'housekeeping',
        'front_desk': 'front_desk',
        'supervisor': 'management',
        'finance': 'finance',
        'sales': 'sales',
        'admin': 'management',
    }
    async for u in db.users.find(
        {'tenant_id': tid, 'is_active': True, 'role': {'$in': list(role_to_dept.keys())}},
        {'_id': 0, 'role': 1},
    ):
        d = role_to_dept.get(u.get('role'))
        if d:
            seen[d] = seen.get(d, 0) + 1
    created: list[dict] = []
    now_iso = datetime.now(UTC).isoformat()
    for code, count in seen.items():
        if code.lower() in existing_codes:
            continue
        item = {
            'id': str(uuid.uuid4()),
            'tenant_id': tid,
            'name': code,
            'code': code.lower().replace(' ', '_')[:40],
            'description': f"Personel verisinden senkronize edildi (sayı={count})",
            'active': True,
            'created_at': now_iso,
            'source': 'staff_sync',
        }
        await db.hr_departments.insert_one(item)
        item.pop('_id', None)
        created.append(item)
    if created:
        await _audit(
            current_user, "hr.department.sync_from_staff", "hr_department", "",
            f"{len(created)} departman personelden senkronize edildi",
            after={'created_codes': [c['code'] for c in created]},
            severity="info",
        )
    return {'success': True, 'created_count': len(created), 'created': created, 'scanned_unique': len(seen)}


@router.get("/hr/positions")
async def list_positions(
    include_inactive: bool = False,
    department: str | None = None,
    current_user: User = Depends(get_current_user),
):
    q: dict = {'tenant_id': current_user.tenant_id}
    if not include_inactive:
        q['$or'] = [{'active': {'$exists': False}}, {'active': True}]
    if department:
        q['department'] = department
    items = await db.hr_positions.find(q, {'_id': 0}).sort('title', 1).to_list(300)
    return {'items': items, 'total': len(items)}


@router.post("/hr/positions")
async def create_position(
    payload: PositionPayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    # v2 Foundation: pozisyon departmana bağlanmalı (FK semantiği).
    if payload.department:
        dept = await db.hr_departments.find_one({
            'tenant_id': current_user.tenant_id,
            '$or': [{'code': payload.department}, {'name': payload.department}],
            '$and': [{'$or': [{'active': {'$exists': False}}, {'active': True}]}],
        })
        if not dept:
            raise HTTPException(
                status_code=400,
                detail=f"Departman '{payload.department}' master data'da bulunamadı. Önce departmanı oluşturun veya 'Personelden Senkronize Et' aksiyonunu çalıştırın.",
            )
    item = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'title': payload.title,
        'department': payload.department,
        'default_hourly_rate': payload.default_hourly_rate,
        'active': True,
        'created_at': datetime.now(UTC).isoformat(),
    }
    await db.hr_positions.insert_one(item)
    item.pop('_id', None)
    await _audit(
        current_user, "hr.position.create", "hr_position", item['id'],
        f"Pozisyon oluşturuldu: {item['title']} (dept={payload.department})",
        after=item, severity="info",
    )
    return {'success': True, 'position': item}


@router.delete("/hr/positions/{pos_id}")
async def delete_position(
    pos_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    """Soft-delete (pasifleştirme). Aktif personeli olan pozisyon silinemez."""
    tid = current_user.tenant_id
    existing = await db.hr_positions.find_one({'tenant_id': tid, 'id': pos_id}, {'_id': 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Pozisyon bulunamadı")
    active_count = await db.staff_members.count_documents({
        'tenant_id': tid,
        'position': existing.get('title'),
        '$or': [{'active': {'$exists': False}}, {'active': True}],
    })
    if active_count > 0:
        raise HTTPException(
            status_code=409,
            detail=f"Pozisyonda {active_count} aktif personel var. Önce personeli başka pozisyona taşıyın veya ayrılış işleyin.",
        )
    await db.hr_positions.update_one(
        {'tenant_id': tid, 'id': pos_id},
        {'$set': {'active': False, 'deactivated_at': datetime.now(UTC).isoformat()}},
    )
    await _audit(
        current_user, "hr.position.deactivate", "hr_position", pos_id,
        f"Pozisyon pasifleştirildi: {existing.get('title')}",
        before=existing, severity="warning",
    )
    return {'success': True, 'soft_deleted': True}


# ============= Staff CRUD (POST/GET/PUT/DELETE/profile) =============

def _scrub_encrypted(value):
    if isinstance(value, str) and value.startswith('aes256gcm:'):
        return ''
    return value or ''


@router.post("/hr/staff")
async def add_staff_member(
    staff_data: dict,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    """Yeni personel ekle.

    Zorunlu: name. Diğer alanlar opsiyonel; UI form akışı eksik bilgiyle
    de minimum personel kaydı oluşturabilsin diye gevşetildi.
    Ek finansal alanlar (hourly_rate, monthly_hours, annual_leave_entitlement)
    payroll ve izin bakiyesi hesaplarında kullanılır.
    """
    if not staff_data.get('name'):
        raise HTTPException(status_code=400, detail="Personel adı zorunludur")
    staff = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'name': staff_data['name'],
        'email': staff_data.get('email'),
        'phone': staff_data.get('phone'),
        'department': staff_data.get('department'),
        'position': staff_data.get('position'),
        'hire_date': staff_data.get('hire_date'),
        'employment_type': staff_data.get('employment_type', 'full_time'),
        'hourly_rate': staff_data.get('hourly_rate'),
        'monthly_hours': staff_data.get('monthly_hours'),
        'annual_leave_entitlement': staff_data.get('annual_leave_entitlement', 14),
        'performance_score': 0.0,
        'active': True,
        'created_at': datetime.now(UTC).isoformat(),
    }
    await db.staff_members.insert_one(staff)
    await _audit(
        current_user, "hr.staff.create", "staff_member", staff['id'],
        f"Personel oluşturuldu: {staff['name']} (dept={staff.get('department')}, pozisyon={staff.get('position')})",
        after={k: v for k, v in staff.items() if k not in _PII_SALARY_FIELDS},
        severity="info",
    )
    return {'success': True, 'staff_id': staff['id']}


@router.get("/hr/staff")
async def get_staff_list(
    department: str | None = None,
    source: Literal['hr', 'users', 'all'] = 'hr',
    include_inactive: bool = False,
    employment_type: str | None = None,
    hire_date_from: str | None = None,
    hire_date_to: str | None = None,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """Personel listesi — v2 Foundation (Task #262) refactor.

    `source` parametresi UI sekme akışını kontrol eder:
      - `hr` (varsayılan): yalnızca `staff_members` koleksiyonu — gerçek
        HR-managed personel; sistem kullanıcıları DAHIL DEĞİL.
      - `users`: yalnızca `users` koleksiyonundan role bazlı türetilmiş kayıt
        (RBAC test login'leri / tenant admin'leri — "Sistem Kullanıcıları" tab).
      - `all`: ikisinin birleşimi (geriye-uyum — v1 davranışı).

    Department-Manager scope: kullanıcının `assigned_department` alanı varsa
    ve `manage_hr` perm'i yoksa, listeleme bu departmana otomatik daraltılır.

    PII maskeleme: response serializer rol-bazlı maskelemeyi uygular
    (`view_hr_payroll`/`manage_hr` perm yoksa phone/IBAN/maaş/TC maskelenir).
    Self-service: kendi kayıt id'siyle eşleşen satırlar unmask edilir.
    """
    tid = current_user.tenant_id
    role_to_dept = {
        'housekeeping': 'housekeeping',
        'front_desk': 'front_desk',
        'supervisor': 'management',
        'finance': 'finance',
        'sales': 'sales',
        'admin': 'management',
    }

    # Department-Manager scope override (kullanıcı manage_hr'a sahip değilse).
    scope_dept = _user_assigned_department(current_user)
    if scope_dept and not _user_has_hr_op(current_user, "manage_hr"):
        department = scope_dept

    explicit: list = []
    if source in ('hr', 'all'):
        explicit_query: dict = {'tenant_id': tid}
        if not include_inactive:
            explicit_query['active'] = True
        if department:
            explicit_query['department'] = department
        if employment_type:
            explicit_query['employment_type'] = employment_type
        if hire_date_from or hire_date_to:
            rng: dict = {}
            if hire_date_from:
                rng['$gte'] = hire_date_from
            if hire_date_to:
                rng['$lte'] = hire_date_to
            explicit_query['hire_date'] = rng
        explicit = await db.staff_members.find(explicit_query, {'_id': 0}).to_list(500)

    seen_emails = {(s.get('email') or '').lower() for s in explicit if s.get('email')}

    derived: list = []
    if source in ('users', 'all'):
        user_query: dict = {
            'tenant_id': tid,
            'is_active': True,
            'role': {'$in': list(role_to_dept.keys())},
        }
        if department:
            roles_in_dept = [r for r, d in role_to_dept.items() if d == department or r == department]
            user_query['role'] = {'$in': roles_in_dept or ['__none__']}

        user_projection = {
            '_id': 0,
            'id': 1, 'name': 1, 'email': 1, 'phone': 1,
            'role': 1, 'created_at': 1,
        }
        cursor = db.users.find(user_query, user_projection).limit(500)
        async for u in cursor:
            email_raw = (u.get('email') or '')
            em = email_raw.lower()
            # `all` modu için dedup; `users` modunda dedup yok (kasıtlı).
            if source == 'all' and em and em in seen_emails:
                continue
            role = u.get('role')
            email_clean = _scrub_encrypted(email_raw)
            phone_clean = _scrub_encrypted(u.get('phone'))
            derived.append({
                'id': u.get('id'),
                'tenant_id': tid,
                'name': u.get('name') or email_clean or 'Personel',
                'email': email_clean,
                'phone': phone_clean,
                'department': role_to_dept.get(role, role or 'other'),
                'position': role or 'staff',
                'hire_date': (u.get('created_at') or '')[:10],
                'employment_type': 'full_time',
                'performance_score': 0.0,
                'active': True,
                'derived_from': 'users',
            })
            if em:
                seen_emails.add(em)

    combined = explicit + derived
    # Rol-bazlı PII maskeleme (post-query, response serializer aşamasında).
    self_id = str(getattr(current_user, 'id', '') or '')
    self_email = str(getattr(current_user, 'email', '') or '')
    masked = [_mask_hr_pii(s, current_user, self_id=self_id, self_email=self_email) for s in combined]
    return {'staff': masked, 'total': len(masked), 'source': source}


@router.get("/hr/system-users")
async def get_system_users(
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """Sistem kullanıcıları (RBAC test login'leri / tenant admin'leri).

    UI'daki "Sistem Kullanıcıları" sekmesi bunu çağırır — `/hr/staff?source=users`
    ile aynı veri ama semantik olarak ayrı endpoint (HR Manager yetkisi yetmez
    daha kısıtlı yapmak istenirse buradan değiştirilebilir).
    """
    # Implementation: delegate to /hr/staff?source=users semantics.
    return await get_staff_list(
        department=None,
        source='users',
        include_inactive=False,
        employment_type=None,
        hire_date_from=None,
        hire_date_to=None,
        current_user=current_user,
        _perm=None,
    )


@router.get("/hr/performance/{staff_id}")
async def get_staff_performance_summary(
    staff_id: str,
    current_user: User = Depends(get_current_user),
):
    """Personel performans özeti (son 10 review + ortalama puan).

    RBAC: performans notları hassas → SADECE `manage_hr` perm'iyle (HR Admin /
    Supervisor / super_admin) görüntülenebilir. Self-service istisnası
    KORUNUR (kendi performansını görebilir). Finance (`view_hr`) erişemez.
    """
    staff = await _verify_staff_in_tenant(staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    _authorize_staff_access(staff, current_user, require_manage=True)
    reviews = await db.performance_reviews.find({
        'staff_id': staff_id,
        'tenant_id': current_user.tenant_id,
    }, {'_id': 0}).sort('reviewed_at', -1).to_list(10)

    avg_score = sum([r.get('overall_score', 0) for r in reviews]) / len(reviews) if reviews else 0

    return {
        'staff_id': staff_id,
        'recent_reviews': reviews,
        'avg_performance_score': round(avg_score, 2),
        'total_reviews': len(reviews),
    }


class StaffUpdatePayload(BaseModel):
    name: str | None = Field(None, max_length=200)
    email: str | None = Field(None, max_length=200)
    phone: str | None = Field(None, max_length=40)
    department: str | None = Field(None, max_length=80)
    position: str | None = Field(None, max_length=120)
    hire_date: str | None = Field(None, pattern=r'^\d{4}-\d{2}-\d{2}$')
    employment_type: Literal[
        'full_time', 'part_time', 'seasonal', 'contract', 'intern'
    ] | None = None
    hourly_rate: float | None = Field(None, ge=0, le=100000)
    monthly_hours: float | None = Field(None, ge=0, le=400)
    annual_leave_entitlement: int | None = Field(None, ge=0, le=365)


@router.put("/hr/staff/{staff_id}")
async def update_staff_member(
    staff_id: str,
    payload: StaffUpdatePayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    """Personel güncelle.

    HR-managed (`staff_members`): tüm alanlar güncellenebilir.
    Users-derived (kullanıcı kaydından türetilmiş): sadece **iletişim**
    alanları (email/phone/name) güncellenebilir; rol/departman/maaş gibi
    alanlar Kullanıcı Yönetimi üzerinden değiştirilir.
    """
    existing = await db.staff_members.find_one(
        {'tenant_id': current_user.tenant_id, 'id': staff_id}
    )
    if existing:
        update = {
            k: v for k, v in payload.model_dump(exclude_unset=True).items()
            if v is not None
        }
        if not update:
            return {'success': True, 'updated_fields': 0}
        update['updated_at'] = datetime.now(UTC).isoformat()
        await db.staff_members.update_one(
            {'tenant_id': current_user.tenant_id, 'id': staff_id},
            {'$set': update}
        )
        # Audit: salary alanı değiştiyse severity=warning, diğerleri info.
        sev = "warning" if any(k in update for k in _PII_SALARY_FIELDS) else "info"
        await _audit(
            current_user, "hr.staff.update", "staff_member", staff_id,
            f"Personel güncellendi (alan sayısı={len(update) - 1})",
            before={k: existing.get(k) for k in update.keys() if k != 'updated_at'},
            after={k: v for k, v in update.items() if k not in _PII_SALARY_FIELDS},
            severity=sev,
        )
        return {'success': True, 'updated_fields': len(update) - 1, 'source': 'staff'}

    # Users-derived path — sadece iletişim alanları
    user = await db.users.find_one(
        {'tenant_id': current_user.tenant_id, 'id': staff_id, 'is_active': True},
        {'_id': 0, 'id': 1}
    )
    if not user:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")

    raw = payload.model_dump(exclude_unset=True)
    allowed = {k: v for k, v in raw.items() if k in {'name', 'email', 'phone'} and v is not None}
    if not allowed:
        raise HTTPException(
            status_code=400,
            detail="Türetilmiş personel için yalnızca isim, e-posta ve telefon güncellenebilir."
        )
    allowed['updated_at'] = datetime.now(UTC).isoformat()
    await db.users.update_one(
        {'tenant_id': current_user.tenant_id, 'id': staff_id},
        {'$set': allowed}
    )
    return {'success': True, 'updated_fields': len(allowed) - 1, 'source': 'users'}


@router.delete("/hr/staff/{staff_id}")
async def delete_staff_member(
    staff_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    """İşten ayrılış (soft deactivate). Personel ASLA hard-delete edilmez —
    bordro / devam / izin geçmişi korunur, listeden çıkar.

    HR-managed: `staff_members.active=False`.
    Users-derived: `users.is_active=False`.
    """
    now_iso = datetime.now(UTC).isoformat()
    res = await db.staff_members.update_one(
        {'tenant_id': current_user.tenant_id, 'id': staff_id},
        {'$set': {
            'active': False,
            'deactivated_at': now_iso,
        }}
    )
    if res.matched_count > 0:
        await _audit(
            current_user, "hr.staff.deactivate", "staff_member", staff_id,
            "Personel pasifleştirildi (HR-managed)",
            severity="warning",
        )
        return {'success': True, 'source': 'staff', 'deactivated_at': now_iso}

    # Users-derived ayrılış
    res2 = await db.users.update_one(
        {'tenant_id': current_user.tenant_id, 'id': staff_id, 'is_active': True},
        {'$set': {
            'is_active': False,
            'deactivated_at': now_iso,
        }}
    )
    if res2.matched_count == 0:
        raise HTTPException(status_code=404, detail="Personel bulunamadı veya zaten ayrılmış")
    await _audit(
        current_user, "hr.staff.deactivate", "user_account", staff_id,
        "Kullanıcı hesabı pasifleştirildi (users-derived staff)",
        severity="warning",
    )
    return {'success': True, 'source': 'users', 'deactivated_at': now_iso}


@router.get("/hr/staff/{staff_id}/profile")
async def get_staff_profile(
    staff_id: str,
    current_user: User = Depends(get_current_user),
):
    """Aggregate profil: kişi + son 30g devam + izinler + performans + bordro + vardiya.

    RBAC: tenant + assigned_department + self-service gate (`_authorize_staff_access`).
    Performans bölümü SADECE `manage_hr` perm'iyle döner (Finance gizlenir).
    PII alanları rol-bazlı maskelenir (`_mask_hr_pii`).
    """
    staff = await _verify_staff_in_tenant(staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    _authorize_staff_access(staff, current_user)
    today = _today_local()
    start_30 = today - timedelta(days=30)

    attendance = await db.attendance_records.find({
        'tenant_id': current_user.tenant_id,
        'staff_id': staff_id,
        'date': {'$gte': start_30.isoformat(), '$lte': today.isoformat()},
    }, {'_id': 0}).sort('date', -1).to_list(100)

    leaves = await db.leave_requests.find({
        'tenant_id': current_user.tenant_id, 'staff_id': staff_id,
    }, {'_id': 0}).sort('created_at', -1).to_list(100)

    reviews = await db.performance_reviews.find({
        'tenant_id': current_user.tenant_id, 'staff_id': staff_id,
    }, {'_id': 0}).sort('reviewed_at', -1).to_list(50)

    payroll = await db.payroll_records.find({
        'tenant_id': current_user.tenant_id, 'staff_id': staff_id,
    }, {'_id': 0}).sort('period_month', -1).to_list(12)

    shifts = await db.shift_schedules.find({
        'tenant_id': current_user.tenant_id, 'staff_id': staff_id,
        'shift_date': {'$gte': today.isoformat()},
    }, {'_id': 0}).sort('shift_date', 1).to_list(20)

    total_hours = round(sum(r.get('total_hours', 0) for r in attendance), 2)
    days_present = len({r['date'] for r in attendance if r.get('clock_out')})
    avg_score = (
        round(sum(r.get('overall_score', 0) for r in reviews) / len(reviews), 2)
        if reviews else 0
    )

    balance = await db.leave_balances.find_one({
        'tenant_id': current_user.tenant_id,
        'staff_id': staff_id,
        'year': today.year,
    }, {'_id': 0})

    # v2 Foundation: rol-bazlı PII maskeleme — staff kayıt + bordro satırları.
    self_id = str(getattr(current_user, 'id', '') or '')
    self_email = str(getattr(current_user, 'email', '') or '')
    staff_masked = _mask_hr_pii(staff, current_user, self_id=self_id, self_email=self_email) or staff
    payroll_masked = [
        _mask_hr_pii(p, current_user, self_id=self_id, self_email=self_email) or p for p in payroll
    ]
    # Performans notları hassas — Finance (`view_hr`) gizlenir, sadece manage_hr
    # veya kendi kaydı görebilir. Aksi halde içerik boşaltılır (UI tab gizler).
    is_self = bool(
        (self_id and str(staff.get('id') or '') == self_id) or
        (self_email and (staff.get('email') or '').strip().lower() == self_email)
    )
    can_view_perf = _user_has_hr_op(current_user, "manage_hr") or is_self
    perf_section = {
        'items': reviews if can_view_perf else [],
        'avg_score': avg_score if can_view_perf else 0,
        'total': len(reviews) if can_view_perf else 0,
        'redacted': not can_view_perf,
    }
    return {
        'staff': staff_masked,
        'attendance': {
            'records': attendance,
            'total_hours_30d': total_hours,
            'days_present_30d': days_present,
        },
        'leaves': {
            'items': leaves,
            'total': len(leaves),
            'pending': sum(1 for leave in leaves if leave.get('status') == 'pending'),
        },
        'leave_balance': balance,
        'performance': perf_section,
        'payroll': {
            'recent': payroll_masked,
            'count': len(payroll),
        },
        'upcoming_shifts': shifts,
    }


# ============= Leave Balance (yıllık izin bakiyesi) =============

class LeaveBalancePayload(BaseModel):
    staff_id: str = Field(..., min_length=1)
    year: int = Field(..., ge=2020, le=2100)
    annual_entitlement: int = Field(..., ge=0, le=365)
    carry_over: int | None = Field(0, ge=0, le=365)
    sick_entitlement: int | None = Field(None, ge=0, le=365)


@router.get("/hr/leave-balance/{staff_id}")
async def get_leave_balance(
    staff_id: str,
    year: int | None = None,
    current_user: User = Depends(get_current_user),
):
    yr = year or _today_local().year
    staff = await _verify_staff_in_tenant(staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    _authorize_staff_access(staff, current_user)
    balance = await db.leave_balances.find_one({
        'tenant_id': current_user.tenant_id,
        'staff_id': staff_id,
        'year': yr,
    }, {'_id': 0})
    # İş Kanunu m.53 default: 14 gün yıllık ücretli izin
    annual_ent = balance.get('annual_entitlement', 14) if balance else 14
    carry = balance.get('carry_over', 0) if balance else 0
    sick_ent = balance.get('sick_entitlement', 5) if balance else 5

    approved = await db.leave_requests.find({
        'tenant_id': current_user.tenant_id,
        'staff_id': staff_id,
        'status': 'approved',
        'start_date': {'$gte': f'{yr}-01-01', '$lte': f'{yr}-12-31'},
    }, {'_id': 0}).to_list(500)
    used_annual = sum(
        leave.get('total_days', 0) for leave in approved
        if leave.get('leave_type') == 'annual'
    )
    used_sick = sum(
        leave.get('total_days', 0) for leave in approved
        if leave.get('leave_type') == 'sick'
    )
    total_annual = annual_ent + carry
    return {
        'staff_id': staff_id,
        'staff_name': staff.get('name'),
        'year': yr,
        'configured': bool(balance),
        'annual': {
            'entitlement': annual_ent,
            'carry_over': carry,
            'total': total_annual,
            'used': used_annual,
            'remaining': max(0, total_annual - used_annual),
        },
        'sick': {
            'entitlement': sick_ent,
            'used': used_sick,
            'remaining': max(0, sick_ent - used_sick),
        },
    }


@router.post("/hr/leave-balance")
async def set_leave_balance(
    payload: LeaveBalancePayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    staff = await _verify_staff_in_tenant(payload.staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    doc = {
        'tenant_id': current_user.tenant_id,
        'staff_id': payload.staff_id,
        'year': payload.year,
        'annual_entitlement': payload.annual_entitlement,
        'carry_over': payload.carry_over or 0,
        'sick_entitlement': (
            payload.sick_entitlement if payload.sick_entitlement is not None else 5
        ),
        'updated_at': datetime.now(UTC).isoformat(),
    }
    await db.leave_balances.update_one(
        {
            'tenant_id': current_user.tenant_id,
            'staff_id': payload.staff_id,
            'year': payload.year,
        },
        {'$set': doc},
        upsert=True,
    )
    return {'success': True}


# ============= Shifts (vardiya planlaması) =============
#
# Task #254 (concurrency follow-up): vardiya çakışma koruması
# `shift_schedule_locks` koleksiyonuna delege edilmiştir. Bu koleksiyon
# (tenant_id, staff_id, shift_date) üzerinde UNIQUE compound index taşır
# (bkz. `backend/bootstrap/phases/d_perf.py`). Her dokümanın `intervals`
# array'i o güne ait AKTİF (cancelled/deleted DEĞİL) vardiya
# entry'lerini tutar (`{shift_id, start_time, end_time}`).
#
# Kontrat:
#   * create  → `_acquire_shift_lock_interval` atomik findOneAndUpdate +
#               upsert. Overlap varsa unique index DuplicateKeyError →
#               HTTPException(409).
#   * delete  → `_release_shift_lock_interval` $pull. Aksi halde silinmiş
#               vardiya hala yeni create'i bloke eder (false 409).
#   * swap    → from_staff lock entry'sini at, target_staff'a ata. Sıra:
#               önce hedef için acquire (çakışma → 409 + abort), sonra
#               shift_schedules update; update başarısızsa target'ten
#               release ile rollback; başarılıysa from_staff'tan release.
# Bu invariant'ı kıran her mutation `shift_schedule_locks` ile drift
# yaratır ve overlap enforcement sessizce eskir.


async def _acquire_shift_lock_interval(
    *,
    tenant_id: str,
    staff_id: str,
    shift_date: str,
    start_time: str,
    end_time: str,
    shift_id: str,
    now_iso: str,
) -> None:
    """Atomic overlap-rejecting reservation for one (staff, day) slot.

    Half-open interval rule: ``new.start < ex.end AND new.end > ex.start``.
    Boundary equality (09:00-13:00 + 13:00-17:00) is NOT overlap.

    Raises HTTPException(409) with a Turkish description if the slot
    overlaps an existing reserved interval. The DuplicateKeyError branch
    is what makes the concurrent case safe — without the unique index in
    `d_perf.py` this collapses back into a TOCTOU race.
    """
    lock_filter = {
        'tenant_id': tenant_id,
        'staff_id': staff_id,
        'shift_date': shift_date,
        '$or': [
            {'intervals': {'$exists': False}},
            {'intervals': {
                '$not': {'$elemMatch': {
                    'start_time': {'$lt': end_time},
                    'end_time': {'$gt': start_time},
                }}
            }},
        ],
    }
    lock_update = {
        '$setOnInsert': {
            'tenant_id': tenant_id,
            'staff_id': staff_id,
            'shift_date': shift_date,
            'created_at': now_iso,
        },
        '$push': {'intervals': {
            'shift_id': shift_id,
            'start_time': start_time,
            'end_time': end_time,
        }},
    }
    # `DuplicateKeyError` iki ayrı senaryoda fırlayabilir:
    #   (a) Gerçek overlap: doc zaten var, filtre overlap nedeniyle
    #       match etmedi → upsert mevcut key'e insert denedi → 409.
    #   (b) Doküman yaratım yarışı: doc henüz yokken iki eşzamanlı
    #       upsert; biri kazanır insert eder, diğeri (b) DuplicateKey
    #       alır AMA aslında overlap yoktu. Bu durumda retry'da doc
    #       artık var ve filtre overlap kontrolünü doğru yapar →
    #       overlap yoksa $push başarılı olur.
    # Bu yüzden bounded retry + post-DuplicateKey "gerçek overlap mı?"
    # doğrulaması yapıyoruz; sadece gerçek overlap'ta 409 atıyoruz.
    last_exc: DuplicateKeyError | None = None
    for _attempt in range(3):
        try:
            await db.shift_schedule_locks.find_one_and_update(
                lock_filter, lock_update, upsert=True,
            )
            return
        except DuplicateKeyError as e:
            last_exc = e
            existing_doc = await db.shift_schedule_locks.find_one(
                {
                    'tenant_id': tenant_id,
                    'staff_id': staff_id,
                    'shift_date': shift_date,
                },
                {'_id': 0, 'intervals': 1},
            )
            conflicting = None
            for iv in (existing_doc or {}).get('intervals') or []:
                if (
                    iv.get('start_time', '') < end_time
                    and iv.get('end_time', '') > start_time
                ):
                    conflicting = iv
                    break
            if conflicting:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"Bu personelin {shift_date} tarihinde "
                        f"{conflicting.get('start_time')}-"
                        f"{conflicting.get('end_time')} saatleri arası "
                        f"başka bir vardiyası var "
                        f"(id={conflicting.get('shift_id')}). Çakışan "
                        f"vardiya oluşturulamaz."
                    ),
                )
            # Overlap yok → creation race'iydi, retry.
            continue
    # 3 retry sonrası hâlâ DuplicateKeyError + overlap yok: patolojik.
    raise HTTPException(
        status_code=503,
        detail=(
            "Vardiya kilidi alınamadı, lütfen tekrar deneyin "
            f"(son hata: {last_exc})"
        ),
    )


async def _acquire_shift_lock_for_shift(
    *,
    tenant_id: str,
    staff_id: str,
    shift_date: str,
    start_time: str,
    end_time: str,
    crosses_midnight: bool,
    shift_id: str,
    now_iso: str,
) -> None:
    """Night-shift aware wrapper over `_acquire_shift_lock_interval`.

    Same-day shifts (crosses_midnight=False) acquire a single interval on
    `shift_date`. Overnight shifts (crosses_midnight=True) acquire TWO
    intervals atomically:
        - `shift_date`           → [start_time, '24:00')
        - `next_date`            → ['00:00', end_time)

    String lex compare is exact within a single day's HH:MM domain, so the
    existing `_acquire_shift_lock_interval` overlap math remains correct on
    each per-day lock document. If the second leg fails (overlap on the
    next day), the first leg is released to avoid orphan reservations.
    """
    if not crosses_midnight:
        await _acquire_shift_lock_interval(
            tenant_id=tenant_id,
            staff_id=staff_id,
            shift_date=shift_date,
            start_time=start_time,
            end_time=end_time,
            shift_id=shift_id,
            now_iso=now_iso,
        )
        return
    next_date = (
        date.fromisoformat(shift_date) + timedelta(days=1)
    ).isoformat()
    # Leg 1: günün geri kalanı (start_time → 24:00 sentinel).
    await _acquire_shift_lock_interval(
        tenant_id=tenant_id,
        staff_id=staff_id,
        shift_date=shift_date,
        start_time=start_time,
        end_time='24:00',
        shift_id=shift_id,
        now_iso=now_iso,
    )
    # Leg 2: ertesi günün başlangıcı (00:00 → end_time). Başarısız olursa
    # birinci leg'i geri al → orphan lock kalmaz.
    try:
        await _acquire_shift_lock_interval(
            tenant_id=tenant_id,
            staff_id=staff_id,
            shift_date=next_date,
            start_time='00:00',
            end_time=end_time,
            shift_id=shift_id,
            now_iso=now_iso,
        )
    except HTTPException:
        await _release_shift_lock_interval(
            tenant_id=tenant_id,
            staff_id=staff_id,
            shift_date=shift_date,
            shift_id=shift_id,
        )
        raise


async def _release_shift_lock_for_shift(
    *,
    tenant_id: str,
    staff_id: str,
    shift_date: str,
    crosses_midnight: bool,
    shift_id: str,
) -> None:
    """Release locks acquired by `_acquire_shift_lock_for_shift`.

    Idempotent: releasing a missing entry is a no-op. For overnight
    shifts, both the start-day and next-day lock entries are pulled.
    """
    await _release_shift_lock_interval(
        tenant_id=tenant_id,
        staff_id=staff_id,
        shift_date=shift_date,
        shift_id=shift_id,
    )
    if crosses_midnight:
        next_date = (
            date.fromisoformat(shift_date) + timedelta(days=1)
        ).isoformat()
        await _release_shift_lock_interval(
            tenant_id=tenant_id,
            staff_id=staff_id,
            shift_date=next_date,
            shift_id=shift_id,
        )


async def _release_shift_lock_interval(
    *,
    tenant_id: str,
    staff_id: str,
    shift_date: str,
    shift_id: str,
) -> None:
    """Remove the locked interval for a shift after delete/swap.

    Idempotent: missing lock doc or missing entry is a no-op. Best-effort
    — failure here must not block the primary mutation; the worst case is
    a stale lock entry that future creates report as overlap (operator
    can re-issue with a different time or delete the lock).
    """
    try:
        await db.shift_schedule_locks.update_one(
            {
                'tenant_id': tenant_id,
                'staff_id': staff_id,
                'shift_date': shift_date,
            },
            {'$pull': {'intervals': {'shift_id': shift_id}}},
        )
    except Exception:  # pragma: no cover - defensive, never block parent
        pass


class ShiftPayload(BaseModel):
    staff_id: str = Field(..., min_length=1)
    shift_date: str = Field(..., pattern=r'^\d{4}-\d{2}-\d{2}$')
    shift_type: Literal['morning', 'afternoon', 'evening', 'night', 'split'] = 'morning'
    start_time: str = Field(..., pattern=r'^\d{2}:\d{2}$')
    end_time: str = Field(..., pattern=r'^\d{2}:\d{2}$')
    crosses_midnight: bool = False
    notes: str | None = Field(None, max_length=300)


# Task #263 — leave↔shift bağı: onaylı (`approved` / `hr_approved`) izin
# günlerinde personele vardiya atanamaz. `_apply_leave_to_shifts` final
# onayda `shift_schedules` üstüne status='on_leave' satırı upsert eder;
# bu helper o satırları (veya klasik leave_requests) tarama yapar.
LEAVE_APPROVED_STATUSES = ('approved', 'hr_approved')


async def _staff_has_approved_leave_on(
    tenant_id: str, staff_id: str, date_iso: str,
) -> dict | None:
    """O gün için onaylı izin var mı? Varsa leave doc'unu döner, yoksa None."""
    doc = await db.leave_requests.find_one(
        {
            'tenant_id': tenant_id,
            'staff_id': staff_id,
            'status': {'$in': list(LEAVE_APPROVED_STATUSES)},
            'start_date': {'$lte': date_iso},
            'end_date': {'$gte': date_iso},
        },
        {'_id': 0, 'id': 1, 'leave_type': 1, 'start_date': 1, 'end_date': 1},
    )
    return doc


async def _staff_is_active(tenant_id: str, staff_id: str) -> bool:
    """Pasif (active=False) staff_members'a vardiya atanmaz. Users-türevli
    derived staff için is_active=True kabul edilir."""
    sm = await db.staff_members.find_one(
        {'id': staff_id, 'tenant_id': tenant_id},
        {'_id': 0, 'active': 1},
    )
    if sm:
        return sm.get('active', True) is not False
    # users türevli kayıt → _verify_staff_in_tenant zaten is_active=True filtreler.
    return True


def _shift_datetimes(shift_date: str, start_time: str, end_time: str,
                     crosses_midnight: bool) -> tuple[datetime, datetime]:
    """Compute (start_dt, end_dt) for a shift record. Naive UTC datetimes;
    only used for relative overlap math, not wall-clock conversions."""
    d = date.fromisoformat(shift_date)
    sh, sm = (int(x) for x in start_time.split(':'))
    eh, em = (int(x) for x in end_time.split(':'))
    start_dt = datetime(d.year, d.month, d.day, sh, sm)
    end_d = d + timedelta(days=1) if crosses_midnight else d
    end_dt = datetime(end_d.year, end_d.month, end_d.day, eh, em)
    return start_dt, end_dt


@router.post("/hr/shifts")
async def create_shift_v2(
    payload: ShiftPayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    staff = await _verify_staff_in_tenant(payload.staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")

    # Task #263: pasif personele vardiya atanamaz.
    if not await _staff_is_active(current_user.tenant_id, payload.staff_id):
        raise HTTPException(
            status_code=400,
            detail="Pasif personele vardiya atanamaz",
        )

    # Task #263: onaylı izin günlerinde vardiya atanamaz (lock acquire
    # öncesi, sahte 409 yaratmamak için).
    leave = await _staff_has_approved_leave_on(
        current_user.tenant_id, payload.staff_id, payload.shift_date,
    )
    if leave:
        raise HTTPException(
            status_code=409,
            detail=(
                f"{payload.shift_date} tarihinde {staff.get('name')} "
                f"onaylı izinde ({leave.get('leave_type')}, "
                f"{leave.get('start_date')} → {leave.get('end_date')}). "
                f"Vardiya atanamaz."
            ),
        )
    if payload.crosses_midnight:
        # Gece vardiyası ertesi güne sarktığı için sonraki günü de izinde
        # geçiriyorsa engellenir.
        next_iso = (
            date.fromisoformat(payload.shift_date) + timedelta(days=1)
        ).isoformat()
        leave_next = await _staff_has_approved_leave_on(
            current_user.tenant_id, payload.staff_id, next_iso,
        )
        if leave_next:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Gece vardiyası {next_iso} gününe sarkıyor ve o gün "
                    f"personel onaylı izinde."
                ),
            )

    # Task #255: Gece vardiyaları artık tek kayıtta planlanabilir
    # (`crosses_midnight=True`). Aksi halde tek-gün kontratı korunur
    # (end > start). crosses_midnight=True iken end > start mantıksızdır
    # (aynı gün biten vardiya overnight olamaz) → reddedilir.
    if payload.crosses_midnight:
        if payload.end_time >= payload.start_time:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"Gece vardiyası işaretlendi ama bitiş saati ({payload.end_time}) "
                    f"başlangıç saatinden ({payload.start_time}) önce olmalı "
                    f"(ör. 22:00 → 06:00). Aynı gün biten vardiyalarda gece "
                    f"vardiyası seçeneğini kapatın."
                ),
            )
    else:
        if payload.end_time <= payload.start_time:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"Vardiya bitiş saati ({payload.end_time}) başlangıç saatinden "
                    f"({payload.start_time}) sonra olmalı. Gece vardiyaları için "
                    f"'gece vardiyası (ertesi güne sarkar)' seçeneğini işaretleyin."
                ),
            )

    # Task #254 (concurrency lock) + #255 (overnight contract): atomic
    # overlap guard `_acquire_shift_lock_for_shift` helper'ına delege edildi.
    # crosses_midnight=True iken helper iki ayrı (start_date, next_date)
    # lock dokümanına yazar — string lex compare aynı gün içinde doğru
    # kalır, ertesi güne sarkan kısım next_date dokümanında ayrı interval
    # olarak kontrol edilir. Böylece TOCTOU race + datetime sınır geçişi
    # birlikte kapatılır. Helper overlap'ta HTTPException(409) atar.
    shift_id = str(uuid.uuid4())
    now_iso = datetime.now(UTC).isoformat()
    await _acquire_shift_lock_for_shift(
        tenant_id=current_user.tenant_id,
        staff_id=payload.staff_id,
        shift_date=payload.shift_date,
        start_time=payload.start_time,
        end_time=payload.end_time,
        crosses_midnight=payload.crosses_midnight,
        shift_id=shift_id,
        now_iso=now_iso,
    )
    d = date.fromisoformat(payload.shift_date)

    end_date = (d + timedelta(days=1)).isoformat() if payload.crosses_midnight \
        else payload.shift_date
    item = {
        'id': shift_id,
        'tenant_id': current_user.tenant_id,
        'staff_id': payload.staff_id,
        'staff_name': staff.get('name'),
        'shift_date': payload.shift_date,
        'shift_type': payload.shift_type,
        'start_time': payload.start_time,
        'end_time': payload.end_time,
        'crosses_midnight': payload.crosses_midnight,
        'end_date': end_date,
        'notes': payload.notes,
        'status': 'scheduled',
        'created_at': datetime.now(UTC).isoformat(),
    }
    # Task #254 follow-up: lock acquired ama insert başarısızsa lock
    # entry'sini geri al → orphan lock + future false 409 olmasın.
    try:
        await db.shift_schedules.insert_one(item)
    except Exception:
        await _release_shift_lock_interval(
            tenant_id=current_user.tenant_id,
            staff_id=payload.staff_id,
            shift_date=payload.shift_date,
            shift_id=shift_id,
        )
        raise
    item.pop('_id', None)
    return {'success': True, 'shift': item}


@router.get("/hr/shifts")
async def list_shifts(
    start: str | None = None,
    end: str | None = None,
    staff_id: str | None = None,
    department: str | None = None,
    current_user: User = Depends(get_current_user),
):
    today = _today_local()
    start_dt = (
        datetime.fromisoformat(start).date() if start
        else today - timedelta(days=7)
    )
    end_dt = (
        datetime.fromisoformat(end).date() if end
        else today + timedelta(days=14)
    )
    # Task #257: overnight (crosses_midnight) vardiyalar bitiş gününde de
    # görünmeli. Frontend ertesi gün hücresinde "← 06:00 (önceki gün)"
    # rozeti çizebilmek için, sorgu penceresinden bir gün önce başlayan
    # gece vardiyalarını da dahil ediyoruz. Çift sayım yapılmaz: tek kayıt
    # döner, sadece tarih penceresi genişletilir.
    pre_start_dt = start_dt - timedelta(days=1)
    query: dict[str, Any] = {
        'tenant_id': current_user.tenant_id,
        '$or': [
            {'shift_date': {'$gte': start_dt.isoformat(), '$lte': end_dt.isoformat()}},
            {
                'shift_date': pre_start_dt.isoformat(),
                'crosses_midnight': True,
            },
        ],
    }
    if staff_id:
        query['staff_id'] = staff_id

    # Task #263: department filter — staff_members.department üzerinden
    # `staff_id $in` listesine indirgenir; users-türevli derived staff
    # (role = department proxy) için role eşleşmesi de katılır.
    if department:
        dept_ids: set[str] = set()
        sm_cursor = db.staff_members.find(
            {'tenant_id': current_user.tenant_id, 'department': department},
            {'_id': 0, 'id': 1},
        )
        async for s in sm_cursor:
            dept_ids.add(s['id'])
        usr_cursor = db.users.find(
            {
                'tenant_id': current_user.tenant_id,
                'is_active': True,
                'role': department,
            },
            {'_id': 0, 'id': 1},
        )
        async for u in usr_cursor:
            dept_ids.add(u['id'])
        if not dept_ids:
            return {
                'items': [], 'total': 0,
                'range': {'start': start_dt.isoformat(), 'end': end_dt.isoformat()},
                'department': department,
            }
        if 'staff_id' in query:
            if query['staff_id'] not in dept_ids:
                return {
                    'items': [], 'total': 0,
                    'range': {'start': start_dt.isoformat(), 'end': end_dt.isoformat()},
                    'department': department,
                }
        else:
            query['staff_id'] = {'$in': list(dept_ids)}

    items = await db.shift_schedules.find(
        query, {'_id': 0}
    ).sort('shift_date', 1).to_list(2000)

    # Personel adıyla zenginleştir (hem staff_members hem türeyen users)
    staff_map = await _get_staff_map(current_user.tenant_id)
    user_cursor = db.users.find({
        'tenant_id': current_user.tenant_id,
        'is_active': True,
        'role': {'$in': [
            'housekeeping', 'front_desk', 'supervisor',
            'finance', 'sales', 'admin'
        ]},
    }, {'_id': 0, 'id': 1, 'name': 1})
    async for u in user_cursor:
        if u['id'] not in staff_map:
            staff_map[u['id']] = {'id': u['id'], 'name': u.get('name') or 'Personel'}
    for item in items:
        if not item.get('staff_name'):
            sm = staff_map.get(item.get('staff_id'), {})
            item['staff_name'] = sm.get('name', item.get('staff_id'))
    return {
        'items': items,
        'total': len(items),
        'range': {'start': start_dt.isoformat(), 'end': end_dt.isoformat()},
    }


@router.delete("/hr/shifts/{shift_id}")
async def delete_shift(
    shift_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    # Task #254 follow-up: önce shift'i okuyup (staff_id, shift_date)
    # parametrelerini al; sonra schedule sil, son olarak overlap lock
    # entry'sini at. Aksi halde silinmiş vardiya `shift_schedule_locks`
    # üzerinde stale kalır ve aynı saat aralığında yeni create false 409
    # döner.
    existing = await db.shift_schedules.find_one(
        {'tenant_id': current_user.tenant_id, 'id': shift_id},
        {'_id': 0, 'staff_id': 1, 'shift_date': 1, 'crosses_midnight': 1},
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Vardiya bulunamadı")
    res = await db.shift_schedules.delete_one(
        {'tenant_id': current_user.tenant_id, 'id': shift_id}
    )
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Vardiya bulunamadı")
    await _release_shift_lock_for_shift(
        tenant_id=current_user.tenant_id,
        staff_id=existing.get('staff_id'),
        shift_date=existing.get('shift_date'),
        crosses_midnight=bool(existing.get('crosses_midnight', False)),
        shift_id=shift_id,
    )
    return {'success': True}


# ============= Bulk shift create (Task #263) =============

class BulkShiftPayload(BaseModel):
    """Toplu vardiya oluştur — (staff_ids × dates) × tek şablon."""
    staff_ids: list[str] = Field(..., min_length=1, max_length=200)
    dates: list[str] = Field(..., min_length=1, max_length=62)
    shift_type: Literal['morning', 'afternoon', 'evening', 'night', 'split'] = 'morning'
    start_time: str = Field(..., pattern=r'^\d{2}:\d{2}$')
    end_time: str = Field(..., pattern=r'^\d{2}:\d{2}$')
    crosses_midnight: bool = False
    notes: str | None = Field(None, max_length=300)

    @field_validator('dates')
    @classmethod
    def _valid_dates(cls, v: list[str]) -> list[str]:
        for d in v:
            try:
                date.fromisoformat(d)
            except Exception as exc:
                raise ValueError(f"Geçersiz tarih: {d}") from exc
        return v


@router.post("/hr/shifts/bulk")
async def create_shifts_bulk(
    payload: BulkShiftPayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    """(staff × date) toplu vardiya oluşturma. Idempotent değildir; her
    başarılı satır kendi `shift_id`'sini alır. Çakışan / pasif / izinli
    satırlar atlanır ve `skipped` listesinde sebep ile döner. Her
    iterasyon mevcut tekil `create_shift_v2` ile aynı invariant'ları
    çalıştırır (active check → leave check → overnight contract → lock
    acquire → insert). Lock entry başarılı olup insert başarısız olursa
    aynı rollback uygulanır."""
    # Aynı isteğin overnight + same-day kontratını payload seviyesinde doğrula:
    if payload.crosses_midnight:
        if payload.end_time >= payload.start_time:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"Gece vardiyası işaretlendi ama bitiş saati "
                    f"({payload.end_time}) başlangıç saatinden "
                    f"({payload.start_time}) önce olmalı."
                ),
            )
    else:
        if payload.end_time <= payload.start_time:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"Vardiya bitiş saati ({payload.end_time}) başlangıç "
                    f"saatinden ({payload.start_time}) sonra olmalı."
                ),
            )

    created: list[dict] = []
    skipped: list[dict] = []
    now_iso_base = datetime.now(UTC).isoformat()

    for staff_id in payload.staff_ids:
        staff = await _verify_staff_in_tenant(staff_id, current_user.tenant_id)
        if not staff:
            for d_iso in payload.dates:
                skipped.append({
                    'staff_id': staff_id, 'shift_date': d_iso,
                    'reason': 'staff_not_found',
                })
            continue
        if not await _staff_is_active(current_user.tenant_id, staff_id):
            for d_iso in payload.dates:
                skipped.append({
                    'staff_id': staff_id, 'shift_date': d_iso,
                    'reason': 'staff_inactive',
                })
            continue

        for d_iso in payload.dates:
            # Leave check (start + opsiyonel next day).
            leave = await _staff_has_approved_leave_on(
                current_user.tenant_id, staff_id, d_iso,
            )
            if leave:
                skipped.append({
                    'staff_id': staff_id, 'shift_date': d_iso,
                    'reason': 'on_approved_leave',
                    'leave_type': leave.get('leave_type'),
                })
                continue
            if payload.crosses_midnight:
                next_iso = (
                    date.fromisoformat(d_iso) + timedelta(days=1)
                ).isoformat()
                leave_n = await _staff_has_approved_leave_on(
                    current_user.tenant_id, staff_id, next_iso,
                )
                if leave_n:
                    skipped.append({
                        'staff_id': staff_id, 'shift_date': d_iso,
                        'reason': 'overnight_overlaps_leave',
                    })
                    continue

            shift_id = str(uuid.uuid4())
            try:
                await _acquire_shift_lock_for_shift(
                    tenant_id=current_user.tenant_id,
                    staff_id=staff_id,
                    shift_date=d_iso,
                    start_time=payload.start_time,
                    end_time=payload.end_time,
                    crosses_midnight=payload.crosses_midnight,
                    shift_id=shift_id,
                    now_iso=now_iso_base,
                )
            except HTTPException as exc:
                skipped.append({
                    'staff_id': staff_id, 'shift_date': d_iso,
                    'reason': 'conflict' if exc.status_code == 409 else 'lock_failed',
                    'detail': str(exc.detail),
                })
                continue

            d_obj = date.fromisoformat(d_iso)
            end_date_iso = (
                (d_obj + timedelta(days=1)).isoformat()
                if payload.crosses_midnight else d_iso
            )
            item = {
                'id': shift_id,
                'tenant_id': current_user.tenant_id,
                'staff_id': staff_id,
                'staff_name': staff.get('name'),
                'shift_date': d_iso,
                'shift_type': payload.shift_type,
                'start_time': payload.start_time,
                'end_time': payload.end_time,
                'crosses_midnight': payload.crosses_midnight,
                'end_date': end_date_iso,
                'notes': payload.notes,
                'status': 'scheduled',
                'created_at': datetime.now(UTC).isoformat(),
                'created_via': 'bulk',
            }
            try:
                await db.shift_schedules.insert_one(item)
            except Exception:
                await _release_shift_lock_for_shift(
                    tenant_id=current_user.tenant_id,
                    staff_id=staff_id,
                    shift_date=d_iso,
                    crosses_midnight=payload.crosses_midnight,
                    shift_id=shift_id,
                )
                skipped.append({
                    'staff_id': staff_id, 'shift_date': d_iso,
                    'reason': 'insert_failed',
                })
                continue
            item.pop('_id', None)
            created.append({
                'id': shift_id, 'staff_id': staff_id, 'shift_date': d_iso,
            })

    return {
        'success': True,
        'created_count': len(created),
        'skipped_count': len(skipped),
        'created': created,
        'skipped': skipped,
    }


# ============= Recruitment Applicants =============

class ApplicantPayload(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    email: str | None = Field(None, max_length=200)
    phone: str | None = Field(None, max_length=40)
    notes: str | None = Field(None, max_length=2000)
    cv_url: str | None = Field(None, max_length=500)


class ApplicantStatusPayload(BaseModel):
    status: Literal['new', 'screening', 'interview', 'offer', 'hired', 'rejected']
    note: str | None = Field(None, max_length=500)


@router.post("/hr/job-postings/{job_id}/applicants")
async def add_applicant(
    job_id: str,
    payload: ApplicantPayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    job = await db.job_postings.find_one(
        {'tenant_id': current_user.tenant_id, 'id': job_id}
    )
    if not job:
        raise HTTPException(status_code=404, detail="İş ilanı bulunamadı")
    if job.get('status') != 'active':
        raise HTTPException(
            status_code=400,
            detail="Sadece aktif (yayında) ilanlara aday eklenebilir",
        )
    item = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'job_id': job_id,
        'job_title': job.get('title'),
        'name': payload.name,
        'email': payload.email,
        'phone': payload.phone,
        'notes': payload.notes,
        'cv_url': payload.cv_url,
        'status': 'new',
        'created_at': datetime.now(UTC).isoformat(),
    }
    await db.job_applicants.insert_one(item)
    await db.job_postings.update_one(
        {'tenant_id': current_user.tenant_id, 'id': job_id},
        {'$inc': {'applicants_count': 1}},
    )
    item.pop('_id', None)
    return {'success': True, 'applicant': item}


@router.get("/hr/job-postings/{job_id}/applicants")
async def list_applicants(
    job_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    items = await db.job_applicants.find({
        'tenant_id': current_user.tenant_id,
        'job_id': job_id,
    }, {'_id': 0}).sort('created_at', -1).to_list(500)
    counts: dict[str, int] = {}
    for it in items:
        s = it.get('status', 'new')
        counts[s] = counts.get(s, 0) + 1
    return {'items': items, 'total': len(items), 'counts': counts}


@router.post("/hr/applicants/{applicant_id}/status")
async def update_applicant_status(
    applicant_id: str,
    payload: ApplicantStatusPayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    res = await db.job_applicants.update_one(
        {'tenant_id': current_user.tenant_id, 'id': applicant_id},
        {'$set': {
            'status': payload.status,
            'status_note': payload.note,
            'status_updated_by': getattr(current_user, 'id', None),
            'status_updated_at': datetime.now(UTC).isoformat(),
        }},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Aday bulunamadı")
    return {'success': True, 'status': payload.status}


# ============= Performance Templates =============

class CompetencyItem(BaseModel):
    name: str = Field(..., max_length=100)
    weight: float = Field(1.0, ge=0, le=10)


class PerformanceTemplatePayload(BaseModel):
    name: str = Field(..., min_length=1, max_length=120)
    description: str | None = Field(None, max_length=1000)
    competencies: list[CompetencyItem] = Field(default_factory=list)


@router.get("/hr/performance-templates")
async def list_performance_templates(
    current_user: User = Depends(get_current_user),
):
    items = await db.performance_templates.find({
        'tenant_id': current_user.tenant_id,
    }, {'_id': 0}).sort('name', 1).to_list(100)
    return {'items': items, 'total': len(items)}


@router.post("/hr/performance-templates")
async def create_performance_template(
    payload: PerformanceTemplatePayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    item = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'name': payload.name,
        'description': payload.description,
        'competencies': [c.model_dump() for c in payload.competencies],
        'created_at': datetime.now(UTC).isoformat(),
    }
    await db.performance_templates.insert_one(item)
    item.pop('_id', None)
    return {'success': True, 'template': item}


@router.delete("/hr/performance-templates/{tpl_id}")
async def delete_performance_template(
    tpl_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    res = await db.performance_templates.delete_one({
        'tenant_id': current_user.tenant_id,
        'id': tpl_id,
    })
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Şablon bulunamadı")
    return {'success': True}


# ============================================================================
# HR FAZ 2 — Eklentiler (Mayıs 2026)
#   1. Mesai (overtime) talep + onay/red + 270h/yıl üst sınırı
#   2. Vardiya değişim talebi + onay (atomic swap)
#   3. Maaş geçmişi / zam kaydı
#   4. İşten ayrılma süreci (Türk İş K. m.14 kıdem tazminatı hesabı)
#   5. Eğitim/sertifika takibi (expiry alarmı)
#   6. Personel belgeleri (multipart upload, MongoDB inline base64, max 5MB)
#   7. Performans hedef-ilerleme check-in
# ============================================================================

# TR İş Kanunu m.41/3: yıllık fazla mesai üst sınırı 270 saat.
TR_ANNUAL_OVERTIME_CAP_HOURS = 270
# m.14 / m.17: kıdem tazminatı = 30 günlük brüt × tam yıl. Yasal tavan
# (kıdem tazminatı tavanı) günlük üst sınır olarak uygulanır — 2026 yarıyıl
# için yaklaşık 53.919,68 TL/aylık (1.797,32 TL/gün). Tenant override için
# settings.severance_daily_cap kullanılabilir.
SEVERANCE_DAYS_PER_YEAR = 30
TR_SEVERANCE_DAILY_CAP_DEFAULT = 1797.32  # 2026 H1 — günlük brüt tavanı
HR_ELEVATED_ROLES = {'admin', 'owner', 'supervisor', 'manager', 'hr', 'finance'}
# Belge boyut sınırı (MongoDB document ~16MB üst sınırı; güvenli marj).
DOC_MAX_BYTES = 5 * 1024 * 1024


# ============= 1. Mesai Talebi =============

class OvertimeRequestPayload(BaseModel):
    staff_id: str = Field(..., min_length=1, max_length=128)
    work_date: str = Field(..., pattern=r'^\d{4}-\d{2}-\d{2}$')
    hours: float = Field(..., gt=0, le=12)
    reason: str = Field(..., min_length=3, max_length=1000)


class OvertimeDecisionPayload(BaseModel):
    # Task #263: 2-aşamalı onay. pending → dept_approve → dept_approved →
    # approve → approved. Reject note ZORUNLU.
    action: Literal['dept_approve', 'approve', 'reject']
    note: str | None = Field(None, max_length=500)


async def _yearly_overtime_hours(tenant_id: str, staff_id: str, year: int) -> float:
    """Onaylanmış (status=approved) yıllık fazla mesai toplamı."""
    cursor = db.overtime_requests.find({
        'tenant_id': tenant_id,
        'staff_id': staff_id,
        'status': 'approved',
        'work_date': {'$gte': f'{year}-01-01', '$lte': f'{year}-12-31'},
    }, {'_id': 0, 'hours': 1})
    total = 0.0
    async for d in cursor:
        total += float(d.get('hours') or 0)
    return total


@router.post("/hr/overtime-request")
async def create_overtime_request(
    payload: OvertimeRequestPayload,
    current_user: User = Depends(get_current_user),
):
    staff = await _verify_staff_in_tenant(payload.staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    # Self-service guard: yetkisi olmayan kullanıcı sadece kendi adına talep açabilir.
    if getattr(current_user, 'role', None) not in HR_ELEVATED_ROLES:
        user_email = (getattr(current_user, 'email', None) or '').lower()
        staff_email = (staff.get('email') or '').lower()
        if not user_email or user_email != staff_email:
            raise HTTPException(status_code=403, detail="Sadece kendi adınıza talep açabilirsiniz")
    item = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'staff_id': payload.staff_id,
        'staff_name': staff.get('name'),
        'work_date': payload.work_date,
        'hours': payload.hours,
        'reason': payload.reason,
        'status': 'pending',
        'requested_by': getattr(current_user, 'id', None),
        'requested_at': datetime.now(UTC).isoformat(),
    }
    await db.overtime_requests.insert_one(item)
    item.pop('_id', None)
    await _notify_hr_managers(
        current_user.tenant_id,
        kind='overtime_request',
        title=f"Mesai talebi: {staff.get('name')}",
        body=f"{payload.work_date} — {payload.hours:g} saat. Sebep: {payload.reason[:120]}",
        link=f"/hr-complete?tab=overtime",
        ref_id=item['id'],
    )
    return {'success': True, 'overtime_request': item}


@router.get("/hr/overtime-requests")
async def list_overtime_requests(
    staff_id: str | None = None,
    status: str | None = None,
    current_user: User = Depends(get_current_user),
):
    query: dict[str, Any] = {'tenant_id': current_user.tenant_id}
    if staff_id:
        query['staff_id'] = staff_id
    if status:
        query['status'] = status
    items = await db.overtime_requests.find(query, {'_id': 0}).sort('requested_at', -1).to_list(500)
    counts = {'pending': 0, 'approved': 0, 'rejected': 0}
    for it in items:
        counts[it.get('status', 'pending')] = counts.get(it.get('status', 'pending'), 0) + 1
    return {'items': items, 'total': len(items), 'counts': counts}


@router.post("/hr/overtime-request/{req_id}/decision")
async def decide_overtime_request(
    req_id: str,
    payload: OvertimeDecisionPayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    req = await db.overtime_requests.find_one({
        'tenant_id': current_user.tenant_id, 'id': req_id,
    })
    if not req:
        raise HTTPException(status_code=404, detail="Talep bulunamadı")

    current_status = req.get('status', 'pending')

    if payload.action == 'reject':
        if current_status not in ('pending', 'dept_approved'):
            raise HTTPException(
                status_code=400,
                detail=f"Talep zaten karara bağlanmış (status={current_status})",
            )
        if not (payload.note and payload.note.strip()):
            raise HTTPException(
                status_code=400,
                detail="Red için gerekçe (note) zorunludur",
            )
        new_status = 'rejected'
        notify_kind = 'overtime_rejected'
        notify_title = "Mesai talebiniz reddedildi"
    elif payload.action == 'dept_approve':
        if current_status != 'pending':
            raise HTTPException(
                status_code=400,
                detail=f"Departman onayı sadece pending durumdan verilebilir (mevcut: {current_status})",
            )
        new_status = 'dept_approved'
        notify_kind = 'overtime_dept_approved'
        notify_title = "Mesai talebiniz departman onayı aldı (final onay bekleniyor)"
    else:  # 'approve' (final HR/Finance) — Task #263: strict chain enforcement
        if current_status != 'dept_approved':
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Final onay sadece departman onayından sonra verilebilir "
                    f"(mevcut: {current_status}). Önce 'dept_approve' aşaması gerekli."
                ),
            )
        # 270h/yıl kontrolü (İş K. m.41/3) — final aşamada
        year = int(req['work_date'][:4])
        already = await _yearly_overtime_hours(
            current_user.tenant_id, req['staff_id'], year,
        )
        proposed = float(req.get('hours') or 0)
        if already + proposed > TR_ANNUAL_OVERTIME_CAP_HOURS:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Yıllık 270 saat fazla mesai üst sınırı aşılır "
                    f"(mevcut: {already:g}h, talep: {proposed:g}h, sınır: 270h)"
                ),
            )
        new_status = 'approved'
        notify_kind = 'overtime_approved'
        notify_title = "Mesai talebiniz onaylandı"

    update_set: dict[str, Any] = {
        'status': new_status,
        'decision_note': payload.note,
        'decided_at': datetime.now(UTC).isoformat(),
    }
    history = req.get('decision_history') or []
    history.append({
        'stage': payload.action,
        'status': new_status,
        'by': getattr(current_user, 'id', None),
        'at': datetime.now(UTC).isoformat(),
        'note': payload.note,
    })
    update_set['decision_history'] = history
    if new_status == 'dept_approved':
        update_set['dept_approved_by'] = getattr(current_user, 'id', None)
        update_set['dept_approved_at'] = datetime.now(UTC).isoformat()
    elif new_status == 'approved':
        update_set['decided_by'] = getattr(current_user, 'id', None)
        update_set['approved_by'] = getattr(current_user, 'id', None)
        update_set['approved_at'] = datetime.now(UTC).isoformat()
        # Bordro entegrasyon kontratı: payroll_ready=True (append-only)
        update_set['payroll_ready'] = True
    else:
        update_set['decided_by'] = getattr(current_user, 'id', None)

    await db.overtime_requests.update_one(
        {'tenant_id': current_user.tenant_id, 'id': req_id},
        {'$set': update_set},
    )

    requester = req.get('requested_by')
    if requester:
        await _notify_user(
            current_user.tenant_id, user_id=requester,
            kind=notify_kind, title=notify_title,
            body=f"{req['work_date']} — {req['hours']:g}h. " + (payload.note or ''),
            ref_id=req_id,
        )
    return {'success': True, 'status': new_status}


@router.get("/hr/overtime/ready-for-payroll")
async def overtime_ready_for_payroll(
    month: str = Query(..., pattern=r'^\d{4}-\d{2}$'),
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """Bordro modülü için onaylı mesai özetini döndüren read-only kontrat.
    Append-only: bordro task'ı bu endpoint'i TÜKETIR, değiştirmez. Şema:
        {month, total_hours, total_requests, by_staff: [...], items: [...]}
    Her satır: {staff_id, staff_name, hours, work_date, request_id,
    approved_at, multiplier=1.5 (TR varsayılan)}."""
    yyyy, mm = month.split('-')
    start_iso = f'{yyyy}-{mm}-01'
    nm = date(int(yyyy), int(mm), 28) + timedelta(days=4)
    end_iso = (nm - timedelta(days=nm.day)).isoformat()
    cursor = db.overtime_requests.find({
        'tenant_id': current_user.tenant_id,
        'status': 'approved',
        'work_date': {'$gte': start_iso, '$lte': end_iso},
    }, {'_id': 0}).sort('work_date', 1)
    items: list[dict] = []
    by_staff: dict[str, dict] = {}
    total_hours = 0.0
    async for r in cursor:
        hours = float(r.get('hours') or 0)
        total_hours += hours
        sid = r['staff_id']
        by_staff.setdefault(sid, {
            'staff_id': sid,
            'staff_name': r.get('staff_name'),
            'hours': 0.0,
            'requests': 0,
        })
        by_staff[sid]['hours'] = round(by_staff[sid]['hours'] + hours, 2)
        by_staff[sid]['requests'] += 1
        items.append({
            'request_id': r['id'],
            'staff_id': sid,
            'staff_name': r.get('staff_name'),
            'work_date': r['work_date'],
            'hours': hours,
            'approved_at': r.get('approved_at') or r.get('decided_at'),
            'multiplier': TR_DEFAULT_OVERTIME_MULTIPLIER,
        })
    return {
        'month': month,
        'range': {'start': start_iso, 'end': end_iso},
        'total_hours': round(total_hours, 2),
        'total_requests': len(items),
        'overtime_multiplier': TR_DEFAULT_OVERTIME_MULTIPLIER,
        'currency': TR_CURRENCY,
        'by_staff': sorted(by_staff.values(), key=lambda x: x['staff_name'] or ''),
        'items': items,
    }


# ============= 2. Vardiya Değişim Talebi =============

class ShiftSwapRequestPayload(BaseModel):
    shift_id: str = Field(..., min_length=1)
    target_staff_id: str = Field(..., min_length=1)
    reason: str | None = Field(None, max_length=500)


class ShiftSwapDecisionPayload(BaseModel):
    action: Literal['approve', 'reject']
    note: str | None = Field(None, max_length=500)


@router.post("/hr/shift-swap-request")
async def create_shift_swap(
    payload: ShiftSwapRequestPayload,
    current_user: User = Depends(get_current_user),
):
    shift = await db.shift_schedules.find_one({
        'tenant_id': current_user.tenant_id, 'id': payload.shift_id,
    })
    if not shift:
        raise HTTPException(status_code=404, detail="Vardiya bulunamadı")
    target = await _verify_staff_in_tenant(payload.target_staff_id, current_user.tenant_id)
    if not target:
        raise HTTPException(status_code=404, detail="Hedef personel bulunamadı")
    if shift['staff_id'] == payload.target_staff_id:
        raise HTTPException(status_code=400, detail="Hedef personel mevcut sahibinin aynısı olamaz")
    # Self-service guard: yetkisiz kullanıcı sadece kendi vardiyasını değişime sunabilir.
    if getattr(current_user, 'role', None) not in HR_ELEVATED_ROLES:
        user_email = (getattr(current_user, 'email', None) or '').lower()
        owner = await db.staff.find_one(
            {'tenant_id': current_user.tenant_id, 'id': shift['staff_id']},
            {'_id': 0, 'email': 1},
        )
        owner_email = ((owner or {}).get('email') or '').lower()
        if not user_email or user_email != owner_email:
            raise HTTPException(status_code=403, detail="Sadece kendi vardiyanızı değişime sunabilirsiniz")
    item = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'shift_id': payload.shift_id,
        'shift_date': shift.get('shift_date'),
        'shift_type': shift.get('shift_type'),
        'from_staff_id': shift['staff_id'],
        'from_staff_name': shift.get('staff_name'),
        'target_staff_id': payload.target_staff_id,
        'target_staff_name': target.get('name'),
        'reason': payload.reason,
        'status': 'pending',
        # Çift onay akışı: önce hedef personelin rızası, sonra İK kararı.
        'target_consent_status': 'pending',
        'target_consent_at': None,
        'target_consent_note': None,
        'requested_by': getattr(current_user, 'id', None),
        'requested_at': datetime.now(UTC).isoformat(),
    }
    await db.shift_swap_requests.insert_one(item)
    item.pop('_id', None)
    # Hedef personele bildirim — rıza bekleniyor.
    target_user = await db.users.find_one(
        {'tenant_id': current_user.tenant_id, 'email': (target.get('email') or '').lower()},
        {'_id': 0, 'id': 1},
    ) if target.get('email') else None
    if target_user and target_user.get('id'):
        await _notify_user(
            current_user.tenant_id, user_id=target_user['id'],
            kind='shift_swap_consent_request',
            title="Vardiya devralma talebi",
            body=f"{shift.get('staff_name')} {shift.get('shift_date')} {shift.get('shift_type')} vardiyasını size devretmek istiyor",
            link="/hr/shifts",
            ref_id=item['id'],
        )
    await _notify_hr_managers(
        current_user.tenant_id,
        kind='shift_swap_request',
        title="Vardiya değişim talebi (hedef onayı bekliyor)",
        body=f"{shift.get('staff_name')} → {target.get('name')} ({shift.get('shift_date')} {shift.get('shift_type')})",
        link="/hr/shifts",
        ref_id=item['id'],
    )
    return {'success': True, 'request': item}


class ShiftSwapConsentPayload(BaseModel):
    action: Literal['approve', 'reject']
    note: str | None = Field(None, max_length=500)


@router.post("/hr/shift-swap-request/{req_id}/consent")
async def consent_shift_swap(
    req_id: str,
    payload: ShiftSwapConsentPayload,
    current_user: User = Depends(get_current_user),
):
    """Hedef personelin vardiyayı devralmaya rıza/red bildirmesi."""
    req = await db.shift_swap_requests.find_one({
        'tenant_id': current_user.tenant_id, 'id': req_id,
    })
    if not req:
        raise HTTPException(status_code=404, detail="Talep bulunamadı")
    if req.get('status') != 'pending':
        raise HTTPException(status_code=400, detail="Talep zaten karara bağlanmış")
    if req.get('target_consent_status') != 'pending':
        raise HTTPException(status_code=400, detail="Bu talebe daha önce yanıt verilmiş")

    # Sadece hedef personel rıza verebilir (yöneticiler dahil değil — etik gereği bizzat hedef).
    target = await db.staff.find_one(
        {'tenant_id': current_user.tenant_id, 'id': req['target_staff_id']},
        {'_id': 0, 'email': 1},
    )
    user_email = (getattr(current_user, 'email', None) or '').lower()
    target_email = ((target or {}).get('email') or '').lower()
    if not user_email or user_email != target_email:
        raise HTTPException(status_code=403, detail="Bu talep size ait değil — sadece hedef personel yanıtlayabilir")

    new_consent = 'approved' if payload.action == 'approve' else 'rejected'
    new_status = req.get('status') if new_consent == 'approved' else 'rejected'
    # Atomik update: target_consent_status hâlâ 'pending' iken yaz; aksi halde başka bir
    # eşzamanlı çağrı yanıt vermiş demektir (race koruması).
    res = await db.shift_swap_requests.update_one(
        {
            'tenant_id': current_user.tenant_id,
            'id': req_id,
            'target_consent_status': 'pending',
        },
        {'$set': {
            'target_consent_status': new_consent,
            'target_consent_at': datetime.now(UTC).isoformat(),
            'target_consent_note': payload.note,
            'status': new_status,
        }},
    )
    if res.modified_count == 0:
        raise HTTPException(status_code=409, detail="Bu talebe bu sırada başka bir yanıt verildi")
    # Talep sahibine bildir.
    if req.get('requested_by'):
        await _notify_user(
            current_user.tenant_id, user_id=req['requested_by'],
            kind=f'shift_swap_consent_{new_consent}',
            title=f"Hedef personel vardiyayı {('kabul etti' if new_consent == 'approved' else 'reddetti')}",
            body=f"{req.get('shift_date')} {req.get('shift_type')} — " + (payload.note or ''),
            ref_id=req_id,
        )
    return {'success': True, 'target_consent_status': new_consent, 'status': new_status}


@router.get("/hr/shift-swap-requests")
async def list_shift_swap_requests(
    status: str | None = None,
    current_user: User = Depends(get_current_user),
):
    query: dict[str, Any] = {'tenant_id': current_user.tenant_id}
    if status:
        query['status'] = status
    items = await db.shift_swap_requests.find(query, {'_id': 0}).sort('requested_at', -1).to_list(500)
    return {'items': items, 'total': len(items)}


@router.post("/hr/shift-swap-request/{req_id}/decision")
async def decide_shift_swap(
    req_id: str,
    payload: ShiftSwapDecisionPayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    req = await db.shift_swap_requests.find_one({
        'tenant_id': current_user.tenant_id, 'id': req_id,
    })
    if not req:
        raise HTTPException(status_code=404, detail="Talep bulunamadı")
    if req.get('status') != 'pending':
        raise HTTPException(status_code=400, detail="Bu talep zaten karara bağlanmış")

    if payload.action == 'approve':
        # Çift onay gate: hedef personelin rızası alınmadan İK onaylayamaz.
        if req.get('target_consent_status') != 'approved':
            raise HTTPException(
                status_code=409,
                detail="Hedef personel henüz rıza vermedi — önce rıza beklenmeli",
            )
        # Atomik swap: shift'i target_staff_id'ye devret
        target = await _verify_staff_in_tenant(req['target_staff_id'], current_user.tenant_id)
        if not target:
            raise HTTPException(status_code=400, detail="Hedef personel artık aktif değil")
        # Task #254 follow-up: swap, overlap lock'ı `from_staff` → `target`
        # taşır. Sıra önemli:
        #   1) Mevcut shift'i oku (start/end/shift_date) — lock filtresi
        #      için lazım.
        #   2) Target için aynı saat aralığını acquire et: target'ın o
        #      gününde başka çakışan vardiyası varsa 409 + abort.
        #   3) shift_schedules update'ini atomik filtreli yap; matched=0
        #      ise (yarış: silindi/değişti) az önce acquire ettiğimiz
        #      target lock entry'sini release ile geri al → orphan lock yok.
        #   4) Başarılıysa from_staff lock entry'sini release et.
        original_shift = await db.shift_schedules.find_one(
            {
                'tenant_id': current_user.tenant_id,
                'id': req['shift_id'],
                'staff_id': req['from_staff_id'],
            },
            {'_id': 0, 'shift_date': 1, 'start_time': 1, 'end_time': 1,
             'crosses_midnight': 1},
        )
        if not original_shift:
            raise HTTPException(
                status_code=409,
                detail="Vardiya bu sürede değişti veya silindi — talebi yeniden değerlendirin",
            )
        swap_now_iso = datetime.now(UTC).isoformat()
        orig_crosses = bool(original_shift.get('crosses_midnight', False))
        await _acquire_shift_lock_for_shift(
            tenant_id=current_user.tenant_id,
            staff_id=req['target_staff_id'],
            shift_date=original_shift.get('shift_date'),
            start_time=original_shift.get('start_time'),
            end_time=original_shift.get('end_time'),
            crosses_midnight=orig_crosses,
            shift_id=req['shift_id'],
            now_iso=swap_now_iso,
        )
        # Atomik filtre: vardiyanın hâlâ orijinal sahibinde olduğunu doğrula.
        # Aksi halde başka bir swap zaten uygulanmış demektir (race koruması).
        upd = await db.shift_schedules.update_one(
            {
                'tenant_id': current_user.tenant_id,
                'id': req['shift_id'],
                'staff_id': req['from_staff_id'],
            },
            {'$set': {
                'staff_id': req['target_staff_id'],
                'staff_name': target.get('name'),
                'swapped_from': req['from_staff_id'],
                'swapped_at': swap_now_iso,
            }},
        )
        if upd.matched_count == 0:
            # Rollback: target için almış olduğumuz lock'ı bırak.
            await _release_shift_lock_for_shift(
                tenant_id=current_user.tenant_id,
                staff_id=req['target_staff_id'],
                shift_date=original_shift.get('shift_date'),
                crosses_midnight=orig_crosses,
                shift_id=req['shift_id'],
            )
            raise HTTPException(
                status_code=409,
                detail="Vardiya bu sürede değişti veya silindi — talebi yeniden değerlendirin",
            )
        # Sahiplik transferi başarılı: from_staff lock entry'sini bırak.
        await _release_shift_lock_for_shift(
            tenant_id=current_user.tenant_id,
            staff_id=req['from_staff_id'],
            shift_date=original_shift.get('shift_date'),
            crosses_midnight=orig_crosses,
            shift_id=req['shift_id'],
        )
        new_status = 'approved'
    else:
        new_status = 'rejected'

    await db.shift_swap_requests.update_one(
        {'tenant_id': current_user.tenant_id, 'id': req_id},
        {'$set': {
            'status': new_status,
            'decision_note': payload.note,
            'decided_by': getattr(current_user, 'id', None),
            'decided_at': datetime.now(UTC).isoformat(),
        }},
    )
    if req.get('requested_by'):
        await _notify_user(
            current_user.tenant_id, user_id=req['requested_by'],
            kind=f'shift_swap_{new_status}',
            title=f"Vardiya değişimi {('onaylandı' if new_status == 'approved' else 'reddedildi')}",
            body=f"{req.get('shift_date')} {req.get('shift_type')} — " + (payload.note or ''),
            ref_id=req_id,
        )
    return {'success': True, 'status': new_status}


# ============= 3. Maaş Geçmişi / Zam Kaydı =============

class SalaryChangePayload(BaseModel):
    new_hourly_rate: float = Field(..., gt=0, le=100000)
    effective_date: str = Field(..., pattern=r'^\d{4}-\d{2}-\d{2}$')
    change_type: Literal['raise', 'promotion', 'correction', 'demotion'] = 'raise'
    reason: str | None = Field(None, max_length=500)


@router.post("/hr/staff/{staff_id}/salary-change")
async def create_salary_change(
    staff_id: str,
    payload: SalaryChangePayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    staff = await _verify_staff_in_tenant(staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    old_rate = float(staff.get('hourly_rate') or TR_DEFAULT_HOURLY_RATE)
    delta_pct = round(((payload.new_hourly_rate - old_rate) / old_rate) * 100, 2) if old_rate else 0
    record = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'staff_id': staff_id,
        'staff_name': staff.get('name'),
        'old_hourly_rate': old_rate,
        'new_hourly_rate': payload.new_hourly_rate,
        'delta_pct': delta_pct,
        'effective_date': payload.effective_date,
        'change_type': payload.change_type,
        'reason': payload.reason,
        'created_by': getattr(current_user, 'id', None),
        'created_at': datetime.now(UTC).isoformat(),
    }
    await db.salary_history.insert_one(record)
    # Aktüel ücreti güncelle (basit yaklaşım: effective_date geçmişse hemen uygula)
    if payload.effective_date <= _today_local().isoformat():
        await db.staff_members.update_one(
            {'tenant_id': current_user.tenant_id, 'id': staff_id},
            {'$set': {
                'hourly_rate': payload.new_hourly_rate,
                'updated_at': datetime.now(UTC).isoformat(),
            }},
        )
    record.pop('_id', None)
    return {'success': True, 'record': record, 'applied_now': payload.effective_date <= _today_local().isoformat()}


@router.get("/hr/staff/{staff_id}/salary-history")
async def list_salary_history(
    staff_id: str,
    current_user: User = Depends(get_current_user),
):
    """Maaş geçmişi — RBAC: tenant + dept scope + self-service (`_authorize_staff_access`).

    Route-level `require_op` YOK — self-service kullanıcının kendi maaş
    geçmişine erişebilmesi için. Authz tek noktada (`_authorize_staff_access`)
    yapılır; o kullanıcıya manage_hr yoksa PII maskelenir (`_mask_hr_pii`).
    """
    staff = await _verify_staff_in_tenant(staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    _authorize_staff_access(staff, current_user)
    items = await db.salary_history.find({
        'tenant_id': current_user.tenant_id, 'staff_id': staff_id,
    }, {'_id': 0}).sort('effective_date', -1).to_list(500)
    # v2 Foundation: maaş alanları rol-bazlı maskelenir — sadece manage_hr unmask.
    self_id = str(getattr(current_user, 'id', '') or '')
    self_email = str(getattr(current_user, 'email', '') or '')
    masked = [_mask_hr_pii(it, current_user, self_id=self_id, self_email=self_email) or it for it in items]
    return {'items': masked, 'total': len(masked)}


# ============= 4. İşten Ayrılma Süreci =============

class TerminationPayload(BaseModel):
    reason: Literal['resign', 'dismiss', 'mutual', 'retire', 'end_of_contract', 'death']
    last_day: str = Field(..., pattern=r'^\d{4}-\d{2}-\d{2}$')
    notice_period_days: int = Field(0, ge=0, le=180)
    exit_interview_notes: str | None = Field(None, max_length=4000)
    severance_override: float | None = Field(None, ge=0, le=10_000_000)
    eligible_for_rehire: bool = True


def _calc_severance_tr(
    hire_date_str: str | None,
    last_day_str: str,
    monthly_gross: float,
    daily_cap: float | None = None,
) -> dict:
    """Türk İş K. m.14: kıdem tazminatı = 30 günlük brüt × tam yıl + orantılı.

    `daily_cap` parametresi yasal tavanı (2026 H1 için ~1.797,32 TL/gün)
    uygular. Çalışanın günlük brüt ücreti tavanı aşıyorsa hesap tavandan
    yapılır. Bu sayede yüksek maaşlı pozisyonlarda şişirilmiş tazminat
    çıkmaz. İhbar tazminatı + birikmiş izin alacağı ayrı hesaplanır.
    """
    if not hire_date_str:
        return {'years_of_service': 0, 'severance_amount': 0, 'note': 'İşe başlama tarihi yok'}
    try:
        hire = date.fromisoformat(hire_date_str)
        last = date.fromisoformat(last_day_str)
    except ValueError:
        return {'years_of_service': 0, 'severance_amount': 0, 'note': 'Geçersiz tarih'}
    days = (last - hire).days
    if days < 365:
        return {
            'years_of_service': round(days / 365, 2),
            'severance_amount': 0,
            'note': '1 yıldan az kıdem — kıdem tazminatına hak kazanılmaz',
        }
    years = days / 365.0
    raw_daily = monthly_gross / 30.0
    cap = daily_cap if (daily_cap is not None and daily_cap > 0) else TR_SEVERANCE_DAILY_CAP_DEFAULT
    capped = raw_daily > cap
    daily_gross = min(raw_daily, cap)
    severance = round(years * SEVERANCE_DAYS_PER_YEAR * daily_gross, 2)
    note = f'{years:.2f} yıl × 30 gün × {daily_gross:.2f} TL/gün'
    if capped:
        note += f' (yasal tavan uygulandı; ham günlük: {raw_daily:.2f} TL)'
    return {
        'years_of_service': round(years, 2),
        'daily_gross': round(daily_gross, 2),
        'raw_daily_gross': round(raw_daily, 2),
        'daily_cap_applied': capped,
        'daily_cap': round(cap, 2),
        'severance_amount': severance,
        'note': note,
    }


@router.post("/hr/staff/{staff_id}/terminate")
async def terminate_staff(
    staff_id: str,
    payload: TerminationPayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    staff = await _verify_staff_in_tenant(staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    if staff.get('terminated_at'):
        raise HTTPException(status_code=400, detail="Personel zaten ayrılmış")

    monthly_hours = float(staff.get('monthly_hours') or TR_DEFAULT_MONTHLY_HOURS)
    hourly_rate = float(staff.get('hourly_rate') or TR_DEFAULT_HOURLY_RATE)
    monthly_gross = monthly_hours * hourly_rate

    # Tenant'a özel kıdem tavan override'ı (settings.severance_daily_cap) — yoksa default.
    tenant_settings = await db.tenant_settings.find_one(
        {'tenant_id': current_user.tenant_id}, {'_id': 0, 'severance_daily_cap': 1},
    ) or {}
    severance_calc = _calc_severance_tr(
        staff.get('hire_date'), payload.last_day, monthly_gross,
        daily_cap=tenant_settings.get('severance_daily_cap'),
    )
    final_severance = (
        payload.severance_override
        if payload.severance_override is not None
        else severance_calc['severance_amount']
    )

    record = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'staff_id': staff_id,
        'staff_name': staff.get('name'),
        'reason': payload.reason,
        'last_day': payload.last_day,
        'notice_period_days': payload.notice_period_days,
        'exit_interview_notes': payload.exit_interview_notes,
        'monthly_gross_at_exit': round(monthly_gross, 2),
        'severance_calc': severance_calc,
        'severance_paid': final_severance,
        'eligible_for_rehire': payload.eligible_for_rehire,
        'processed_by': getattr(current_user, 'id', None),
        'processed_at': datetime.now(UTC).isoformat(),
    }
    await db.staff_terminations.insert_one(record)

    await db.staff_members.update_one(
        {'tenant_id': current_user.tenant_id, 'id': staff_id},
        {'$set': {
            'active': False,
            'terminated_at': datetime.now(UTC).isoformat(),
            'termination_reason': payload.reason,
            'last_day': payload.last_day,
        }},
    )
    record.pop('_id', None)
    return {'success': True, 'termination': record}


@router.get("/hr/staff/{staff_id}/termination")
async def get_termination_record(
    staff_id: str,
    current_user: User = Depends(get_current_user),
):
    """İşten ayrılma kaydı — RBAC: dept scope + self bypass (`_authorize_staff_access`)."""
    staff = await _verify_staff_in_tenant(staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    _authorize_staff_access(staff, current_user)
    rec = await db.staff_terminations.find_one(
        {'tenant_id': current_user.tenant_id, 'staff_id': staff_id},
        {'_id': 0},
        sort=[('processed_at', -1)],
    )
    return {'record': rec}


# ============= 5. Eğitim/Sertifika Takibi =============

class CertificationPayload(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    issuer: str | None = Field(None, max_length=200)
    issue_date: str = Field(..., pattern=r'^\d{4}-\d{2}-\d{2}$')
    expiry_date: str | None = Field(None, pattern=r'^\d{4}-\d{2}-\d{2}$')
    certificate_no: str | None = Field(None, max_length=100)
    file_url: str | None = Field(None, max_length=500)
    notes: str | None = Field(None, max_length=1000)


@router.post("/hr/staff/{staff_id}/certifications")
async def add_certification(
    staff_id: str,
    payload: CertificationPayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    staff = await _verify_staff_in_tenant(staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    item = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'staff_id': staff_id,
        'staff_name': staff.get('name'),
        'name': payload.name,
        'issuer': payload.issuer,
        'issue_date': payload.issue_date,
        'expiry_date': payload.expiry_date,
        'certificate_no': payload.certificate_no,
        'file_url': payload.file_url,
        'notes': payload.notes,
        'created_at': datetime.now(UTC).isoformat(),
    }
    await db.staff_certifications.insert_one(item)
    item.pop('_id', None)
    return {'success': True, 'certification': item}


@router.get("/hr/staff/{staff_id}/certifications")
async def list_staff_certifications(
    staff_id: str,
    current_user: User = Depends(get_current_user),
):
    staff = await _verify_staff_in_tenant(staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    _authorize_staff_access(staff, current_user)
    items = await db.staff_certifications.find({
        'tenant_id': current_user.tenant_id, 'staff_id': staff_id,
    }, {'_id': 0}).sort('issue_date', -1).to_list(500)
    today = _today_local().isoformat()
    expiring = sum(1 for it in items if it.get('expiry_date') and it['expiry_date'] >= today)
    expired = sum(1 for it in items if it.get('expiry_date') and it['expiry_date'] < today)
    return {'items': items, 'total': len(items), 'active': expiring, 'expired': expired}


@router.delete("/hr/certifications/{cert_id}")
async def delete_certification(
    cert_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    res = await db.staff_certifications.delete_one({
        'tenant_id': current_user.tenant_id, 'id': cert_id,
    })
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sertifika bulunamadı")
    return {'success': True}


@router.get("/hr/certifications/expiring")
async def list_expiring_certifications(
    days_ahead: int = Query(90, ge=1, le=365),
    current_user: User = Depends(get_current_user),
):
    """Önümüzdeki N gün içinde süresi dolan sertifikalar (compliance dashboard)."""
    today = _today_local()
    end = (today + timedelta(days=days_ahead)).isoformat()
    items = await db.staff_certifications.find({
        'tenant_id': current_user.tenant_id,
        'expiry_date': {'$gte': today.isoformat(), '$lte': end},
    }, {'_id': 0}).sort('expiry_date', 1).to_list(500)
    return {'items': items, 'total': len(items), 'window_days': days_ahead}


# ============= 6. Personel Belgeleri (PDF/sözleşme yükleme) =============

ALLOWED_DOC_TYPES = {'contract', 'id', 'diploma', 'health', 'insurance', 'tax', 'other'}
ALLOWED_DOC_MIME = {
    'application/pdf', 'image/png', 'image/jpeg', 'image/webp',
    'application/msword',
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
}


@router.post("/hr/staff/{staff_id}/documents")
async def upload_staff_document(
    staff_id: str,
    file: UploadFile = File(...),
    doc_type: str = Query('other', max_length=20),
    label: str = Query('', max_length=200),
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    staff = await _verify_staff_in_tenant(staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    if doc_type not in ALLOWED_DOC_TYPES:
        raise HTTPException(status_code=400, detail=f"Geçersiz belge türü. İzinli: {sorted(ALLOWED_DOC_TYPES)}")
    if file.content_type not in ALLOWED_DOC_MIME:
        raise HTTPException(status_code=400, detail=f"Geçersiz dosya türü: {file.content_type}")

    content = await file.read()
    if len(content) > DOC_MAX_BYTES:
        raise HTTPException(status_code=413, detail=f"Dosya çok büyük (>5MB)")
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="Boş dosya")

    # GridFS'e yaz — büyük dosyalar memory'de tutulmaz, koleksiyon liste sorguları hızlı kalır.
    gridfs_id = await _hr_docs_bucket.upload_from_stream(
        file.filename or 'document',
        content,
        metadata={
            'tenant_id': current_user.tenant_id,
            'staff_id': staff_id,
            'doc_type': doc_type,
            'content_type': file.content_type,
        },
    )
    item = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'staff_id': staff_id,
        'staff_name': staff.get('name'),
        'doc_type': doc_type,
        'label': label or file.filename or 'Belge',
        'filename': file.filename,
        'content_type': file.content_type,
        'size_bytes': len(content),
        'gridfs_id': str(gridfs_id),
        'storage': 'gridfs',
        'uploaded_by': getattr(current_user, 'id', None),
        'uploaded_at': datetime.now(UTC).isoformat(),
    }
    await db.staff_documents.insert_one(item)
    response = {k: v for k, v in item.items() if k != '_id'}
    return {'success': True, 'document': response}


@router.get("/hr/staff/{staff_id}/documents")
async def list_staff_documents(
    staff_id: str,
    current_user: User = Depends(get_current_user),
):
    """Personel belge listesi — RBAC: dept scope + self bypass."""
    staff = await _verify_staff_in_tenant(staff_id, current_user.tenant_id)
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    _authorize_staff_access(staff, current_user)
    items = await db.staff_documents.find(
        {'tenant_id': current_user.tenant_id, 'staff_id': staff_id},
        {'_id': 0, 'data_b64': 0},  # Legacy binary alanı liste yanıtında olmasın
    ).sort('uploaded_at', -1).to_list(500)
    return {'items': items, 'total': len(items)}


@router.get("/hr/documents/{doc_id}/download")
async def download_staff_document(
    doc_id: str,
    current_user: User = Depends(get_current_user),
):
    """Belge indir — tenant + dept scope + self bypass.

    Route-level `require_op` YOK; authz `_authorize_staff_access` ile doc'un
    bağlı olduğu staff üzerinden yapılır. Bu, kendi sözleşmesini/diplomasını
    indirebilen self-service kullanıcı durumunu kapsar.
    """
    doc = await db.staff_documents.find_one({
        'tenant_id': current_user.tenant_id, 'id': doc_id,
    })
    if not doc:
        raise HTTPException(status_code=404, detail="Belge bulunamadı")
    # Doc'un sahibi olan staff'ı yükleyip per-record authz uygula.
    doc_staff_id = doc.get('staff_id')
    if doc_staff_id:
        doc_staff = await _verify_staff_in_tenant(doc_staff_id, current_user.tenant_id)
        _authorize_staff_access(doc_staff, current_user)

    # GridFS'ten oku; eski (data_b64) kayıtlar için geriye dönük destek.
    if doc.get('gridfs_id'):
        try:
            gridfs_oid = ObjectId(doc['gridfs_id'])
        except Exception:
            raise HTTPException(status_code=404, detail="Belge depolamada bulunamadı")
        # Defense-in-depth: GridFS bucket tenant-aware proxy bypass'lı (_raw_db) olduğundan
        # metadata.tenant_id'yi tekrar doğrula. Meta kaydı zaten scoped ama belge bütünlüğü
        # için cross-check yapıyoruz.
        gf_meta = await _hr_docs_bucket.find({
            '_id': gridfs_oid,
            'metadata.tenant_id': current_user.tenant_id,
        }).to_list(1)
        if not gf_meta:
            raise HTTPException(status_code=404, detail="Belge depolamada bulunamadı")
        try:
            stream = await _hr_docs_bucket.open_download_stream(gridfs_oid)
            raw = await stream.read()
        except Exception:
            raise HTTPException(status_code=404, detail="Belge depolamada bulunamadı")
    elif doc.get('data_b64'):
        raw = base64.b64decode(doc['data_b64'])
    else:
        raise HTTPException(status_code=410, detail="Belge içeriği yok")

    headers = {
        'Content-Disposition': f'attachment; filename="{doc.get("filename", doc_id)}"',
        'Content-Length': str(len(raw)),
    }
    return StreamingResponse(
        iter([raw]),
        media_type=doc.get('content_type', 'application/octet-stream'),
        headers=headers,
    )


@router.delete("/hr/documents/{doc_id}")
async def delete_staff_document(
    doc_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    doc = await db.staff_documents.find_one({
        'tenant_id': current_user.tenant_id, 'id': doc_id,
    })
    if not doc:
        raise HTTPException(status_code=404, detail="Belge bulunamadı")
    # Önce GridFS chunk'larını temizle (varsa); sonra meta kaydını sil.
    if doc.get('gridfs_id'):
        try:
            await _hr_docs_bucket.delete(ObjectId(doc['gridfs_id']))
        except Exception:
            pass  # Çoktan silinmiş olabilir — meta silmeyi engellemesin.
    await db.staff_documents.delete_one({
        'tenant_id': current_user.tenant_id, 'id': doc_id,
    })
    return {'success': True}


# ============= 6b. Kıdem Tazminatı Tavanı Ayarı =============

class SeveranceCapPayload(BaseModel):
    daily_cap: float = Field(..., gt=0, le=100000, description="Günlük brüt tavanı TL")


@router.get("/hr/settings/severance-cap")
async def get_severance_cap(
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    s = await db.tenant_settings.find_one(
        {'tenant_id': current_user.tenant_id}, {'_id': 0, 'severance_daily_cap': 1, 'severance_cap_updated_at': 1},
    ) or {}
    return {
        'daily_cap': s.get('severance_daily_cap') or TR_SEVERANCE_DAILY_CAP_DEFAULT,
        'is_default': not s.get('severance_daily_cap'),
        'updated_at': s.get('severance_cap_updated_at'),
        'monthly_cap_estimate': round((s.get('severance_daily_cap') or TR_SEVERANCE_DAILY_CAP_DEFAULT) * 30, 2),
        'note': "Hazine yarıyıl açıklamasıyla (Ocak/Temmuz) güncellenmeli. Boş bırakılırsa default tavan uygulanır.",
    }


@router.put("/hr/settings/severance-cap")
async def set_severance_cap(
    payload: SeveranceCapPayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    await db.tenant_settings.update_one(
        {'tenant_id': current_user.tenant_id},
        {'$set': {
            'severance_daily_cap': payload.daily_cap,
            'severance_cap_updated_at': datetime.now(UTC).isoformat(),
            'severance_cap_updated_by': getattr(current_user, 'id', None),
        }},
        upsert=True,
    )
    return {'success': True, 'daily_cap': payload.daily_cap}


# ============= 7. Performans Hedef Check-in =============

class GoalCheckinPayload(BaseModel):
    goal_text: str = Field(..., min_length=1, max_length=1000)
    progress_pct: int = Field(..., ge=0, le=100)
    status: Literal['on_track', 'at_risk', 'blocked', 'done'] = 'on_track'
    note: str | None = Field(None, max_length=2000)
    checkin_date: str | None = Field(None, pattern=r'^\d{4}-\d{2}-\d{2}$')


@router.post("/hr/performance/{review_id}/checkin")
async def add_goal_checkin(
    review_id: str,
    payload: GoalCheckinPayload,
    current_user: User = Depends(get_current_user),
):
    review = await db.performance_reviews.find_one({
        'tenant_id': current_user.tenant_id, 'id': review_id,
    })
    if not review:
        raise HTTPException(status_code=404, detail="Performans değerlendirmesi bulunamadı")
    item = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'review_id': review_id,
        'staff_id': review.get('staff_id'),
        'goal_text': payload.goal_text,
        'progress_pct': payload.progress_pct,
        'status': payload.status,
        'note': payload.note,
        'checkin_date': payload.checkin_date or _today_local().isoformat(),
        'created_by': getattr(current_user, 'id', None),
        'created_at': datetime.now(UTC).isoformat(),
    }
    await db.performance_checkins.insert_one(item)
    item.pop('_id', None)
    return {'success': True, 'checkin': item}


@router.get("/hr/performance/{review_id}/checkins")
async def list_goal_checkins(
    review_id: str,
    current_user: User = Depends(get_current_user),
):
    items = await db.performance_checkins.find({
        'tenant_id': current_user.tenant_id, 'review_id': review_id,
    }, {'_id': 0}).sort('checkin_date', -1).to_list(500)
    return {'items': items, 'total': len(items)}


@router.delete("/hr/performance/checkins/{checkin_id}")
async def delete_goal_checkin(
    checkin_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    res = await db.performance_checkins.delete_one({
        'tenant_id': current_user.tenant_id, 'id': checkin_id,
    })
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Check-in bulunamadı")
    return {'success': True}


# ============= Coverage Rules + Weekly Hours (Task #263 / T002) =============

class CoverageRulePayload(BaseModel):
    """Departman-bazlı minimum kapasite kuralı.

    weekday: 0-6 (Pzt-Pzr) veya -1 (her gün).
    shift_type: 'morning'|'afternoon'|'evening'|'night'|'split'|'any'.
    """
    department: str = Field(..., min_length=1, max_length=100)
    weekday: int = Field(..., ge=-1, le=6)
    shift_type: Literal['morning', 'afternoon', 'evening', 'night', 'split', 'any'] = 'any'
    min_staff: int = Field(..., ge=1, le=200)
    note: str | None = Field(None, max_length=300)


@router.post("/hr/coverage-rules")
async def create_coverage_rule(
    payload: CoverageRulePayload,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    item = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'department': payload.department,
        'weekday': payload.weekday,
        'shift_type': payload.shift_type,
        'min_staff': payload.min_staff,
        'note': payload.note,
        'created_by': getattr(current_user, 'id', None),
        'created_at': datetime.now(UTC).isoformat(),
    }
    await db.hr_coverage_rules.insert_one(item)
    item.pop('_id', None)
    return {'success': True, 'rule': item}


@router.get("/hr/coverage-rules")
async def list_coverage_rules(
    department: str | None = None,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    query: dict[str, Any] = {'tenant_id': current_user.tenant_id}
    if department:
        query['department'] = department
    items = await db.hr_coverage_rules.find(
        query, {'_id': 0}
    ).sort([('department', 1), ('weekday', 1)]).to_list(500)
    return {'items': items, 'total': len(items)}


@router.delete("/hr/coverage-rules/{rule_id}")
async def delete_coverage_rule(
    rule_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    res = await db.hr_coverage_rules.delete_one({
        'tenant_id': current_user.tenant_id, 'id': rule_id,
    })
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kural bulunamadı")
    return {'success': True}


async def _build_staff_department_map(tenant_id: str) -> dict[str, str]:
    """staff_id → department lookup (staff_members + users.role fallback)."""
    out: dict[str, str] = {}
    async for s in db.staff_members.find(
        {'tenant_id': tenant_id}, {'_id': 0, 'id': 1, 'department': 1},
    ):
        out[s['id']] = s.get('department') or 'unknown'
    async for u in db.users.find(
        {'tenant_id': tenant_id, 'is_active': True},
        {'_id': 0, 'id': 1, 'role': 1},
    ):
        if u['id'] not in out:
            out[u['id']] = u.get('role') or 'staff'
    return out


@router.get("/hr/coverage/check")
async def coverage_check(
    start: str = Query(..., pattern=r'^\d{4}-\d{2}-\d{2}$'),
    end: str = Query(..., pattern=r'^\d{4}-\d{2}-\d{2}$'),
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """Date×dept×shift_type matrisinde gerçekleşen sayım vs kural min_staff
    kıyaslaması. `gaps` listesi coverage altındaki kombinasyonları döner."""
    start_d = date.fromisoformat(start)
    end_d = date.fromisoformat(end)
    if end_d < start_d:
        raise HTTPException(status_code=400, detail="end < start")
    rules = await db.hr_coverage_rules.find(
        {'tenant_id': current_user.tenant_id}, {'_id': 0}
    ).to_list(500)
    if not rules:
        return {'rules_count': 0, 'gaps': [], 'days_checked': (end_d - start_d).days + 1}

    shifts = await db.shift_schedules.find({
        'tenant_id': current_user.tenant_id,
        'shift_date': {'$gte': start, '$lte': end},
        'status': {'$ne': 'on_leave'},
    }, {'_id': 0, 'staff_id': 1, 'shift_date': 1, 'shift_type': 1, 'status': 1}).to_list(20000)
    staff_dept = await _build_staff_department_map(current_user.tenant_id)

    # bucket[(date,dept,shift_type)] = set(staff_ids)
    bucket: dict[tuple[str, str, str], set[str]] = {}
    for s in shifts:
        dept = staff_dept.get(s['staff_id'], 'unknown')
        key = (s['shift_date'], dept, s.get('shift_type') or 'any')
        bucket.setdefault(key, set()).add(s['staff_id'])

    gaps: list[dict] = []
    cur = start_d
    while cur <= end_d:
        d_iso = cur.isoformat()
        wd = cur.weekday()
        for r in rules:
            if r['weekday'] not in (-1, wd):
                continue
            # count = belirli shift_type ise tek bucket; 'any' ise tüm shift_type'lar
            if r['shift_type'] == 'any':
                staff_set: set[str] = set()
                for k, v in bucket.items():
                    if k[0] == d_iso and k[1] == r['department']:
                        staff_set |= v
                actual = len(staff_set)
            else:
                actual = len(bucket.get((d_iso, r['department'], r['shift_type']), set()))
            if actual < r['min_staff']:
                gaps.append({
                    'date': d_iso,
                    'weekday': wd,
                    'department': r['department'],
                    'shift_type': r['shift_type'],
                    'min_staff': r['min_staff'],
                    'actual': actual,
                    'gap': r['min_staff'] - actual,
                    'rule_id': r['id'],
                })
        cur += timedelta(days=1)
    return {
        'range': {'start': start, 'end': end},
        'rules_count': len(rules),
        'gaps': gaps,
        'gaps_count': len(gaps),
    }


@router.get("/hr/shifts/weekly-hours")
async def shifts_weekly_hours(
    week_start: str = Query(..., pattern=r'^\d{4}-\d{2}-\d{2}$'),
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """Hafta başlangıcından 7 gün için per-staff toplam planlı saat ve
    >45h üzeri 'overtime estimate' uyarı flag'i."""
    start_d = date.fromisoformat(week_start)
    end_d = start_d + timedelta(days=6)
    items = await db.shift_schedules.find({
        'tenant_id': current_user.tenant_id,
        'shift_date': {'$gte': start_d.isoformat(), '$lte': end_d.isoformat()},
        'status': {'$ne': 'on_leave'},
    }, {'_id': 0, 'staff_id': 1, 'staff_name': 1, 'shift_date': 1,
        'start_time': 1, 'end_time': 1, 'crosses_midnight': 1}).to_list(5000)

    def _hours_of(sh: dict) -> float:
        try:
            sdt, edt = _shift_datetimes(
                sh['shift_date'], sh['start_time'], sh['end_time'],
                bool(sh.get('crosses_midnight', False)),
            )
            return round((edt - sdt).total_seconds() / 3600.0, 2)
        except Exception:
            return 0.0

    agg: dict[str, dict] = {}
    for sh in items:
        sid = sh['staff_id']
        agg.setdefault(sid, {
            'staff_id': sid,
            'staff_name': sh.get('staff_name'),
            'total_hours': 0.0,
            'shifts_count': 0,
        })
        agg[sid]['total_hours'] = round(agg[sid]['total_hours'] + _hours_of(sh), 2)
        agg[sid]['shifts_count'] += 1
    out: list[dict] = []
    for v in agg.values():
        v['overtime_estimate'] = round(max(0.0, v['total_hours'] - 45.0), 2)
        v['exceeds_legal_week'] = v['total_hours'] > 45.0
        out.append(v)
    out.sort(key=lambda x: (-x['total_hours'], x['staff_name'] or ''))
    return {
        'week_start': start_d.isoformat(),
        'week_end': end_d.isoformat(),
        'legal_weekly_hours': 45,
        'items': out,
        'total_staff': len(out),
    }


# ============= Attendance v2 — Missing Clockout + Late/Early + Dept Summary (Task #263 / T005) =============

@router.post("/hr/attendance/flag-missing")
async def flag_missing_clockouts(
    cutoff_date: str | None = None,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("manage_hr")),
):
    """Cron tetikleyici: belirtilen tarihten (varsayılan: dün) önceki
    `clock_out=null` kayıtlarını `missing_clockout=True` ile flag'ler.
    Bugünün açık kayıtları (henüz mesai bitmedi) DOKUNULMAZ."""
    cutoff = (
        date.fromisoformat(cutoff_date) if cutoff_date
        else _today_local() - timedelta(days=1)
    )
    res = await db.attendance_records.update_many(
        {
            'tenant_id': current_user.tenant_id,
            'date': {'$lte': cutoff.isoformat()},
            'clock_out': None,
            '$or': [
                {'missing_clockout': {'$exists': False}},
                {'missing_clockout': False},
            ],
        },
        {'$set': {
            'missing_clockout': True,
            'missing_clockout_flagged_at': datetime.now(UTC).isoformat(),
        }},
    )
    return {
        'success': True,
        'flagged_count': res.modified_count,
        'cutoff_date': cutoff.isoformat(),
    }


def _parse_hhmm_on_date(d_iso: str, hhmm: str) -> datetime:
    sh, sm = (int(x) for x in hhmm.split(':'))
    dd = date.fromisoformat(d_iso)
    return datetime(dd.year, dd.month, dd.day, sh, sm, tzinfo=UTC)


def _derive_late_early(record: dict, shift: dict | None) -> dict:
    """Vardiya start/end ile karşılaştırarak late/early dakika hesabı."""
    out = {'late_minutes': 0, 'early_minutes': 0, 'shift_matched': False}
    if not shift:
        return out
    try:
        sched_start = _parse_hhmm_on_date(shift['shift_date'], shift['start_time'])
        end_dt = date.fromisoformat(shift['shift_date']) + (
            timedelta(days=1) if shift.get('crosses_midnight') else timedelta()
        )
        sched_end = datetime(
            end_dt.year, end_dt.month, end_dt.day,
            int(shift['end_time'].split(':')[0]),
            int(shift['end_time'].split(':')[1]),
            tzinfo=UTC,
        )
    except Exception:
        return out
    out['shift_matched'] = True
    ci = record.get('clock_in')
    co = record.get('clock_out')
    if ci:
        try:
            ci_dt = datetime.fromisoformat(ci.replace('Z', '+00:00'))
            delta_min = int((ci_dt - sched_start).total_seconds() / 60)
            if delta_min > 0:
                out['late_minutes'] = delta_min
        except Exception:
            pass
    if co:
        try:
            co_dt = datetime.fromisoformat(co.replace('Z', '+00:00'))
            delta_min = int((sched_end - co_dt).total_seconds() / 60)
            if delta_min > 0:
                out['early_minutes'] = delta_min
        except Exception:
            pass
    return out


@router.get("/hr/attendance/department-summary")
async def attendance_department_summary(
    start: str | None = None,
    end: str | None = None,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """Departman bazlı çalışma saati özet kartları."""
    start_dt, end_dt = _parse_date_range(start, end, days=30)
    records = await db.attendance_records.find({
        'tenant_id': current_user.tenant_id,
        'date': {'$gte': start_dt.isoformat(), '$lte': end_dt.isoformat()},
    }, {'_id': 0}).to_list(5000)
    staff_dept = await _build_staff_department_map(current_user.tenant_id)
    agg: dict[str, dict] = {}
    for r in records:
        dept = staff_dept.get(r['staff_id'], 'unknown')
        agg.setdefault(dept, {
            'department': dept,
            'total_hours': 0.0,
            'records_count': 0,
            'staff_set': set(),
            'missing_clockouts': 0,
        })
        agg[dept]['total_hours'] = round(
            agg[dept]['total_hours'] + float(r.get('total_hours') or 0), 2,
        )
        agg[dept]['records_count'] += 1
        agg[dept]['staff_set'].add(r['staff_id'])
        if r.get('missing_clockout'):
            agg[dept]['missing_clockouts'] += 1
    out = []
    for v in agg.values():
        v['unique_staff'] = len(v['staff_set'])
        del v['staff_set']
        out.append(v)
    out.sort(key=lambda x: -x['total_hours'])
    return {
        'range': {'start': start_dt.isoformat(), 'end': end_dt.isoformat()},
        'departments': out,
    }


# ============= Excel Exports (Task #263 / T005) =============

def _xlsx_stream(workbook) -> StreamingResponse:
    """openpyxl Workbook → StreamingResponse helper. Filename caller'a ait."""
    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/hr/shifts/export/xlsx")
async def export_shifts_xlsx(
    start: str = Query(..., pattern=r'^\d{4}-\d{2}-\d{2}$'),
    end: str = Query(..., pattern=r'^\d{4}-\d{2}-\d{2}$'),
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """Vardiya planı Excel — departman başına ayrı sayfa."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    items = await db.shift_schedules.find({
        'tenant_id': current_user.tenant_id,
        'shift_date': {'$gte': start, '$lte': end},
    }, {'_id': 0}).sort('shift_date', 1).to_list(20000)
    staff_dept = await _build_staff_department_map(current_user.tenant_id)
    wb = Workbook()
    wb.remove(wb.active)
    by_dept: dict[str, list[dict]] = {}
    for it in items:
        d = staff_dept.get(it['staff_id'], 'unknown')
        by_dept.setdefault(d, []).append(it)
    if not by_dept:
        ws = wb.create_sheet('Boş')
        ws['A1'] = 'Bu aralıkta vardiya yok'
    else:
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        for dept, rows in sorted(by_dept.items()):
            sheet_name = (dept or 'Bölüm')[:30]
            ws = wb.create_sheet(sheet_name)
            headers = ['Tarih', 'Personel', 'Vardiya', 'Başlangıç', 'Bitiş', 'Gece', 'Durum', 'Not']
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font = header_font
                c.fill = header_fill
            for r_idx, it in enumerate(rows, 2):
                ws.cell(row=r_idx, column=1, value=it.get('shift_date'))
                ws.cell(row=r_idx, column=2, value=it.get('staff_name') or it.get('staff_id'))
                ws.cell(row=r_idx, column=3, value=it.get('shift_type'))
                ws.cell(row=r_idx, column=4, value=it.get('start_time'))
                ws.cell(row=r_idx, column=5, value=it.get('end_time'))
                ws.cell(row=r_idx, column=6, value='Evet' if it.get('crosses_midnight') else 'Hayır')
                ws.cell(row=r_idx, column=7, value=it.get('status'))
                ws.cell(row=r_idx, column=8, value=(it.get('notes') or '')[:200])
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col_idx)].width = 16
    resp = _xlsx_stream(wb)
    resp.headers['Content-Disposition'] = f'attachment; filename="vardiyalar_{start}_{end}.xlsx"'
    return resp


@router.get("/hr/attendance/export/xlsx")
async def export_attendance_xlsx(
    month: str = Query(..., pattern=r'^\d{4}-\d{2}$'),
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """Puantaj aylık: personel × gün matrisi (saat değerleri)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    yyyy, mm = month.split('-')
    start = date(int(yyyy), int(mm), 1)
    nm = start.replace(day=28) + timedelta(days=4)
    end = nm - timedelta(days=nm.day)
    records = await db.attendance_records.find({
        'tenant_id': current_user.tenant_id,
        'date': {'$gte': start.isoformat(), '$lte': end.isoformat()},
    }, {'_id': 0}).to_list(10000)
    staff_map = await _get_staff_map(current_user.tenant_id)
    days_in_month = (end - start).days + 1
    # matrix[staff_id][day_num] = hours
    matrix: dict[str, dict[int, float]] = {}
    name_map: dict[str, str] = {}
    for r in records:
        sid = r['staff_id']
        d_num = int(r['date'][-2:])
        matrix.setdefault(sid, {})
        matrix[sid][d_num] = round(matrix[sid].get(d_num, 0.0) + float(r.get('total_hours') or 0), 2)
        if sid not in name_map:
            sm = staff_map.get(sid)
            name_map[sid] = (sm or {}).get('name') or r.get('staff_name') or sid

    wb = Workbook()
    ws = wb.active
    ws.title = f'Puantaj {month}'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    ws.cell(row=1, column=1, value='Personel').font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    for d in range(1, days_in_month + 1):
        c = ws.cell(row=1, column=1 + d, value=d)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
    total_col = days_in_month + 2
    tc = ws.cell(row=1, column=total_col, value='Toplam')
    tc.font = header_font
    tc.fill = header_fill
    row_idx = 2
    for sid in sorted(matrix.keys(), key=lambda s: name_map.get(s, '')):
        ws.cell(row=row_idx, column=1, value=name_map.get(sid, sid))
        total = 0.0
        for d in range(1, days_in_month + 1):
            v = matrix[sid].get(d, 0)
            if v:
                ws.cell(row=row_idx, column=1 + d, value=v)
                total += v
        ws.cell(row=row_idx, column=total_col, value=round(total, 2)).font = Font(bold=True)
        row_idx += 1
    ws.column_dimensions['A'].width = 28
    resp = _xlsx_stream(wb)
    resp.headers['Content-Disposition'] = f'attachment; filename="puantaj_{month}.xlsx"'
    return resp


@router.get("/hr/leave/export/xlsx")
async def export_leave_xlsx(
    start: str = Query(..., pattern=r'^\d{4}-\d{2}-\d{2}$'),
    end: str = Query(..., pattern=r'^\d{4}-\d{2}-\d{2}$'),
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """İzin geçmişi Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    # Task #263 review carryover: range overlap doctrini —
    # `start_date<=end AND end_date>=start` (eski $or pencereyi tamamen
    # kapsayan izinleri kaçırıyordu).
    items = await db.leave_requests.find({
        'tenant_id': current_user.tenant_id,
        'start_date': {'$lte': end},
        'end_date': {'$gte': start},
    }, {'_id': 0}).sort('start_date', -1).to_list(5000)
    wb = Workbook()
    ws = wb.active
    ws.title = 'İzin Geçmişi'
    headers = ['Personel', 'Tip', 'Başlangıç', 'Bitiş', 'Gün', 'Durum', 'Sebep', 'Onay Notu']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    for r_idx, it in enumerate(items, 2):
        ws.cell(row=r_idx, column=1, value=it.get('staff_name') or it.get('staff_id'))
        ws.cell(row=r_idx, column=2, value=it.get('leave_type'))
        ws.cell(row=r_idx, column=3, value=it.get('start_date'))
        ws.cell(row=r_idx, column=4, value=it.get('end_date'))
        ws.cell(row=r_idx, column=5, value=it.get('total_days'))
        ws.cell(row=r_idx, column=6, value=it.get('status'))
        ws.cell(row=r_idx, column=7, value=(it.get('reason') or '')[:200])
        ws.cell(row=r_idx, column=8, value=(it.get('decision_note') or '')[:200])
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18
    resp = _xlsx_stream(wb)
    resp.headers['Content-Disposition'] = f'attachment; filename="izinler_{start}_{end}.xlsx"'
    return resp


@router.get("/hr/overtime/export/xlsx")
async def export_overtime_xlsx(
    year: int = Query(..., ge=2020, le=2100),
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_hr")),
):
    """Mesai geçmişi Excel — yıl bazlı."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    items = await db.overtime_requests.find({
        'tenant_id': current_user.tenant_id,
        'work_date': {'$gte': f'{year}-01-01', '$lte': f'{year}-12-31'},
    }, {'_id': 0}).sort('work_date', -1).to_list(5000)
    wb = Workbook()
    ws = wb.active
    ws.title = f'Mesai {year}'
    headers = ['Personel', 'Tarih', 'Saat', 'Sebep', 'Durum', 'Onay Notu', 'Bordro Hazır']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    for r_idx, it in enumerate(items, 2):
        ws.cell(row=r_idx, column=1, value=it.get('staff_name') or it.get('staff_id'))
        ws.cell(row=r_idx, column=2, value=it.get('work_date'))
        ws.cell(row=r_idx, column=3, value=float(it.get('hours') or 0))
        ws.cell(row=r_idx, column=4, value=(it.get('reason') or '')[:200])
        ws.cell(row=r_idx, column=5, value=it.get('status'))
        ws.cell(row=r_idx, column=6, value=(it.get('decision_note') or '')[:200])
        ws.cell(row=r_idx, column=7, value='Evet' if it.get('payroll_ready') else 'Hayır')
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18
    resp = _xlsx_stream(wb)
    resp.headers['Content-Disposition'] = f'attachment; filename="mesai_{year}.xlsx"'
    return resp

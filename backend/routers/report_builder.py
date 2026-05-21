"""
Report Builder Router - Özel Rapor Oluşturucu (Tur 28+ hardened).

Kullanıcıların dinamik rapor oluşturmasını, filtrelemesini ve dışa aktarmasını
sağlar. P0/P1 düzeltmeleri:
  - DB alan adlarıyla uyumlu projection (room_number/room_type/base_price/
    capacity, charge_category, booking_source, vip_status vb.) + alias
    fallback (eski tenant verilerinde alternatif alan adları).
  - Tarih filtresi end-of-day çevirisi (T23:59:59).
  - Filtre değer tip dönüşümü (number/boolean/date) — string karşılaştırma
    sessizce boş döndüğü için.
  - sort_by allow-list (yalnızca tanımlı sütunlar) + limit cap (1..5000).
  - Permission gate: tüm endpoint'ler `view_reports`; PII sütunları
    (id_number, passport_number, email, phone) admin/manager/granted
    olmayan kullanıcılarda otomatik maskelenir.
  - Şablon silmede created_by kontrolü (sadece sahibi veya admin/manager).
  - Nested ObjectId temizliği (recursive).
  - Excel TOPLAM yanlış sütuna düşmesi (col 1 numeric ise üzerine yazıyordu).
  - PDF sütun sayısı çok fazla ise font otomatik küçülür.
"""
import io
import logging
import uuid
from datetime import UTC, date, datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer
from pydantic import BaseModel

from modules.pms_core.role_permission_service import require_op  # v92 DW

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/reports/builder", tags=["report-builder"])


# ─── Models ───────────────────────────────────────────────────────────────

class ReportFilter(BaseModel):
    field: str
    operator: str  # eq, ne, gt, gte, lt, lte, in, contains
    value: object


class ReportConfig(BaseModel):
    data_source: str  # reservations, revenue, guests, rooms, housekeeping, folios
    columns: list[str]
    filters: list[ReportFilter] | None = []
    sort_by: str | None = None
    sort_order: str | None = "desc"
    date_from: str | None = None
    date_to: str | None = None
    group_by: str | None = None  # ileride kullanılmak üzere; şu an no-op
    limit: int | None = 500


class SavedTemplate(BaseModel):
    name: str
    description: str | None = ""
    config: ReportConfig


# ─── Dependencies (will be injected from server.py) ──────────────────────

_db = None
_get_current_user = None


def init_report_builder(db, get_current_user_dep):
    global _db, _get_current_user
    _db = db
    _get_current_user = get_current_user_dep


def get_db():
    return _db


# ─── Data Source Definitions ──────────────────────────────────────────────

DATA_SOURCES = {
    "reservations": {
        "label": "Rezervasyonlar",
        "collection": "bookings",
        "columns": {
            "guest_name": {"label": "Misafir Adı", "type": "text"},
            "room_number": {"label": "Oda No", "type": "text"},
            "room_type": {"label": "Oda Tipi", "type": "text"},
            "check_in": {"label": "Giriş Tarihi", "type": "date"},
            "check_out": {"label": "Çıkış Tarihi", "type": "date"},
            "status": {"label": "Durum", "type": "select", "options": ["confirmed", "checked_in", "checked_out", "cancelled", "no_show"]},
            "total_amount": {"label": "Toplam Tutar", "type": "currency"},
            "source": {"label": "Kaynak", "type": "select", "options": ["direct", "ota", "corporate", "walk_in", "booking_com", "hotelrunner", "exely"]},
            "nights": {"label": "Gece Sayısı", "type": "number"},
            "adults": {"label": "Yetişkin", "type": "number"},
            "children": {"label": "Çocuk", "type": "number"},
            "rate_code": {"label": "Ücret Kodu", "type": "text"},
            "market_segment": {"label": "Pazar Segmenti", "type": "text"},
            "created_at": {"label": "Oluşturma Tarihi", "type": "date"},
            "notes": {"label": "Notlar", "type": "text"},
        },
        "date_field": "check_in",
    },
    "revenue": {
        "label": "Gelir",
        "collection": "folio_charges",
        "columns": {
            "description": {"label": "Açıklama", "type": "text"},
            "amount": {"label": "Tutar", "type": "currency"},
            "total": {"label": "Toplam", "type": "currency"},
            "charge_type": {"label": "Masraf Tipi", "type": "select", "options": ["room", "fnb", "minibar", "laundry", "spa", "parking", "phone", "other"]},
            "date": {"label": "Tarih", "type": "date"},
            "room_number": {"label": "Oda No", "type": "text"},
            "folio_id": {"label": "Folio ID", "type": "text"},
            "quantity": {"label": "Adet", "type": "number"},
            "unit_price": {"label": "Birim Fiyat", "type": "currency"},
            "voided": {"label": "İptal Edildi", "type": "boolean"},
            "posted_by": {"label": "İşlemi Yapan", "type": "text"},
        },
        "date_field": "date",
    },
    "guests": {
        "label": "Misafirler",
        "collection": "guests",
        "columns": {
            "name": {"label": "Ad Soyad", "type": "text"},
            "email": {"label": "E-posta", "type": "text"},
            "phone": {"label": "Telefon", "type": "text"},
            "nationality": {"label": "Uyruk", "type": "text"},
            "id_number": {"label": "TC/Pasaport No", "type": "text"},
            "vip": {"label": "VIP", "type": "boolean"},
            "gender": {"label": "Cinsiyet", "type": "text"},
            "total_stays": {"label": "Toplam Konaklama", "type": "number"},
            "total_revenue": {"label": "Toplam Harcama", "type": "currency"},
            "created_at": {"label": "Kayıt Tarihi", "type": "date"},
            "notes": {"label": "Notlar", "type": "text"},
        },
        "date_field": "created_at",
    },
    "rooms": {
        "label": "Odalar",
        "collection": "rooms",
        "columns": {
            "number": {"label": "Oda No", "type": "text"},
            "type": {"label": "Oda Tipi", "type": "text"},
            "floor": {"label": "Kat", "type": "number"},
            "status": {"label": "Durum", "type": "select", "options": ["available", "occupied", "dirty", "maintenance", "out_of_order"]},
            "housekeeping_status": {"label": "HK Durumu", "type": "text"},
            "base_rate": {"label": "Taban Fiyat", "type": "currency"},
            "max_occupancy": {"label": "Max Kapasite", "type": "number"},
            "amenities": {"label": "Olanaklar", "type": "text"},
            "is_active": {"label": "Aktif", "type": "boolean"},
        },
        "date_field": None,
    },
    "housekeeping": {
        "label": "Kat Hizmetleri",
        "collection": "housekeeping_tasks",
        "columns": {
            "room_number": {"label": "Oda No", "type": "text"},
            "task_type": {"label": "Görev Tipi", "type": "select", "options": ["checkout_clean", "stayover_clean", "deep_clean", "turndown", "inspection"]},
            "status": {"label": "Durum", "type": "select", "options": ["pending", "in_progress", "completed", "inspected"]},
            "assigned_to": {"label": "Atanan Kişi", "type": "text"},
            "priority": {"label": "Öncelik", "type": "select", "options": ["low", "medium", "high", "urgent"]},
            "started_at": {"label": "Başlangıç", "type": "date"},
            "completed_at": {"label": "Tamamlanma", "type": "date"},
            "duration_minutes": {"label": "Süre (dk)", "type": "number"},
            "notes": {"label": "Notlar", "type": "text"},
        },
        "date_field": "created_at",
    },
    "folios": {
        "label": "Foliolar",
        "collection": "folios",
        "columns": {
            "folio_number": {"label": "Folio No", "type": "text"},
            "guest_name": {"label": "Misafir Adı", "type": "text"},
            "room_number": {"label": "Oda No", "type": "text"},
            "status": {"label": "Durum", "type": "select", "options": ["open", "closed", "settled"]},
            "total_charges": {"label": "Toplam Masraf", "type": "currency"},
            "total_payments": {"label": "Toplam Ödeme", "type": "currency"},
            "balance": {"label": "Bakiye", "type": "currency"},
            "check_in": {"label": "Giriş", "type": "date"},
            "check_out": {"label": "Çıkış", "type": "date"},
            "created_at": {"label": "Oluşturma", "type": "date"},
            "payment_method": {"label": "Ödeme Yöntemi", "type": "text"},
        },
        "date_field": "created_at",
    },
}


# ─── Field aliases: column key → list of candidate DB field names ────────
# İlk dolu olan değer kullanılır. Eski tenant verilerinde alternatif alan
# isimleri olabilir; bu sayede projection sessizce boş dönmez.
SOURCE_FIELD_MAP: dict[str, dict[str, list[str]]] = {
    "reservations": {
        "source": ["booking_source", "source_channel", "source", "channel"],
        "rate_code": ["rate_plan", "rate_code"],
        "id_number": ["id_number", "national_id"],
    },
    "guests": {
        "name": ["name", "full_name"],
        "id_number": ["id_number", "national_id", "passport_number"],
        "vip": ["vip_status", "vip"],
        "phone": ["phone", "phone_number"],
        "total_stays": ["total_stays", "stay_count"],
        "total_revenue": ["total_revenue", "lifetime_value"],
    },
    "rooms": {
        "number": ["room_number", "number"],
        "type": ["room_type", "type"],
        "base_rate": ["base_price", "base_rate", "price_per_night"],
        "max_occupancy": ["capacity", "max_occupancy"],
    },
    "revenue": {
        "charge_type": ["charge_category", "charge_type"],
    },
    "folios": {
        "total_charges": ["total_charges", "charges_total"],
        "total_payments": ["total_payments", "payments_total"],
    },
}

# PII column keys per source (column-level masking when no PII access).
PII_COLUMNS: dict[str, set[str]] = {
    "reservations": {"id_number", "passport_number", "guest_email", "guest_phone"},
    "guests": {"id_number", "email", "phone"},
    "folios": {"guest_name"},  # name itself is sensitive in some KVKK contexts
}

# Maximum allowed result rows per request (DoS guard).
MAX_LIMIT = 5000

# Maximum columns for landscape PDF (above this, font shrinks).
PDF_FIT_COLUMNS = 8


# ─── PII helpers ─────────────────────────────────────────────────────────

def _user_has_pii_access(user) -> bool:
    role = getattr(user, 'role', None)
    role_str = getattr(role, 'value', None) or str(role or '')
    if role_str in ('admin', 'super_admin', 'manager', 'general_manager'):
        return True
    granted = getattr(user, 'granted_permissions', None) or []
    return 'view_guest_pii' in granted


def _mask_pii(value):
    if value is None or value == "":
        return value
    s = str(value)
    if len(s) <= 4:
        return '*' * len(s)
    return s[:2] + '*' * (len(s) - 4) + s[-2:]


# ─── Type coercion for filter values ─────────────────────────────────────

def _coerce_value(value, col_type: str | None):
    """Backend filter value tipi her zaman string gelir (HTML input). DB
    karşılaştırması doğru tip ile yapılmazsa sessiz boş döner.
    """
    if value is None:
        return value
    if isinstance(value, list):
        return [_coerce_value(v, col_type) for v in value]
    s = str(value).strip()
    if s == "":
        return s
    try:
        if col_type == "number":
            if "." in s:
                return float(s)
            return int(s)
        if col_type == "currency":
            return float(s)
        if col_type == "boolean":
            return s.lower() in ("true", "1", "yes", "evet", "var")
        if col_type == "date":
            # Tarih filtresinde tam ISO string ile karşılaştırma yeterli.
            return s
        return value
    except (ValueError, TypeError):
        return value


# ─── Mongo filter / projection ───────────────────────────────────────────

def _resolve_db_field(source_key: str, col_key: str) -> str:
    """API column key → birinci DB alan adı (legacy fallback için tüm
    listeyi `_resolve_db_fields` ile alın).
    """
    return _resolve_db_fields(source_key, col_key)[0]


def _resolve_db_fields(source_key: str, col_key: str) -> list[str]:
    aliases = SOURCE_FIELD_MAP.get(source_key, {}).get(col_key)
    if aliases:
        return aliases
    return [col_key]


def build_mongo_filter(config: ReportConfig, tenant_id: str) -> dict:
    """Build MongoDB query filter from ReportConfig.
    Column-key referansları DB alan adına çözülür ve değerler tipe göre
    cast edilir.
    """
    query: dict = {"tenant_id": tenant_id}
    source_key = config.data_source
    source_def = DATA_SOURCES.get(source_key, {})
    cols = source_def.get("columns", {})

    # Date range filter (end-of-day düzeltmesi).
    date_field = source_def.get("date_field")
    if date_field and (config.date_from or config.date_to):
        date_q: dict = {}
        if config.date_from:
            df = str(config.date_from).strip()
            # `T` yoksa gün başlangıcına çevir.
            if "T" not in df:
                df = df + "T00:00:00"
            date_q["$gte"] = df
        if config.date_to:
            dt = str(config.date_to).strip()
            if "T" not in dt:
                dt = dt + "T23:59:59"
            date_q["$lte"] = dt
        query[date_field] = date_q

    # Custom filters (allow-list field, type-aware value).
    for f in (config.filters or []):
        if f.field not in cols:
            # Sadece tanımlı sütunlar üzerinde filtre. Bilinmeyen alanları
            # sessizce düşür (silent drop yerine reddedilebilir; sızıntı
            # vektörünü kapatmak için bilinmeyen alan kabul edilmiyor).
            continue
        col_type = cols.get(f.field, {}).get("type")
        db_field = _resolve_db_field(source_key, f.field)
        op = f.operator
        val = _coerce_value(f.value, col_type)

        # `contains` haricinde boş değer filtreyi düşürmesin — boş string
        # yine `eq ""` olabilir; ama None ise atla.
        if val is None:
            continue

        if op == "eq":
            query[db_field] = val
        elif op == "ne":
            query[db_field] = {"$ne": val}
        elif op == "gt":
            query[db_field] = {"$gt": val}
        elif op == "gte":
            query[db_field] = {"$gte": val}
        elif op == "lt":
            query[db_field] = {"$lt": val}
        elif op == "lte":
            query[db_field] = {"$lte": val}
        elif op == "in":
            query[db_field] = {"$in": val if isinstance(val, list) else [val]}
        elif op == "contains":
            from security.query_safety import safe_search_term
            _v = safe_search_term(str(val))
            if _v:
                query[db_field] = {"$regex": _v, "$options": "i"}
            # Aksi halde filtreyi DROP etmiyoruz — eşleşmeyecek bir
            # sentinel kullanıyoruz ki "tüm kayıtlar gelir" sızıntısı
            # olmasın.
            else:
                query[db_field] = {"$exists": False, "$type": "null"}

    return query


def _projection_fields(source_key: str, columns: list[str]) -> set[str]:
    """Tüm gerekli DB alanlarını (alias ve compute girdileri dahil) toplar."""
    fields: set[str] = set()
    for col in columns:
        for alias in _resolve_db_fields(source_key, col):
            fields.add(alias)
        # Compute girdileri:
        if source_key == "reservations" and col == "nights":
            fields.update({"check_in", "check_out", "nights"})
        if source_key == "guests" and col == "name":
            fields.update({"first_name", "last_name"})
    return fields


def build_projection(source_key: str, columns: list[str]) -> dict:
    proj = {"_id": 0}
    for f in _projection_fields(source_key, columns):
        proj[f] = 1
    return proj


# ─── Cleaning helpers ────────────────────────────────────────────────────

def _clean_value(v):
    """Recursive ObjectId / datetime serialize-safe çevirme."""
    if v is None:
        return None
    t = type(v).__name__
    if t == "ObjectId":
        return str(v)
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, dict):
        return {k: _clean_value(val) for k, val in v.items()}
    if isinstance(v, list):
        return [_clean_value(x) for x in v]
    return v


def _compute_nights(doc: dict) -> int | None:
    if doc.get("nights"):
        try:
            return int(doc["nights"])
        except (TypeError, ValueError):
            pass
    ci, co = doc.get("check_in"), doc.get("check_out")
    if not ci or not co:
        return None
    try:
        ci_d = ci if isinstance(ci, datetime) else datetime.fromisoformat(str(ci).replace("Z", "+00:00"))
        co_d = co if isinstance(co, datetime) else datetime.fromisoformat(str(co).replace("Z", "+00:00"))
        n = (co_d.date() - ci_d.date()).days
        return max(n, 1)
    except (ValueError, TypeError):
        return None


def _compose_guest_name(doc: dict) -> str | None:
    n = doc.get("name") or doc.get("full_name")
    if n:
        return n
    fn, ln = doc.get("first_name") or "", doc.get("last_name") or ""
    composed = f"{fn} {ln}".strip()
    return composed or None


# ─── Fetch ──────────────────────────────────────────────────────────────

async def fetch_report_data(config: ReportConfig, tenant_id: str, has_pii: bool) -> list:
    db = get_db()
    source_def = DATA_SOURCES.get(config.data_source)
    if not source_def:
        raise HTTPException(status_code=400, detail=f"Geçersiz veri kaynağı: {config.data_source}")

    cols_def = source_def.get("columns", {})
    # Bilinmeyen sütunları reddet.
    for c in config.columns:
        if c not in cols_def:
            raise HTTPException(status_code=400, detail=f"Bilinmeyen sütun: {c}")

    # sort_by allow-list.
    sort_field_key = config.sort_by if config.sort_by in cols_def else None
    sort_field_db = _resolve_db_field(config.data_source, sort_field_key) if sort_field_key else (source_def.get("date_field") or "_id")
    sort_dir = -1 if (config.sort_order or "desc") == "desc" else 1

    # Limit cap.
    raw_limit = config.limit or 500
    try:
        raw_limit = int(raw_limit)
    except (TypeError, ValueError):
        raw_limit = 500
    safe_limit = max(1, min(raw_limit, MAX_LIMIT))

    collection = db[source_def["collection"]]
    query = build_mongo_filter(config, tenant_id)
    projection = build_projection(config.data_source, config.columns)

    cursor = collection.find(query, projection).sort(sort_field_db, sort_dir).limit(safe_limit)
    raw = await cursor.to_list(length=safe_limit)

    pii_keys = PII_COLUMNS.get(config.data_source, set())

    cleaned: list[dict] = []
    for doc in raw:
        row: dict = {}
        for col in config.columns:
            # Compute virtual fields.
            if config.data_source == "reservations" and col == "nights":
                v = _compute_nights(doc)
            elif config.data_source == "guests" and col == "name":
                v = _compose_guest_name(doc)
            else:
                v = None
                for alias in _resolve_db_fields(config.data_source, col):
                    cand = doc.get(alias)
                    if cand not in (None, ""):
                        v = cand
                        break
            v = _clean_value(v)
            if (not has_pii) and col in pii_keys:
                v = _mask_pii(v)
            row[col] = v
        cleaned.append(row)

    return cleaned


# ─── Endpoints ────────────────────────────────────────────────────────────

@router.get("/config")
async def get_builder_config(
    credentials=Depends(HTTPBearer()),
    _perm=Depends(require_op("view_reports")),
):
    """Rapor oluşturucu için mevcut veri kaynaklarını ve sütun tanımlarını döndürür."""
    await _get_current_user(credentials)
    sources = {}
    for key, src in DATA_SOURCES.items():
        sources[key] = {
            "label": src["label"],
            "columns": src["columns"],
            "date_field": src.get("date_field"),
            "pii_columns": sorted(PII_COLUMNS.get(key, set())),
        }
    return {"data_sources": sources, "max_limit": MAX_LIMIT}


@router.post("/generate")
async def generate_report(
    config: ReportConfig,
    credentials=Depends(HTTPBearer()),
    _perm=Depends(require_op("view_reports")),
):
    """Özel rapor verisini üretir."""
    current_user = await _get_current_user(credentials)
    tenant_id = getattr(current_user, 'tenant_id', None)
    if not tenant_id:
        raise HTTPException(status_code=400, detail="Tenant bilgisi bulunamadı")
    has_pii = _user_has_pii_access(current_user)

    data = await fetch_report_data(config, tenant_id, has_pii)

    source_def = DATA_SOURCES.get(config.data_source, {})
    column_labels = {col: source_def.get("columns", {}).get(col, {}).get("label", col) for col in config.columns}

    summary = {}
    for col in config.columns:
        col_type = source_def.get("columns", {}).get(col, {}).get("type")
        if col_type in ("number", "currency"):
            values = [row.get(col) for row in data if isinstance(row.get(col), (int, float)) and not isinstance(row.get(col), bool)]
            if values:
                summary[col] = {
                    "sum": round(sum(values), 2),
                    "avg": round(sum(values) / len(values), 2),
                    "min": round(min(values), 2),
                    "max": round(max(values), 2),
                    "count": len(values),
                }

    return {
        "data": data,
        "total_count": len(data),
        "column_labels": column_labels,
        "summary": summary,
        "pii_masked": not has_pii,
        "generated_at": datetime.now(UTC).isoformat(),
    }


def _first_numeric_col_index(source_def: dict, columns: list[str]) -> int | None:
    """1-tabanlı sütun indeksi (Excel için)."""
    for i, c in enumerate(columns, 1):
        t = source_def.get("columns", {}).get(c, {}).get("type")
        if t in ("number", "currency"):
            return i
    return None


# Task #253 (tur-2 fix): openpyxl rejects C0 control chars (0x00-0x08,
# 0x0B-0x0C, 0x0E-0x1F) anywhere in a cell value with `IllegalCharacterError`,
# and any cell longer than 32767 chars also raises. Stress tenants accumulate
# residue across rounds (F8B complaints, free-text descriptions, byte arrays
# decoded with errors='replace') that can carry these characters even though
# the canonical factory seeds clean ASCII. We sanitize at the coerce boundary
# so every export path (builder + departments) inherits the guard.
_XLSX_MAX_CELL_LEN = 32767


def _xlsx_sanitize_str(s: str) -> str:
    """Strip openpyxl-illegal control chars and cap cell length.

    Idempotent. Returns "" for empty input. Length cap leaves an ellipsis
    sentinel so truncated cells are visible during audits instead of silently
    cut off mid-token.
    """
    if not s:
        return s
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    cleaned = ILLEGAL_CHARACTERS_RE.sub("", s)
    if len(cleaned) > _XLSX_MAX_CELL_LEN:
        cleaned = cleaned[: _XLSX_MAX_CELL_LEN - 1] + "…"
    return cleaned


def _coerce_excel_value(val):
    """Defensive value coercion for openpyxl cell writes (Task #246, #253).

    Stress tenants surface heterogeneous BSON values (Decimal, native
    datetime/date, ObjectId, NaN/Inf floats, bytes) that crash
    `xlsx_safe(str(val))` or `Cell.value =` silently. Normalizes to
    (typed_value, is_numeric) where typed_value is safe to assign.

    - None / missing → ""
    - bool → "Evet"/"Hayır"
    - int/float (finite, non-bool) → numeric, preserved for formatting
    - int/float (NaN/Inf) → string fallback
    - Decimal → float when finite, else string
    - datetime → ISO string (avoids tz-aware/naive mix crashes)
    - date → ISO string
    - bytes → utf-8 string (errors='replace')
    - list/tuple/set → comma-join of coerced items
    - dict → str(dict) string
    - other → str(val) string

    Task #253: every string-producing branch is run through
    `_xlsx_sanitize_str` so openpyxl's `IllegalCharacterError` and the
    32767-char hard limit can never escalate to a 500 on export. Numeric
    returns are unchanged (number_format compatibility preserved).
    """
    import math
    from decimal import Decimal
    from datetime import date as _date

    if val is None:
        return ("", False)
    if isinstance(val, bool):
        return ("Evet" if val else "Hayır", False)
    if isinstance(val, (int, float)):
        if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
            return (_xlsx_sanitize_str(str(val)), False)
        return (val, True)
    if isinstance(val, Decimal):
        try:
            f = float(val)
            if math.isnan(f) or math.isinf(f):
                return (_xlsx_sanitize_str(str(val)), False)
            return (f, True)
        except Exception:
            return (_xlsx_sanitize_str(str(val)), False)
    if isinstance(val, datetime):
        try:
            return (_xlsx_sanitize_str(val.isoformat()), False)
        except Exception:
            return (_xlsx_sanitize_str(str(val)), False)
    if isinstance(val, _date):
        try:
            return (_xlsx_sanitize_str(val.isoformat()), False)
        except Exception:
            return (_xlsx_sanitize_str(str(val)), False)
    if isinstance(val, bytes):
        try:
            return (_xlsx_sanitize_str(val.decode("utf-8", errors="replace")), False)
        except Exception:
            return ("", False)
    if isinstance(val, str):
        return (_xlsx_sanitize_str(val), False)
    if isinstance(val, (list, tuple, set)):
        try:
            joined = ", ".join(str(_coerce_excel_value(v)[0]) for v in val)
            return (_xlsx_sanitize_str(joined), False)
        except Exception:
            return (_xlsx_sanitize_str(str(val)), False)
    if isinstance(val, dict):
        return (_xlsx_sanitize_str(str(val)), False)
    # Fallback: any other object (ObjectId, UUID, custom types) → str()
    try:
        return (_xlsx_sanitize_str(str(val)), False)
    except Exception:
        return ("", False)


@router.post("/export/excel")
async def export_report_excel(
    config: ReportConfig,
    credentials=Depends(HTTPBearer()),
    _perm=Depends(require_op("view_reports")),
):
    """Özel raporu Excel formatında dışa aktarır."""
    current_user = await _get_current_user(credentials)
    tenant_id = getattr(current_user, 'tenant_id', None)
    if not tenant_id:
        raise HTTPException(status_code=400, detail="Tenant bilgisi bulunamadı")
    has_pii = _user_has_pii_access(current_user)

    try:
        return await _build_excel_response(config, tenant_id, has_pii)
    except HTTPException:
        raise
    except Exception:
        logger.exception(
            "report_builder excel export failed tenant=%s data_source=%s cols=%s",
            tenant_id, config.data_source, config.columns,
        )
        raise HTTPException(status_code=500, detail="report_export_failed")


async def _build_excel_response(config: "ReportConfig", tenant_id: str, has_pii: bool):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    data = await fetch_report_data(config, tenant_id, has_pii)
    source_def = DATA_SOURCES.get(config.data_source, {})

    wb = Workbook()
    ws = wb.active
    ws.title = (source_def.get("label", "Rapor") or "Rapor")[:30]

    headers = [source_def.get("columns", {}).get(col, {}).get("label", col) for col in config.columns]
    n_cols = max(len(headers), 1)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{source_def.get('label', 'Rapor')} - Özel Rapor"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Date / meta row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    date_cell = ws.cell(row=2, column=1)
    parts = []
    if config.date_from:
        parts.append(f"Başlangıç: {config.date_from}")
    if config.date_to:
        parts.append(f"Bitiş: {config.date_to}")
    if not has_pii:
        parts.append("PII alanları maskelenmiştir")
    date_cell.value = " | ".join(parts) if parts else f"Oluşturma: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M')}"
    date_cell.font = Font(size=10, italic=True, color="666666")
    date_cell.alignment = Alignment(horizontal="center")

    # Headers
    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = max(len(header) + 4, 14)

    # Data rows (Task #246: defensive coerce — heterogeneous BSON types in
    # stress tenants would crash `str(val)` / `xlsx_safe()` / Cell.value
    # assignment silently in CI).
    from core.csv_safe import xlsx_safe
    light_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    for row_num, row_data in enumerate(data, 4):
        for col_num, col_key in enumerate(config.columns, 1):
            cell = ws.cell(row=row_num, column=col_num)
            raw_val = row_data.get(col_key, "")
            col_type = source_def.get("columns", {}).get(col_key, {}).get("type")
            coerced, is_numeric = _coerce_excel_value(raw_val)

            if col_type == "currency" and is_numeric:
                cell.value = coerced
                cell.number_format = '#,##0.00 ₺'
            elif col_type == "number" and is_numeric:
                cell.value = coerced
                cell.number_format = '#,##0'
            else:
                # String path — apply xlsx_safe to defend against formula
                # injection (Bug AN). xlsx_safe handles None internally.
                cell.value = xlsx_safe(coerced) if coerced != "" else ""

            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if (row_num - 4) % 2 == 1:
                cell.fill = light_fill

    # Summary row — TOPLAM yalnızca uygun sütuna düşsün.
    if data:
        summary_row = len(data) + 5
        # 1. sütun numeric değilse TOPLAM etiketini oraya, değilse ilk text
        # sütununa, hiç text yoksa A sütunu üzerine yazma.
        label_col = None
        for i, c in enumerate(config.columns, 1):
            t = source_def.get("columns", {}).get(c, {}).get("type")
            if t not in ("number", "currency"):
                label_col = i
                break
        if label_col is not None:
            cell = ws.cell(row=summary_row, column=label_col, value="TOPLAM")
            cell.font = Font(bold=True, size=11)
            cell.border = Border(top=Side(style='double'))

        for col_num, col_key in enumerate(config.columns, 1):
            col_type = source_def.get("columns", {}).get(col_key, {}).get("type")
            if col_type in ("number", "currency"):
                # Task #246: also coerce summary-source values so Decimal/etc.
                # don't silently exclude rows that should be summed.
                numeric_values = []
                for r in data:
                    coerced_v, is_num = _coerce_excel_value(r.get(col_key))
                    if is_num:
                        numeric_values.append(coerced_v)
                if numeric_values:
                    cell = ws.cell(row=summary_row, column=col_num)
                    cell.value = sum(numeric_values)
                    cell.font = Font(bold=True, size=11)
                    cell.border = Border(top=Side(style='double'))
                    if col_type == "currency":
                        cell.number_format = '#,##0.00 ₺'

    # Task #253 (tur-2): belt-and-suspenders save retry. If, despite the
    # _coerce_excel_value sanitization, openpyxl still raises (future seed
    # surface, third-party data shape we didn't anticipate), scrub every
    # string cell once more and retry. This must NEVER mask a real error
    # silently — the second exception propagates and the outer try/except
    # logs the traceback so the next stress run surfaces the true cause.
    output = io.BytesIO()
    try:
        wb.save(output)
    except Exception:
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        logger.exception(
            "report_builder excel save failed, retrying with full re-scrub tenant=%s data_source=%s rows=%s cols=%s",
            tenant_id, config.data_source, len(data), len(config.columns),
        )
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = ILLEGAL_CHARACTERS_RE.sub("", cell.value)[:_XLSX_MAX_CELL_LEN]
        output = io.BytesIO()
        wb.save(output)
    output.seek(0)

    filename = f"ozel_rapor_{config.data_source}_{datetime.now(UTC).strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/export/pdf")
async def export_report_pdf(
    config: ReportConfig,
    credentials=Depends(HTTPBearer()),
    _perm=Depends(require_op("view_reports")),
):
    """Özel raporu PDF formatında dışa aktarır."""
    current_user = await _get_current_user(credentials)
    tenant_id = getattr(current_user, 'tenant_id', None)
    if not tenant_id:
        raise HTTPException(status_code=400, detail="Tenant bilgisi bulunamadı")
    has_pii = _user_has_pii_access(current_user)

    data = await fetch_report_data(config, tenant_id, has_pii)
    source_def = DATA_SOURCES.get(config.data_source, {})

    headers = [source_def.get("columns", {}).get(col, {}).get("label", col) for col in config.columns]

    import html as _html_mod

    def _e(v) -> str:
        return _html_mod.escape("" if v is None else str(v), quote=True)

    # Sütun sayısına göre font scale (A4 yatay max ~10 sütun rahat).
    n_cols = max(len(config.columns), 1)
    if n_cols <= PDF_FIT_COLUMNS:
        body_size, header_size, cell_pad = 10, 10, "6px 8px"
    elif n_cols <= 12:
        body_size, header_size, cell_pad = 8, 8, "4px 6px"
    else:
        body_size, header_size, cell_pad = 7, 7, "3px 4px"

    rows_html = ""
    for i, row in enumerate(data):
        bg = "#f8fafc" if i % 2 == 0 else "#ffffff"
        cells = ""
        for col_key in config.columns:
            val = row.get(col_key, "")
            col_type = source_def.get("columns", {}).get(col_key, {}).get("type")
            if col_type == "currency" and isinstance(val, (int, float)) and not isinstance(val, bool):
                display = f"₺{val:,.2f}"
            elif col_type == "boolean":
                display = "Evet" if val else "Hayır"
            elif isinstance(val, list):
                display = ", ".join(str(v) for v in val)
            else:
                display = str(val) if val is not None else ""
            cells += f'<td style="padding:{cell_pad};border-bottom:1px solid #e2e8f0;font-size:{body_size}px;">{_e(display)}</td>'
        rows_html += f'<tr style="background:{bg}">{cells}</tr>'

    header_cells = "".join(
        f'<th style="padding:8px;background:#0F172A;color:white;font-size:{header_size}px;text-align:left;border-bottom:2px solid #0d2137;">{_e(h)}</th>'
        for h in headers
    )

    date_info = ""
    if config.date_from or config.date_to or not has_pii:
        parts = []
        if config.date_from:
            parts.append(f"Başlangıç: {_e(config.date_from)}")
        if config.date_to:
            parts.append(f"Bitiş: {_e(config.date_to)}")
        if not has_pii:
            parts.append("<i>PII alanları maskelenmiştir</i>")
        date_info = f'<p style="color:#64748b;font-size:11px;margin:4px 0 12px;">{" | ".join(parts)}</p>'

    title_label = _e(source_def.get('label', 'Rapor'))
    page_size = "A4 landscape" if n_cols <= PDF_FIT_COLUMNS else "A3 landscape"

    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
  @page {{ size: {page_size}; margin: 1.2cm; }}
  body {{ font-family: Arial, Helvetica, sans-serif; color: #1e293b; margin:0; padding:0; }}
  .header {{ background: #0F172A; color: white; padding: 16px 20px; margin-bottom: 12px; }}
  .header h1 {{ margin: 0; font-size: 16px; }}
  .header p {{ margin: 4px 0 0; font-size: 11px; opacity: 0.8; }}
  table {{ width: 100%; border-collapse: collapse; font-size: {body_size}px; table-layout: auto; }}
  .footer {{ text-align: center; font-size: 9px; color: #94a3b8; margin-top: 16px; padding-top: 8px; border-top: 1px solid #e2e8f0; }}
</style></head><body>
<div class="header">
  <h1>{title_label} - Özel Rapor</h1>
  <p>Toplam {len(data)} kayıt | Oluşturma: {datetime.now(UTC).strftime('%d.%m.%Y %H:%M')}</p>
</div>
{date_info}
<table><thead><tr>{header_cells}</tr></thead><tbody>{rows_html}</tbody></table>
<div class="footer">Syroce PMS - Otomatik Oluşturulmuş Rapor</div>
</body></html>"""

    try:
        from weasyprint import HTML

        def _safe_fetcher(url: str, timeout=10, ssl_context=None):
            if not url.lower().startswith("https://"):
                raise ValueError(f"blocked URL scheme: {url[:40]}")
            from weasyprint import default_url_fetcher  # type: ignore
            return default_url_fetcher(url, timeout=timeout, ssl_context=ssl_context)

        pdf_bytes = HTML(string=html, url_fetcher=_safe_fetcher).write_pdf()
        output = io.BytesIO(pdf_bytes)
    except Exception:
        output = io.BytesIO(html.encode('utf-8'))

    output.seek(0)
    filename = f"ozel_rapor_{config.data_source}_{datetime.now(UTC).strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── Template CRUD ────────────────────────────────────────────────────────

@router.get("/templates")
async def list_templates(
    credentials=Depends(HTTPBearer()),
    _perm=Depends(require_op("view_reports")),
):
    """Kayıtlı rapor şablonlarını listeler (tenant scoped)."""
    current_user = await _get_current_user(credentials)
    db = get_db()
    tenant_id = getattr(current_user, 'tenant_id', None)
    templates = await db.report_templates.find(
        {"tenant_id": tenant_id},
        {"_id": 0},
    ).sort("created_at", -1).to_list(100)
    return {"templates": templates}


@router.post("/templates")
async def save_template(
    template: SavedTemplate,
    credentials=Depends(HTTPBearer()),
    _perm=Depends(require_op("view_reports")),
):
    """Rapor şablonunu kaydeder."""
    current_user = await _get_current_user(credentials)
    db = get_db()
    tenant_id = getattr(current_user, 'tenant_id', None)
    user_id = getattr(current_user, 'id', None)

    # Sütun ve veri kaynağı doğrulaması.
    if template.config.data_source not in DATA_SOURCES:
        raise HTTPException(status_code=400, detail="Geçersiz veri kaynağı")
    cols_def = DATA_SOURCES[template.config.data_source].get("columns", {})
    for c in template.config.columns or []:
        if c not in cols_def:
            raise HTTPException(status_code=400, detail=f"Bilinmeyen sütun: {c}")

    doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "created_by": user_id,
        "name": template.name.strip()[:120],
        "description": (template.description or "").strip()[:500],
        "config": template.config.dict(),
        "created_at": datetime.now(UTC).isoformat(),
        "updated_at": datetime.now(UTC).isoformat(),
    }
    await db.report_templates.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.delete("/templates/{template_id}")
async def delete_template(
    template_id: str,
    credentials=Depends(HTTPBearer()),
    _perm=Depends(require_op("view_reports")),
):
    """Rapor şablonunu siler. Sadece şablonu oluşturan kullanıcı veya
    admin/manager silebilir.
    """
    current_user = await _get_current_user(credentials)
    db = get_db()
    tenant_id = getattr(current_user, 'tenant_id', None)
    user_id = getattr(current_user, 'id', None)

    role = getattr(current_user, 'role', None)
    role_str = getattr(role, 'value', None) or str(role or '')
    is_admin = role_str in ('admin', 'super_admin', 'manager', 'general_manager')

    existing = await db.report_templates.find_one(
        {"id": template_id, "tenant_id": tenant_id},
        {"_id": 0, "created_by": 1},
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Şablon bulunamadı")
    if not is_admin and existing.get("created_by") and existing.get("created_by") != user_id:
        raise HTTPException(status_code=403, detail="Bu şablonu silme yetkiniz yok")

    result = await db.report_templates.delete_one({"id": template_id, "tenant_id": tenant_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Şablon bulunamadı")
    return {"message": "Şablon silindi"}

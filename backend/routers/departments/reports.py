"""
reports

Auto-split sub-router (shared imports/classes inlined).
"""
"""
Department-Specific Endpoints Router
Front Office, Housekeeping Manager, Finance, Revenue, F&B, Maintenance,
Sales, HR, IT/Security department dashboards.
Extracted from server.py for modularity.
"""
import logging

logger = logging.getLogger(__name__)
from datetime import UTC, datetime, timedelta
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.security import HTTPBearer

from core.database import db
from core.helpers import require_module
from core.security import get_current_user
from core.utils import calculate_folio_balance, create_excel_workbook, excel_response
from models.schemas import User
from modules.pms_core.role_permission_service import RolePermissionService, require_op

_role_perm = RolePermissionService()


def _enforce(role: str, op: str):
    """Bug CU (v60) — Departments/Reports/Rates/POS RBAC zorunlu."""
    _role_perm.enforce_permission(role, op)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: F401
except ImportError:
    Workbook = None

try:
    from cache_manager import cache, cached
except ImportError:
    cache = None  # type: ignore
    def cached(ttl=300, key_prefix=""):
        def decorator(func):
            return func
        return decorator

security = HTTPBearer()


# ==================== DEPARTMENT-SPECIFIC ENDPOINTS ====================

# rbac-allow: cache-rbac — FO dashboard operasyonel, hotel staff geneli görür (FO/HK/manager/admin)

# rbac-allow: cache-rbac — HK dashboard operasyonel, FO/HK/manager/admin görür








# NOTE: /ai/dashboard/briefing duplicate removed (R10b) — canonical implementation
# lives in `domains/ai/endpoints.py::get_daily_briefing` with @cached(ttl=300) and
# parallel `_asyncio.gather` over 4 collections.




# rbac-allow: cache-rbac — booking için müsait odalar operasyonel (FO/HK/manager)



# rbac-allow: cache-rbac — HK aktif temizlik timer'ları operasyonel (HK/FO/manager)











































# rbac-allow: cache-rbac — task kanban operasyonel cross-role (FO/HK/maintenance/manager)

router = APIRouter(prefix="/api", tags=["departments"])


# ── GET /reports/market-segment ──
@router.get("/reports/market-segment")
@cached(ttl=900, key_prefix="report_market_segment")  # Cache for 15 minutes
async def get_market_segment_report(
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_finance_reports")),  # v84 DT: market segment performance
    _nocache: bool = Query(False, alias="nocache"),
):
    """Market Segment & Rate Type Performance Report"""
    start = datetime.fromisoformat(start_date)
    end = datetime.fromisoformat(end_date)

    # Get all bookings in date range
    bookings = await db.bookings.find({
        'tenant_id': current_user.tenant_id,
        'check_in': {'$gte': start.isoformat()},
        'check_out': {'$lte': end.isoformat()}
    }).to_list(10000)

    # Aggregate by market segment
    segment_data = {}
    rate_type_data = {}

    for booking in bookings:
        segment = booking.get('market_segment', 'other')
        rate_type = booking.get('rate_type', 'bar')

        # Calculate nights
        check_in = datetime.fromisoformat(booking['check_in'])
        check_out = datetime.fromisoformat(booking['check_out'])
        nights = (check_out - check_in).days
        revenue = booking.get('total_amount', 0)

        # Market segment aggregation
        if segment not in segment_data:
            segment_data[segment] = {'bookings': 0, 'nights': 0, 'revenue': 0}
        segment_data[segment]['bookings'] += 1
        segment_data[segment]['nights'] += nights
        segment_data[segment]['revenue'] += revenue

        # Rate type aggregation
        if rate_type not in rate_type_data:
            rate_type_data[rate_type] = {'bookings': 0, 'nights': 0, 'revenue': 0}
        rate_type_data[rate_type]['bookings'] += 1
        rate_type_data[rate_type]['nights'] += nights
        rate_type_data[rate_type]['revenue'] += revenue

    # Calculate averages
    for segment in segment_data:
        segment_data[segment]['adr'] = round(
            segment_data[segment]['revenue'] / segment_data[segment]['nights'], 2
        ) if segment_data[segment]['nights'] > 0 else 0

    for rate_type in rate_type_data:
        rate_type_data[rate_type]['adr'] = round(
            rate_type_data[rate_type]['revenue'] / rate_type_data[rate_type]['nights'], 2
        ) if rate_type_data[rate_type]['nights'] > 0 else 0

    return {
        'start_date': start_date,
        'end_date': end_date,
        'total_bookings': len(bookings),
        'market_segments': segment_data,
        'rate_types': rate_type_data
    }
# ── GET /reports/market-segment/excel ──
@router.get("/reports/market-segment/excel")
@cached(ttl=900, key_prefix="report_market_segment_excel")  # Cache for 15 min
async def export_market_segment_excel(
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_finance_reports")),  # v84 DT: market segment export
):
    """Export Market Segment Report to Excel"""
    report_data = await get_market_segment_report(start_date, end_date, current_user)

    # Create workbook with multiple sheets
    wb = Workbook()

    # Sheet 1: Market Segments
    ws1 = wb.active
    ws1.title = "Market Segments"

    headers1 = ["Segment", "Bookings", "Nights", "Revenue", "ADR"]
    data1 = []
    for segment, stats in report_data['market_segments'].items():
        data1.append([
            segment.title(),
            stats['bookings'],
            stats['nights'],
            f"${stats['revenue']:,.2f}",
            f"${stats['adr']:,.2f}"
        ])

    # Add title and headers
    ws1.merge_cells('A1:E1')
    title_cell = ws1['A1']
    title_cell.value = f"Market Segment Report ({start_date} to {end_date})"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    for col_num, header in enumerate(headers1, 1):
        cell = ws1.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    for row_num, row_data in enumerate(data1, 3):
        for col_num, value in enumerate(row_data, 1):
            ws1.cell(row=row_num, column=col_num, value=value)

    # Sheet 2: Rate Types
    ws2 = wb.create_sheet("Rate Types")

    headers2 = ["Rate Type", "Bookings", "Nights", "Revenue", "ADR"]
    data2 = []
    for rate_type, stats in report_data['rate_types'].items():
        data2.append([
            rate_type.upper(),
            stats['bookings'],
            stats['nights'],
            f"${stats['revenue']:,.2f}",
            f"${stats['adr']:,.2f}"
        ])

    ws2.merge_cells('A1:E1')
    title_cell = ws2['A1']
    title_cell.value = f"Rate Type Report ({start_date} to {end_date})"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    for col_num, header in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    for row_num, row_data in enumerate(data2, 3):
        for col_num, value in enumerate(row_data, 1):
            ws2.cell(row=row_num, column=col_num, value=value)

    filename = f"market_segment_report_{start_date}_to_{end_date}.xlsx"
    return excel_response(wb, filename)
# ── Helper: safe date coercion (Task #246) ──
def _coerce_to_date(v):
    """Coerce mixed-type `created_at` values to date. Stress env may store
    `created_at` as native BSON datetime (motor returns datetime obj) rather
    than ISO string — `datetime.fromisoformat()` raises TypeError in that
    case. Returns None for unparseable/missing values so caller can skip.
    """
    from datetime import date as _date
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, _date):
        return v
    if isinstance(v, str):
        try:
            # Tolerate trailing 'Z' (UTC suffix); fromisoformat accepts it on 3.11+
            return datetime.fromisoformat(v.replace('Z', '+00:00')).date()
        except (ValueError, TypeError):
            return None
    return None


# ── Pure helper: company-aging compute (Task #246) ──
async def _compute_company_aging(tenant_id: str) -> dict:
    """Pure async data builder for Company AR Aging Report.

    Extracted from `get_company_aging_report` so the Excel export route can
    re-use the computation without invoking another `@cached`+`@router.get`
    decorated route handler directly (anti-pattern that broke under stress
    when decorator semantics shifted, and obscured tracebacks).
    """
    today = datetime.now(UTC).date()

    folios = await db.folios.find({
        'tenant_id': tenant_id,
        'folio_type': 'company',
        'status': 'open'
    }).to_list(10000)

    company_balances: dict[str, dict] = {}

    for folio in folios:
        try:
            balance = await calculate_folio_balance(folio['id'], tenant_id)
        except Exception:
            logger.exception("calculate_folio_balance failed for folio=%s tenant=%s", folio.get('id'), tenant_id)
            continue

        if balance <= 0:
            continue

        company_id = folio.get('company_id')
        if not company_id:
            continue

        # Task #246: tenant_id scope on company lookup (architect review #3) —
        # company ID collisions across tenants would otherwise leak metadata.
        company = await db.companies.find_one(
            {'id': company_id, 'tenant_id': tenant_id}, {'_id': 0}
        )
        if not company:
            continue

        folio_created = _coerce_to_date(folio.get('created_at'))
        if folio_created is None:
            # Unparseable date → treat as oldest bucket; do not crash export.
            age_days = 99999
        else:
            age_days = (today - folio_created).days

        if age_days <= 7:
            aging_bucket = '0-7 days'
        elif age_days <= 14:
            aging_bucket = '8-14 days'
        elif age_days <= 30:
            aging_bucket = '15-30 days'
        else:
            aging_bucket = '30+ days'

        if company_id not in company_balances:
            company_balances[company_id] = {
                'company_name': company.get('name', 'N/A'),
                'corporate_code': company.get('corporate_code', 'N/A'),
                'total_balance': 0,
                'aging': {
                    '0-7 days': 0,
                    '8-14 days': 0,
                    '15-30 days': 0,
                    '30+ days': 0,
                },
                'folio_count': 0,
            }

        company_balances[company_id]['total_balance'] += balance
        company_balances[company_id]['aging'][aging_bucket] += balance
        company_balances[company_id]['folio_count'] += 1

    sorted_companies = sorted(
        company_balances.values(),
        key=lambda x: x['total_balance'],
        reverse=True,
    )

    total_ar = sum(c['total_balance'] for c in sorted_companies)

    return {
        'report_date': today.isoformat(),
        'total_ar': round(total_ar, 2),
        'company_count': len(sorted_companies),
        'companies': sorted_companies,
    }


# ── GET /reports/company-aging ──
@router.get("/reports/company-aging")
@cached(ttl=900, key_prefix="report_company_aging")  # Cache for 15 min
async def get_company_aging_report(current_user: User = Depends(get_current_user),
    _perm: None = Depends(require_op("view_finance_reports")),
):
    """Company Accounts Receivable Aging Report"""
    _enforce(current_user.role, "view_finance_reports")  # Bug CU
    try:
        return await _compute_company_aging(current_user.tenant_id)
    except HTTPException:
        raise
    except Exception:
        logger.exception("company_aging report failed for tenant=%s", current_user.tenant_id)
        raise HTTPException(status_code=500, detail="report_failed")


# ── GET /reports/company-aging/excel ──
# Task #246: @cached REMOVED — decorator serializes the StreamingResponse
# object via `repr()` ("<starlette.responses.StreamingResponse object at 0x…>")
# so the second call within the TTL returned non-XLSX garbage with HTTP 200.
# Excel-shape rendering is cheap enough to recompute; underlying data builder
# (_compute_company_aging) is itself called by the cached JSON route and can
# be wrapped separately if hot-path caching becomes necessary.
@router.get("/reports/company-aging/excel")
async def export_company_aging_excel(
    current_user: User = Depends(get_current_user),
    _perm: None = Depends(require_op("view_finance_reports")),
):
    """Export Company Aging Report to Excel"""
    _enforce(current_user.role, "view_finance_reports")  # Bug CU
    try:
        # Use shared pure helper instead of calling cached route handler directly
        # (Task #246 — `@cached` decorated route handler MUST NOT be invoked
        # directly; cache key derivation + Depends sentinel handling are fragile).
        report_data = await _compute_company_aging(current_user.tenant_id)

        headers = ["Company", "Corporate Code", "Total Balance", "0-7 Days", "8-14 Days", "15-30 Days", "30+ Days", "Folios"]
        data = []

        for company in report_data['companies']:
            data.append([
                company['company_name'],
                company['corporate_code'],
                f"${company['total_balance']:,.2f}",
                f"${company['aging']['0-7 days']:,.2f}",
                f"${company['aging']['8-14 days']:,.2f}",
                f"${company['aging']['15-30 days']:,.2f}",
                f"${company['aging']['30+ days']:,.2f}",
                company['folio_count']
            ])

        # Add total row
        data.append([
            "TOTAL",
            "",
            f"${report_data['total_ar']:,.2f}",
            "",
            "",
            "",
            "",
            ""
        ])

        wb = create_excel_workbook(
            title=f"Company Aging Report - {report_data['report_date']}",
            headers=headers,
            data=data,
            sheet_name="Company Aging"
        )

        # Task #253 (tur-2): belt-and-suspenders save retry. create_excel_workbook
        # already sanitizes via _xlsx_sanitize_str, but if any future code path
        # bypasses it (e.g. company.get('name') with a control char that lands
        # post-sanitize, or any non-string field that openpyxl rejects), retry
        # once with full re-scrub. Outer try/except logs the true traceback.
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        import io as _io
        from core.utils import _XLSX_MAX_CELL_LEN
        buf = _io.BytesIO()
        try:
            wb.save(buf)
            buf.seek(0)
        except Exception:
            logger.exception(
                "company_aging excel save failed, retrying with full re-scrub tenant=%s rows=%s",
                current_user.tenant_id, len(report_data.get('companies', [])),
            )
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            cell.value = ILLEGAL_CHARACTERS_RE.sub("", cell.value)[:_XLSX_MAX_CELL_LEN]
            buf = _io.BytesIO()
            wb.save(buf)
            buf.seek(0)

        from fastapi.responses import StreamingResponse
        filename = f"company_aging_report_{report_data['report_date']}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception:
        logger.exception(
            "company_aging excel export failed for tenant=%s",
            current_user.tenant_id,
        )
        raise HTTPException(status_code=500, detail="report_export_failed")
# ── GET /reports/finance-snapshot ──
@router.get("/reports/finance-snapshot")
@cached(ttl=600, key_prefix="report_finance_snapshot")  # Cache for 10 min
async def get_finance_snapshot(current_user: User = Depends(get_current_user),
    _perm: None = Depends(require_op("view_finance_reports")),
):
    """
    Finance Snapshot for GM Dashboard
    Returns: Total Pending AR, Overdue Invoices (categorized), Today's Collections
    """
    _enforce(current_user.role, "view_finance_reports")  # Bug CU
    today = datetime.now(UTC).date()
    today_start = datetime.combine(today, datetime.min.time()).replace(tzinfo=UTC)
    today_end = datetime.combine(today, datetime.max.time()).replace(tzinfo=UTC)

    # 1. Calculate Total Pending AR from company folios
    company_folios = await db.folios.find({
        'tenant_id': current_user.tenant_id,
        'folio_type': 'company',
        'status': 'open'
    }).to_list(10000)

    total_pending_ar = 0
    overdue_0_30 = 0
    overdue_30_60 = 0
    overdue_60_plus = 0
    overdue_invoices_count = 0

    for folio in company_folios:
        balance = await calculate_folio_balance(folio['id'], current_user.tenant_id)

        if balance > 0:
            total_pending_ar += balance

            # Calculate aging
            folio_created = datetime.fromisoformat(folio['created_at']).date()
            age_days = (today - folio_created).days

            if age_days > 0:  # Any overdue
                overdue_invoices_count += 1

                if age_days <= 30:
                    overdue_0_30 += balance
                elif age_days <= 60:
                    overdue_30_60 += balance
                else:
                    overdue_60_plus += balance

    # 2. Calculate Today's Collections (payments received today)
    todays_payments = await db.payments.find({
        'tenant_id': current_user.tenant_id,
        'processed_at': {
            '$gte': today_start.isoformat(),
            '$lte': today_end.isoformat()
        }
    }).to_list(10000)

    todays_collections = sum(payment.get('amount', 0) for payment in todays_payments)
    todays_payment_count = len(todays_payments)

    # 3. Calculate MTD (Month-to-Date) Collections
    month_start = today.replace(day=1)
    month_start_dt = datetime.combine(month_start, datetime.min.time()).replace(tzinfo=UTC)

    mtd_payments = await db.payments.find({
        'tenant_id': current_user.tenant_id,
        'processed_at': {
            '$gte': month_start_dt.isoformat(),
            '$lte': today_end.isoformat()
        }
    }).to_list(10000)

    mtd_collections = sum(payment.get('amount', 0) for payment in mtd_payments)

    # 4. Calculate Collection Rate (MTD Collections / MTD Revenue)
    mtd_charges = await db.folio_charges.find({
        'tenant_id': current_user.tenant_id,
        'date': {
            '$gte': month_start_dt.isoformat(),
            '$lte': today_end.isoformat()
        },
        'voided': False
    }).to_list(10000)

    mtd_revenue = sum(charge.get('total', 0) for charge in mtd_charges)
    collection_rate = (mtd_collections / mtd_revenue * 100) if mtd_revenue > 0 else 0

    # 5. Get Accounting Invoices (E-Fatura ready)
    pending_invoices = await db.accounting_invoices.find({
        'tenant_id': current_user.tenant_id,
        'status': {'$in': ['pending', 'partial']}
    }).to_list(1000)

    pending_invoice_total = sum(inv.get('total', 0) for inv in pending_invoices)
    pending_invoice_count = len(pending_invoices)

    return {
        'report_date': today.isoformat(),
        'pending_ar': {
            'total': round(total_pending_ar, 2),
            'overdue_breakdown': {
                '0-30_days': round(overdue_0_30, 2),
                '30-60_days': round(overdue_30_60, 2),
                '60_plus_days': round(overdue_60_plus, 2)
            },
            'overdue_invoices_count': overdue_invoices_count
        },
        'todays_collections': {
            'amount': round(todays_collections, 2),
            'payment_count': todays_payment_count
        },
        'mtd_collections': {
            'amount': round(mtd_collections, 2),
            'collection_rate_percentage': round(collection_rate, 2)
        },
        'accounting_invoices': {
            'pending_count': pending_invoice_count,
            'pending_total': round(pending_invoice_total, 2)
        }
    }
# ── GET /reports/revenue-detail/excel ──
@router.get("/reports/revenue-detail/excel")
async def export_revenue_detail_excel(
    start_date: str | None = None,
    end_date: str | None = None,
    current_user: User = Depends(get_current_user),
    _: None = Depends(require_module("reports")),
):
    if not end_date:
        end_date = datetime.now(UTC).date().isoformat()
    if not start_date:
        start_date = (datetime.now(UTC) - timedelta(days=30)).date().isoformat()
    """Detailed room revenue by date, room type and rate code.

    NOTE: Uses bookings collection and groups by date, room_type and rate_code-like fields.
    """
    _enforce(current_user.role, "view_finance_reports")  # Bug CU
    start = datetime.fromisoformat(start_date)
    end = datetime.fromisoformat(end_date)

    if start.tzinfo is None:
        start = start.replace(tzinfo=UTC)
    if end.tzinfo is None:
        end = end.replace(tzinfo=UTC)

    # Fetch bookings in range
    bookings = await db.bookings.find(
        {
            'tenant_id': current_user.tenant_id,
            'status': {'$in': ['confirmed', 'guaranteed', 'checked_in', 'checked_out']},
            '$or': [
                {'check_in': {'$gte': start.isoformat(), '$lte': end.isoformat()}},
                {'check_out': {'$gte': start.isoformat(), '$lte': end.isoformat()}},
                {'check_in': {'$lte': start.isoformat()}, 'check_out': {'$gte': end.isoformat()}},
            ],
        },
        {
            '_id': 0,
            'check_in': 1,
            'check_out': 1,
            'total_amount': 1,
            'room_type': 1,
            'rate_plan': 1,
            'market_segment': 1,
        },
    ).to_list(10000)

    # Aggregate per stay-date

    for b in bookings:
        try:
            ci = datetime.fromisoformat(b['check_in'])
            co = datetime.fromisoformat(b['check_out'])
        except Exception:
            continue

        # Normalize to date range
        ci_date = max(ci.date(), start.date())
        co_date = min(co.date(), end.date())

        days = (co_date - ci_date).days or 1
        (b.get('total_amount') or 0) / days
# ── GET /reports/forecast-detail/excel ──
@router.get("/reports/forecast-detail/excel")
async def export_forecast_detail_excel(
    start_date: str | None = None,
    end_date: str | None = None,
    current_user: User = Depends(get_current_user),
    _: None = Depends(require_module("reports")),
):
    if not start_date:
        start_date = datetime.now(UTC).date().isoformat()
    if not end_date:
        end_date = (datetime.now(UTC) + timedelta(days=30)).date().isoformat()
    """Forecasted occupancy and revenue detail by date using existing forecast logic.

    NOTE: This uses get_forecast endpoint internally if available.
    """
    _enforce(current_user.role, "view_finance_reports")  # Bug CU
    # Reuse get_forecast if defined (lazy-loaded to avoid circular imports)
    try:
        from routers.reports import get_forecast as _get_forecast
        forecast_response = await _get_forecast(
            days=(datetime.fromisoformat(end_date) - datetime.fromisoformat(start_date)).days + 1,
            current_user=current_user,
            _=None,
        )
    except Exception:
        forecast_response = {}

    headers = ['Date', 'Expected Occupancy %', 'Expected Revenue']
    data: list[list[Any]] = []

    if isinstance(forecast_response, list):
        items = forecast_response
    elif isinstance(forecast_response, dict):
        items = forecast_response.get('days') or forecast_response.get('forecast') or []
    else:
        items = []
    for item in items:
        if not isinstance(item, dict):
            continue
        data.append([
            item.get('date'),
            item.get('expected_occupancy_pct', item.get('occupancy_pct', 0)),
            item.get('expected_revenue', item.get('revenue', 0)),
        ])

    title = f"Forecast Detail {start_date} to {end_date}"
    wb = create_excel_workbook(
        title=title,
        headers=headers,
        data=data,
        sheet_name="Forecast Detail",
    )

    filename = f"forecast_detail_{start_date}_to_{end_date}.xlsx"
    return excel_response(wb, filename)
# ── GET /reports/operations-daily-summary/excel ──
@router.get("/reports/operations-daily-summary/excel")
async def export_operations_daily_summary_excel(
    date: str,
    current_user: User = Depends(get_current_user),
    _: None = Depends(require_module("reports")),
):
    """Daily operations summary: arrivals, departures, in-house guests."""
    target = datetime.fromisoformat(date)
    if target.tzinfo is None:
        target = target.replace(tzinfo=UTC)

    day_start = datetime.combine(target.date(), datetime.min.time()).replace(tzinfo=UTC)
    day_end = datetime.combine(target.date(), datetime.max.time()).replace(tzinfo=UTC)

    arrivals = await db.bookings.count_documents({
        'tenant_id': current_user.tenant_id,
        'check_in': {'$gte': day_start.isoformat(), '$lte': day_end.isoformat()},
        'status': {'$in': ['confirmed', 'guaranteed', 'checked_in']},
    })

    departures = await db.bookings.count_documents({
        'tenant_id': current_user.tenant_id,
        'check_out': {'$gte': day_start.isoformat(), '$lte': day_end.isoformat()},
        'status': {'$in': ['checked_in', 'checked_out']},
    })

    in_house = await db.bookings.count_documents({
        'tenant_id': current_user.tenant_id,
        'status': 'checked_in',
    })

    headers = ['Metric', 'Value']
    data = [
        ['Date', target.date().isoformat()],
        ['', ''],
        ['Arrivals', arrivals],
        ['Departures', departures],
        ['In-House Guests', in_house],
    ]

    title = f"Operations Daily Summary {target.date().isoformat()}"
    wb = create_excel_workbook(
        title=title,
        headers=headers,
        data=data,
        sheet_name="Operations Summary",
    )

    filename = f"operations_daily_summary_{target.date().isoformat()}.xlsx"
    return excel_response(wb, filename)
# ── GET /reports/channel-distribution/excel ──
@router.get("/reports/channel-distribution/excel")
async def export_channel_distribution_excel(
    start_date: str | None = None,
    end_date: str | None = None,
    current_user: User = Depends(get_current_user),
    _: None = Depends(require_module("reports")),
):
    if not end_date:
        end_date = datetime.now(UTC).date().isoformat()
    if not start_date:
        start_date = (datetime.now(UTC) - timedelta(days=30)).date().isoformat()
    """Sales channel distribution report (OTA, Direct, Corporate, etc.)."""
    _enforce(current_user.role, "view_finance_reports")  # Bug CU
    start = datetime.fromisoformat(start_date)
    end = datetime.fromisoformat(end_date)

    if start.tzinfo is None:
        start = start.replace(tzinfo=UTC)
    if end.tzinfo is None:
        end = end.replace(tzinfo=UTC)

    bookings = await db.bookings.find(
        {
            'tenant_id': current_user.tenant_id,
            'status': {'$in': ['confirmed', 'guaranteed', 'checked_in', 'checked_out']},
            'check_in': {'$gte': start.isoformat(), '$lte': end.isoformat()},
        },
        {
            '_id': 0,
            'total_amount': 1,
            'channel': 1,
            'market_segment': 1,
        },
    ).to_list(20000)

    channel_stats: dict[str, dict[str, Any]] = {}

    for b in bookings:
        channel = str(b.get('channel') or 'DIRECT')
        if channel not in channel_stats:
            channel_stats[channel] = {
                'channel': channel,
                'bookings': 0,
                'revenue': 0.0,
            }
        channel_stats[channel]['bookings'] += 1
        channel_stats[channel]['revenue'] += b.get('total_amount') or 0.0

    total_revenue = sum(v['revenue'] for v in channel_stats.values()) or 1.0

    headers = [
        'Channel',
        'Bookings',
        'Revenue',
        'Share %',
    ]

    data: list[list[Any]] = []
    for key, row in sorted(channel_stats.items(), key=lambda x: x[0]):
        share = (row['revenue'] / total_revenue) * 100.0
        data.append([
            row['channel'],
            row['bookings'],
            round(row['revenue'], 2),
            round(share, 2),
        ])

    title = f"Channel Distribution {start_date} to {end_date}"
    wb = create_excel_workbook(
        title=title,
        headers=headers,
        data=data,
        sheet_name="Channels",
    )

    filename = f"channel_distribution_{start_date}_to_{end_date}.xlsx"
    return excel_response(wb, filename)
# ── GET /reports/cost-summary ──
@router.get("/reports/cost-summary")
@cached(ttl=600, key_prefix="report_cost_summary")  # Cache for 10 min
async def get_cost_summary(current_user: User = Depends(get_current_user),
    _perm: None = Depends(require_op("view_finance_reports")),
):
    """
    Cost Summary Report for GM Dashboard
    Returns: MTD costs by category, top cost categories, per-room cost, cost vs RevPAR
    """
    _enforce(current_user.role, "view_finance_reports")  # Bug CU
    today = datetime.now(UTC).date()
    month_start = today.replace(day=1)
    month_start_dt = datetime.combine(month_start, datetime.min.time()).replace(tzinfo=UTC)
    today_end = datetime.combine(today, datetime.max.time()).replace(tzinfo=UTC)

    # 1. Get all Purchase Orders from Marketplace for this month (approved/received status)
    await db.purchase_orders.find({
        'tenant_id': current_user.tenant_id,
        'status': {'$in': ['approved', 'received', 'completed']},
        'created_at': {
            '$gte': month_start_dt.isoformat(),
            '$lte': today_end.isoformat()
        }
    }).to_list(10000)
# ── GET /finance/mtd-cost-summary ──
@router.get("/finance/mtd-cost-summary")
@cached(ttl=600, key_prefix="finance_mtd_cost_summary")
async def get_mtd_cost_summary(current_user: User = Depends(get_current_user),
    _perm: None = Depends(require_op("view_finance_reports")),
):
    """Month-to-date cost summary by category with per-room metrics."""
    _enforce(current_user.role, "view_finance_reports")  # Bug CU
    today = datetime.now(UTC).date()
    month_start = today.replace(day=1)
    month_start_dt = datetime.combine(month_start, datetime.min.time()).replace(tzinfo=UTC)
    today_end = datetime.combine(today, datetime.max.time()).replace(tzinfo=UTC)

    purchase_orders = await db.purchase_orders.find({
        'tenant_id': current_user.tenant_id,
        'created_at': {
            '$gte': month_start_dt.isoformat(),
            '$lte': today_end.isoformat()
        }
    }).to_list(10000)

    # Map purchase order categories to cost categories
    category_mapping = {
        'cleaning': 'Housekeeping',
        'linens': 'Housekeeping',
        'amenities': 'Housekeeping',
        'food': 'F&B',
        'beverage': 'F&B',
        'kitchen': 'F&B',
        'maintenance': 'Technical',
        'electrical': 'Technical',
        'plumbing': 'Technical',
        'hvac': 'Technical',
        'furniture': 'General Expenses',
        'office': 'General Expenses',
        'it': 'General Expenses',
        'other': 'General Expenses'
    }

    # Aggregate costs by category
    cost_categories = {
        'Housekeeping': 0,
        'F&B': 0,
        'Technical': 0,
        'General Expenses': 0
    }

    total_mtd_costs = 0

    for po in purchase_orders:
        category = po.get('category', 'other')
        cost_category = category_mapping.get(category, 'General Expenses')
        total_amount = po.get('total_amount', 0)

        cost_categories[cost_category] += total_amount
        total_mtd_costs += total_amount

    # 2. Sort categories to get top 3
    sorted_categories = sorted(
        [{'name': k, 'amount': v} for k, v in cost_categories.items()],
        key=lambda x: x['amount'],
        reverse=True
    )

    top_3_categories = sorted_categories[:3]

    # 3. Calculate per-room cost (total costs / occupied room nights MTD)
    # Get all bookings for MTD that were checked-in
    bookings = await db.bookings.find({
        'tenant_id': current_user.tenant_id,
        'status': {'$in': ['checked_in', 'checked_out']},
        'check_in': {
            '$gte': month_start.isoformat(),
            '$lte': today.isoformat()
        }
    }).to_list(10000)

    # Calculate total occupied room nights
    total_room_nights = 0
    for booking in bookings:
        checkin = datetime.fromisoformat(booking['check_in']).date()
        checkout_str = booking.get('check_out', booking['check_in'])
        checkout = datetime.fromisoformat(checkout_str).date()

        # Calculate nights (minimum 1)
        nights = max((checkout - checkin).days, 1)
        total_room_nights += nights

    per_room_cost = (total_mtd_costs / total_room_nights) if total_room_nights > 0 else 0

    # 4. Get RevPAR from daily flash report for comparison
    # Calculate MTD RevPAR
    total_revenue = 0
    total_available_room_days = 0

    # Get all rooms
    rooms = await db.rooms.find({'tenant_id': current_user.tenant_id}).to_list(1000)
    total_rooms_count = len(rooms)

    # Get MTD charges
    mtd_charges = await db.folio_charges.find({
        'tenant_id': current_user.tenant_id,
        'date': {
            '$gte': month_start_dt.isoformat(),
            '$lte': today_end.isoformat()
        },
        'voided': False,
        'charge_category': 'room'
    }).to_list(10000)

    total_revenue = sum(charge.get('total', 0) for charge in mtd_charges)

    # Calculate days in month so far
    days_in_month_so_far = (today - month_start).days + 1
    total_available_room_days = total_rooms_count * days_in_month_so_far

    mtd_revpar = (total_revenue / total_available_room_days) if total_available_room_days > 0 else 0

    # 5. Calculate Cost to Revenue Ratio
    cost_to_revenue_ratio = (total_mtd_costs / total_revenue * 100) if total_revenue > 0 else 0

    # 6. Calculate profit margin
    gross_profit = total_revenue - total_mtd_costs
    profit_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0

    return {
        'report_date': today.isoformat(),
        'period': f'{month_start.isoformat()} to {today.isoformat()}',
        'total_mtd_costs': round(total_mtd_costs, 2),
        'cost_categories': {
            'housekeeping': round(cost_categories['Housekeeping'], 2),
            'fnb': round(cost_categories['F&B'], 2),
            'technical': round(cost_categories['Technical'], 2),
            'general_expenses': round(cost_categories['General Expenses'], 2)
        },
        'top_3_categories': [
            {
                'name': cat['name'],
                'amount': round(cat['amount'], 2),
                'percentage': round((cat['amount'] / total_mtd_costs * 100), 1) if total_mtd_costs > 0 else 0
            }
            for cat in top_3_categories
        ],
        'per_room_metrics': {
            'total_room_nights': total_room_nights,
            'cost_per_room_night': round(per_room_cost, 2),
            'mtd_revpar': round(mtd_revpar, 2),
            'cost_to_revpar_ratio': round((per_room_cost / mtd_revpar * 100), 1) if mtd_revpar > 0 else 0
        },
        'financial_metrics': {
            'mtd_revenue': round(total_revenue, 2),
            'mtd_costs': round(total_mtd_costs, 2),
            'gross_profit': round(gross_profit, 2),
            'profit_margin_percentage': round(profit_margin, 1),
            'cost_to_revenue_ratio': round(cost_to_revenue_ratio, 1)
        }
    }

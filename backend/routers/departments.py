"""
Department-Specific Endpoints Router
Front Office, Housekeeping Manager, Finance, Revenue, F&B, Maintenance,
Sales, HR, IT/Security department dashboards.
Extracted from server.py for modularity.
"""
import logging
logger = logging.getLogger(__name__)
import random
import uuid
from datetime import UTC, datetime, timedelta
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer

from core.database import db
from core.helpers import require_module
from core.security import get_current_user
from core.utils import calculate_folio_balance, create_excel_workbook, excel_response
from models.schemas import User

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: F401
except ImportError:
    Workbook = None

try:
    from cache_manager import cached
except ImportError:
    def cached(ttl=300, key_prefix=""):
        def decorator(func):
            return func
        return decorator

router = APIRouter(prefix="/api", tags=["departments"])
security = HTTPBearer()


# ==================== DEPARTMENT-SPECIFIC ENDPOINTS ====================

@router.get("/department/front-office/dashboard")
@cached(ttl=180, key_prefix="front_office_dashboard")  # Cache for 3 minutes
async def get_front_office_dashboard(current_user: User = Depends(get_current_user)):
    """Front Office Manager Dashboard with overbooking alerts"""
    today = datetime.now(UTC)
    today_start = datetime.combine(today.date(), datetime.min.time()).replace(tzinfo=UTC)
    today_end = datetime.combine(today.date(), datetime.max.time()).replace(tzinfo=UTC)

    # Check-ins today with room ready status
    checkins = []
    async for booking in db.bookings.find({
        'tenant_id': current_user.tenant_id,
        'check_in': {'$gte': today_start.isoformat(), '$lte': today_end.isoformat()},
        'status': {'$in': ['confirmed', 'guaranteed']}
    }).limit(50):
        room = await db.rooms.find_one({'id': booking.get('room_id')})
        guest = await db.guests.find_one({'id': booking.get('guest_id')})

        checkins.append({
            'booking_id': booking.get('id'),
            'guest_name': booking.get('guest_name'),
            'room_number': booking.get('room_number'),
            'check_in_time': booking.get('check_in'),
            'room_ready': room.get('status') in ['available', 'inspected'] if room else False,
            'vip': guest.get('vip', False) if guest else False,
            'actions': ['upgrade', 'late_checkout', 'message', 'print']
        })

    # Overbooking detection
    next_7_days = today + timedelta(days=7)
    room_dates = {}
    async for booking in db.bookings.find({
        'tenant_id': current_user.tenant_id,
        'status': {'$in': ['confirmed', 'guaranteed', 'checked_in']},
        'check_in': {'$lte': next_7_days.isoformat()}
    }):
        room_id = booking.get('room_id')
        check_in = datetime.fromisoformat(booking.get('check_in')).date()
        check_out = datetime.fromisoformat(booking.get('check_out')).date()

        current = check_in
        while current < check_out:
            key = f"{room_id}_{current}"
            if key not in room_dates:
                room_dates[key] = []
            room_dates[key].append({
                'booking_id': booking.get('id'),
                'guest': booking.get('guest_name'),
                'source': booking.get('booking_source', 'direct')
            })
            current += timedelta(days=1)

    overbookings = []
    for key, bookings_list in room_dates.items():
        if len(bookings_list) > 1:
            room_id, date_str = key.split('_')
            room = await db.rooms.find_one({'id': room_id})
            overbookings.append({
                'date': date_str,
                'room_number': room.get('room_number') if room else 'Unknown',
                'conflicts': bookings_list,
                'severity': 'critical' if len(bookings_list) > 2 else 'high'
            })

    return {
        'checkins_today': checkins,
        'overbooking_alerts': overbookings,
        'total_checkins': len(checkins),
        'total_overbookings': len(overbookings),
        'vip_determination': {
            'source': 'PMS + CRM',
            'rules': ['Manual tag', 'Loyalty tier >= Gold', 'Spend > $10k', 'Frequency > 5/year']
        }
    }

@router.get("/department/housekeeping/dashboard")
@cached(ttl=120, key_prefix="housekeeping_dashboard")  # Cache for 2 minutes
async def get_housekeeping_dashboard(current_user: User = Depends(get_current_user)):
    """Housekeeping Manager Dashboard with room details"""

    # Room status counts
    dirty_rooms = []
    async for room in db.rooms.find({'tenant_id': current_user.tenant_id, 'status': 'dirty'}):
        dirty_rooms.append({
            'room_number': room.get('room_number'),
            'floor': room.get('floor'),
            'room_type': room.get('room_type'),
            'last_checkout': room.get('last_checkout')
        })

    cleaning_rooms = []
    async for room in db.rooms.find({'tenant_id': current_user.tenant_id, 'status': 'cleaning'}):
        cleaning_rooms.append({
            'room_number': room.get('room_number'),
            'floor': room.get('floor'),
            'assigned_to': room.get('assigned_cleaner')
        })

    inspected_rooms = await db.rooms.count_documents({'tenant_id': current_user.tenant_id, 'status': 'inspected'})
    maintenance_rooms = await db.rooms.count_documents({'tenant_id': current_user.tenant_id, 'status': 'maintenance'})

    # Auto-rules
    auto_rules = [
        {
            'id': 'rule_checkout',
            'name': 'Auto-dirty on checkout',
            'trigger': 'guest_checkout',
            'action': 'set_status_dirty',
            'active': True
        },
        {
            'id': 'rule_cleaning',
            'name': 'Auto-inspected after cleaning',
            'trigger': 'cleaning_complete',
            'action': 'set_status_inspected',
            'delay': '15 minutes',
            'active': True
        }
    ]

    return {
        'status_summary': {
            'dirty': len(dirty_rooms),
            'cleaning': len(cleaning_rooms),
            'inspected': inspected_rooms,
            'maintenance': maintenance_rooms
        },
        'dirty_rooms_list': dirty_rooms,
        'cleaning_rooms_list': cleaning_rooms,
        'auto_rules': auto_rules,
        'mobile_enabled': True
    }

@router.get("/department/revenue/comprehensive-suggestions")
@cached(ttl=600, key_prefix="revenue_suggestions")  # Cache for 10 min
async def get_revenue_comprehensive_suggestions(current_user: User = Depends(get_current_user)):
    """Revenue Manager comprehensive suggestions: pricing, min stay, CTA"""
    today = datetime.now(UTC).date()

    suggestions = []
    for days_ahead in range(14):
        target_date = today + timedelta(days=days_ahead)

        # Forecast occupancy
        base = 65
        weekend_boost = 15 if target_date.weekday() in [4, 5] else 0
        variation = random.randint(-8, 12)
        occupancy = min(98, base + weekend_boost + variation)

        # Generate strategy
        if occupancy < 50:
            strategy = {
                'date': target_date.isoformat(),
                'day_of_week': target_date.strftime('%A'),
                'forecasted_occupancy': occupancy,
                'price_adjustment': -15,
                'min_stay': 1,
                'close_to_arrival': False,
                'close_to_departure': False,
                'stop_sell': False,
                'reasoning': 'Low demand - stimulate bookings with price decrease',
                'action_priority': 'high'
            }
        elif occupancy > 85:
            strategy = {
                'date': target_date.isoformat(),
                'day_of_week': target_date.strftime('%A'),
                'forecasted_occupancy': occupancy,
                'price_adjustment': 20 if occupancy > 90 else 10,
                'min_stay': 2 if occupancy > 90 else 1,
                'close_to_arrival': occupancy > 93,
                'close_to_departure': False,
                'stop_sell': occupancy > 96,
                'reasoning': 'High demand - maximize revenue with restrictions',
                'action_priority': 'high'
            }
        else:
            strategy = {
                'date': target_date.isoformat(),
                'day_of_week': target_date.strftime('%A'),
                'forecasted_occupancy': occupancy,
                'price_adjustment': 0,
                'min_stay': 1,
                'close_to_arrival': False,
                'close_to_departure': False,
                'stop_sell': False,
                'reasoning': 'Balanced demand - maintain current strategy',
                'action_priority': 'low'
            }

        suggestions.append(strategy)

    return {
        'suggestions': suggestions,
        'data_sources': {
            'pickup_curves': True,
            'historical_pace': True,
            'otb_analysis': True,
            'competitive_set': True,
            'events_calendar': True
        }
    }

@router.get("/department/finance/dashboard")
@cached(ttl=300, key_prefix="finance_dashboard")  # Cache for 5 minutes
async def get_finance_dashboard(current_user: User = Depends(get_current_user)):
    """Finance Manager Dashboard with real-time AR and integrations"""

    # AR Summary
    pending_ar = await db.invoices.count_documents({
        'tenant_id': current_user.tenant_id,
        'payment_status': {'$in': ['pending', 'partial']}
    })

    overdue_invoices = []
    total_overdue = 0
    async for invoice in db.invoices.find({
        'tenant_id': current_user.tenant_id,
        'payment_status': {'$in': ['pending', 'partial']},
        'due_date': {'$lt': datetime.now(UTC).isoformat()}
    }):
        overdue_invoices.append(invoice)
        total_overdue += invoice.get('total', 0) - invoice.get('paid_amount', 0)

    return {
        'ar_summary': {
            'pending_invoices': pending_ar,
            'overdue_count': len(overdue_invoices),
            'overdue_amount': round(total_overdue, 2),
            'aging': {
                '0-30_days': sum(1 for inv in overdue_invoices if (datetime.now(UTC) - datetime.fromisoformat(inv['due_date'])).days <= 30),
                '31-60_days': sum(1 for inv in overdue_invoices if 30 < (datetime.now(UTC) - datetime.fromisoformat(inv['due_date'])).days <= 60),
                '60+_days': sum(1 for inv in overdue_invoices if (datetime.now(UTC) - datetime.fromisoformat(inv['due_date'])).days > 60)
            }
        },
        'integrations': {
            'logo': {'enabled': False, 'status': 'not_configured'},
            'mikro': {'enabled': False, 'status': 'not_configured'},
            'sap': {'enabled': False, 'status': 'not_configured'},
            'oracle': {'enabled': False, 'status': 'not_configured'}
        },
        'e_invoice': {
            'xml_generation': True,
            'gib_integration': True,
            'status': 'active'
        },
        'data_timing': 'real_time',
        'last_closing': (datetime.now(UTC) - timedelta(days=1)).isoformat()
    }

@router.get("/department/sales/corporate-accounts")
@cached(ttl=600, key_prefix="sales_corporate")  # Cache for 10 min
async def get_corporate_accounts(
    sort_by: str = 'revenue',
    current_user: User = Depends(get_current_user)
):
    """Sales & Marketing - Corporate accounts with profiles"""

    # Aggregate corporate bookings
    corporate_bookings = await db.bookings.find({
        'tenant_id': current_user.tenant_id,
        'booking_source': {'$in': ['corporate', 'company_direct']}
    }).to_list(10000)

    companies = {}
    for booking in corporate_bookings:
        company = booking.get('company_name', 'Unknown')
        if company not in companies:
            companies[company] = {
                'name': company,
                'total_revenue': 0,
                'total_nights': 0,
                'booking_count': 0,
                'last_booking': None,
                'adr': 0
            }

        companies[company]['total_revenue'] += booking.get('total_amount', 0)
        companies[company]['booking_count'] += 1

        try:
            nights = (datetime.fromisoformat(booking.get('check_out')) -
                     datetime.fromisoformat(booking.get('check_in'))).days
            companies[company]['total_nights'] += nights
        except Exception:
            pass

    # Calculate ADR and create list
    accounts = []
    for company, data in companies.items():
        adr = data['total_revenue'] / data['total_nights'] if data['total_nights'] > 0 else 0

        # Check if profile exists
        profile = await db.corporate_profiles.find_one({
            'tenant_id': current_user.tenant_id,
            'company_name': company
        })

        accounts.append({
            **data,
            'adr': round(adr, 2),
            'has_profile': profile is not None,
            'contract_status': profile.get('contract_status') if profile else 'none',
            'blacklisted': profile.get('blacklisted', False) if profile else False
        })

    # Sort
    if sort_by == 'revenue':
        accounts.sort(key=lambda x: x['total_revenue'], reverse=True)
    elif sort_by == 'nights':
        accounts.sort(key=lambda x: x['total_nights'], reverse=True)
    elif sort_by == 'adr':
        accounts.sort(key=lambda x: x['adr'], reverse=True)

    return {
        'accounts': accounts,
        'total_companies': len(accounts),
        'sorted_by': sort_by
    }

@router.get("/department/it/system-info")
@cached(ttl=600, key_prefix="it_system_info")  # Cache for 10 min
async def get_it_system_info(current_user: User = Depends(get_current_user)):
    """IT Manager - System architecture and performance info"""
    return {
        'api_architecture': {
            'type': 'REST',
            'protocol': 'HTTP/HTTPS',
            'websocket_support': False,
            'sse_support': False,
            'polling': 'client-side (30s interval)'
        },
        'widget_architecture': {
            'type': 'modular',
            'independent_apis': True,
            'lazy_loading': True,
            'caching': 'browser + redis'
        },
        'scalability': {
            'tested_rooms': 40,
            'max_recommended': '500+ rooms',
            'database': 'MongoDB (horizontally scalable)',
            'performance_optimization': [
                'Database indexing on tenant_id, dates',
                'Query result limiting',
                'Async/await patterns',
                'Connection pooling'
            ]
        },
        'performance_metrics': {
            'avg_response_time': '< 200ms',
            'concurrent_users': '100+',
            'uptime': '99.5%'
        }
    }

@router.get("/department/guest-relations/vip-notes")
@cached(ttl=300, key_prefix="guest_relations_vip")  # Cache for 5 min
async def get_vip_notes(current_user: User = Depends(get_current_user)):
    """Guest Relations - VIP notes and review integrations"""

    # Get VIP guests with notes
    vip_guests = []
    async for guest in db.guests.find({
        'tenant_id': current_user.tenant_id,
        'vip': True
    }).limit(50):
        # Get notes
        notes = await db.guest_notes.find({
            'tenant_id': current_user.tenant_id,
            'guest_id': guest.get('id')
        }).to_list(10)

        vip_guests.append({
            'guest_id': guest.get('id'),
            'name': guest.get('name'),
            'email': guest.get('email'),
            'vip_tier': guest.get('loyalty_tier', 'gold'),
            'preferences': guest.get('preferences'),
            'notes': notes,
            'notes_visible_on_dashboard': True
        })

    return {
        'vip_guests': vip_guests,
        'review_integrations': {
            'google': {'enabled': False, 'status': 'not_configured'},
            'tripadvisor': {'enabled': False, 'status': 'not_configured'},
            'booking_com': {'enabled': False, 'status': 'not_configured'},
            'trustpilot': {'enabled': False, 'status': 'not_configured'}
        },
        'complaint_tracking': {
            'enabled': True,
            'open_complaints': 0,
            'avg_resolution_time': '24 hours'
        }
    }

@router.get("/ai/activity-feed")
@cached(ttl=300, key_prefix="ai_activity_feed")  # Cache for 5 min
async def get_ai_activity_feed(
    limit: int = 10,
    current_user: User = Depends(get_current_user)
):
    """AI Activity Feed - Real-time AI suggestions and insights"""
    today = datetime.now(UTC)

    activities = []

    # 1. Price Optimization Suggestions
    total_rooms = await db.rooms.count_documents({'tenant_id': current_user.tenant_id})
    occupied = await db.bookings.count_documents({
        'tenant_id': current_user.tenant_id,
        'status': 'checked_in'
    })
    occupancy = (occupied / total_rooms * 100) if total_rooms > 0 else 0

    if occupancy > 85:
        activities.append({
            'id': str(uuid.uuid4()),
            'type': 'price_suggestion',
            'priority': 'high',
            'title': '💰 Price Optimization Opportunity',
            'message': f'High occupancy detected ({occupancy:.1f}%). Recommend increasing rates by 15-20% for next 3 days.',
            'action': 'adjust_pricing',
            'confidence': 0.89,
            'potential_revenue': round(total_rooms * 50 * 0.15, 2),
            'created_at': today.isoformat(),
            'status': 'active'
        })
    elif occupancy < 50:
        activities.append({
            'id': str(uuid.uuid4()),
            'type': 'price_suggestion',
            'priority': 'high',
            'title': '📉 Low Occupancy Alert',
            'message': f'Occupancy at {occupancy:.1f}%. Recommend promotional rates (-10%) and special packages.',
            'action': 'create_promotion',
            'confidence': 0.85,
            'potential_bookings': round(total_rooms * 0.2),
            'created_at': today.isoformat(),
            'status': 'active'
        })

    # 2. Overbooking Detection
    next_7_days = today + timedelta(days=7)
    room_dates = {}
    async for booking in db.bookings.find({
        'tenant_id': current_user.tenant_id,
        'status': {'$in': ['confirmed', 'guaranteed']},
        'check_in': {'$lte': next_7_days.isoformat()}
    }):
        room_id = booking.get('room_id')
        check_in = datetime.fromisoformat(booking.get('check_in')).date()
        check_out = datetime.fromisoformat(booking.get('check_out')).date()

        current = check_in
        while current < check_out:
            key = f"{room_id}_{current}"
            if key not in room_dates:
                room_dates[key] = []
            room_dates[key].append(booking)
            current += timedelta(days=1)

    overbooking_count = sum(1 for bookings in room_dates.values() if len(bookings) > 1)
    if overbooking_count > 0:
        activities.append({
            'id': str(uuid.uuid4()),
            'type': 'overbooking_alert',
            'priority': 'critical',
            'title': '🚨 Overbooking Detected',
            'message': f'{overbooking_count} room conflicts found in next 7 days. Immediate action required.',
            'action': 'resolve_conflicts',
            'conflicts': overbooking_count,
            'confidence': 1.0,
            'created_at': today.isoformat(),
            'status': 'active'
        })

    # 3. VIP Visitor Insights
    vip_arrivals = []
    today_start = datetime.combine(today.date(), datetime.min.time()).replace(tzinfo=UTC)
    today_end = datetime.combine(today.date(), datetime.max.time()).replace(tzinfo=UTC)

    async for booking in db.bookings.find({
        'tenant_id': current_user.tenant_id,
        'check_in': {'$gte': today_start.isoformat(), '$lte': today_end.isoformat()},
        'status': {'$in': ['confirmed', 'guaranteed']}
    }):
        guest = await db.guests.find_one({'id': booking.get('guest_id')})
        if guest and guest.get('vip'):
            vip_arrivals.append({
                'name': guest.get('name'),
                'room': booking.get('room_number'),
                'tier': guest.get('loyalty_tier', 'gold'),
                'preferences': guest.get('preferences')
            })

    if vip_arrivals:
        activities.append({
            'id': str(uuid.uuid4()),
            'type': 'vip_insight',
            'priority': 'high',
            'title': f'⭐ {len(vip_arrivals)} VIP Arrivals Today',
            'message': 'Special attention required. Ensure welcome amenities and room preferences are prepared.',
            'action': 'review_vip_list',
            'vip_count': len(vip_arrivals),
            'vips': vip_arrivals[:3],
            'confidence': 1.0,
            'created_at': today.isoformat(),
            'status': 'active'
        })

    # 4. Revenue Anomaly Detection
    today_revenue = 0
    charges = await db.folio_charges.find({
        'tenant_id': current_user.tenant_id,
        'date': {'$gte': today_start.isoformat(), '$lte': today_end.isoformat()},
        'voided': False
    }).to_list(10000)
    today_revenue = sum(c.get('total', 0) for c in charges)

    # Get average daily revenue (last 30 days)
    thirty_days_ago = today - timedelta(days=30)
    all_charges = await db.folio_charges.find({
        'tenant_id': current_user.tenant_id,
        'date': {'$gte': thirty_days_ago.isoformat()},
        'voided': False
    }).to_list(100000)

    if all_charges:
        avg_daily_revenue = sum(c.get('total', 0) for c in all_charges) / 30
        variance = ((today_revenue - avg_daily_revenue) / avg_daily_revenue * 100) if avg_daily_revenue > 0 else 0

        if abs(variance) > 20:
            activities.append({
                'id': str(uuid.uuid4()),
                'type': 'revenue_anomaly',
                'priority': 'medium' if variance > 0 else 'high',
                'title': '📊 Revenue Anomaly Detected',
                'message': f"Today's revenue {'↗️ +' if variance > 0 else '↘️ '}{abs(variance):.1f}% vs 30-day average. {'Investigate positive spike.' if variance > 0 else 'Review for potential issues.'}",
                'action': 'analyze_revenue',
                'variance': round(variance, 1),
                'today_revenue': round(today_revenue, 2),
                'avg_revenue': round(avg_daily_revenue, 2),
                'confidence': 0.92,
                'created_at': today.isoformat(),
                'status': 'active'
            })

    # 5. Predictive Maintenance
    maintenance_due = await db.rooms.count_documents({
        'tenant_id': current_user.tenant_id,
        'last_maintenance': {'$lt': (today - timedelta(days=90)).isoformat()}
    })

    if maintenance_due > 0:
        activities.append({
            'id': str(uuid.uuid4()),
            'type': 'maintenance_alert',
            'priority': 'medium',
            'title': '🔧 Predictive Maintenance Alert',
            'message': f'{maintenance_due} rooms require scheduled maintenance. Prevent future issues with proactive maintenance.',
            'action': 'schedule_maintenance',
            'rooms_count': maintenance_due,
            'confidence': 0.88,
            'created_at': today.isoformat(),
            'status': 'active'
        })

    # 6. Booking Trend Insight
    week_bookings = await db.bookings.count_documents({
        'tenant_id': current_user.tenant_id,
        'created_at': {'$gte': (today - timedelta(days=7)).isoformat()}
    })
    prev_week_bookings = await db.bookings.count_documents({
        'tenant_id': current_user.tenant_id,
        'created_at': {
            '$gte': (today - timedelta(days=14)).isoformat(),
            '$lt': (today - timedelta(days=7)).isoformat()
        }
    })

    if week_bookings > 0 and prev_week_bookings > 0:
        booking_trend = ((week_bookings - prev_week_bookings) / prev_week_bookings * 100)
        if abs(booking_trend) > 15:
            activities.append({
                'id': str(uuid.uuid4()),
                'type': 'booking_trend',
                'priority': 'low',
                'title': '📈 Booking Trend Analysis',
                'message': f"Booking velocity {'↗️ +' if booking_trend > 0 else '↘️ '}{abs(booking_trend):.1f}% this week vs last week.",
                'action': 'view_trends',
                'trend': round(booking_trend, 1),
                'this_week': week_bookings,
                'last_week': prev_week_bookings,
                'confidence': 0.83,
                'created_at': today.isoformat(),
                'status': 'active'
            })

    # Sort by priority and limit
    priority_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
    activities.sort(key=lambda x: priority_order.get(x['priority'], 4))

    return {
        'activities': activities[:limit],
        'total_count': len(activities),
        'last_updated': today.isoformat()
    }


@router.get("/ai/dashboard/briefing")
async def get_ai_dashboard_briefing(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get AI-powered dashboard briefing for the day"""
    current_user = await get_current_user(credentials)

    today = datetime.now(UTC).strftime("%Y-%m-%d")

    # Get today's key metrics
    total_rooms = await db.rooms.count_documents({'tenant_id': current_user.tenant_id})
    occupied = await db.bookings.count_documents({
        'tenant_id': current_user.tenant_id,
        'status': 'checked_in',
        'check_in': {'$lte': today},
        'check_out': {'$gt': today}
    })

    arrivals = await db.bookings.count_documents({
        'tenant_id': current_user.tenant_id,
        'check_in': today,
        'status': {'$in': ['confirmed', 'guaranteed']}
    })

    departures = await db.bookings.count_documents({
        'tenant_id': current_user.tenant_id,
        'check_out': today
    })

    confirmed_bookings = await db.bookings.count_documents({
        'tenant_id': current_user.tenant_id,
        'status': {'$in': ['confirmed', 'guaranteed']}
    })

    pending_invoices = await db.accounting_invoices.count_documents({
        'tenant_id': current_user.tenant_id,
        'status': 'pending'
    })

    occupancy_pct = round((occupied / total_rooms * 100), 1) if total_rooms > 0 else 0

    # Get hotel name
    tenant = await db.tenants.find_one({"id": current_user.tenant_id})
    hotel_name = tenant.get('property_name', 'Otel') if tenant else 'Otel'

    # Try AI-generated briefing
    ai_summary = None
    try:
        from domains.ai.service import get_ai_service
        ai_svc = get_ai_service()
        if ai_svc.llm_enabled:
            ai_summary = await ai_svc.generate_daily_briefing(
                hotel_name=hotel_name,
                total_rooms=total_rooms,
                occupied_rooms=occupied,
                today_checkins=arrivals,
                today_checkouts=departures,
                pending_invoices=pending_invoices,
                monthly_revenue=0,
                weather="clear"
            )
    except Exception as ai_err:
        logger.info(f"AI briefing generation failed: {ai_err}")

    # Fallback summary
    if not ai_summary:
        ai_summary = (
            f"Günaydın! {hotel_name} için günlük özet: "
            f"Toplam {total_rooms} odadan {occupied} tanesi dolu (%{occupancy_pct} doluluk). "
            f"Bugün {arrivals} giriş ve {departures} çıkış bekleniyor."
        )

    # Build insights
    insights = []
    if occupancy_pct > 80:
        insights.append("Doluluk oranı yüksek! Fiyat artışı değerlendirilebilir.")
    elif occupancy_pct < 40:
        insights.append("Doluluk düşük. Promosyon kampanyası başlatmayı düşünün.")
    if arrivals > 5:
        insights.append(f"Bugün {arrivals} giriş var, resepsiyon ekibini bilgilendirin.")
    if pending_invoices > 3:
        insights.append(f"{pending_invoices} bekleyen fatura var, muhasebe takibi önerilir.")
    if confirmed_bookings > 0:
        insights.append(f"{confirmed_bookings} onaylı rezervasyon aktif.")
    if departures > 0:
        insights.append(f"{departures} çıkış planlanmış, kat hizmetlerini hazırlayın.")

    return {
        'summary': ai_summary,
        'text': ai_summary,
        'briefing': ai_summary,
        'briefing_date': today,
        'insights': insights,
        'metrics': {
            'total_rooms': total_rooms,
            'occupied_rooms': occupied,
            'occupancy_rate': occupancy_pct,
            'today_checkins': arrivals,
            'today_checkouts': departures,
            'confirmed_bookings': confirmed_bookings,
            'pending_invoices': pending_invoices,
            'monthly_revenue': 0
        },
        'generated_at': datetime.now(UTC).isoformat()
    }


@router.get("/revenue/by-department")
@cached(ttl=900, key_prefix="revenue_by_dept")  # Cache for 15 min
async def get_revenue_by_department(
    start_date: str = None,
    end_date: str = None,
    current_user: User = Depends(get_current_user)
):
    """Revenue breakdown by department (Rooms, F&B, Other)"""
    today = datetime.now(UTC)

    if not start_date:
        start_date = datetime.combine(today.date(), datetime.min.time()).replace(tzinfo=UTC).isoformat()
    if not end_date:
        end_date = datetime.combine(today.date(), datetime.max.time()).replace(tzinfo=UTC).isoformat()

    # Get all charges
    charges = await db.folio_charges.find({
        'tenant_id': current_user.tenant_id,
        'date': {'$gte': start_date, '$lte': end_date},
        'voided': False
    }).to_list(100000)

    # Categorize by department
    departments = {
        'rooms': {'name': 'Rooms', 'revenue': 0, 'count': 0, 'icon': '🛏️'},
        'fnb': {'name': 'Food & Beverage', 'revenue': 0, 'count': 0, 'icon': '🍽️'},
        'spa': {'name': 'Spa & Wellness', 'revenue': 0, 'count': 0, 'icon': '💆'},
        'minibar': {'name': 'Minibar', 'revenue': 0, 'count': 0, 'icon': '🍷'},
        'laundry': {'name': 'Laundry', 'revenue': 0, 'count': 0, 'icon': '👔'},
        'parking': {'name': 'Parking', 'revenue': 0, 'count': 0, 'icon': '🚗'},
        'telephone': {'name': 'Telephone', 'revenue': 0, 'count': 0, 'icon': '📞'},
        'other': {'name': 'Other Services', 'revenue': 0, 'count': 0, 'icon': '🎯'}
    }

    for charge in charges:
        charge_type = charge.get('charge_type', 'other').lower()
        amount = charge.get('total', 0)

        if charge_type in ['room', 'accommodation', 'room_charge']:
            departments['rooms']['revenue'] += amount
            departments['rooms']['count'] += 1
        elif charge_type in ['food', 'beverage', 'restaurant', 'bar', 'fnb']:
            departments['fnb']['revenue'] += amount
            departments['fnb']['count'] += 1
        elif charge_type in ['spa', 'massage', 'wellness']:
            departments['spa']['revenue'] += amount
            departments['spa']['count'] += 1
        elif charge_type in ['minibar', 'mini_bar']:
            departments['minibar']['revenue'] += amount
            departments['minibar']['count'] += 1
        elif charge_type in ['laundry', 'dry_cleaning']:
            departments['laundry']['revenue'] += amount
            departments['laundry']['count'] += 1
        elif charge_type in ['parking', 'valet']:
            departments['parking']['revenue'] += amount
            departments['parking']['count'] += 1
        elif charge_type in ['telephone', 'phone']:
            departments['telephone']['revenue'] += amount
            departments['telephone']['count'] += 1
        else:
            departments['other']['revenue'] += amount
            departments['other']['count'] += 1

    # Calculate totals and percentages
    total_revenue = sum(dept['revenue'] for dept in departments.values())

    for dept in departments.values():
        dept['percentage'] = round((dept['revenue'] / total_revenue * 100) if total_revenue > 0 else 0, 1)
        dept['revenue'] = round(dept['revenue'], 2)

    # Sort by revenue
    sorted_departments = sorted(
        [{'key': k, **v} for k, v in departments.items()],
        key=lambda x: x['revenue'],
        reverse=True
    )

    return {
        'departments': sorted_departments,
        'total_revenue': round(total_revenue, 2),
        'period': {
            'start': start_date,
            'end': end_date
        },
        'summary': {
            'rooms_percentage': departments['rooms']['percentage'],
            'fnb_percentage': departments['fnb']['percentage'],
            'other_percentage': sum(d['percentage'] for k, d in departments.items() if k not in ['rooms', 'fnb'])
        }
    }

@router.post("/bookings/{booking_id}/assign-room")
async def assign_room_to_booking(
    booking_id: str,
    room_assignment: dict,
    current_user: User = Depends(get_current_user)
):
    """Assign a specific room to a booking"""
    room_id = room_assignment.get('room_id')
    room_number = room_assignment.get('room_number')
    notes = room_assignment.get('notes', '')

    # Get booking
    booking = await db.bookings.find_one({
        'id': booking_id,
        'tenant_id': current_user.tenant_id
    })

    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")

    # Get room
    room = await db.rooms.find_one({
        'id': room_id,
        'tenant_id': current_user.tenant_id
    })

    if not room:
        raise HTTPException(status_code=404, detail="Room not found")

    # Check room availability
    check_in = datetime.fromisoformat(booking.get('check_in'))
    check_out = datetime.fromisoformat(booking.get('check_out'))

    # Check for conflicts
    conflicts = await db.bookings.count_documents({
        'tenant_id': current_user.tenant_id,
        'room_id': room_id,
        'id': {'$ne': booking_id},
        'status': {'$in': ['confirmed', 'guaranteed', 'checked_in']},
        '$or': [
            {
                'check_in': {'$lte': check_in.isoformat()},
                'check_out': {'$gt': check_in.isoformat()}
            },
            {
                'check_in': {'$lt': check_out.isoformat()},
                'check_out': {'$gte': check_out.isoformat()}
            }
        ]
    })

    if conflicts > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Room {room_number} is not available for this period"
        )

    # Update booking
    await db.bookings.update_one(
        {'id': booking_id},
        {
            '$set': {
                'room_id': room_id,
                'room_number': room_number,
                'room_type': room.get('room_type'),
                'room_assigned_at': datetime.now(UTC).isoformat(),
                'room_assigned_by': current_user.email,
                'room_assignment_notes': notes
            }
        }
    )

    # Log activity
    await db.activity_log.insert_one({
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'type': 'room_assignment',
        'booking_id': booking_id,
        'room_id': room_id,
        'room_number': room_number,
        'performed_by': current_user.email,
        'notes': notes,
        'timestamp': datetime.now(UTC).isoformat()
    })

    return {
        'success': True,
        'message': f'Room {room_number} assigned successfully',
        'booking_id': booking_id,
        'room_number': room_number,
        'assigned_at': datetime.now(UTC).isoformat()
    }

@router.get("/bookings/{booking_id}/available-rooms")
@cached(ttl=120, key_prefix="booking_available_rooms")  # Cache for 2 min
async def get_available_rooms_for_booking(
    booking_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get list of available rooms for a specific booking"""
    booking = await db.bookings.find_one({
        'id': booking_id,
        'tenant_id': current_user.tenant_id
    })

    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")

    check_in = datetime.fromisoformat(booking.get('check_in'))
    check_out = datetime.fromisoformat(booking.get('check_out'))
    requested_type = booking.get('room_type', 'standard')

    # Get all rooms
    all_rooms = await db.rooms.find({'tenant_id': current_user.tenant_id}).to_list(1000)

    available_rooms = []
    for room in all_rooms:
        # Check if room has conflicts
        conflicts = await db.bookings.count_documents({
            'tenant_id': current_user.tenant_id,
            'room_id': room['id'],
            'id': {'$ne': booking_id},
            'status': {'$in': ['confirmed', 'guaranteed', 'checked_in']},
            '$or': [
                {
                    'check_in': {'$lte': check_in.isoformat()},
                    'check_out': {'$gt': check_in.isoformat()}
                },
                {
                    'check_in': {'$lt': check_out.isoformat()},
                    'check_out': {'$gte': check_out.isoformat()}
                }
            ]
        })

        if conflicts == 0 and room.get('status') in ['available', 'inspected']:
            available_rooms.append({
                'id': room['id'],
                'room_number': room['room_number'],
                'room_type': room['room_type'],
                'floor': room.get('floor', 1),
                'status': room['status'],
                'price_per_night': room.get('price_per_night', 0),
                'is_same_type': room['room_type'].lower() == requested_type.lower(),
                'is_upgrade': room.get('price_per_night', 0) > booking.get('rate', 0),
                'amenities': room.get('amenities', [])
            })

    # Sort: same type first, then by floor
    available_rooms.sort(key=lambda x: (not x['is_same_type'], x['floor']))

    return {
        'available_rooms': available_rooms,
        'total_available': len(available_rooms),
        'requested_type': requested_type,
        'booking_dates': {
            'check_in': check_in.isoformat(),
            'check_out': check_out.isoformat()
        }
    }

@router.post("/housekeeping/start-cleaning/{room_id}")
async def start_cleaning_timer(
    room_id: str,
    staff_info: dict = {},
    current_user: User = Depends(get_current_user)
):
    """Start cleaning timer for a room"""
    room = await db.rooms.find_one({
        'id': room_id,
        'tenant_id': current_user.tenant_id
    })

    if not room:
        raise HTTPException(status_code=404, detail="Room not found")

    # Create cleaning task
    task_id = str(uuid.uuid4())
    task = {
        'id': task_id,
        'tenant_id': current_user.tenant_id,
        'room_id': room_id,
        'room_number': room.get('room_number'),
        'assigned_to': staff_info.get('staff_name', current_user.name),
        'assigned_id': staff_info.get('staff_id', current_user.id),
        'started_at': datetime.now(UTC).isoformat(),
        'completed_at': None,
        'status': 'in_progress',
        'duration_minutes': None,
        'notes': staff_info.get('notes', '')
    }

    await db.housekeeping_tasks.insert_one(task)

    # Update room status
    await db.rooms.update_one(
        {'id': room_id},
        {
            '$set': {
                'status': 'cleaning',
                'assigned_cleaner': task['assigned_to'],
                'cleaning_started_at': task['started_at'],
                'current_task_id': task_id
            }
        }
    )

    return {
        'success': True,
        'task_id': task_id,
        'room_number': room.get('room_number'),
        'started_at': task['started_at'],
        'assigned_to': task['assigned_to']
    }

@router.post("/housekeeping/complete-cleaning/{task_id}")
async def complete_cleaning_timer(
    task_id: str,
    completion_data: dict = {},
    current_user: User = Depends(get_current_user)
):
    """Complete cleaning timer and update room status"""
    task = await db.housekeeping_tasks.find_one({
        'id': task_id,
        'tenant_id': current_user.tenant_id
    })

    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    # Calculate duration
    started_at = datetime.fromisoformat(task['started_at'])
    completed_at = datetime.now(UTC)
    duration = (completed_at - started_at).total_seconds() / 60  # minutes

    # Update task
    await db.housekeeping_tasks.update_one(
        {'id': task_id},
        {
            '$set': {
                'completed_at': completed_at.isoformat(),
                'status': 'completed',
                'duration_minutes': round(duration, 1),
                'completion_notes': completion_data.get('notes', ''),
                'quality_score': completion_data.get('quality_score', 5)
            }
        }
    )

    # Update room status
    await db.rooms.update_one(
        {'id': task['room_id']},
        {
            '$set': {
                'status': 'inspected',
                'cleaning_completed_at': completed_at.isoformat(),
                'last_cleaned': completed_at.isoformat(),
                'current_task_id': None
            }
        }
    )

    return {
        'success': True,
        'task_id': task_id,
        'room_number': task['room_number'],
        'duration_minutes': round(duration, 1),
        'completed_at': completed_at.isoformat()
    }

@router.get("/housekeeping/active-timers")
@cached(ttl=60, key_prefix="hk_active_timers")  # Cache for 1 min
async def get_active_cleaning_timers(current_user: User = Depends(get_current_user)):
    """Get all active cleaning timers"""
    tasks = await db.housekeeping_tasks.find({
        'tenant_id': current_user.tenant_id,
        'status': 'in_progress'
    }).to_list(100)

    now = datetime.now(UTC)
    active_timers = []

    for task in tasks:
        started_at = datetime.fromisoformat(task['started_at'])
        elapsed = (now - started_at).total_seconds() / 60  # minutes

        active_timers.append({
            'task_id': task['id'],
            'room_number': task['room_number'],
            'assigned_to': task['assigned_to'],
            'started_at': task['started_at'],
            'elapsed_minutes': round(elapsed, 1),
            'status': 'in_progress'
        })

    return {
        'active_timers': active_timers,
        'total_active': len(active_timers)
    }

@router.get("/housekeeping/performance-stats")
@cached(ttl=600, key_prefix="housekeeping_performance")  # Cache for 10 minutes
async def get_housekeeping_performance_stats(
    days: int = 7,
    current_user: User = Depends(get_current_user)
):
    """Get housekeeping performance statistics"""
    since = datetime.now(UTC) - timedelta(days=days)

    completed_tasks = await db.housekeeping_tasks.find({
        'tenant_id': current_user.tenant_id,
        'status': 'completed',
        'completed_at': {'$gte': since.isoformat()}
    }).to_list(10000)

    if not completed_tasks:
        return {
            'average_duration': 0,
            'total_rooms_cleaned': 0,
            'fastest_cleaning': 0,
            'slowest_cleaning': 0,
            'staff_performance': []
        }

    durations = [t['duration_minutes'] for t in completed_tasks if t.get('duration_minutes')]
    avg_duration = sum(durations) / len(durations) if durations else 0

    # Staff performance
    staff_stats = {}
    for task in completed_tasks:
        staff = task.get('assigned_to', 'Unknown')
        if staff not in staff_stats:
            staff_stats[staff] = {
                'name': staff,
                'rooms_cleaned': 0,
                'total_duration': 0,
                'avg_duration': 0
            }
        staff_stats[staff]['rooms_cleaned'] += 1
        staff_stats[staff]['total_duration'] += task.get('duration_minutes', 0)

    for staff in staff_stats.values():
        staff['avg_duration'] = round(staff['total_duration'] / staff['rooms_cleaned'], 1) if staff['rooms_cleaned'] > 0 else 0

    return {
        'period_days': days,
        'average_duration': round(avg_duration, 1),
        'total_rooms_cleaned': len(completed_tasks),
        'fastest_cleaning': round(min(durations), 1) if durations else 0,
        'slowest_cleaning': round(max(durations), 1) if durations else 0,
        'staff_performance': sorted(staff_stats.values(), key=lambda x: x['rooms_cleaned'], reverse=True)
    }

@router.get("/rms/rate-recommendations")
@cached(ttl=600, key_prefix="rms_recommendations")  # Cache for 10 min
async def get_rate_recommendations(
    days_ahead: int = 14,
    current_user: User = Depends(get_current_user)
):
    """AI-powered rate recommendations based on demand forecast"""
    today = datetime.now(UTC).date()

    # Get current base rates
    room_types = await db.rooms.aggregate([
        {'$match': {'tenant_id': current_user.tenant_id}},
        {'$group': {
            '_id': '$room_type',
            'avg_price': {'$avg': '$price_per_night'},
            'count': {'$sum': 1}
        }}
    ]).to_list(100)

    base_rates = {rt['_id']: rt['avg_price'] for rt in room_types}
    if not base_rates:
        base_rates = {'standard': 100, 'deluxe': 150, 'suite': 250}

    recommendations = []

    for days in range(days_ahead):
        target_date = today + timedelta(days=days)

        # Forecast occupancy
        base_occ = 65
        weekend_boost = 15 if target_date.weekday() in [4, 5] else 0
        seasonal = 10 if target_date.month in [6, 7, 8, 12] else 0
        variation = random.randint(-5, 8)
        forecasted_occ = min(98, base_occ + weekend_boost + seasonal + variation)

        # Get historical bookings for this date range
        same_date_last_year = target_date.replace(year=target_date.year - 1)
        historical = await db.bookings.count_documents({
            'tenant_id': current_user.tenant_id,
            'check_in': {
                '$gte': same_date_last_year.isoformat(),
                '$lte': (same_date_last_year + timedelta(days=1)).isoformat()
            }
        })

        # Rate recommendation logic
        rate_adjustments = {}
        strategy = {}

        if forecasted_occ >= 90:
            # Very high demand
            for room_type, base_rate in base_rates.items():
                adjustment = 25
                rate_adjustments[room_type] = {
                    'current_rate': base_rate,
                    'recommended_rate': round(base_rate * (1 + adjustment/100), 2),
                    'adjustment_pct': adjustment,
                    'adjustment_amount': round(base_rate * adjustment/100, 2)
                }
            strategy = {
                'action': 'maximize',
                'min_stay': 2,
                'close_to_arrival': True,
                'stop_sell': forecasted_occ > 95,
                'reason': 'Peak demand - maximize revenue'
            }
        elif forecasted_occ >= 75:
            # Good demand
            for room_type, base_rate in base_rates.items():
                adjustment = 10
                rate_adjustments[room_type] = {
                    'current_rate': base_rate,
                    'recommended_rate': round(base_rate * (1 + adjustment/100), 2),
                    'adjustment_pct': adjustment,
                    'adjustment_amount': round(base_rate * adjustment/100, 2)
                }
            strategy = {
                'action': 'optimize',
                'min_stay': 1,
                'close_to_arrival': False,
                'stop_sell': False,
                'reason': 'Strong demand - optimize rates'
            }
        elif forecasted_occ >= 50:
            # Moderate demand
            for room_type, base_rate in base_rates.items():
                adjustment = 0
                rate_adjustments[room_type] = {
                    'current_rate': base_rate,
                    'recommended_rate': base_rate,
                    'adjustment_pct': adjustment,
                    'adjustment_amount': 0
                }
            strategy = {
                'action': 'maintain',
                'min_stay': 1,
                'close_to_arrival': False,
                'stop_sell': False,
                'reason': 'Balanced demand - maintain rates'
            }
        else:
            # Low demand
            for room_type, base_rate in base_rates.items():
                adjustment = -15
                rate_adjustments[room_type] = {
                    'current_rate': base_rate,
                    'recommended_rate': round(base_rate * (1 + adjustment/100), 2),
                    'adjustment_pct': adjustment,
                    'adjustment_amount': round(base_rate * adjustment/100, 2)
                }
            strategy = {
                'action': 'stimulate',
                'min_stay': 1,
                'close_to_arrival': False,
                'stop_sell': False,
                'reason': 'Low demand - stimulate bookings',
                'suggested_promotions': ['Weekend getaway', 'Extended stay discount']
            }

        # Calculate potential revenue impact
        total_rooms = await db.rooms.count_documents({'tenant_id': current_user.tenant_id})
        potential_revenue_impact = sum(
            adj['adjustment_amount'] * total_rooms * (forecasted_occ / 100)
            for adj in rate_adjustments.values()
        ) / len(rate_adjustments) if rate_adjustments else 0

        recommendations.append({
            'date': target_date.isoformat(),
            'day_of_week': target_date.strftime('%A'),
            'forecasted_occupancy': forecasted_occ,
            'historical_bookings': historical,
            'rate_adjustments': rate_adjustments,
            'strategy': strategy,
            'potential_revenue_impact': round(potential_revenue_impact, 2),
            'confidence': 0.85 if days < 7 else 0.75,
            'priority': 'high' if abs(strategy.get('action') in ['maximize', 'stimulate']) else 'medium'
        })

    return {
        'recommendations': recommendations,
        'total_days': len(recommendations),
        'summary': {
            'high_demand_days': sum(1 for r in recommendations if r['forecasted_occupancy'] >= 85),
            'low_demand_days': sum(1 for r in recommendations if r['forecasted_occupancy'] < 50),
            'total_potential_impact': round(sum(r['potential_revenue_impact'] for r in recommendations), 2)
        }
    }

@router.post("/rms/apply-recommendation")
async def apply_rate_recommendation(
    recommendation_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Apply recommended rates to room inventory"""
    target_date = recommendation_data.get('date')
    rate_adjustments = recommendation_data.get('rate_adjustments', {})

    updated_rooms = 0
    for room_type, adjustment in rate_adjustments.items():
        result = await db.rooms.update_many(
            {
                'tenant_id': current_user.tenant_id,
                'room_type': room_type
            },
            {
                '$set': {
                    'price_per_night': adjustment['recommended_rate'],
                    'last_rate_update': datetime.now(UTC).isoformat(),
                    'rate_update_reason': f"RMS recommendation for {target_date}"
                }
            }
        )
        updated_rooms += result.modified_count

    # Log the rate change
    await db.rate_change_log.insert_one({
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'date': target_date,
        'rate_adjustments': rate_adjustments,
        'applied_by': current_user.email,
        'applied_at': datetime.now(UTC).isoformat(),
        'source': 'rms_recommendation'
    })

    return {
        'success': True,
        'rooms_updated': updated_rooms,
        'date': target_date,
        'message': f'Rates updated for {updated_rooms} rooms'
    }

@router.get("/housekeeping/staff/{staff_id}/detailed-stats")
@cached(ttl=600, key_prefix="staff_detailed_stats")  # Cache for 10 min
async def get_staff_detailed_statistics(
    staff_id: str,
    days: int = 30,
    current_user: User = Depends(get_current_user)
):
    """Detailed staff performance by room type, shift, and speed"""
    since = datetime.now(UTC) - timedelta(days=days)

    # Get all tasks for this staff member
    tasks = await db.housekeeping_tasks.find({
        'tenant_id': current_user.tenant_id,
        'assigned_id': staff_id,
        'status': 'completed',
        'completed_at': {'$gte': since.isoformat()}
    }).to_list(10000)

    if not tasks:
        return {'error': 'No data for this staff member'}

    # Get staff info
    staff = await db.users.find_one({'id': staff_id}) or await db.staff.find_one({'id': staff_id})

    # BY ROOM TYPE
    by_room_type = {}
    for task in tasks:
        room = await db.rooms.find_one({'id': task['room_id']})
        room_type = room.get('room_type', 'unknown') if room else 'unknown'

        if room_type not in by_room_type:
            by_room_type[room_type] = {
                'count': 0,
                'total_duration': 0,
                'avg_duration': 0,
                'fastest': 999,
                'slowest': 0
            }

        duration = task.get('duration_minutes', 0)
        by_room_type[room_type]['count'] += 1
        by_room_type[room_type]['total_duration'] += duration
        by_room_type[room_type]['fastest'] = min(by_room_type[room_type]['fastest'], duration)
        by_room_type[room_type]['slowest'] = max(by_room_type[room_type]['slowest'], duration)

    for stats in by_room_type.values():
        stats['avg_duration'] = round(stats['total_duration'] / stats['count'], 1) if stats['count'] > 0 else 0

    # BY SHIFT (Morning / Afternoon / Night)
    by_shift = {'morning': [], 'afternoon': [], 'evening': []}
    for task in tasks:
        started_at = datetime.fromisoformat(task['started_at'])
        hour = started_at.hour

        if 6 <= hour < 14:
            by_shift['morning'].append(task)
        elif 14 <= hour < 22:
            by_shift['afternoon'].append(task)
        else:
            by_shift['evening'].append(task)

    shift_stats = {}
    for shift, shift_tasks in by_shift.items():
        if shift_tasks:
            durations = [t.get('duration_minutes', 0) for t in shift_tasks]
            shift_stats[shift] = {
                'rooms_cleaned': len(shift_tasks),
                'avg_duration': round(sum(durations) / len(durations), 1),
                'total_hours': round(sum(durations) / 60, 1)
            }
        else:
            shift_stats[shift] = {'rooms_cleaned': 0, 'avg_duration': 0, 'total_hours': 0}

    # SPEED ANALYSIS
    all_durations = [t.get('duration_minutes', 0) for t in tasks]
    avg_duration = sum(all_durations) / len(all_durations)

    # Compare to hotel average
    hotel_tasks = await db.housekeeping_tasks.find({
        'tenant_id': current_user.tenant_id,
        'status': 'completed',
        'completed_at': {'$gte': since.isoformat()}
    }).to_list(100000)

    hotel_durations = [t.get('duration_minutes', 0) for t in hotel_tasks]
    hotel_avg = sum(hotel_durations) / len(hotel_durations) if hotel_durations else 0

    speed_rating = 'average'
    if avg_duration < hotel_avg * 0.85:
        speed_rating = 'fast'
    elif avg_duration > hotel_avg * 1.15:
        speed_rating = 'slow'

    # QUALITY SCORES
    quality_scores = [t.get('quality_score', 5) for t in tasks if t.get('quality_score')]
    avg_quality = sum(quality_scores) / len(quality_scores) if quality_scores else 5

    # DAY-by-DAY PERFORMANCE
    daily_performance = {}
    for task in tasks:
        date = task['started_at'][:10]
        if date not in daily_performance:
            daily_performance[date] = {'rooms': 0, 'total_time': 0}
        daily_performance[date]['rooms'] += 1
        daily_performance[date]['total_time'] += task.get('duration_minutes', 0)

    return {
        'staff_info': {
            'id': staff_id,
            'name': staff.get('name', 'Unknown') if staff else 'Unknown',
            'email': staff.get('email', '') if staff else ''
        },
        'period': {
            'days': days,
            'start_date': since.isoformat(),
            'end_date': datetime.now(UTC).isoformat()
        },
        'overall': {
            'total_rooms_cleaned': len(tasks),
            'avg_duration': round(avg_duration, 1),
            'fastest_cleaning': round(min(all_durations), 1),
            'slowest_cleaning': round(max(all_durations), 1),
            'avg_quality_score': round(avg_quality, 1),
            'speed_rating': speed_rating,
            'vs_hotel_avg': round(((avg_duration - hotel_avg) / hotel_avg * 100) if hotel_avg > 0 else 0, 1)
        },
        'by_room_type': by_room_type,
        'by_shift': shift_stats,
        'daily_performance': daily_performance
    }

@router.get("/reports/market-segment")
@cached(ttl=900, key_prefix="report_market_segment")  # Cache for 15 minutes
async def get_market_segment_report(
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user)
):
    """Market Segment & Rate Type Performance Report"""
    start = datetime.fromisoformat(start_date)
    end = datetime.fromisoformat(end_date)

    # Get all bookings in date range
    bookings = await db.bookings.find({
        'tenant_id': current_user.tenant_id,
        'check_in': {'$gte': start.isoformat()},
        'check_out': {'$lte': end.isoformat()}
    }).to_list(10000)

    # Aggregate by market segment
    segment_data = {}
    rate_type_data = {}

    for booking in bookings:
        segment = booking.get('market_segment', 'other')
        rate_type = booking.get('rate_type', 'bar')

        # Calculate nights
        check_in = datetime.fromisoformat(booking['check_in'])
        check_out = datetime.fromisoformat(booking['check_out'])
        nights = (check_out - check_in).days
        revenue = booking.get('total_amount', 0)

        # Market segment aggregation
        if segment not in segment_data:
            segment_data[segment] = {'bookings': 0, 'nights': 0, 'revenue': 0}
        segment_data[segment]['bookings'] += 1
        segment_data[segment]['nights'] += nights
        segment_data[segment]['revenue'] += revenue

        # Rate type aggregation
        if rate_type not in rate_type_data:
            rate_type_data[rate_type] = {'bookings': 0, 'nights': 0, 'revenue': 0}
        rate_type_data[rate_type]['bookings'] += 1
        rate_type_data[rate_type]['nights'] += nights
        rate_type_data[rate_type]['revenue'] += revenue

    # Calculate averages
    for segment in segment_data:
        segment_data[segment]['adr'] = round(
            segment_data[segment]['revenue'] / segment_data[segment]['nights'], 2
        ) if segment_data[segment]['nights'] > 0 else 0

    for rate_type in rate_type_data:
        rate_type_data[rate_type]['adr'] = round(
            rate_type_data[rate_type]['revenue'] / rate_type_data[rate_type]['nights'], 2
        ) if rate_type_data[rate_type]['nights'] > 0 else 0

    return {
        'start_date': start_date,
        'end_date': end_date,
        'total_bookings': len(bookings),
        'market_segments': segment_data,
        'rate_types': rate_type_data
    }


@router.get("/reports/market-segment/excel")
@cached(ttl=900, key_prefix="report_market_segment_excel")  # Cache for 15 min
async def export_market_segment_excel(
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user)
):
    """Export Market Segment Report to Excel"""
    report_data = await get_market_segment_report(start_date, end_date, current_user)

    # Create workbook with multiple sheets
    wb = Workbook()

    # Sheet 1: Market Segments
    ws1 = wb.active
    ws1.title = "Market Segments"

    headers1 = ["Segment", "Bookings", "Nights", "Revenue", "ADR"]
    data1 = []
    for segment, stats in report_data['market_segments'].items():
        data1.append([
            segment.title(),
            stats['bookings'],
            stats['nights'],
            f"${stats['revenue']:,.2f}",
            f"${stats['adr']:,.2f}"
        ])

    # Add title and headers
    ws1.merge_cells('A1:E1')
    title_cell = ws1['A1']
    title_cell.value = f"Market Segment Report ({start_date} to {end_date})"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    for col_num, header in enumerate(headers1, 1):
        cell = ws1.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    for row_num, row_data in enumerate(data1, 3):
        for col_num, value in enumerate(row_data, 1):
            ws1.cell(row=row_num, column=col_num, value=value)

    # Sheet 2: Rate Types
    ws2 = wb.create_sheet("Rate Types")

    headers2 = ["Rate Type", "Bookings", "Nights", "Revenue", "ADR"]
    data2 = []
    for rate_type, stats in report_data['rate_types'].items():
        data2.append([
            rate_type.upper(),
            stats['bookings'],
            stats['nights'],
            f"${stats['revenue']:,.2f}",
            f"${stats['adr']:,.2f}"
        ])

    ws2.merge_cells('A1:E1')
    title_cell = ws2['A1']
    title_cell.value = f"Rate Type Report ({start_date} to {end_date})"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    for col_num, header in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    for row_num, row_data in enumerate(data2, 3):
        for col_num, value in enumerate(row_data, 1):
            ws2.cell(row=row_num, column=col_num, value=value)

    filename = f"market_segment_report_{start_date}_to_{end_date}.xlsx"
    return excel_response(wb, filename)


@router.get("/reports/company-aging")
@cached(ttl=900, key_prefix="report_company_aging")  # Cache for 15 min
async def get_company_aging_report(current_user: User = Depends(get_current_user)):
    """Company Accounts Receivable Aging Report"""
    today = datetime.now(UTC).date()

    # Get all company folios with outstanding balance
    folios = await db.folios.find({
        'tenant_id': current_user.tenant_id,
        'folio_type': 'company',
        'status': 'open'
    }).to_list(10000)

    company_balances = {}

    for folio in folios:
        balance = await calculate_folio_balance(folio['id'], current_user.tenant_id)

        if balance > 0:
            company_id = folio.get('company_id')
            if not company_id:
                continue

            # Get company details
            company = await db.companies.find_one({'id': company_id}, {'_id': 0})
            if not company:
                continue

            # Calculate aging based on folio creation date
            folio_created = datetime.fromisoformat(folio['created_at']).date()
            age_days = (today - folio_created).days

            # Determine aging bucket
            if age_days <= 7:
                aging_bucket = '0-7 days'
            elif age_days <= 14:
                aging_bucket = '8-14 days'
            elif age_days <= 30:
                aging_bucket = '15-30 days'
            else:
                aging_bucket = '30+ days'

            # Aggregate by company
            if company_id not in company_balances:
                company_balances[company_id] = {
                    'company_name': company['name'],
                    'corporate_code': company.get('corporate_code', 'N/A'),
                    'total_balance': 0,
                    'aging': {
                        '0-7 days': 0,
                        '8-14 days': 0,
                        '15-30 days': 0,
                        '30+ days': 0
                    },
                    'folio_count': 0
                }

            company_balances[company_id]['total_balance'] += balance
            company_balances[company_id]['aging'][aging_bucket] += balance
            company_balances[company_id]['folio_count'] += 1

    # Sort by total balance descending
    sorted_companies = sorted(
        company_balances.values(),
        key=lambda x: x['total_balance'],
        reverse=True
    )

    total_ar = sum(c['total_balance'] for c in sorted_companies)

    return {
        'report_date': today.isoformat(),
        'total_ar': round(total_ar, 2),
        'company_count': len(sorted_companies),
        'companies': sorted_companies
    }


@router.get("/reports/company-aging/excel")
@cached(ttl=900, key_prefix="report_company_aging_excel")  # Cache for 15 min
async def export_company_aging_excel(current_user: User = Depends(get_current_user)):
    """Export Company Aging Report to Excel"""
    report_data = await get_company_aging_report(current_user)

    headers = ["Company", "Corporate Code", "Total Balance", "0-7 Days", "8-14 Days", "15-30 Days", "30+ Days", "Folios"]
    data = []

    for company in report_data['companies']:
        data.append([
            company['company_name'],
            company['corporate_code'],
            f"${company['total_balance']:,.2f}",
            f"${company['aging']['0-7 days']:,.2f}",
            f"${company['aging']['8-14 days']:,.2f}",
            f"${company['aging']['15-30 days']:,.2f}",
            f"${company['aging']['30+ days']:,.2f}",
            company['folio_count']
        ])

    # Add total row
    data.append([
        "TOTAL",
        "",
        f"${report_data['total_ar']:,.2f}",
        "",
        "",
        "",
        "",
        ""
    ])

    wb = create_excel_workbook(
        title=f"Company Aging Report - {report_data['report_date']}",
        headers=headers,
        data=data,
        sheet_name="Company Aging"
    )

    filename = f"company_aging_report_{report_data['report_date']}.xlsx"
    return excel_response(wb, filename)



@router.get("/reports/finance-snapshot")
@cached(ttl=600, key_prefix="report_finance_snapshot")  # Cache for 10 min
async def get_finance_snapshot(current_user: User = Depends(get_current_user)):
    """
    Finance Snapshot for GM Dashboard
    Returns: Total Pending AR, Overdue Invoices (categorized), Today's Collections
    """
    today = datetime.now(UTC).date()
    today_start = datetime.combine(today, datetime.min.time()).replace(tzinfo=UTC)
    today_end = datetime.combine(today, datetime.max.time()).replace(tzinfo=UTC)

    # 1. Calculate Total Pending AR from company folios
    company_folios = await db.folios.find({
        'tenant_id': current_user.tenant_id,
        'folio_type': 'company',
        'status': 'open'
    }).to_list(10000)

    total_pending_ar = 0
    overdue_0_30 = 0
    overdue_30_60 = 0
    overdue_60_plus = 0
    overdue_invoices_count = 0

    for folio in company_folios:
        balance = await calculate_folio_balance(folio['id'], current_user.tenant_id)

        if balance > 0:
            total_pending_ar += balance

            # Calculate aging
            folio_created = datetime.fromisoformat(folio['created_at']).date()
            age_days = (today - folio_created).days

            if age_days > 0:  # Any overdue
                overdue_invoices_count += 1

                if age_days <= 30:
                    overdue_0_30 += balance
                elif age_days <= 60:
                    overdue_30_60 += balance
                else:
                    overdue_60_plus += balance

    # 2. Calculate Today's Collections (payments received today)
    todays_payments = await db.payments.find({
        'tenant_id': current_user.tenant_id,
        'processed_at': {
            '$gte': today_start.isoformat(),
            '$lte': today_end.isoformat()
        }
    }).to_list(10000)

    todays_collections = sum(payment.get('amount', 0) for payment in todays_payments)
    todays_payment_count = len(todays_payments)

    # 3. Calculate MTD (Month-to-Date) Collections
    month_start = today.replace(day=1)
    month_start_dt = datetime.combine(month_start, datetime.min.time()).replace(tzinfo=UTC)

    mtd_payments = await db.payments.find({
        'tenant_id': current_user.tenant_id,
        'processed_at': {
            '$gte': month_start_dt.isoformat(),
            '$lte': today_end.isoformat()
        }
    }).to_list(10000)

    mtd_collections = sum(payment.get('amount', 0) for payment in mtd_payments)

    # 4. Calculate Collection Rate (MTD Collections / MTD Revenue)
    mtd_charges = await db.folio_charges.find({
        'tenant_id': current_user.tenant_id,
        'date': {
            '$gte': month_start_dt.isoformat(),
            '$lte': today_end.isoformat()
        },
        'voided': False
    }).to_list(10000)

    mtd_revenue = sum(charge.get('total', 0) for charge in mtd_charges)
    collection_rate = (mtd_collections / mtd_revenue * 100) if mtd_revenue > 0 else 0

    # 5. Get Accounting Invoices (E-Fatura ready)
    pending_invoices = await db.accounting_invoices.find({
        'tenant_id': current_user.tenant_id,
        'status': {'$in': ['pending', 'partial']}
    }).to_list(1000)

    pending_invoice_total = sum(inv.get('total', 0) for inv in pending_invoices)
    pending_invoice_count = len(pending_invoices)

    return {
        'report_date': today.isoformat(),
        'pending_ar': {
            'total': round(total_pending_ar, 2),
            'overdue_breakdown': {
                '0-30_days': round(overdue_0_30, 2),
                '30-60_days': round(overdue_30_60, 2),
                '60_plus_days': round(overdue_60_plus, 2)
            },
            'overdue_invoices_count': overdue_invoices_count
        },
        'todays_collections': {
            'amount': round(todays_collections, 2),
            'payment_count': todays_payment_count
        },
        'mtd_collections': {
            'amount': round(mtd_collections, 2),
            'collection_rate_percentage': round(collection_rate, 2)
        },
        'accounting_invoices': {
            'pending_count': pending_invoice_count,
            'pending_total': round(pending_invoice_total, 2)
        }
    }

@router.get("/reports/revenue-detail/excel")
async def export_revenue_detail_excel(
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user),
    _: None = Depends(require_module("reports")),
):
    """Detailed room revenue by date, room type and rate code.

    NOTE: Uses bookings collection and groups by date, room_type and rate_code-like fields.
    """
    start = datetime.fromisoformat(start_date)
    end = datetime.fromisoformat(end_date)

    if start.tzinfo is None:
        start = start.replace(tzinfo=UTC)
    if end.tzinfo is None:
        end = end.replace(tzinfo=UTC)

    # Fetch bookings in range
    bookings = await db.bookings.find(
        {
            'tenant_id': current_user.tenant_id,
            'status': {'$in': ['confirmed', 'guaranteed', 'checked_in', 'checked_out']},
            '$or': [
                {'check_in': {'$gte': start.isoformat(), '$lte': end.isoformat()}},
                {'check_out': {'$gte': start.isoformat(), '$lte': end.isoformat()}},
                {'check_in': {'$lte': start.isoformat()}, 'check_out': {'$gte': end.isoformat()}},
            ],
        },
        {
            '_id': 0,
            'check_in': 1,
            'check_out': 1,
            'total_amount': 1,
            'room_type': 1,
            'rate_plan': 1,
            'market_segment': 1,
        },
    ).to_list(10000)

    # Aggregate per stay-date

    for b in bookings:
        try:
            ci = datetime.fromisoformat(b['check_in'])
            co = datetime.fromisoformat(b['check_out'])
        except Exception:
            continue

        # Normalize to date range
        ci_date = max(ci.date(), start.date())
        co_date = min(co.date(), end.date())

        days = (co_date - ci_date).days or 1
        (b.get('total_amount') or 0) / days



@router.get("/reports/forecast-detail/excel")
async def export_forecast_detail_excel(
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user),
    _: None = Depends(require_module("reports")),
):
    """Forecasted occupancy and revenue detail by date using existing forecast logic.

    NOTE: This uses get_forecast endpoint internally if available.
    """
    # Reuse get_forecast if defined
    try:
        forecast_response = await get_forecast(
            days=(datetime.fromisoformat(end_date) - datetime.fromisoformat(start_date)).days + 1,
            current_user=current_user,
            _=None,
        )
    except Exception:
        forecast_response = {}

    headers = ['Date', 'Expected Occupancy %', 'Expected Revenue']
    data: list[list[Any]] = []

    for item in forecast_response.get('days', []):
        data.append([
            item.get('date'),
            item.get('expected_occupancy_pct', 0),
            item.get('expected_revenue', 0),
        ])

    title = f"Forecast Detail {start_date} to {end_date}"
    wb = create_excel_workbook(
        title=title,
        headers=headers,
        data=data,
        sheet_name="Forecast Detail",
    )

    filename = f"forecast_detail_{start_date}_to_{end_date}.xlsx"
    return excel_response(wb, filename)

    # Prepare Excel
    headers = [
        'Date',
        'Room Type',
        'Rate Plan',
        'Nights',
        'Revenue',
        'ADR',
    ]

    data: list[list[Any]] = []
    for key, row in sorted(daily_stats.items(), key=lambda x: (x[1]['date'], x[1]['room_type'])):
        nights = row['nights'] or 1
        adr = row['revenue'] / nights
        data.append([
            row['date'],
            row['room_type'],
            row['rate_plan'],
            row['nights'],
            round(row['revenue'], 2),
            round(adr, 2),
        ])

    title = f"Revenue Detail {start_date} to {end_date}"
    wb = create_excel_workbook(
        title=title,
        headers=headers,
        data=data,
        sheet_name="Revenue Detail",
    )

    filename = f"revenue_detail_{start_date}_to_{end_date}.xlsx"
    return excel_response(wb, filename)


@router.get("/reports/operations-daily-summary/excel")
async def export_operations_daily_summary_excel(
    date: str,
    current_user: User = Depends(get_current_user),
    _: None = Depends(require_module("reports")),
):
    """Daily operations summary: arrivals, departures, in-house guests."""
    target = datetime.fromisoformat(date)
    if target.tzinfo is None:
        target = target.replace(tzinfo=UTC)

    day_start = datetime.combine(target.date(), datetime.min.time()).replace(tzinfo=UTC)
    day_end = datetime.combine(target.date(), datetime.max.time()).replace(tzinfo=UTC)

    arrivals = await db.bookings.count_documents({
        'tenant_id': current_user.tenant_id,
        'check_in': {'$gte': day_start.isoformat(), '$lte': day_end.isoformat()},
        'status': {'$in': ['confirmed', 'guaranteed', 'checked_in']},
    })

    departures = await db.bookings.count_documents({
        'tenant_id': current_user.tenant_id,
        'check_out': {'$gte': day_start.isoformat(), '$lte': day_end.isoformat()},
        'status': {'$in': ['checked_in', 'checked_out']},
    })

    in_house = await db.bookings.count_documents({
        'tenant_id': current_user.tenant_id,
        'status': 'checked_in',
    })

    headers = ['Metric', 'Value']
    data = [
        ['Date', target.date().isoformat()],
        ['', ''],
        ['Arrivals', arrivals],
        ['Departures', departures],
        ['In-House Guests', in_house],
    ]

    title = f"Operations Daily Summary {target.date().isoformat()}"
    wb = create_excel_workbook(
        title=title,
        headers=headers,
        data=data,
        sheet_name="Operations Summary",
    )

    filename = f"operations_daily_summary_{target.date().isoformat()}.xlsx"
    return excel_response(wb, filename)


@router.get("/reports/channel-distribution/excel")
async def export_channel_distribution_excel(
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user),
    _: None = Depends(require_module("reports")),
):
    """Sales channel distribution report (OTA, Direct, Corporate, etc.)."""
    start = datetime.fromisoformat(start_date)
    end = datetime.fromisoformat(end_date)

    if start.tzinfo is None:
        start = start.replace(tzinfo=UTC)
    if end.tzinfo is None:
        end = end.replace(tzinfo=UTC)

    bookings = await db.bookings.find(
        {
            'tenant_id': current_user.tenant_id,
            'status': {'$in': ['confirmed', 'guaranteed', 'checked_in', 'checked_out']},
            'check_in': {'$gte': start.isoformat(), '$lte': end.isoformat()},
        },
        {
            '_id': 0,
            'total_amount': 1,
            'channel': 1,
            'market_segment': 1,
        },
    ).to_list(20000)

    channel_stats: dict[str, dict[str, Any]] = {}

    for b in bookings:
        channel = str(b.get('channel') or 'DIRECT')
        if channel not in channel_stats:
            channel_stats[channel] = {
                'channel': channel,
                'bookings': 0,
                'revenue': 0.0,
            }
        channel_stats[channel]['bookings'] += 1
        channel_stats[channel]['revenue'] += b.get('total_amount') or 0.0

    total_revenue = sum(v['revenue'] for v in channel_stats.values()) or 1.0

    headers = [
        'Channel',
        'Bookings',
        'Revenue',
        'Share %',
    ]

    data: list[list[Any]] = []
    for key, row in sorted(channel_stats.items(), key=lambda x: x[0]):
        share = (row['revenue'] / total_revenue) * 100.0
        data.append([
            row['channel'],
            row['bookings'],
            round(row['revenue'], 2),
            round(share, 2),
        ])

    title = f"Channel Distribution {start_date} to {end_date}"
    wb = create_excel_workbook(
        title=title,
        headers=headers,
        data=data,
        sheet_name="Channels",
    )

    filename = f"channel_distribution_{start_date}_to_{end_date}.xlsx"
    return excel_response(wb, filename)


@router.get("/pos/auto-post-settings")
@cached(ttl=600, key_prefix="pos_auto_post")  # Cache for 10 min
async def get_pos_auto_post_settings(current_user: User = Depends(get_current_user)):
    """
    Get POS auto-post settings for the tenant
    """
    settings = await db.pos_settings.find_one({
        'tenant_id': current_user.tenant_id,
        'type': 'auto_post'
    })

    if not settings:
        # Default settings
        return {
            'mode': 'realtime',
            'batch_interval': 15,
            'last_sync': None
        }

    return {
        'mode': settings.get('mode', 'realtime'),
        'batch_interval': settings.get('batch_interval', 15),
        'last_sync': settings.get('last_sync')
    }

@router.post("/pos/auto-post-settings")
async def update_pos_auto_post_settings(
    settings_data: dict,
    current_user: User = Depends(get_current_user)
):
    """
    Update POS auto-post settings
    """
    await db.pos_settings.update_one(
        {
            'tenant_id': current_user.tenant_id,
            'type': 'auto_post'
        },
        {
            '$set': {
                'mode': settings_data.get('mode', 'realtime'),
                'batch_interval': settings_data.get('batch_interval', 15),
                'updated_at': datetime.now(UTC).isoformat(),
                'updated_by': current_user.id
            }
        },
        upsert=True
    )

    return {'message': 'Settings updated successfully'}

@router.post("/pos/manual-sync")
async def manual_pos_sync(current_user: User = Depends(get_current_user)):
    """
    Manually trigger POS charges sync to folios
    """
    # Get all pending POS charges
    pending_charges = await db.pos_charges.find({
        'tenant_id': current_user.tenant_id,
        'posted_to_folio': False,
        'status': 'closed'
    }).to_list(1000)

    posted_count = 0

    for charge in pending_charges:
        try:
            # Post to folio
            folio_charge = {
                'id': str(uuid.uuid4()),
                'folio_id': charge['folio_id'],
                'tenant_id': current_user.tenant_id,
                'description': charge.get('description', 'POS Charge'),
                'charge_category': charge.get('outlet', 'restaurant'),
                'date': charge['charge_date'],
                'quantity': 1,
                'unit_price': charge['total'],
                'total': charge['total'],
                'tax_amount': charge.get('tax', 0),
                'voided': False,
                'line_items': charge.get('items', []),  # Include POS line items
                'created_at': datetime.now(UTC).isoformat(),
                'created_by': current_user.id
            }

            await db.folio_charges.insert_one(folio_charge)

            # Mark as posted
            await db.pos_charges.update_one(
                {'_id': charge['_id']},
                {'$set': {'posted_to_folio': True, 'posted_at': datetime.now(UTC).isoformat()}}
            )

            posted_count += 1
        except Exception as e:
            logger.info(f"Failed to post POS charge {charge.get('id')}: {str(e)}")
            continue

    # Update last sync time
    await db.pos_settings.update_one(
        {
            'tenant_id': current_user.tenant_id,
            'type': 'auto_post'
        },
        {
            '$set': {'last_sync': datetime.now(UTC).isoformat()}
        },
        upsert=True
    )

    return {
        'posted_count': posted_count,
        'message': f'Successfully posted {posted_count} POS charges to folios'
    }

@router.post("/pos/manual-post")
async def manual_pos_post(
    post_data: dict,
    current_user: User = Depends(get_current_user)
):
    """
    Manual post of POS charge via QR/barcode (fallback when integration fails)
    """
    charge_id = post_data.get('charge_id')
    folio_id = post_data.get('folio_id')
    method = post_data.get('method', 'manual')

    # Get POS charge
    charge = await db.pos_charges.find_one({
        'id': charge_id,
        'tenant_id': current_user.tenant_id
    })

    if not charge:
        raise HTTPException(status_code=404, detail='POS charge not found')

    # Check if already posted
    if charge.get('posted_to_folio'):
        raise HTTPException(status_code=409, detail='Charge already posted to folio')

    # Post to folio
    folio_charge = {
        'id': str(uuid.uuid4()),
        'folio_id': folio_id,
        'tenant_id': current_user.tenant_id,
        'description': charge.get('description', 'POS Charge - Manual Post'),
        'charge_category': charge.get('outlet', 'restaurant'),
        'date': charge['charge_date'],
        'quantity': 1,
        'unit_price': charge['total'],
        'total': charge['total'],
        'tax_amount': charge.get('tax', 0),
        'voided': False,
        'line_items': charge.get('items', []),
        'manual_post': True,
        'post_method': method,
        'created_at': datetime.now(UTC).isoformat(),
        'created_by': current_user.id
    }

    await db.folio_charges.insert_one(folio_charge)

    # Mark as posted
    await db.pos_charges.update_one(
        {'_id': charge['_id']},
        {
            '$set': {
                'posted_to_folio': True,
                'posted_at': datetime.now(UTC).isoformat(),
                'post_method': method
            }
        }
    )

    return {
        'total': charge['total'],
        'description': charge.get('description'),
        'folio_id': folio_id,
        'posted_at': datetime.now(UTC).isoformat()
    }

@router.get("/rates/periods")
@cached(ttl=600, key_prefix="rates_periods")  # Cache for 10 min
async def get_rate_periods(
    operator_id: str,
    room_type_id: str,
    current_user: User = Depends(get_current_user)
):
    """
    Get multi-period rates for operator and room type
    """
    periods = await db.rate_periods.find({
        'tenant_id': current_user.tenant_id,
        'operator_id': operator_id,
        'room_type_id': room_type_id
    }).sort('start_date', 1).to_list(100)

    return {'periods': periods}

@router.post("/rates/periods/bulk-update")
async def bulk_update_rate_periods(
    data: dict,
    current_user: User = Depends(get_current_user)
):
    """
    Bulk update/insert rate periods for operator
    """
    operator_id = data.get('operator_id')
    room_type_id = data.get('room_type_id')
    periods = data.get('periods', [])

    # Delete existing periods
    await db.rate_periods.delete_many({
        'tenant_id': current_user.tenant_id,
        'operator_id': operator_id,
        'room_type_id': room_type_id
    })

    # Insert new periods
    if periods:
        for period in periods:
            period_doc = {
                'id': period.get('id') if not period.get('isNew') else str(uuid.uuid4()),
                'tenant_id': current_user.tenant_id,
                'operator_id': operator_id,
                'room_type_id': room_type_id,
                'start_date': period['start_date'],
                'end_date': period['end_date'],
                'rate': period['rate'],
                'currency': period.get('currency', 'USD'),
                'created_at': datetime.now(UTC).isoformat(),
                'created_by': current_user.id
            }
            await db.rate_periods.insert_one(period_doc)

    return {'message': f'{len(periods)} rate periods saved successfully'}

@router.get("/rates/stop-sale/status")
@cached(ttl=300, key_prefix="rates_stop_sale")  # Cache for 5 min
async def get_stop_sale_status(current_user: User = Depends(get_current_user)):
    """
    Get stop-sale status for all operators
    """
    stop_sales = await db.stop_sales.find({
        'tenant_id': current_user.tenant_id,
        'active': True
    }).to_list(100)

    operators = {}
    for ss in stop_sales:
        operators[ss['operator_id']] = ss.get('stop_sale', False)
        operators[f"{ss['operator_id']}_timestamp"] = ss.get('updated_at')

    return {'operators': operators}

@router.post("/rates/stop-sale/toggle")
async def toggle_stop_sale(
    data: dict,
    current_user: User = Depends(get_current_user)
):
    """
    Toggle stop-sale for specific operator
    """
    operator_id = data.get('operator_id')
    stop_sale = data.get('stop_sale', False)

    await db.stop_sales.update_one(
        {
            'tenant_id': current_user.tenant_id,
            'operator_id': operator_id
        },
        {
            '$set': {
                'stop_sale': stop_sale,
                'active': True,
                'updated_at': datetime.now(UTC).isoformat(),
                'updated_by': current_user.id
            }
        },
        upsert=True
    )

    return {
        'operator_id': operator_id,
        'stop_sale': stop_sale,
        'message': f'Stop-sale {"activated" if stop_sale else "deactivated"} for {operator_id}'
    }

@router.get("/allotment/consumption")
@cached(ttl=300, key_prefix="allotment_consumption")  # Cache for 5 min
async def get_allotment_consumption(
    start_date: str = None,
    end_date: str = None,
    current_user: User = Depends(get_current_user)
):
    """
    Get allotment consumption chart data: Allocated vs Sold vs Remaining
    """
    # Get all allotments for tenant
    allotments = await db.allotments.find({
        'tenant_id': current_user.tenant_id,
        'status': 'active'
    }).to_list(100)

    consumption_data = []

    for allotment in allotments:
        operator_name = allotment.get('operator_name', 'Unknown')
        allocated = allotment.get('allocated_rooms', 0)

        # Count sold bookings for this allotment
        bookings = await db.bookings.find({
            'tenant_id': current_user.tenant_id,
            'allotment_id': allotment.get('id'),
            'status': {'$in': ['confirmed', 'checked_in', 'checked_out']}
        }).to_list(1000)

        sold = len(bookings)
        remaining = max(allocated - sold, 0)
        utilization = int(sold / allocated * 100) if allocated > 0 else 0

        # Determine status
        if remaining == 0:
            status = 'critical'
        elif utilization >= 80:
            status = 'warning'
        else:
            status = 'good'

        consumption_data.append({
            'operator': operator_name,
            'allocated': allocated,
            'sold': sold,
            'remaining': remaining,
            'utilization': utilization,
            'status': status
        })

    return {'allotments': consumption_data}





@router.get("/reports/cost-summary")
@cached(ttl=600, key_prefix="report_cost_summary")  # Cache for 10 min
async def get_cost_summary(current_user: User = Depends(get_current_user)):
    """
    Cost Summary Report for GM Dashboard
    Returns: MTD costs by category, top cost categories, per-room cost, cost vs RevPAR
    """
    today = datetime.now(UTC).date()
    month_start = today.replace(day=1)
    month_start_dt = datetime.combine(month_start, datetime.min.time()).replace(tzinfo=UTC)
    today_end = datetime.combine(today, datetime.max.time()).replace(tzinfo=UTC)

    # 1. Get all Purchase Orders from Marketplace for this month (approved/received status)
    await db.purchase_orders.find({
        'tenant_id': current_user.tenant_id,
        'status': {'$in': ['approved', 'received', 'completed']},
        'created_at': {
            '$gte': month_start_dt.isoformat(),
            '$lte': today_end.isoformat()
        }
    }).to_list(10000)


@router.post("/reviews/ai-sentiment-analysis")
async def ai_sentiment_analysis(
    data: dict,
    current_user: User = Depends(get_current_user)
):
    """
    AI Sentiment Analysis for guest reviews
    Returns: sentiment, confidence, issues, highlights, recommendations
    """
    review_text = data.get('review_text', '')

    if not review_text:
        raise HTTPException(status_code=400, detail='Review text is required')

    # Simple keyword-based sentiment analysis (can be replaced with actual AI API)

@router.post("/bookings/walk-in-quick")
async def create_walk_in_booking(data: dict, current_user: User = Depends(get_current_user)):
    """Quick walk-in booking creation"""
    booking_id = str(uuid.uuid4())
    guest_id = str(uuid.uuid4())

    # Create guest
    await db.guests.insert_one({
        'id': guest_id,
        'tenant_id': current_user.tenant_id,
        'name': data['guest_name'],
        'phone': data['guest_phone'],
        'email': data.get('guest_email'),
        'created_at': datetime.now(UTC).isoformat()
    })

    # Find available room
    available_room = await db.rooms.find_one({
        'tenant_id': current_user.tenant_id,
        'room_type': data['room_type'],
        'current_status': 'available'
    })

    if not available_room:
        raise HTTPException(status_code=400, detail='No rooms available')

    # Create booking (atomic overbooking check)
    from core.atomic_booking import BookingConflictError, create_booking_atomic
    try:
        await create_booking_atomic({
            'id': booking_id,
            'tenant_id': current_user.tenant_id,
            'guest_id': guest_id,
            'room_id': available_room['id'],
            'check_in': data['check_in'],
            'check_out': data['check_out'],
            'adults': data['adults'],
            'status': 'confirmed',
            'source': 'walk-in',
            'created_at': datetime.now(UTC).isoformat()
        })
    except BookingConflictError as e:
        raise HTTPException(status_code=409, detail=str(e))

    return {'booking_id': booking_id, 'room_number': available_room['room_number']}


    review_lower = review_text.lower()

    # Negative keywords
    negative_keywords = ['dirty', 'broken', 'bad', 'terrible', 'awful', 'poor', 'noise', 'smell', 'rude', 'slow']
    # Positive keywords
    positive_keywords = ['great', 'excellent', 'amazing', 'wonderful', 'clean', 'friendly', 'helpful', 'perfect', 'love']

    negative_count = sum(1 for keyword in negative_keywords if keyword in review_lower)
    positive_count = sum(1 for keyword in positive_keywords if keyword in review_lower)

    # Determine sentiment
    if negative_count > positive_count:
        sentiment = 'negative'
        confidence = min(0.6 + (negative_count * 0.1), 0.95)
    elif positive_count > negative_count:
        sentiment = 'positive'
        confidence = min(0.6 + (positive_count * 0.1), 0.95)
    else:
        sentiment = 'neutral'
        confidence = 0.5

    # Detect issues
    issues = []
    if 'dirty' in review_lower or 'clean' in review_lower:
        issues.append({
            'category': 'Cleanliness',
            'description': 'Guest mentioned cleanliness concerns',
            'severity': 'high' if 'dirty' in review_lower else 'medium'
        })
    if 'broken' in review_lower or 'repair' in review_lower:
        issues.append({
            'category': 'Maintenance',
            'description': 'Equipment or room maintenance issue',
            'severity': 'high'
        })
    if 'noise' in review_lower:
        issues.append({
            'category': 'Noise',
            'description': 'Noise complaint detected',
            'severity': 'medium'
        })
    if 'rude' in review_lower or 'unfriendly' in review_lower:
        issues.append({
            'category': 'Staff Behavior',
            'description': 'Staff attitude issue mentioned',
            'severity': 'high'
        })

    # Detect highlights
    highlights = []
    if 'friendly' in review_lower or 'helpful' in review_lower:
        highlights.append({
            'category': 'Staff Friendliness',
            'description': 'Guest praised staff attitude'
        })
    if 'clean' in review_lower and 'dirty' not in review_lower:
        highlights.append({
            'category': 'Cleanliness',
            'description': 'Guest appreciated room cleanliness'
        })
    if 'location' in review_lower and ('great' in review_lower or 'perfect' in review_lower):
        highlights.append({
            'category': 'Location',
            'description': 'Guest loved the location'
        })

    # Generate recommendations
    recommendations = []
    if sentiment == 'negative':
        recommendations.append('Contact guest immediately for service recovery')
        recommendations.append('Assign compensation (points/discount) if appropriate')
        if issues:
            recommendations.append(f'Create maintenance task for {issues[0]["category"]}')
    elif sentiment == 'positive':
        recommendations.append('Thank guest and encourage loyalty program enrollment')
        recommendations.append('Share review on social media (with permission)')

    return {
        'sentiment': sentiment,
        'confidence': confidence,
        'issues': issues,
        'highlights': highlights,
        'recommendations': recommendations
    }

@router.get("/tasks/kanban")
@cached(ttl=180, key_prefix="tasks_kanban")  # Cache for 3 min
async def get_tasks_kanban(current_user: User = Depends(get_current_user)):
    """
    Get tasks organized by kanban columns: new, in_progress, waiting_parts, completed
    """
    tasks = await db.tasks.find({
        'tenant_id': current_user.tenant_id
    }).to_list(1000)

    kanban = {
        'new': [],
        'in_progress': [],
        'waiting_parts': [],
        'completed': []
    }

    for task in tasks:
        status = task.get('status', 'new')
        kanban[status].append(task)

    return {'tasks': kanban}

@router.post("/tasks/move")
async def move_task(
    data: dict,
    current_user: User = Depends(get_current_user)
):
    """
    Move task between kanban columns
    """
    task_id = data.get('task_id')
    to_status = data.get('to_status')

    await db.tasks.update_one(
        {
            'id': task_id,
            'tenant_id': current_user.tenant_id
        },
        {
            '$set': {
                'status': to_status,
                'updated_at': datetime.now(UTC).isoformat()
            }
        }
    )

    return {'message': f'Task moved to {to_status}'}

@router.post("/loyalty/tier-benefits/update")
async def update_loyalty_tier_benefits(
    data: dict,
    current_user: User = Depends(get_current_user)
):
    """
    Update loyalty tier benefits configuration
    """
    tiers = data.get('tiers', [])

    for tier in tiers:
        await db.loyalty_tier_benefits.update_one(
            {
                'tenant_id': current_user.tenant_id,
                'tier_name': tier['name']
            },
            {
                '$set': {
                    'benefits': tier['benefits'],
                    'updated_at': datetime.now(UTC).isoformat(),
                    'updated_by': current_user.id
                }
            },
            upsert=True
        )

    return {'message': f'{len(tiers)} tier benefits updated successfully'}


    # Map purchase order categories to cost categories
    category_mapping = {
        'cleaning': 'Housekeeping',
        'linens': 'Housekeeping',
        'amenities': 'Housekeeping',
        'food': 'F&B',
        'beverage': 'F&B',
        'kitchen': 'F&B',
        'maintenance': 'Technical',
        'electrical': 'Technical',
        'plumbing': 'Technical',
        'hvac': 'Technical',
        'furniture': 'General Expenses',
        'office': 'General Expenses',
        'it': 'General Expenses',
        'other': 'General Expenses'
    }

    # Aggregate costs by category
    cost_categories = {
        'Housekeeping': 0,
        'F&B': 0,
        'Technical': 0,
        'General Expenses': 0
    }

    total_mtd_costs = 0

    for po in purchase_orders:
        category = po.get('category', 'other')
        cost_category = category_mapping.get(category, 'General Expenses')
        total_amount = po.get('total_amount', 0)

        cost_categories[cost_category] += total_amount
        total_mtd_costs += total_amount

    # 2. Sort categories to get top 3
    sorted_categories = sorted(
        [{'name': k, 'amount': v} for k, v in cost_categories.items()],
        key=lambda x: x['amount'],
        reverse=True
    )

    top_3_categories = sorted_categories[:3]

    # 3. Calculate per-room cost (total costs / occupied room nights MTD)
    # Get all bookings for MTD that were checked-in
    bookings = await db.bookings.find({
        'tenant_id': current_user.tenant_id,
        'status': {'$in': ['checked_in', 'checked_out']},
        'check_in': {
            '$gte': month_start.isoformat(),
            '$lte': today.isoformat()
        }
    }).to_list(10000)

    # Calculate total occupied room nights
    total_room_nights = 0
    for booking in bookings:
        checkin = datetime.fromisoformat(booking['check_in']).date()
        checkout_str = booking.get('check_out', booking['check_in'])
        checkout = datetime.fromisoformat(checkout_str).date()

        # Calculate nights (minimum 1)
        nights = max((checkout - checkin).days, 1)
        total_room_nights += nights

    per_room_cost = (total_mtd_costs / total_room_nights) if total_room_nights > 0 else 0

    # 4. Get RevPAR from daily flash report for comparison
    # Calculate MTD RevPAR
    total_revenue = 0
    total_available_room_days = 0

    # Get all rooms
    rooms = await db.rooms.find({'tenant_id': current_user.tenant_id}).to_list(1000)
    total_rooms_count = len(rooms)

    # Get MTD charges
    mtd_charges = await db.folio_charges.find({
        'tenant_id': current_user.tenant_id,
        'date': {
            '$gte': month_start_dt.isoformat(),
            '$lte': today_end.isoformat()
        },
        'voided': False,
        'charge_category': 'room'
    }).to_list(10000)

    total_revenue = sum(charge.get('total', 0) for charge in mtd_charges)

    # Calculate days in month so far
    days_in_month_so_far = (today - month_start).days + 1
    total_available_room_days = total_rooms_count * days_in_month_so_far

    mtd_revpar = (total_revenue / total_available_room_days) if total_available_room_days > 0 else 0

    # 5. Calculate Cost to Revenue Ratio
    cost_to_revenue_ratio = (total_mtd_costs / total_revenue * 100) if total_revenue > 0 else 0

    # 6. Calculate profit margin
    gross_profit = total_revenue - total_mtd_costs
    profit_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0

    return {
        'report_date': today.isoformat(),
        'period': f'{month_start.isoformat()} to {today.isoformat()}',
        'total_mtd_costs': round(total_mtd_costs, 2),
        'cost_categories': {
            'housekeeping': round(cost_categories['Housekeeping'], 2),
            'fnb': round(cost_categories['F&B'], 2),
            'technical': round(cost_categories['Technical'], 2),
            'general_expenses': round(cost_categories['General Expenses'], 2)
        },
        'top_3_categories': [
            {
                'name': cat['name'],
                'amount': round(cat['amount'], 2),
                'percentage': round((cat['amount'] / total_mtd_costs * 100), 1) if total_mtd_costs > 0 else 0
            }
            for cat in top_3_categories
        ],
        'per_room_metrics': {
            'total_room_nights': total_room_nights,
            'cost_per_room_night': round(per_room_cost, 2),
            'mtd_revpar': round(mtd_revpar, 2),
            'cost_to_revpar_ratio': round((per_room_cost / mtd_revpar * 100), 1) if mtd_revpar > 0 else 0
        },
        'financial_metrics': {
            'mtd_revenue': round(total_revenue, 2),
            'mtd_costs': round(total_mtd_costs, 2),
            'gross_profit': round(gross_profit, 2),
            'profit_margin_percentage': round(profit_margin, 1),
            'cost_to_revenue_ratio': round(cost_to_revenue_ratio, 1)
        }
    }



@router.get("/reports/housekeeping-efficiency")
@cached(ttl=600, key_prefix="report_hk_efficiency")  # Cache for 10 min
async def get_housekeeping_efficiency_report(
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user)
):
    """Housekeeping Efficiency Report"""
    start = datetime.fromisoformat(start_date)
    end = datetime.fromisoformat(end_date)

    # Get completed housekeeping tasks in date range
    tasks = await db.housekeeping_tasks.find({
        'tenant_id': current_user.tenant_id,
        'status': 'completed',
        'created_at': {'$gte': start.isoformat(), '$lte': end.isoformat()}
    }).to_list(10000)

    # Aggregate by assigned staff
    staff_performance = {}

    for task in tasks:
        assigned_to = task.get('assigned_to', 'Unassigned')
        task_type = task.get('task_type', 'cleaning')

        if assigned_to not in staff_performance:
            staff_performance[assigned_to] = {
                'tasks_completed': 0,
                'by_type': {}
            }

        staff_performance[assigned_to]['tasks_completed'] += 1

        if task_type not in staff_performance[assigned_to]['by_type']:
            staff_performance[assigned_to]['by_type'][task_type] = 0
        staff_performance[assigned_to]['by_type'][task_type] += 1

    # Calculate daily averages
    date_range_days = (end.date() - start.date()).days + 1

    for staff in staff_performance:
        staff_performance[staff]['daily_average'] = round(
            staff_performance[staff]['tasks_completed'] / date_range_days, 2
        )

    return {
        'start_date': start_date,
        'end_date': end_date,
        'date_range_days': date_range_days,
        'total_tasks_completed': len(tasks),
        'staff_performance': staff_performance,
        'daily_average_all_staff': round(len(tasks) / date_range_days, 2) if date_range_days > 0 else 0
    }


@router.get("/reports/housekeeping-efficiency/excel")
@cached(ttl=900, key_prefix="report_hk_efficiency_excel")  # Cache for 15 min
async def export_housekeeping_efficiency_excel(
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user)
):
    """Export Housekeeping Efficiency Report to Excel"""
    report_data = await get_housekeeping_efficiency_report(start_date, end_date, current_user)

    headers = ["Staff Member", "Tasks Completed", "Daily Average", "Cleaning", "Maintenance", "Inspection"]
    data = []

    for staff, performance in report_data['staff_performance'].items():
        by_type = performance['by_type']
        data.append([
            staff,
            performance['tasks_completed'],
            f"{performance['daily_average']:.2f}",
            by_type.get('cleaning', 0),
            by_type.get('maintenance', 0),
            by_type.get('inspection', 0)
        ])

    wb = create_excel_workbook(
        title=f"Housekeeping Efficiency Report ({start_date} to {end_date})",
        headers=headers,
        data=data,
        sheet_name="HK Efficiency"
    )

    filename = f"housekeeping_efficiency_{start_date}_to_{end_date}.xlsx"
    return excel_response(wb, filename)



"""Auto-split from finance.py — section: folio."""
import logging

logger = logging.getLogger(__name__)
import asyncio
from datetime import UTC, datetime, timedelta
from typing import Any

from fastapi import APIRouter, Body, Depends, HTTPException, Request
from fastapi.security import HTTPBearer

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: F401
    from openpyxl.utils import get_column_letter  # noqa: F401
except ImportError:
    Workbook = None

from core.database import db
from core.helpers import create_audit_log
from core.pagination import PaginationParams, paginate
from core.security import get_current_user
from core.utils import calculate_folio_balance, excel_response
from models.enums import ChargeCategory, FolioOperationType
from models.schemas import (
    ChargeCreate,
    Folio,
    FolioCharge,
    FolioCreate,
    FolioOperation,
    FolioOperationCreate,
    Payment,
    PaymentCreate,
    User,
)
from modules.folio.services.folio_balance_read_service import FolioBalanceReadService
from modules.folio.services.open_folio_service import OpenFolioService
from modules.pms_core.role_permission_service import require_op
from shared_kernel.shadow_metrics import compare_folio_payloads, run_shadow_compare

try:
    from cache_manager import cache, cached
except ImportError:
    cache = None  # type: ignore
    def cached(ttl=300, key_prefix=""):
        def decorator(func):
            return func
        return decorator

router = APIRouter()
security = HTTPBearer()
folio_balance_read_service = FolioBalanceReadService()
open_folio_service = OpenFolioService()

@router.post("/folio/create", response_model=Folio)
async def create_folio(
    folio_data: FolioCreate,
    request: Request,
    current_user: User = Depends(get_current_user),
):
    """Create a new folio for a booking"""
    from modules.pms_core.role_permission_service import RolePermissionService  # Bug CQ-R2
    RolePermissionService().enforce_permission(current_user.role, "post_charge")
    return await open_folio_service.create(folio_data, current_user, request)




@router.get("/folio/list")
async def list_folios(
    status: str | None = None,
    p: PaginationParams = Depends(paginate(default_limit=50, max_limit=500)),
    current_user: User = Depends(get_current_user)
):
    """List all folios for the current tenant with optional status filter."""
    limit, offset = p.limit, p.offset
    query = {'tenant_id': current_user.tenant_id}
    if status:
        query['status'] = status

    # Sprint 33: parallelize list + count, then BATCH-fetch bookings via $in
    # (was N+1 — one find_one per folio causing 6.6s on 50 rows).
    import asyncio as _asyncio
    folios_q = db.folios.find(query, {'_id': 0}).sort(
        'created_at', -1
    ).skip(offset).limit(limit).to_list(limit)
    total_q = db.folios.count_documents(query)
    folios, total = await _asyncio.gather(folios_q, total_q)

    booking_ids = [f.get('booking_id') for f in folios if f.get('booking_id')]
    booking_map = {}
    if booking_ids:
        bookings = await db.bookings.find(
            {'id': {'$in': booking_ids}, 'tenant_id': current_user.tenant_id},
            {'_id': 0, 'id': 1, 'guest_name': 1, 'room_number': 1,
             'room_id': 1, 'check_in': 1, 'check_out': 1}
        ).to_list(len(booking_ids))
        booking_map = {b['id']: b for b in bookings}

    for folio in folios:
        booking = booking_map.get(folio.get('booking_id'))
        if booking:
            folio['guest_name'] = booking.get('guest_name', '')
            folio['room_number'] = booking.get('room_number', '')
            folio['check_in'] = booking.get('check_in', '')
            folio['check_out'] = booking.get('check_out', '')

    return {
        'folios': folios,
        'total': total,
        'limit': limit,
        'offset': offset,
    }


@router.get("/folio/dashboard-stats")
@cached(ttl=300, key_prefix="folio_dashboard_stats")  # Cache for 5 minutes
async def get_folio_dashboard_stats(
    current_user=Depends(get_current_user),  # v68 Bug DE: tenant-scoped cache key
    _perm=Depends(require_op("view_finance_reports")),  # v70 Bug DG: HK/FO finance leak
):
    """Get folio statistics for dashboard"""

    try:
        # v95 — Parallel queries + server-side $sum (was to_list(1000) + Python sum, sequential)
        tid = current_user.tenant_id
        yesterday = (datetime.now(UTC) - timedelta(days=1)).isoformat()

        open_folios_pipeline = [
            {'$match': {'tenant_id': tid, 'status': 'open'}},
            {'$group': {
                '_id': None,
                'count': {'$sum': 1},
                'total_balance': {'$sum': {'$ifNull': ['$balance', 0]}},
            }},
        ]
        open_folios_q = db.folios.aggregate(open_folios_pipeline).to_list(1)
        charges_q = db.folio_charges.count_documents({
            'tenant_id': tid, 'date': {'$gte': yesterday}, 'voided': False
        })
        payments_q = db.payments.count_documents({
            'tenant_id': tid, 'date': {'$gte': yesterday}
        })
        open_agg, recent_charges, recent_payments = await asyncio.gather(
            open_folios_q, charges_q, payments_q
        )

        total_open = open_agg[0]['count'] if open_agg else 0
        total_outstanding = open_agg[0]['total_balance'] if open_agg else 0.0

        return {
            'total_open_folios': total_open,
            'total_outstanding_balance': round(total_outstanding, 2),
            'recent_charges_24h': recent_charges,
            'recent_payments_24h': recent_payments
        }
    except Exception as e:
        logger.info(f"Error in folio dashboard stats: {str(e)}")
        import traceback
        traceback.print_exc()
        # Return default values instead of raising
        return {
            'total_open_folios': 0,
            'total_outstanding_balance': 0.0,
            'recent_charges_24h': 0,
            'recent_payments_24h': 0
        }


@router.get("/folio/pending-ar")
@cached(ttl=600, key_prefix="folio_pending_ar")  # Cache for 10 min
async def get_pending_ar(
    current_user=Depends(get_current_user),  # v68 Bug DE: tenant-scoped cache key
    _perm=Depends(require_op("view_finance_reports")),  # v70 Bug DG: AR leak HK/FO
):
    """Get pending accounts receivable (company folios with outstanding balances)"""

    try:
        # v95 — Batch lookup: was N+1 (1 query per company × ~100 companies = 4.5s).
        # Now: 1 companies query + 1 folios query, then in-memory grouping.
        tenant_id = current_user.tenant_id
        companies = await db.companies.find(
            {'tenant_id': tenant_id}, {'_id': 0}
        ).to_list(1000)
        if not companies:
            return []

        company_ids = [c['id'] for c in companies]
        company_map = {c['id']: c for c in companies}

        # Single query for all open folios with balance > 0
        all_folios = await db.folios.find(
            {
                'tenant_id': tenant_id,
                'company_id': {'$in': company_ids},
                'status': 'open',
                'balance': {'$gt': 0},
            },
            {'_id': 0, 'company_id': 1, 'balance': 1, 'created_at': 1},
        ).to_list(10000)

        if not all_folios:
            return []

        # Group folios by company
        folios_by_company: dict[str, list] = {}
        for f in all_folios:
            folios_by_company.setdefault(f['company_id'], []).append(f)

        now = datetime.now(UTC)
        ar_data = []

        for cid, folios in folios_by_company.items():
            company = company_map.get(cid)
            if not company:
                continue

            total_outstanding = sum(f.get('balance', 0) for f in folios)
            if total_outstanding <= 0:
                continue

            # Aging calculation
            aging = {'0-7': 0, '8-14': 0, '15-30': 0, '30+': 0}
            oldest_iso = None
            oldest_days = 0
            for folio in folios:
                created_at = folio.get('created_at') or now.isoformat()
                try:
                    folio_dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                except (ValueError, AttributeError):
                    folio_dt = now
                days = (now - folio_dt).days
                balance = folio.get('balance', 0)
                if days <= 7:
                    aging['0-7'] += balance
                elif days <= 14:
                    aging['8-14'] += balance
                elif days <= 30:
                    aging['15-30'] += balance
                else:
                    aging['30+'] += balance
                if oldest_iso is None or created_at < oldest_iso:
                    oldest_iso = created_at
                    oldest_days = days

            ar_data.append({
                'company_id': cid,
                'company_name': company.get('name', 'Unknown'),
                'corporate_code': company.get('corporate_code', ''),
                'contact_person': company.get('contact_person', ''),
                'contact_email': company.get('contact_email', ''),
                'contact_phone': company.get('contact_phone', ''),
                'payment_terms': company.get('payment_terms', 'Net 30'),
                'total_outstanding': round(total_outstanding, 2),
                'open_folios_count': len(folios),
                'oldest_invoice_date': oldest_iso,
                'days_outstanding': oldest_days,
                'aging': aging,
            })

        ar_data.sort(key=lambda x: x['days_outstanding'], reverse=True)
        return ar_data

    except Exception as e:
        logger.info(f"Error in get_pending_ar: {str(e)}")
        import traceback
        traceback.print_exc()
        return []


@router.get("/folio/booking/{booking_id}", response_model=list[Folio])
@cached(ttl=180, key_prefix="folio_by_booking")  # Cache for 3 min
async def get_booking_folios(
    booking_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_finance_reports")),  # v70 Bug DG
):
    """Get all folios for a booking"""
    folios = await db.folios.find({
        'booking_id': booking_id,
        'tenant_id': current_user.tenant_id
    }, {'_id': 0}).to_list(1000)

    # Calculate current balance + backfill legacy folios that pre-date the
    # folio_number / folio_type schema fields (Pydantic response model would
    # otherwise raise 500 ResponseValidationError on rows missing these keys).
    for folio in folios:
        folio['balance'] = await calculate_folio_balance(folio['id'], current_user.tenant_id)
        if not folio.get('folio_number'):
            folio['folio_number'] = f"F-{(folio.get('id') or '')[:8] or 'LEGACY'}"
        if not folio.get('folio_type'):
            folio['folio_type'] = 'guest'

    return folios


@router.get("/folio/{folio_id}", response_model=dict[str, Any])
@cached(ttl=180, key_prefix="folio_details")  # Cache for 3 min
async def get_folio_details(
    folio_id: str,
    request: Request,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_finance_reports")),  # v70 Bug DG
):
    """Get folio with charges and payments"""
    semantic_response = await folio_balance_read_service.get_folio_details(
        tenant_id=current_user.tenant_id,
        folio_id=folio_id,
    )
    asyncio.create_task(
        run_shadow_compare(
            endpoint="folio",
            tenant_id=current_user.tenant_id,
            property_id=request.headers.get("x-property-id"),
            correlation_id=request.headers.get("x-correlation-id"),
            semantic_payload=semantic_response,
            legacy_loader=lambda: _legacy_get_folio_details(
                tenant_id=current_user.tenant_id,
                folio_id=folio_id,
            ),
            comparator=compare_folio_payloads,
            entity_id=folio_id,
        )
    )
    return semantic_response


async def _legacy_get_folio_details(tenant_id: str, folio_id: str):
    folio = await db.folios.find_one({
        'id': folio_id,
        'tenant_id': tenant_id
    }, {'_id': 0})

    if not folio:
        raise HTTPException(status_code=404, detail="Folio not found")

    charges = await db.folio_charges.find({
        'folio_id': folio_id,
        'tenant_id': tenant_id
    }, {'_id': 0}).to_list(1000)

    payments = await db.payments.find({
        'folio_id': folio_id,
        'tenant_id': tenant_id
    }, {'_id': 0}).to_list(1000)

    balance = await calculate_folio_balance(folio_id, tenant_id)
    folio['balance'] = balance

    return {
        'folio': folio,
        'charges': charges,
        'payments': payments,
        'balance': balance
    }



@router.get("/folio/{folio_id}/excel")
@cached(ttl=600, key_prefix="folio_excel")  # Cache for 10 min
async def export_folio_excel(
    folio_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_finance_reports")),  # v70 Bug DG
):
    """Export Folio to Excel"""
    folio_data = await _legacy_get_folio_details(current_user.tenant_id, folio_id)

    folio = folio_data['folio']
    charges = folio_data['charges']
    payments = folio_data['payments']
    balance = folio_data['balance']

    wb = Workbook()
    ws = wb.active
    ws.title = "Folio"

    # Folio header
    ws['A1'] = "GUEST FOLIO"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:E1')

    ws['A3'] = "Folio Number:"
    ws['B3'] = folio.get('folio_number', 'N/A')
    ws['A4'] = "Type:"
    ws['B4'] = folio.get('folio_type', 'guest').title()
    ws['A5'] = "Status:"
    ws['B5'] = folio.get('status', 'open').upper()
    ws['A6'] = "Created:"
    ws['B6'] = folio.get('created_at', '')[:10]

    # Charges section
    ws['A9'] = "CHARGES"
    ws['A9'].font = Font(size=14, bold=True)

    charge_headers = ["Date", "Description", "Qty", "Subtotal", "Discount", "Net", "VAT %", "VAT", "City Tax", "Total"]
    for col_num, header in enumerate(charge_headers, 1):
        cell = ws.cell(row=10, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Bug AN: charge.description is user-controlled; openpyxl would parse a
    # leading '=' as a formula. xlsx_safe() prepends apostrophe to neutralize.
    from core.csv_safe import xlsx_safe
    row = 11
    total_charges = 0
    for charge in charges:
        if not charge.get('voided', False):
            net = float(charge.get('amount', 0) or 0)
            disc = float(charge.get('discount_amount', 0) or 0)
            sub = float(charge.get('subtotal') or (net + disc))  # geriye uyumlu: eski kayıtlarda subtotal=amount+discount=amount
            ws.cell(row=row, column=1, value=(charge.get('posted_at') or charge.get('date') or '')[:10])
            ws.cell(row=row, column=2, value=xlsx_safe(charge.get('description', '')))
            ws.cell(row=row, column=3, value=charge.get('quantity', 1))
            ws.cell(row=row, column=4, value=round(sub, 2))
            ws.cell(row=row, column=5, value=round(disc, 2))
            ws.cell(row=row, column=6, value=round(net, 2))
            ws.cell(row=row, column=7, value=round(float(charge.get('vat_rate', 0) or 0), 2))
            ws.cell(row=row, column=8, value=round(float(charge.get('vat_amount', 0) or 0), 2))
            ws.cell(row=row, column=9, value=round(float(charge.get('tax_amount', 0) or 0), 2))
            ws.cell(row=row, column=10, value=round(float(charge.get('total', 0) or 0), 2))
            total_charges += float(charge.get('total', 0) or 0)
            row += 1

    ws.cell(row=row, column=9, value="Total Charges:")
    ws.cell(row=row, column=9).font = Font(bold=True)
    ws.cell(row=row, column=10, value=round(total_charges, 2))
    ws.cell(row=row, column=10).font = Font(bold=True)

    # Payments section
    row += 2
    ws.cell(row=row, column=1, value="PAYMENTS")
    ws.cell(row=row, column=1).font = Font(size=14, bold=True)
    row += 1

    payment_headers = ["Date", "Method", "Type", "Amount"]
    for col_num, header in enumerate(payment_headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    row += 1
    total_payments = 0
    for payment in payments:
        ws.cell(row=row, column=1, value=payment.get('processed_at', '')[:10])
        ws.cell(row=row, column=2, value=payment.get('payment_method', '').title())
        ws.cell(row=row, column=3, value=payment.get('payment_type', '').title())
        ws.cell(row=row, column=4, value=f"${payment.get('amount', 0):,.2f}")
        total_payments += payment.get('amount', 0)
        row += 1

    ws.cell(row=row, column=3, value="Total Payments:")
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=4, value=f"${total_payments:,.2f}")
    ws.cell(row=row, column=4).font = Font(bold=True)

    # Balance
    row += 2
    ws.cell(row=row, column=5, value="BALANCE DUE:")
    ws.cell(row=row, column=5).font = Font(size=14, bold=True)
    ws.cell(row=row, column=6, value=f"${balance:,.2f}")
    ws.cell(row=row, column=6).font = Font(size=14, bold=True)
    ws.cell(row=row, column=6).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    filename = f"folio_{folio.get('folio_number', folio_id)}.xlsx"
    return excel_response(wb, filename)



@router.post("/folio/{folio_id}/charge", response_model=FolioCharge)
async def post_charge_to_folio(
    folio_id: str,
    charge_data: ChargeCreate,
    current_user: User = Depends(get_current_user)
):
    """Post a charge to folio"""
    # Role / permission enforcement (Bug CP fix)
    from modules.pms_core.role_permission_service import RolePermissionService
    RolePermissionService().enforce_permission(current_user.role, "post_charge")

    folio = await db.folios.find_one({
        'id': folio_id,
        'tenant_id': current_user.tenant_id,
        'status': 'open'
    })

    if not folio:
        raise HTTPException(status_code=404, detail="Folio not found or closed")

    # Calculate amounts with proper rounding
    subtotal = round(charge_data.amount * charge_data.quantity, 2)
    discount = round(max(0.0, float(charge_data.discount_amount or 0.0)), 2)
    if discount > subtotal:
        raise HTTPException(status_code=400, detail="İndirim ara toplamı aşamaz")
    net = round(subtotal - discount, 2)
    vat_rate = max(0.0, min(100.0, float(charge_data.vat_rate or 0.0)))
    vat_amount = round(net * vat_rate / 100.0, 2)
    tax_amount = 0.0

    # Auto-calculate city tax if requested
    if charge_data.auto_calculate_tax and charge_data.charge_category == ChargeCategory.ROOM:
        # Get city tax rule
        tax_rule = await db.city_tax_rules.find_one({
            'tenant_id': current_user.tenant_id,
            'active': True
        })
        if tax_rule:
            if tax_rule.get('flat_amount'):
                tax_amount = round(tax_rule['flat_amount'], 2)
            else:
                # v95.7: rate_percent yeni canonical alan; tax_percentage legacy fallback.
                rate_pct = tax_rule.get('rate_percent', tax_rule.get('tax_percentage', 0))
                tax_amount = round(net * (float(rate_pct) / 100), 2)

    total = round(net + vat_amount + tax_amount, 2)

    charge = FolioCharge(
        tenant_id=current_user.tenant_id,
        folio_id=folio_id,
        booking_id=folio['booking_id'],
        charge_category=charge_data.charge_category,
        description=charge_data.description,
        unit_price=charge_data.amount,
        quantity=charge_data.quantity,
        amount=net,
        subtotal=subtotal,
        discount_amount=discount,
        discount_reason=(charge_data.discount_reason or None) if discount > 0 else None,
        vat_rate=vat_rate,
        vat_amount=vat_amount,
        tax_amount=tax_amount,
        total=total,
        posted_by=current_user.id
    )

    charge_dict = charge.model_dump()
    charge_dict['date'] = charge_dict['date'].isoformat()
    await db.folio_charges.insert_one(charge_dict)

    # Update folio balance
    balance = await calculate_folio_balance(folio_id, current_user.tenant_id)
    await db.folios.update_one(
        {'id': folio_id},
        {'$set': {'balance': balance}}
    )

    # Audit log
    await create_audit_log(
        tenant_id=current_user.tenant_id,
        user=current_user,
        action="POST_CHARGE",
        entity_type="folio_charge",
        entity_id=charge.id,
        changes={
            'charge_category': charge_data.charge_category,
            'subtotal': subtotal,
            'discount': discount,
            'vat_rate': vat_rate,
            'vat_amount': vat_amount,
            'city_tax': tax_amount,
            'total': total,
            'folio_id': folio_id,
        }
    )

    # v95.1 — revenue raporu cache'ini geçersiz kıl (yeni charge eklenince)
    if cache:
        cache.invalidate_tenant_cache(current_user.tenant_id, "folio_revenue_by_category")

    # Acente webhook: rezervasyon güncellendi (yeni charge → toplam değişti)
    from routers.webhook_retry_service import schedule_emit_reservation_updated
    schedule_emit_reservation_updated(
        current_user.tenant_id, folio['booking_id'], "charge_added",
        {"charge_id": charge.id, "amount": total, "category": str(charge_data.charge_category)},
    )

    return charge


@router.post("/folio/{folio_id}/payment", response_model=Payment)
async def post_payment_to_folio(
    folio_id: str,
    payment_data: PaymentCreate,
    current_user: User = Depends(get_current_user)
):
    """Post a payment to folio"""
    # Role / permission enforcement (Bug CP fix)
    from modules.pms_core.role_permission_service import RolePermissionService
    RolePermissionService().enforce_permission(current_user.role, "post_payment")

    folio = await db.folios.find_one({
        'id': folio_id,
        'tenant_id': current_user.tenant_id
    })

    if not folio:
        raise HTTPException(status_code=404, detail="Folio not found")

    # Vardiya kontrolü: nakit ödemede aktif vardiya zorunlu (5y kasa standardı)
    from domains.pms.cashier_service import ensure_active_shift, record_cash_transaction
    method_str = payment_data.method.value if hasattr(payment_data.method, 'value') else str(payment_data.method)
    await ensure_active_shift(current_user.tenant_id, method_str)

    payment = Payment(
        tenant_id=current_user.tenant_id,
        folio_id=folio_id,
        booking_id=folio['booking_id'],
        processed_by=current_user.id,
        **payment_data.model_dump()
    )

    payment_dict = payment.model_dump()
    payment_dict['processed_at'] = payment_dict['processed_at'].isoformat()
    payment_dict['processed_by_name'] = current_user.name  # Add user name
    await db.payments.insert_one(payment_dict)

    # Update folio balance
    balance = await calculate_folio_balance(folio_id, current_user.tenant_id)
    await db.folios.update_one(
        {'id': folio_id},
        {'$set': {'balance': balance}}
    )

    # Kasa hareketine yaz (idempotent — aynı payment.id için tek kayıt)
    # Cash için vardiya zorunlu: TOCTOU race'inde 409 → ödemeyi geri al
    is_cash = method_str.lower() == "cash"
    try:
        await record_cash_transaction(
            tenant_id=current_user.tenant_id,
            amount=payment.amount,
            method=method_str,
            direction="in",
            description=f"Folio ödemesi - {folio.get('folio_number') or folio_id[:8]}",
            txn_type="folio_payment",
            ref_type="payment",
            ref_id=payment.id,
            created_by=current_user.email,
            created_by_name=getattr(current_user, 'name', None) or current_user.email,
            idempotency_key=f"payment:{payment.id}",
            require_open_shift=is_cash,
        )
    except HTTPException as he:
        if is_cash and he.status_code == 409:
            # vardiya kapanmış → ödemeyi rollback et
            try:
                await db.payments.delete_one({'id': payment.id, 'tenant_id': current_user.tenant_id})
                new_balance = await calculate_folio_balance(folio_id, current_user.tenant_id)
                await db.folios.update_one({'id': folio_id}, {'$set': {'balance': new_balance}})
            except Exception:
                import logging as _lg
                _lg.getLogger(__name__).exception("payment rollback failed after cashier 409")
        raise
    except Exception:
        # Kasa kayıt arızası (kart/banka) ödemeyi düşürmesin (loglanır)
        import logging as _lg
        _lg.getLogger(__name__).exception("cashier txn record failed")

    # Acente webhook: rezervasyon güncellendi (ödeme alındı → bakiye değişti)
    from routers.webhook_retry_service import schedule_emit_reservation_updated
    schedule_emit_reservation_updated(
        current_user.tenant_id, folio['booking_id'], "payment_added",
        {"payment_id": payment.id, "amount": float(payment.amount), "method": method_str},
    )

    return payment


@router.get("/folio/reports/revenue-by-category")
@cached(ttl=300, key_prefix="folio_revenue_by_category")  # 5 dk cache; tarih+tenant key
async def revenue_by_category(
    date_from: str | None = None,
    date_to: str | None = None,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_finance_reports")),
):
    """Tarih aralığında kategori bazlı gelir raporu (void hariç).
    Varsayılan: son 30 gün.
    """
    if not date_to:
        date_to = datetime.now(UTC).date().isoformat()
    if not date_from:
        date_from = (datetime.now(UTC).date() - timedelta(days=30)).isoformat()
    try:
        dt_from = datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=UTC)
        dt_to = datetime.strptime(date_to, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59, microsecond=999999, tzinfo=UTC
        )
    except ValueError:
        raise HTTPException(status_code=400, detail="Tarih formatı YYYY-MM-DD olmalı")
    if dt_from > dt_to:
        raise HTTPException(status_code=400, detail="date_from > date_to olamaz")

    # charge.date hem ISO string hem BSON datetime olarak depolanmış olabilir.
    # ISO 8601 string'ler lexicographic olarak sıralanabilir; type-mismatch'ten
    # kaçınmak için $expr + $cond ile her iki tipi de tek pipeline'da karşıla.
    pipeline = [
        {'$match': {
            'tenant_id': current_user.tenant_id,
            'voided': {'$ne': True},
            '$expr': {
                '$let': {
                    'vars': {
                        'd': {
                            '$cond': [
                                {'$eq': [{'$type': '$date'}, 'string']},
                                {'$dateFromString': {'dateString': '$date', 'onError': None, 'onNull': None}},
                                '$date',
                            ]
                        }
                    },
                    'in': {'$and': [
                        {'$ne': ['$$d', None]},
                        {'$gte': ['$$d', dt_from]},
                        {'$lte': ['$$d', dt_to]},
                    ]}
                }
            }
        }},
        {'$group': {
            '_id': '$charge_category',
            'count': {'$sum': 1},
            'subtotal': {'$sum': {'$ifNull': ['$subtotal', '$amount']}},
            'discount': {'$sum': {'$ifNull': ['$discount_amount', 0]}},
            'net': {'$sum': '$amount'},
            'vat': {'$sum': {'$ifNull': ['$vat_amount', 0]}},
            'city_tax': {'$sum': {'$ifNull': ['$tax_amount', 0]}},
            'total': {'$sum': {'$ifNull': ['$total', '$amount']}},
        }},
    ]
    rows_raw = await db.folio_charges.aggregate(pipeline).to_list(100)

    rows = []
    totals = {'count': 0, 'subtotal': 0.0, 'discount': 0.0, 'net': 0.0, 'vat': 0.0, 'city_tax': 0.0, 'total': 0.0}
    for r in rows_raw:
        item = {
            'category': r['_id'] or 'other',
            'count': int(r.get('count') or 0),
            'subtotal': round(float(r.get('subtotal') or 0.0), 2),
            'discount': round(float(r.get('discount') or 0.0), 2),
            'net': round(float(r.get('net') or 0.0), 2),
            'vat': round(float(r.get('vat') or 0.0), 2),
            'city_tax': round(float(r.get('city_tax') or 0.0), 2),
            'total': round(float(r.get('total') or 0.0), 2),
        }
        rows.append(item)
        totals['count'] += item['count']
        for k in ('subtotal', 'discount', 'net', 'vat', 'city_tax', 'total'):
            totals[k] = round(totals[k] + item[k], 2)

    rows.sort(key=lambda x: x['total'], reverse=True)
    return {
        'date_from': date_from,
        'date_to': date_to,
        'rows': rows,
        'totals': totals,
    }


@router.post("/folio/transfer", response_model=FolioOperation)
async def transfer_charges(
    operation_data: FolioOperationCreate,
    current_user: User = Depends(get_current_user)
):
    """Transfer charges from one folio to another"""
    from modules.pms_core.role_permission_service import RolePermissionService  # Bug CQ-R2
    RolePermissionService().enforce_permission(current_user.role, "transfer_folio")
    if operation_data.operation_type != FolioOperationType.TRANSFER:
        raise HTTPException(status_code=400, detail="Invalid operation type")

    if not operation_data.to_folio_id:
        raise HTTPException(status_code=400, detail="Destination folio required for transfer")

    # Verify both folios exist
    from_folio = await db.folios.find_one({
        'id': operation_data.from_folio_id,
        'tenant_id': current_user.tenant_id
    })

    to_folio = await db.folios.find_one({
        'id': operation_data.to_folio_id,
        'tenant_id': current_user.tenant_id,
        'status': 'open'
    })

    if not from_folio or not to_folio:
        raise HTTPException(status_code=404, detail="Folio not found")

    if not operation_data.charge_ids:
        raise HTTPException(status_code=400, detail="En az bir charge_id gerekli")

    # Önce hedef charge'ların hepsi gerçekten kaynak folio'da ve void değil mi doğrula
    existing = await db.folio_charges.find(
        {
            'id': {'$in': operation_data.charge_ids},
            'folio_id': operation_data.from_folio_id,
            'tenant_id': current_user.tenant_id,
            'voided': {'$ne': True},
        },
        {'_id': 0, 'id': 1}
    ).to_list(len(operation_data.charge_ids))
    found_ids = {c['id'] for c in existing}
    missing = [cid for cid in operation_data.charge_ids if cid not in found_ids]
    if missing:
        raise HTTPException(
            status_code=409,
            detail=f"Aktarılamayan kayıtlar (kaynakta yok / iptal edilmiş / başka tenant): {missing}"
        )

    # Tek bulk update — modified_count tüm hedeflerle eşit olmalı
    res = await db.folio_charges.update_many(
        {
            'id': {'$in': operation_data.charge_ids},
            'folio_id': operation_data.from_folio_id,
            'tenant_id': current_user.tenant_id,
            'voided': {'$ne': True},
        },
        {'$set': {'folio_id': operation_data.to_folio_id}}
    )
    if getattr(res, 'modified_count', 0) != len(operation_data.charge_ids):
        raise HTTPException(
            status_code=500,
            detail=f"Aktarım kısmi: {res.modified_count}/{len(operation_data.charge_ids)} işlem güncellendi"
        )

    # Create operation record
    operation = FolioOperation(
        tenant_id=current_user.tenant_id,
        performed_by=current_user.id,
        **operation_data.model_dump()
    )

    operation_dict = operation.model_dump()
    operation_dict['performed_at'] = operation_dict['performed_at'].isoformat()
    await db.folio_operations.insert_one(operation_dict)

    # Update balances
    from_balance = await calculate_folio_balance(operation_data.from_folio_id, current_user.tenant_id)
    to_balance = await calculate_folio_balance(operation_data.to_folio_id, current_user.tenant_id)

    await db.folios.update_one(
        {'id': operation_data.from_folio_id},
        {'$set': {'balance': from_balance}}
    )
    await db.folios.update_one(
        {'id': operation_data.to_folio_id},
        {'$set': {'balance': to_balance}}
    )

    return operation


@router.post("/folio/{folio_id}/void-charge/{charge_id}")
async def void_charge(
    folio_id: str,
    charge_id: str,
    void_reason: str,
    current_user: User = Depends(get_current_user)
):
    """Void a charge"""
    from modules.pms_core.role_permission_service import RolePermissionService  # Bug CQ-R2
    RolePermissionService().enforce_permission(current_user.role, "void_charge")
    charge = await db.folio_charges.find_one({
        'id': charge_id,
        'folio_id': folio_id,
        'tenant_id': current_user.tenant_id,
        'voided': False
    })

    if not charge:
        raise HTTPException(status_code=404, detail="Charge not found or already voided")

    await db.folio_charges.update_one(
        {'id': charge_id},
        {'$set': {
            'voided': True,
            'void_reason': void_reason,
            'voided_by': current_user.id,
            'voided_at': datetime.now(UTC).isoformat()
        }}
    )

    # Update folio balance
    balance = await calculate_folio_balance(folio_id, current_user.tenant_id)
    await db.folios.update_one(
        {'id': folio_id},
        {'$set': {'balance': balance}}
    )

    # Create operation record
    operation = FolioOperation(
        tenant_id=current_user.tenant_id,
        operation_type=FolioOperationType.VOID,
        from_folio_id=folio_id,
        charge_ids=[charge_id],
        amount=charge['total'],
        reason=void_reason,
        performed_by=current_user.id
    )

    operation_dict = operation.model_dump()
    operation_dict['performed_at'] = operation_dict['performed_at'].isoformat()
    await db.folio_operations.insert_one(operation_dict)

    # v95.1 — revenue raporu cache'ini geçersiz kıl (charge void edilince)
    if cache:
        cache.invalidate_tenant_cache(current_user.tenant_id, "folio_revenue_by_category")

    # Acente webhook: rezervasyon güncellendi (charge iptal → toplam değişti)
    if charge.get('booking_id'):
        from routers.webhook_retry_service import schedule_emit_reservation_updated
        schedule_emit_reservation_updated(
            current_user.tenant_id, charge['booking_id'], "charge_voided",
            {"charge_id": charge_id, "amount": charge.get('total'), "reason": void_reason},
        )

    return {"message": "Charge voided successfully"}


@router.post("/folio/{folio_id}/payment/{payment_id}/void")
async def void_payment(
    folio_id: str,
    payment_id: str,
    body: dict = Body(...),
    current_user: User = Depends(get_current_user),
):
    """
    Ödemeyi iade eder (void). Kasaya negatif kayıt (refund) düşer.
    Body: { reason: str }
    """
    from modules.pms_core.role_permission_service import RolePermissionService
    RolePermissionService().enforce_permission(current_user.role, "post_payment")

    reason = (body or {}).get("reason", "").strip()
    if not reason:
        raise HTTPException(status_code=400, detail="İade nedeni zorunlu")

    payment = await db.payments.find_one({
        'id': payment_id,
        'folio_id': folio_id,
        'tenant_id': current_user.tenant_id,
    })
    if not payment:
        raise HTTPException(status_code=404, detail="Ödeme bulunamadı")
    if payment.get('voided'):
        raise HTTPException(status_code=400, detail="Ödeme zaten iade edilmiş")

    method_str = payment.get('method') or 'cash'
    if hasattr(method_str, 'value'):
        method_str = method_str.value
    method_str = str(method_str).lower()
    is_cash = method_str == "cash"

    # Vardiya kontrolü: nakit iadesi için aktif vardiya zorunlu
    from domains.pms.cashier_service import ensure_active_shift, record_cash_transaction
    await ensure_active_shift(current_user.tenant_id, method_str)

    # CAS: voided != True koşullu update — eşzamanlı iki istek ikinci kez geçemez
    now = datetime.now(UTC)
    cas_res = await db.payments.update_one(
        {
            'id': payment_id,
            'tenant_id': current_user.tenant_id,
            '$or': [{'voided': {'$exists': False}}, {'voided': False}],
        },
        {'$set': {
            'voided': True,
            'void_reason': reason,
            'voided_by': current_user.id,
            'voided_by_name': current_user.name,
            'voided_at': now.isoformat(),
        }}
    )
    if cas_res.modified_count == 0:
        raise HTTPException(status_code=409, detail="Ödeme zaten iade edilmiş veya değiştirildi")

    new_balance = await calculate_folio_balance(folio_id, current_user.tenant_id)
    await db.folios.update_one(
        {'id': folio_id},
        {'$set': {'balance': new_balance}}
    )

    async def _rollback_void():
        await db.payments.update_one(
            {'id': payment_id, 'tenant_id': current_user.tenant_id},
            {'$set': {'voided': False},
             '$unset': {'void_reason': '', 'voided_by': '', 'voided_by_name': '', 'voided_at': ''}}
        )
        restored = await calculate_folio_balance(folio_id, current_user.tenant_id)
        await db.folios.update_one({'id': folio_id}, {'$set': {'balance': restored}})

    # Kasaya iade kaydı (negatif/out) — başarısızsa void'i geri al (compensation)
    try:
        await record_cash_transaction(
            tenant_id=current_user.tenant_id,
            amount=float(payment.get('amount') or 0),
            method=method_str,
            direction="out",
            description=f"Ödeme iadesi - {reason}",
            txn_type="refund",
            ref_type="payment",
            ref_id=payment_id,
            created_by=current_user.email,
            created_by_name=getattr(current_user, 'name', None) or current_user.email,
            idempotency_key=f"void:{payment_id}",
            require_open_shift=is_cash,
        )
    except HTTPException:
        await _rollback_void()
        raise
    except Exception as e:
        import logging as _lg
        _lg.getLogger(__name__).exception("cashier refund record failed")
        await _rollback_void()
        raise HTTPException(status_code=500, detail=f"Kasa iade kaydı başarısız, iade geri alındı: {e}")

    # Acente webhook: rezervasyon güncellendi (ödeme iade → bakiye değişti)
    if payment.get('booking_id'):
        from routers.webhook_retry_service import schedule_emit_reservation_updated
        schedule_emit_reservation_updated(
            current_user.tenant_id, payment['booking_id'], "payment_voided",
            {"payment_id": payment_id, "amount": payment.get('amount'), "method": method_str, "reason": reason},
        )

    return {
        "message": "Ödeme iade edildi",
        "payment_id": payment_id,
        "amount": payment.get('amount'),
        "method": method_str,
        "new_balance": new_balance,
    }


@router.post("/folio/{folio_id}/close")
async def close_folio(
    folio_id: str,
    current_user: User = Depends(get_current_user)
):
    """Close a folio"""
    from modules.pms_core.role_permission_service import RolePermissionService  # Bug CQ-R2
    RolePermissionService().enforce_permission(current_user.role, "close_folio")
    folio = await db.folios.find_one({
        'id': folio_id,
        'tenant_id': current_user.tenant_id,
        'status': 'open'
    })

    if not folio:
        raise HTTPException(status_code=404, detail="Folio not found or already closed")

    # Check balance
    balance = await calculate_folio_balance(folio_id, current_user.tenant_id)

    if balance > 0.01:  # Allow small rounding differences
        raise HTTPException(
            status_code=400,
            detail=f"Cannot close folio with outstanding balance: {balance}"
        )

    await db.folios.update_one(
        {'id': folio_id},
        {'$set': {
            'status': 'closed',
            'balance': 0.0,
            'closed_at': datetime.now(UTC).isoformat()
        }}
    )

    return {"message": "Folio closed successfully"}


@router.get("/folio/{folio_id}/activity-log")
async def get_folio_activity_log(
    folio_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get activity log for a folio"""
    folio = await db.folios.find_one({
        'id': folio_id,
        'tenant_id': current_user.tenant_id
    }, {'_id': 0})

    if not folio:
        raise HTTPException(status_code=404, detail="Folio not found")

    # Get all activities
    activities = []

    # Charges
    charges = await db.folio_charges.find({
        'folio_id': folio_id,
        'tenant_id': current_user.tenant_id
    }, {'_id': 0}).to_list(1000)

    for charge in charges:
        activities.append({
            'type': 'charge',
            'action': 'voided' if charge.get('voided') else 'added',
            'timestamp': charge.get('posted_at'),
            'description': charge.get('description'),
            'amount': charge.get('total', charge.get('amount', 0)),
            'user': charge.get('posted_by'),
            'details': charge
        })

    # Payments
    payments = await db.payments.find({
        'folio_id': folio_id,
        'tenant_id': current_user.tenant_id
    }, {'_id': 0}).to_list(1000)

    for payment in payments:
        activities.append({
            'type': 'payment',
            'action': 'voided' if payment.get('voided') else 'processed',
            'timestamp': payment.get('voided_at') if payment.get('voided') else payment.get('processed_at'),
            'description': f"{payment.get('method', 'Payment')} - {payment.get('payment_type', '')}",
            'amount': payment.get('amount', 0),
            'user': payment.get('voided_by') if payment.get('voided') else payment.get('processed_by'),
            'details': payment
        })

    # Operations (transfers, etc)
    operations = await db.folio_operations.find({
        '$or': [
            {'from_folio_id': folio_id},
            {'to_folio_id': folio_id}
        ],
        'tenant_id': current_user.tenant_id
    }, {'_id': 0}).to_list(1000)

    for op in operations:
        activities.append({
            'type': 'operation',
            'action': op.get('operation_type'),
            'timestamp': op.get('performed_at'),
            'description': f"Operation: {op.get('operation_type')}",
            'user': op.get('performed_by'),
            'details': op
        })

    # Sort by timestamp
    activities.sort(key=lambda x: x['timestamp'] if x['timestamp'] else '', reverse=True)

    return {
        'folio': folio,
        'activities': activities,
        'total_count': len(activities)
    }




@router.get("/folio/{folio_id}/operations")
async def get_folio_operations(
    folio_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_finance_reports")),
):
    """Folio'ya ait transfer/işlem geçmişi (from veya to == folio_id)."""
    folio = await db.folios.find_one(
        {'id': folio_id, 'tenant_id': current_user.tenant_id},
        {'_id': 0, 'folio_number': 1}
    )
    if not folio:
        raise HTTPException(status_code=404, detail="Folio bulunamadı")

    ops = await db.folio_operations.find({
        'tenant_id': current_user.tenant_id,
        '$or': [{'from_folio_id': folio_id}, {'to_folio_id': folio_id}]
    }, {'_id': 0}).to_list(500)

    # performed_by → kullanıcı adı çöz
    user_ids = list({op.get('performed_by') for op in ops if op.get('performed_by')})
    users = {}
    if user_ids:
        async for u in db.users.find(
            {'tenant_id': current_user.tenant_id, 'id': {'$in': user_ids}},
            {'_id': 0, 'id': 1, 'name': 1, 'email': 1}
        ):
            users[u['id']] = u.get('name') or u.get('email') or u['id']

    # Karşı folio numaralarını çöz
    other_ids = set()
    for op in ops:
        for k in ('from_folio_id', 'to_folio_id'):
            v = op.get(k)
            if v and v != folio_id:
                other_ids.add(v)
    other_folios = {}
    if other_ids:
        async for f in db.folios.find(
            {'tenant_id': current_user.tenant_id, 'id': {'$in': list(other_ids)}},
            {'_id': 0, 'id': 1, 'folio_number': 1}
        ):
            other_folios[f['id']] = f.get('folio_number') or f['id'][:8]

    items = []
    for op in ops:
        items.append({
            'id': op.get('id'),
            'operation_type': op.get('operation_type'),
            'from_folio_id': op.get('from_folio_id'),
            'from_folio_number': folio.get('folio_number') if op.get('from_folio_id') == folio_id else other_folios.get(op.get('from_folio_id')),
            'to_folio_id': op.get('to_folio_id'),
            'to_folio_number': folio.get('folio_number') if op.get('to_folio_id') == folio_id else other_folios.get(op.get('to_folio_id')),
            'direction': 'out' if op.get('from_folio_id') == folio_id else 'in',
            'charge_ids': op.get('charge_ids') or [],
            'amount': op.get('amount'),
            'reason': op.get('reason'),
            'performed_by': op.get('performed_by'),
            'performed_by_name': users.get(op.get('performed_by')) if op.get('performed_by') else None,
            'performed_at': op.get('performed_at'),
        })
    items.sort(key=lambda x: x.get('performed_at') or '', reverse=True)
    return {'folio_id': folio_id, 'folio_number': folio.get('folio_number'), 'operations': items, 'count': len(items)}


@router.post("/folio/{folio_id}/proforma")
async def generate_folio_proforma(
    folio_id: str,
    current_user: User = Depends(get_current_user),
    _perm=Depends(require_op("view_finance_reports")),
):
    """Proforma fatura verisi: KDV oranı bazlı kırılım + toplamlar (folio'yu kapatmaz)."""
    folio = await db.folios.find_one(
        {'id': folio_id, 'tenant_id': current_user.tenant_id}, {'_id': 0}
    )
    if not folio:
        raise HTTPException(status_code=404, detail="Folio bulunamadı")

    charges = await db.folio_charges.find(
        {'folio_id': folio_id, 'tenant_id': current_user.tenant_id, 'voided': False},
        {'_id': 0}
    ).to_list(2000)

    payments = await db.payments.find(
        {'folio_id': folio_id, 'tenant_id': current_user.tenant_id, 'voided': False},
        {'_id': 0}
    ).to_list(2000)

    # KDV oranı bazlı gruplama
    vat_groups: dict[str, dict[str, float]] = {}
    subtotal_sum = 0.0
    discount_sum = 0.0
    net_sum = 0.0
    vat_sum = 0.0
    city_tax_sum = 0.0
    grand_total = 0.0

    for c in charges:
        sub = float(c.get('subtotal') or c.get('amount') or 0.0)
        disc = float(c.get('discount_amount') or 0.0)
        net = float(c.get('amount') or max(0.0, sub - disc))
        rate = float(c.get('vat_rate') or 0.0)
        vat = float(c.get('vat_amount') or 0.0)
        city = float(c.get('tax_amount') or 0.0)
        tot = float(c.get('total') or (net + vat + city))

        subtotal_sum += sub
        discount_sum += disc
        net_sum += net
        vat_sum += vat
        city_tax_sum += city
        grand_total += tot

        key = f"{rate:.2f}"
        g = vat_groups.setdefault(key, {'vat_rate': rate, 'net': 0.0, 'vat_amount': 0.0, 'count': 0})
        g['net'] = round(g['net'] + net, 2)
        g['vat_amount'] = round(g['vat_amount'] + vat, 2)
        g['count'] += 1

    payments_total = round(sum(float(p.get('amount') or 0.0) for p in payments), 2)
    balance = round(grand_total - payments_total, 2)

    # Misafir ve booking bilgileri
    guest = None
    if folio.get('guest_id'):
        guest = await db.guests.find_one(
            {'id': folio['guest_id'], 'tenant_id': current_user.tenant_id},
            {'_id': 0, 'name': 1, 'email': 1, 'phone': 1, 'tc_no': 1, 'address': 1}
        )
    booking = await db.bookings.find_one(
        {'id': folio['booking_id'], 'tenant_id': current_user.tenant_id},
        {'_id': 0, 'check_in': 1, 'check_out': 1, 'room_id': 1, 'room_number': 1, 'adults': 1, 'children': 1}
    )

    # Tenant ad
    tenant = await db.tenants.find_one(
        {'id': current_user.tenant_id}, {'_id': 0, 'name': 1, 'tax_no': 1, 'tax_office': 1, 'address': 1}
    ) or {}

    return {
        'status': 'draft',
        'document_type': 'proforma',
        'generated_at': datetime.now(UTC).isoformat(),
        'folio': {
            'id': folio.get('id'),
            'folio_number': folio.get('folio_number'),
            'folio_type': folio.get('folio_type'),
            'status': folio.get('status'),
        },
        'hotel': tenant,
        'guest': guest or {},
        'booking': booking or {},
        'charges': sorted(charges, key=lambda c: c.get('date') or c.get('created_at') or ''),
        'vat_breakdown': sorted(vat_groups.values(), key=lambda g: g['vat_rate']),
        'totals': {
            'subtotal': round(subtotal_sum, 2),
            'discount_total': round(discount_sum, 2),
            'net_total': round(net_sum, 2),
            'vat_total': round(vat_sum, 2),
            'city_tax_total': round(city_tax_sum, 2),
            'grand_total': round(grand_total, 2),
            'payments_total': payments_total,
            'balance_due': balance,
        },
    }

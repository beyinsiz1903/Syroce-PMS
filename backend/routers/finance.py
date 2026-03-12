"""
FINANCE Router - Extracted from server.py
"""
import asyncio
import uuid
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Dict, Any
from enum import Enum

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel, Field, ConfigDict, EmailStr

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: F401
    from openpyxl.utils import get_column_letter  # noqa: F401
except ImportError:
    Workbook = None

from core.database import db
from core.security import get_current_user
from core.helpers import (
    require_module, create_audit_log,
)
from models.enums import (
    PaymentStatus,
    ChargeCategory, FolioOperationType,
)
from models.schemas import (
    User, Folio, FolioCreate, FolioCharge, ChargeCreate,
    Payment, PaymentCreate, FolioOperation, FolioOperationCreate,
    Invoice, InvoiceCreate, Expense, CashFlow, BankAccount,
    CityLedgerTransaction,
    GenerateInvoiceFromFolioRequest, ConvertCurrencyRequest,
    CreateCurrencyRateRequest, CreateMultiCurrencyInvoiceRequest,
)

from core.utils import (
    calculate_folio_balance,
    excel_response,
)
from modules.folio.services.folio_balance_read_service import FolioBalanceReadService
from modules.folio.services.open_folio_service import OpenFolioService
from shared_kernel.shadow_metrics import compare_folio_payloads, run_shadow_compare

try:
    from cache_manager import cached
except ImportError:
    def cached(ttl=300, key_prefix=""):
        def decorator(func):
            return func
        return decorator

router = APIRouter(prefix="/api", tags=["finance"])
security = HTTPBearer()
folio_balance_read_service = FolioBalanceReadService()
open_folio_service = OpenFolioService()


# ── ERP Connectors (mock) ──

class LogoConnector:
    """Mock Logo/Netsis connector for ERP sync"""
    def __init__(self):
        import os
        self.base_url = os.environ.get('LOGO_API_URL', 'https://logo.example/api')

    async def send_invoice(self, invoice):
        import asyncio
        await asyncio.sleep(0.1)
        return {'external_id': f"LOGO-{invoice['id'][:8]}", 'status': 'synced', 'message': 'Invoice pushed to Logo'}

    async def send_payment(self, payment):
        import asyncio
        await asyncio.sleep(0.1)
        return {'external_id': f"LOGO-PAY-{payment['id'][:8]}", 'status': 'synced', 'message': 'Payment pushed to Logo'}


class NetsisConnector:
    """Mock Netsis connector"""
    def __init__(self):
        import os
        self.base_url = os.environ.get('NETSIS_API_URL', 'https://netsis.example/api')

    async def send_invoice(self, invoice):
        import asyncio
        await asyncio.sleep(0.1)
        return {'external_id': f"NETSIS-{invoice['id'][:8]}", 'status': 'synced', 'message': 'Invoice pushed to Netsis'}


async def _gather_invoices(tenant_id: str, since=None):
    query = {'tenant_id': tenant_id}
    if since:
        query['created_at'] = {'$gte': since}
    return await db.finance_invoices.find(query, {'_id': 0}).sort('created_at', -1).to_list(500)


async def _gather_payments(tenant_id: str, since=None):
    query = {'tenant_id': tenant_id}
    if since:
        query['created_at'] = {'$gte': since}
    return await db.finance_payments.find(query, {'_id': 0}).sort('created_at', -1).to_list(500)


async def _log_accounting_sync(tenant_id: str, payload: dict):
    record = {
        'id': str(uuid.uuid4()),
        'tenant_id': tenant_id,
        **payload,
        'created_at': datetime.now(timezone.utc).isoformat(),
    }
    await db.accounting_sync_logs.insert_one(record)
    return record


@router.post("/finance/logo-integration/sync")
async def sync_with_logo(sync_data: dict = None, current_user: User = Depends(get_current_user)):
    """Sync finance data with Logo ERP"""
    connector = LogoConnector()
    since = sync_data.get('since') if sync_data else None
    invoices = await _gather_invoices(current_user.tenant_id, since)
    payments = await _gather_payments(current_user.tenant_id, since)
    
    synced_invoices = []
    for invoice in invoices:
        result = await connector.send_invoice(invoice)
        synced_invoices.append({**invoice, **result})
    
    synced_payments = []
    for payment in payments:
        result = await connector.send_payment(payment)
        synced_payments.append({**payment, **result})
    
    log_entry = await _log_accounting_sync(current_user.tenant_id, {
        'provider': 'logo',
        'synced_invoices': len(synced_invoices),
        'synced_payments': len(synced_payments),
        'synced_at': datetime.now(timezone.utc).isoformat(),
        'status': 'success'
    })
    
    return {
        'success': True,
        'synced_invoices': len(synced_invoices),
        'synced_payments': len(synced_payments),
        'log_id': log_entry['id']
    }



@router.post("/finance/netsis-integration/sync")
async def sync_with_netsis(sync_data: dict = None, current_user: User = Depends(get_current_user)):
    connector = NetsisConnector()
    since = sync_data.get('since') if sync_data else None
    invoices = await _gather_invoices(current_user.tenant_id, since)
    
    synced = []
    for invoice in invoices:
        result = await connector.send_invoice(invoice)
        synced.append({**invoice, **result})
    
    log_entry = await _log_accounting_sync(current_user.tenant_id, {
        'provider': 'netsis',
        'synced_invoices': len(synced),
        'synced_payments': 0,
        'synced_at': datetime.now(timezone.utc).isoformat(),
        'status': 'success'
    })
    
    return {
        'success': True,
        'synced_invoices': len(synced),
        'log_id': log_entry['id']
    }



@router.get("/finance/integration/logs")
async def get_integration_logs(limit: int = 20, current_user: User = Depends(get_current_user)):
    logs = await db.accounting_sync_logs.find(
        {'tenant_id': current_user.tenant_id},
        {'_id': 0}
    ).sort('created_at', -1).limit(limit).to_list(limit)
    return {'logs': logs, 'count': len(logs)}


@router.get("/finance/budget-vs-actual")
async def budget_vs_actual(month: str, current_user: User = Depends(get_current_user)):
    # Simulated budget data
    budget = {'rooms': 150000, 'fnb': 50000, 'other': 20000, 'total': 220000}
    actual = {'rooms': 165000, 'fnb': 48000, 'other': 22000, 'total': 235000}
    variance = {k: actual[k] - budget[k] for k in budget}
    variance_pct = {k: round((variance[k] / budget[k] * 100), 1) if budget[k] > 0 else 0 for k in budget}
    return {
        'month': month, 'budget': budget, 'actual': actual, 
        'variance': variance, 'variance_pct': variance_pct
    }


@router.post("/folio/create", response_model=Folio)
async def create_folio(
    folio_data: FolioCreate,
    request: Request,
    current_user: User = Depends(get_current_user),
):
    """Create a new folio for a booking"""
    return await open_folio_service.create(folio_data, current_user, request)




@router.get("/folio/list")
async def list_folios(
    status: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    current_user: User = Depends(get_current_user)
):
    """List all folios for the current tenant with optional status filter."""
    query = {'tenant_id': current_user.tenant_id}
    if status:
        query['status'] = status

    folios = await db.folios.find(query, {'_id': 0}).sort(
        'created_at', -1
    ).skip(offset).limit(limit).to_list(limit)

    total = await db.folios.count_documents(query)

    # Enrich with guest/booking info
    for folio in folios:
        booking = await db.bookings.find_one(
            {'id': folio.get('booking_id'), 'tenant_id': current_user.tenant_id},
            {'_id': 0, 'guest_name': 1, 'room_number': 1, 'room_id': 1, 'check_in': 1, 'check_out': 1}
        )
        if booking:
            folio['guest_name'] = booking.get('guest_name', '')
            folio['room_number'] = booking.get('room_number', '')
            folio['check_in'] = booking.get('check_in', '')
            folio['check_out'] = booking.get('check_out', '')

    return {
        'folios': folios,
        'total': total,
        'limit': limit,
        'offset': offset,
    }


@router.get("/folio/dashboard-stats")
@cached(ttl=300, key_prefix="folio_dashboard_stats")  # Cache for 5 minutes
async def get_folio_dashboard_stats(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get folio statistics for dashboard"""
    current_user = await get_current_user(credentials)
    
    try:
        # Get all open folios
        open_folios = await db.folios.find({
            'tenant_id': current_user.tenant_id,
            'status': 'open'
        }, {'_id': 0}).to_list(1000)
        
        # Calculate total outstanding balance from folio balance field
        total_outstanding = 0.0
        for folio in open_folios:
            # Use the balance field directly instead of calculating
            total_outstanding += folio.get('balance', 0)
        
        # Get recent charges (last 24 hours)
        yesterday = (datetime.now(timezone.utc) - timedelta(days=1)).isoformat()
        recent_charges = await db.folio_charges.count_documents({
            'tenant_id': current_user.tenant_id,
            'date': {'$gte': yesterday},
            'voided': False
        })
        
        # Get recent payments (last 24 hours)
        recent_payments = await db.payments.count_documents({
            'tenant_id': current_user.tenant_id,
            'date': {'$gte': yesterday}
        })
        
        return {
            'total_open_folios': len(open_folios),
            'total_outstanding_balance': round(total_outstanding, 2),
            'recent_charges_24h': recent_charges,
            'recent_payments_24h': recent_payments
        }
    except Exception as e:
        print(f"Error in folio dashboard stats: {str(e)}")
        import traceback
        traceback.print_exc()
        # Return default values instead of raising
        return {
            'total_open_folios': 0,
            'total_outstanding_balance': 0.0,
            'recent_charges_24h': 0,
            'recent_payments_24h': 0
        }


@router.get("/folio/pending-ar")
@cached(ttl=600, key_prefix="folio_pending_ar")  # Cache for 10 min
async def get_pending_ar(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get pending accounts receivable (company folios with outstanding balances)"""
    current_user = await get_current_user(credentials)
    
    try:
        # Get all companies
        companies = await db.companies.find({
            'tenant_id': current_user.tenant_id
        }, {'_id': 0}).to_list(1000)
        
        ar_data = []
        
        for company in companies:
            # Get company's open folios with balance
            company_folios = await db.folios.find({
                'tenant_id': current_user.tenant_id,
                'company_id': company['id'],
                'status': 'open'
            }, {'_id': 0}).to_list(1000)
            
            # Use balance field directly
            folios_with_balance = [f for f in company_folios if f.get('balance', 0) > 0]
            total_outstanding = sum(f.get('balance', 0) for f in folios_with_balance)
            
            if total_outstanding > 0 and folios_with_balance:
                # Find oldest folio
                oldest_folio = min(folios_with_balance, key=lambda f: f.get('created_at', datetime.now(timezone.utc).isoformat()))
                oldest_date = datetime.fromisoformat(oldest_folio['created_at'].replace('Z', '+00:00'))
                days_outstanding = (datetime.now(timezone.utc) - oldest_date).days
                
                # Calculate aging
                aging = {'0-7': 0, '8-14': 0, '15-30': 0, '30+': 0}
                for folio in folios_with_balance:
                    days = (datetime.now(timezone.utc) - datetime.fromisoformat(folio['created_at'].replace('Z', '+00:00'))).days
                    balance = folio.get('balance', 0)
                    if days <= 7:
                        aging['0-7'] += balance
                    elif days <= 14:
                        aging['8-14'] += balance
                    elif days <= 30:
                        aging['15-30'] += balance
                    else:
                        aging['30+'] += balance
                
                ar_data.append({
                    'company_id': company['id'],
                    'company_name': company.get('name', 'Unknown'),
                    'corporate_code': company.get('corporate_code', ''),
                    'contact_person': company.get('contact_person', ''),
                    'contact_email': company.get('contact_email', ''),
                    'contact_phone': company.get('contact_phone', ''),
                    'payment_terms': company.get('payment_terms', 'Net 30'),
                    'total_outstanding': round(total_outstanding, 2),
                    'open_folios_count': len(folios_with_balance),
                    'oldest_invoice_date': oldest_folio['created_at'],
                    'days_outstanding': days_outstanding,
                    'aging': aging
                })
        
        # Sort by days outstanding (oldest first)
        ar_data.sort(key=lambda x: x['days_outstanding'], reverse=True)
        
        return ar_data
        
    except Exception as e:
        print(f"Error in get_pending_ar: {str(e)}")
        import traceback
        traceback.print_exc()
        return []


@router.get("/folio/booking/{booking_id}", response_model=List[Folio])
@cached(ttl=180, key_prefix="folio_by_booking")  # Cache for 3 min
async def get_booking_folios(booking_id: str, current_user: User = Depends(get_current_user)):
    """Get all folios for a booking"""
    folios = await db.folios.find({
        'booking_id': booking_id,
        'tenant_id': current_user.tenant_id
    }, {'_id': 0}).to_list(1000)
    
    # Calculate current balance for each folio
    for folio in folios:
        folio['balance'] = await calculate_folio_balance(folio['id'], current_user.tenant_id)
    
    return folios


@router.get("/folio/{folio_id}", response_model=Dict[str, Any])
@cached(ttl=180, key_prefix="folio_details")  # Cache for 3 min
async def get_folio_details(
    folio_id: str,
    request: Request,
    current_user: User = Depends(get_current_user),
):
    """Get folio with charges and payments"""
    semantic_response = await folio_balance_read_service.get_folio_details(
        tenant_id=current_user.tenant_id,
        folio_id=folio_id,
    )
    asyncio.create_task(
        run_shadow_compare(
            endpoint="folio",
            tenant_id=current_user.tenant_id,
            property_id=request.headers.get("x-property-id"),
            correlation_id=request.headers.get("x-correlation-id"),
            semantic_payload=semantic_response,
            legacy_loader=lambda: _legacy_get_folio_details(
                tenant_id=current_user.tenant_id,
                folio_id=folio_id,
            ),
            comparator=compare_folio_payloads,
            entity_id=folio_id,
        )
    )
    return semantic_response


async def _legacy_get_folio_details(tenant_id: str, folio_id: str):
    folio = await db.folios.find_one({
        'id': folio_id,
        'tenant_id': tenant_id
    }, {'_id': 0})

    if not folio:
        raise HTTPException(status_code=404, detail="Folio not found")

    charges = await db.folio_charges.find({
        'folio_id': folio_id,
        'tenant_id': tenant_id
    }, {'_id': 0}).to_list(1000)

    payments = await db.payments.find({
        'folio_id': folio_id,
        'tenant_id': tenant_id
    }, {'_id': 0}).to_list(1000)

    balance = await calculate_folio_balance(folio_id, tenant_id)
    folio['balance'] = balance

    return {
        'folio': folio,
        'charges': charges,
        'payments': payments,
        'balance': balance
    }



@router.get("/folio/{folio_id}/excel")
@cached(ttl=600, key_prefix="folio_excel")  # Cache for 10 min
async def export_folio_excel(folio_id: str, current_user: User = Depends(get_current_user)):
    """Export Folio to Excel"""
    folio_data = await get_folio_details(folio_id, current_user)
    
    folio = folio_data['folio']
    charges = folio_data['charges']
    payments = folio_data['payments']
    balance = folio_data['balance']
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Folio"
    
    # Folio header
    ws['A1'] = "GUEST FOLIO"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:E1')
    
    ws['A3'] = "Folio Number:"
    ws['B3'] = folio.get('folio_number', 'N/A')
    ws['A4'] = "Type:"
    ws['B4'] = folio.get('folio_type', 'guest').title()
    ws['A5'] = "Status:"
    ws['B5'] = folio.get('status', 'open').upper()
    ws['A6'] = "Created:"
    ws['B6'] = folio.get('created_at', '')[:10]
    
    # Charges section
    ws['A9'] = "CHARGES"
    ws['A9'].font = Font(size=14, bold=True)
    
    charge_headers = ["Date", "Description", "Qty", "Amount", "Tax", "Total"]
    for col_num, header in enumerate(charge_headers, 1):
        cell = ws.cell(row=10, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    row = 11
    total_charges = 0
    for charge in charges:
        if not charge.get('voided', False):
            ws.cell(row=row, column=1, value=charge.get('posted_at', '')[:10])
            ws.cell(row=row, column=2, value=charge.get('description', ''))
            ws.cell(row=row, column=3, value=charge.get('quantity', 1))
            ws.cell(row=row, column=4, value=f"${charge.get('amount', 0):,.2f}")
            ws.cell(row=row, column=5, value=f"${charge.get('tax_amount', 0):,.2f}")
            ws.cell(row=row, column=6, value=f"${charge.get('total', 0):,.2f}")
            total_charges += charge.get('total', 0)
            row += 1
    
    ws.cell(row=row, column=5, value="Total Charges:")
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=6, value=f"${total_charges:,.2f}")
    ws.cell(row=row, column=6).font = Font(bold=True)
    
    # Payments section
    row += 2
    ws.cell(row=row, column=1, value="PAYMENTS")
    ws.cell(row=row, column=1).font = Font(size=14, bold=True)
    row += 1
    
    payment_headers = ["Date", "Method", "Type", "Amount"]
    for col_num, header in enumerate(payment_headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    row += 1
    total_payments = 0
    for payment in payments:
        ws.cell(row=row, column=1, value=payment.get('processed_at', '')[:10])
        ws.cell(row=row, column=2, value=payment.get('payment_method', '').title())
        ws.cell(row=row, column=3, value=payment.get('payment_type', '').title())
        ws.cell(row=row, column=4, value=f"${payment.get('amount', 0):,.2f}")
        total_payments += payment.get('amount', 0)
        row += 1
    
    ws.cell(row=row, column=3, value="Total Payments:")
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=4, value=f"${total_payments:,.2f}")
    ws.cell(row=row, column=4).font = Font(bold=True)
    
    # Balance
    row += 2
    ws.cell(row=row, column=5, value="BALANCE DUE:")
    ws.cell(row=row, column=5).font = Font(size=14, bold=True)
    ws.cell(row=row, column=6, value=f"${balance:,.2f}")
    ws.cell(row=row, column=6).font = Font(size=14, bold=True)
    ws.cell(row=row, column=6).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    filename = f"folio_{folio.get('folio_number', folio_id)}.xlsx"
    return excel_response(wb, filename)



@router.post("/folio/{folio_id}/charge", response_model=FolioCharge)
async def post_charge_to_folio(
    folio_id: str,
    charge_data: ChargeCreate,
    current_user: User = Depends(get_current_user)
):
    """Post a charge to folio"""
    folio = await db.folios.find_one({
        'id': folio_id,
        'tenant_id': current_user.tenant_id,
        'status': 'open'
    })
    
    if not folio:
        raise HTTPException(status_code=404, detail="Folio not found or closed")
    
    # Calculate amounts with proper rounding
    amount = round(charge_data.amount * charge_data.quantity, 2)
    tax_amount = 0.0
    
    # Auto-calculate city tax if requested
    if charge_data.auto_calculate_tax and charge_data.charge_category == ChargeCategory.ROOM:
        # Get city tax rule
        tax_rule = await db.city_tax_rules.find_one({
            'tenant_id': current_user.tenant_id,
            'active': True
        })
        if tax_rule:
            if tax_rule.get('flat_amount'):
                tax_amount = round(tax_rule['flat_amount'], 2)
            else:
                tax_amount = round(amount * (tax_rule['tax_percentage'] / 100), 2)
    
    total = round(amount + tax_amount, 2)
    
    charge = FolioCharge(
        tenant_id=current_user.tenant_id,
        folio_id=folio_id,
        booking_id=folio['booking_id'],
        charge_category=charge_data.charge_category,
        description=charge_data.description,
        unit_price=charge_data.amount,
        quantity=charge_data.quantity,
        amount=amount,
        tax_amount=tax_amount,
        total=total,
        posted_by=current_user.id
    )
    
    charge_dict = charge.model_dump()
    charge_dict['date'] = charge_dict['date'].isoformat()
    await db.folio_charges.insert_one(charge_dict)
    
    # Update folio balance
    balance = await calculate_folio_balance(folio_id, current_user.tenant_id)
    await db.folios.update_one(
        {'id': folio_id},
        {'$set': {'balance': balance}}
    )
    
    # Audit log
    await create_audit_log(
        tenant_id=current_user.tenant_id,
        user=current_user,
        action="POST_CHARGE",
        entity_type="folio_charge",
        entity_id=charge.id,
        changes={'charge_category': charge_data.charge_category, 'amount': total, 'folio_id': folio_id}
    )
    
    return charge


@router.post("/folio/{folio_id}/payment", response_model=Payment)
async def post_payment_to_folio(
    folio_id: str,
    payment_data: PaymentCreate,
    current_user: User = Depends(get_current_user)
):
    """Post a payment to folio"""
    folio = await db.folios.find_one({
        'id': folio_id,
        'tenant_id': current_user.tenant_id
    })
    
    if not folio:
        raise HTTPException(status_code=404, detail="Folio not found")
    
    payment = Payment(
        tenant_id=current_user.tenant_id,
        folio_id=folio_id,
        booking_id=folio['booking_id'],
        processed_by=current_user.id,
        **payment_data.model_dump()
    )
    
    payment_dict = payment.model_dump()
    payment_dict['processed_at'] = payment_dict['processed_at'].isoformat()
    payment_dict['processed_by_name'] = current_user.name  # Add user name
    await db.payments.insert_one(payment_dict)
    
    # Update folio balance
    balance = await calculate_folio_balance(folio_id, current_user.tenant_id)
    await db.folios.update_one(
        {'id': folio_id},
        {'$set': {'balance': balance}}
    )
    
    return payment


@router.post("/folio/transfer", response_model=FolioOperation)
async def transfer_charges(
    operation_data: FolioOperationCreate,
    current_user: User = Depends(get_current_user)
):
    """Transfer charges from one folio to another"""
    if operation_data.operation_type != FolioOperationType.TRANSFER:
        raise HTTPException(status_code=400, detail="Invalid operation type")
    
    if not operation_data.to_folio_id:
        raise HTTPException(status_code=400, detail="Destination folio required for transfer")
    
    # Verify both folios exist
    from_folio = await db.folios.find_one({
        'id': operation_data.from_folio_id,
        'tenant_id': current_user.tenant_id
    })
    
    to_folio = await db.folios.find_one({
        'id': operation_data.to_folio_id,
        'tenant_id': current_user.tenant_id,
        'status': 'open'
    })
    
    if not from_folio or not to_folio:
        raise HTTPException(status_code=404, detail="Folio not found")
    
    # Transfer specified charges
    for charge_id in operation_data.charge_ids:
        await db.folio_charges.update_one(
            {'id': charge_id, 'folio_id': operation_data.from_folio_id},
            {'$set': {'folio_id': operation_data.to_folio_id}}
        )
    
    # Create operation record
    operation = FolioOperation(
        tenant_id=current_user.tenant_id,
        performed_by=current_user.id,
        **operation_data.model_dump()
    )
    
    operation_dict = operation.model_dump()
    operation_dict['performed_at'] = operation_dict['performed_at'].isoformat()
    await db.folio_operations.insert_one(operation_dict)
    
    # Update balances
    from_balance = await calculate_folio_balance(operation_data.from_folio_id, current_user.tenant_id)
    to_balance = await calculate_folio_balance(operation_data.to_folio_id, current_user.tenant_id)
    
    await db.folios.update_one(
        {'id': operation_data.from_folio_id},
        {'$set': {'balance': from_balance}}
    )
    await db.folios.update_one(
        {'id': operation_data.to_folio_id},
        {'$set': {'balance': to_balance}}
    )
    
    return operation


@router.post("/folio/{folio_id}/void-charge/{charge_id}")
async def void_charge(
    folio_id: str,
    charge_id: str,
    void_reason: str,
    current_user: User = Depends(get_current_user)
):
    """Void a charge"""
    charge = await db.folio_charges.find_one({
        'id': charge_id,
        'folio_id': folio_id,
        'tenant_id': current_user.tenant_id,
        'voided': False
    })
    
    if not charge:
        raise HTTPException(status_code=404, detail="Charge not found or already voided")
    
    await db.folio_charges.update_one(
        {'id': charge_id},
        {'$set': {
            'voided': True,
            'void_reason': void_reason,
            'voided_by': current_user.id,
            'voided_at': datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update folio balance
    balance = await calculate_folio_balance(folio_id, current_user.tenant_id)
    await db.folios.update_one(
        {'id': folio_id},
        {'$set': {'balance': balance}}
    )
    
    # Create operation record
    operation = FolioOperation(
        tenant_id=current_user.tenant_id,
        operation_type=FolioOperationType.VOID,
        from_folio_id=folio_id,
        charge_ids=[charge_id],
        amount=charge['total'],
        reason=void_reason,
        performed_by=current_user.id
    )
    
    operation_dict = operation.model_dump()
    operation_dict['performed_at'] = operation_dict['performed_at'].isoformat()
    await db.folio_operations.insert_one(operation_dict)
    
    return {"message": "Charge voided successfully"}


@router.post("/folio/{folio_id}/close")
async def close_folio(
    folio_id: str,
    current_user: User = Depends(get_current_user)
):
    """Close a folio"""
    folio = await db.folios.find_one({
        'id': folio_id,
        'tenant_id': current_user.tenant_id,
        'status': 'open'
    })
    
    if not folio:
        raise HTTPException(status_code=404, detail="Folio not found or already closed")
    
    # Check balance
    balance = await calculate_folio_balance(folio_id, current_user.tenant_id)
    
    if balance > 0.01:  # Allow small rounding differences
        raise HTTPException(
            status_code=400,
            detail=f"Cannot close folio with outstanding balance: {balance}"
        )
    
    await db.folios.update_one(
        {'id': folio_id},
        {'$set': {
            'status': 'closed',
            'balance': 0.0,
            'closed_at': datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Folio closed successfully"}


@router.get("/folio/{folio_id}/activity-log")
async def get_folio_activity_log(
    folio_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get activity log for a folio"""
    folio = await db.folios.find_one({
        'id': folio_id,
        'tenant_id': current_user.tenant_id
    }, {'_id': 0})
    
    if not folio:
        raise HTTPException(status_code=404, detail="Folio not found")
    
    # Get all activities
    activities = []
    
    # Charges
    charges = await db.folio_charges.find({
        'folio_id': folio_id,
        'tenant_id': current_user.tenant_id
    }, {'_id': 0}).to_list(1000)
    
    for charge in charges:
        activities.append({
            'type': 'charge',
            'action': 'voided' if charge.get('voided') else 'added',
            'timestamp': charge.get('posted_at'),
            'description': charge.get('description'),
            'amount': charge.get('total', charge.get('amount', 0)),
            'user': charge.get('posted_by'),
            'details': charge
        })
    
    # Payments
    payments = await db.payments.find({
        'folio_id': folio_id,
        'tenant_id': current_user.tenant_id
    }, {'_id': 0}).to_list(1000)
    
    for payment in payments:
        activities.append({
            'type': 'payment',
            'action': 'voided' if payment.get('voided') else 'processed',
            'timestamp': payment.get('voided_at') if payment.get('voided') else payment.get('processed_at'),
            'description': f"{payment.get('method', 'Payment')} - {payment.get('payment_type', '')}",
            'amount': payment.get('amount', 0),
            'user': payment.get('voided_by') if payment.get('voided') else payment.get('processed_by'),
            'details': payment
        })
    
    # Operations (transfers, etc)
    operations = await db.folio_operations.find({
        '$or': [
            {'from_folio_id': folio_id},
            {'to_folio_id': folio_id}
        ],
        'tenant_id': current_user.tenant_id
    }, {'_id': 0}).to_list(1000)
    
    for op in operations:
        activities.append({
            'type': 'operation',
            'action': op.get('operation_type'),
            'timestamp': op.get('performed_at'),
            'description': f"Operation: {op.get('operation_type')}",
            'user': op.get('performed_by'),
            'details': op
        })
    
    # Sort by timestamp
    activities.sort(key=lambda x: x['timestamp'] if x['timestamp'] else '', reverse=True)
    
    return {
        'folio': folio,
        'activities': activities,
        'total_count': len(activities)
    }


@router.post("/invoices", response_model=Invoice)
async def create_invoice(
    invoice_data: InvoiceCreate,
    current_user: User = Depends(get_current_user),
    _: None = Depends(require_module("invoices")),
):
    count = await db.invoices.count_documents({'tenant_id': current_user.tenant_id})
    invoice_number = f"INV-{count + 1:05d}"
    due_date_dt = datetime.fromisoformat(invoice_data.due_date.replace('Z', '+00:00'))
    invoice = Invoice(tenant_id=current_user.tenant_id, invoice_number=invoice_number, due_date=due_date_dt,
                     **{k: v for k, v in invoice_data.model_dump().items() if k != 'due_date'})
    invoice_dict = invoice.model_dump()
    invoice_dict['issue_date'] = invoice_dict['issue_date'].isoformat()
    invoice_dict['due_date'] = invoice_dict['due_date'].isoformat()
    await db.invoices.insert_one(invoice_dict)
    return invoice


@router.get("/invoices", response_model=List[Invoice])
@cached(ttl=300, key_prefix="invoices_list")  # Cache for 5 min
async def get_invoices(
    current_user: User = Depends(get_current_user),
    _: None = Depends(require_module("invoices")),
):
    invoices = await db.invoices.find({'tenant_id': current_user.tenant_id}, {'_id': 0}).to_list(1000)
    return invoices


@router.put("/invoices/{invoice_id}")
async def update_invoice(
    invoice_id: str,
    updates: Dict[str, Any],
    current_user: User = Depends(get_current_user),
    _: None = Depends(require_module("invoices")),
):
    await db.invoices.update_one({'id': invoice_id, 'tenant_id': current_user.tenant_id}, {'$set': updates})
    invoice_doc = await db.invoices.find_one({'id': invoice_id}, {'_id': 0})
    return invoice_doc


@router.get("/invoices/stats")
@cached(ttl=120, key_prefix="invoices_stats")  # Cache for 2 min - faster refresh
async def get_invoice_stats(current_user: User = Depends(get_current_user)):
    invoices = await db.invoices.find({'tenant_id': current_user.tenant_id}, {'_id': 0}).to_list(1000)
    total_revenue = sum(inv['total'] for inv in invoices if inv['status'] == 'paid')
    pending_amount = sum(inv['total'] for inv in invoices if inv['status'] in ['draft', 'sent'])
    overdue_amount = sum(inv['total'] for inv in invoices if inv['status'] == 'overdue')
    return {'total_invoices': len(invoices), 'total_revenue': total_revenue, 'pending_amount': pending_amount, 'overdue_amount': overdue_amount}


class InvoiceType(str, Enum):
    SALES = "sales"  # Satış faturası
    PURCHASE = "purchase"  # Alış faturası
    PROFORMA = "proforma"  # Proforma
    E_INVOICE = "e_invoice"  # E-Fatura
    E_ARCHIVE = "e_archive"  # E-Arşiv


class ExpenseCategory(str, Enum):
    SALARIES = "salaries"
    UTILITIES = "utilities"
    SUPPLIES = "supplies"
    MAINTENANCE = "maintenance"
    MARKETING = "marketing"
    RENT = "rent"
    INSURANCE = "insurance"
    TAXES = "taxes"
    OTHER = "other"


class Supplier(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    name: str
    tax_office: Optional[str] = None
    tax_number: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    account_balance: float = 0.0
    category: str = "general"
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class BankAccount(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    name: str
    bank_name: str
    account_number: str
    iban: Optional[str] = None
    currency: str = "USD"
    balance: float = 0.0
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class Expense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    expense_number: str
    supplier_id: Optional[str] = None
    category: ExpenseCategory
    description: str
    amount: float
    vat_rate: float = 18.0
    vat_amount: float = 0.0
    total_amount: float
    date: datetime
    payment_status: PaymentStatus = PaymentStatus.PENDING
    payment_method: Optional[str] = None
    receipt_url: Optional[str] = None
    notes: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class InventoryItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    name: str
    sku: Optional[str] = None
    category: str
    unit: str
    quantity: float = 0.0
    unit_cost: float = 0.0
    reorder_level: float = 0.0
    supplier_id: Optional[str] = None
    location: Optional[str] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class StockMovement(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    item_id: str
    movement_type: str  # in, out, adjustment
    quantity: float
    unit_cost: float
    reference: Optional[str] = None
    notes: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


@router.post("/accounting/suppliers")
async def create_supplier(
    name: str,
    tax_office: Optional[str] = None,
    tax_number: Optional[str] = None,
    email: Optional[str] = None,
    phone: Optional[str] = None,
    address: Optional[str] = None,
    category: str = "general",
    current_user: User = Depends(get_current_user)
):
    # Supplier model imported at top
    supplier = Supplier(
        tenant_id=current_user.tenant_id,
        name=name,
        tax_office=tax_office,
        tax_number=tax_number,
        email=email,
        phone=phone,
        address=address,
        category=category
    )
    supplier_dict = supplier.model_dump()
    supplier_dict['created_at'] = supplier_dict['created_at'].isoformat()
    await db.suppliers.insert_one(supplier_dict)
    return supplier



@router.get("/accounting/suppliers")
async def get_suppliers(current_user: User = Depends(get_current_user)):
    suppliers = await db.suppliers.find({'tenant_id': current_user.tenant_id}, {'_id': 0}).to_list(1000)
    return suppliers



@router.put("/accounting/suppliers/{supplier_id}")
async def update_supplier(supplier_id: str, updates: Dict[str, Any], current_user: User = Depends(get_current_user)):
    await db.suppliers.update_one({'id': supplier_id, 'tenant_id': current_user.tenant_id}, {'$set': updates})
    supplier = await db.suppliers.find_one({'id': supplier_id}, {'_id': 0})
    return supplier


@router.post("/accounting/bank-accounts")
async def create_bank_account(
    name: str,
    bank_name: str,
    account_number: str,
    iban: Optional[str] = None,
    currency: str = "USD",
    balance: float = 0.0,
    current_user: User = Depends(get_current_user)
):
    # BankAccount model imported at top
    bank_account = BankAccount(
        tenant_id=current_user.tenant_id,
        name=name,
        bank_name=bank_name,
        account_number=account_number,
        iban=iban,
        currency=currency,
        balance=balance
    )
    account_dict = bank_account.model_dump()
    account_dict['created_at'] = account_dict['created_at'].isoformat()
    await db.bank_accounts.insert_one(account_dict)
    return bank_account



@router.get("/accounting/bank-accounts")
async def get_bank_accounts(current_user: User = Depends(get_current_user)):
    accounts = await db.bank_accounts.find({'tenant_id': current_user.tenant_id}, {'_id': 0}).to_list(1000)
    return accounts



@router.put("/accounting/bank-accounts/{account_id}")
async def update_bank_account(account_id: str, updates: Dict[str, Any], current_user: User = Depends(get_current_user)):
    await db.bank_accounts.update_one({'id': account_id, 'tenant_id': current_user.tenant_id}, {'$set': updates})
    account = await db.bank_accounts.find_one({'id': account_id}, {'_id': 0})
    return account


@router.post("/accounting/expenses")
async def create_expense(
    category: str,
    description: str,
    amount: float,
    vat_rate: float,
    date: str,
    supplier_id: Optional[str] = None,
    payment_method: Optional[str] = None,
    receipt_url: Optional[str] = None,
    notes: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    # Expense model imported at top
    
    count = await db.expenses.count_documents({'tenant_id': current_user.tenant_id})
    expense_number = f"EXP-{count + 1:05d}"
    
    vat_amount = amount * (vat_rate / 100)
    total_amount = amount + vat_amount
    
    expense = Expense(
        tenant_id=current_user.tenant_id,
        expense_number=expense_number,
        supplier_id=supplier_id,
        category=category,
        description=description,
        amount=amount,
        vat_rate=vat_rate,
        vat_amount=vat_amount,
        total_amount=total_amount,
        date=datetime.fromisoformat(date),
        payment_method=payment_method,
        receipt_url=receipt_url,
        notes=notes,
        created_by=current_user.name
    )
    
    expense_dict = expense.model_dump()
    expense_dict['date'] = expense_dict['date'].isoformat()
    expense_dict['created_at'] = expense_dict['created_at'].isoformat()
    await db.expenses.insert_one(expense_dict)
    
    # Update supplier balance if applicable
    if supplier_id:
        await db.suppliers.update_one(
            {'id': supplier_id},
            {'$inc': {'account_balance': total_amount}}
        )
    
    # Create cash flow entry
    # CashFlow model imported at top
    cash_flow = CashFlow(
        tenant_id=current_user.tenant_id,
        transaction_type='expense',
        category=category,
        amount=total_amount,
        description=description,
        reference_id=expense.id,
        reference_type='expense',
        date=datetime.fromisoformat(date),
        created_by=current_user.name
    )
    cf_dict = cash_flow.model_dump()
    cf_dict['date'] = cf_dict['date'].isoformat()
    cf_dict['created_at'] = cf_dict['created_at'].isoformat()
    await db.cash_flow.insert_one(cf_dict)
    
    return expense



@router.get("/accounting/expenses")
async def get_expenses(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    category: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {'tenant_id': current_user.tenant_id}
    if start_date and end_date:
        query['date'] = {'$gte': start_date, '$lte': end_date}
    if category:
        query['category'] = category
    
    expenses = await db.expenses.find(query, {'_id': 0}).sort('date', -1).to_list(1000)
    return expenses



@router.put("/accounting/expenses/{expense_id}")
async def update_expense(expense_id: str, updates: Dict[str, Any], current_user: User = Depends(get_current_user)):
    await db.expenses.update_one({'id': expense_id, 'tenant_id': current_user.tenant_id}, {'$set': updates})
    expense = await db.expenses.find_one({'id': expense_id}, {'_id': 0})
    return expense


@router.post("/accounting/inventory")
async def create_inventory_item(
    name: str,
    category: str,
    unit: str,
    quantity: float = 0.0,
    unit_cost: float = 0.0,
    reorder_level: float = 0.0,
    sku: Optional[str] = None,
    supplier_id: Optional[str] = None,
    location: Optional[str] = None,
    notes: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    # InventoryItem model imported at top
    item = InventoryItem(
        tenant_id=current_user.tenant_id,
        name=name,
        sku=sku,
        category=category,
        unit=unit,
        quantity=quantity,
        unit_cost=unit_cost,
        reorder_level=reorder_level,
        supplier_id=supplier_id,
        location=location,
        notes=notes
    )
    item_dict = item.model_dump()
    item_dict['created_at'] = item_dict['created_at'].isoformat()
    await db.inventory_items.insert_one(item_dict)
    return item



@router.get("/accounting/inventory")
async def get_inventory(current_user: User = Depends(get_current_user)):
    items = await db.inventory_items.find({'tenant_id': current_user.tenant_id}, {'_id': 0}).to_list(1000)
    
    # Get low stock items
    low_stock = [item for item in items if item['quantity'] <= item['reorder_level']]
    
    return {
        'items': items,
        'low_stock_count': len(low_stock),
        'total_value': sum(item['quantity'] * item['unit_cost'] for item in items)
    }



@router.post("/accounting/inventory/movement")
async def create_stock_movement(
    item_id: str,
    movement_type: str,
    quantity: float,
    unit_cost: float,
    reference: Optional[str] = None,
    notes: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    # StockMovement model imported at top
    
    movement = StockMovement(
        tenant_id=current_user.tenant_id,
        item_id=item_id,
        movement_type=movement_type,
        quantity=quantity,
        unit_cost=unit_cost,
        reference=reference,
        notes=notes,
        created_by=current_user.name
    )
    
    movement_dict = movement.model_dump()
    movement_dict['created_at'] = movement_dict['created_at'].isoformat()
    await db.stock_movements.insert_one(movement_dict)
    
    # Update inventory quantity
    if movement_type == 'in':
        await db.inventory_items.update_one(
            {'id': item_id},
            {'$inc': {'quantity': quantity}}
        )
    elif movement_type == 'out':
        await db.inventory_items.update_one(
            {'id': item_id},
            {'$inc': {'quantity': -quantity}}
        )
    else:  # adjustment
        await db.inventory_items.update_one(
            {'id': item_id},
            {'$set': {'quantity': quantity}}
        )
    
    return movement


class AccountingInvoiceCreateRequest(BaseModel):
    invoice_type: str
    customer_name: str
    customer_email: Optional[str] = None
    customer_tax_office: Optional[str] = None
    customer_tax_number: Optional[str] = None
    customer_address: Optional[str] = None
    items: List[Dict[str, Any]] = []
    due_date: str
    booking_id: Optional[str] = None
    notes: Optional[str] = None


@router.post("/accounting/invoices")
async def create_accounting_invoice(
    request: AccountingInvoiceCreateRequest,
    current_user: User = Depends(get_current_user)
):
    # Models are now imported at the top of the file
    
    count = await db.accounting_invoices.count_documents({'tenant_id': current_user.tenant_id})
    invoice_number = f"INV-{datetime.now().year}-{count + 1:05d}"
    
    invoice_items = []
    subtotal = 0.0
    total_vat = 0.0
    vat_withholding = 0.0
    total_additional_taxes = 0.0
    
    for item_data in request.items:
        # Handle additional_taxes parsing
        additional_taxes = []
        if 'additional_taxes' in item_data and item_data['additional_taxes']:
            for tax_data in item_data['additional_taxes']:
                additional_taxes.append(AdditionalTax(**tax_data))
        
        # Create item with parsed additional taxes
        item_dict = {k: v for k, v in item_data.items() if k != 'additional_taxes'}
        item_dict['additional_taxes'] = additional_taxes
        
        item = AccountingInvoiceItem(**item_dict)
        
        invoice_items.append(item)
        subtotal += item.quantity * item.unit_price
        total_vat += item.vat_amount
        
        # Calculate additional taxes if present
        if item.additional_taxes:
            for tax in item.additional_taxes:
                if tax.tax_type == 'withholding':
                    # Withholding tax is deducted from VAT
                    # Calculate based on withholding rate (e.g., "7/10" = 70%)
                    if tax.withholding_rate:
                        rate_parts = tax.withholding_rate.split('/')
                        if len(rate_parts) == 2:
                            rate_percent = (int(rate_parts[0]) / int(rate_parts[1])) * 100
                            withholding_amount = item.vat_amount * (rate_percent / 100)
                            vat_withholding += withholding_amount
                            tax.calculated_amount = withholding_amount
                else:
                    # Other taxes (ÖTV, accommodation, etc.)
                    if tax.is_percentage and tax.rate:
                        tax_amount = (item.quantity * item.unit_price) * (tax.rate / 100)
                        total_additional_taxes += tax_amount
                        tax.calculated_amount = tax_amount
                    elif tax.amount:
                        total_additional_taxes += tax.amount
                        tax.calculated_amount = tax.amount
    
    total = subtotal + total_vat + total_additional_taxes - vat_withholding
    
    invoice = AccountingInvoice(
        tenant_id=current_user.tenant_id,
        invoice_number=invoice_number,
        invoice_type=request.invoice_type,
        customer_name=request.customer_name,
        customer_email=request.customer_email,
        customer_tax_office=request.customer_tax_office,
        customer_tax_number=request.customer_tax_number,
        customer_address=request.customer_address,
        items=invoice_items,
        subtotal=subtotal,
        total_vat=total_vat,
        vat_withholding=vat_withholding,
        total_additional_taxes=total_additional_taxes,
        total=total,
        due_date=datetime.fromisoformat(request.due_date),
        booking_id=request.booking_id,
        notes=request.notes,
        created_by=current_user.name
    )
    
    invoice_dict = invoice.model_dump()
    invoice_dict['issue_date'] = invoice_dict['issue_date'].isoformat()
    invoice_dict['due_date'] = invoice_dict['due_date'].isoformat()
    invoice_dict['created_at'] = invoice_dict['created_at'].isoformat()
    await db.accounting_invoices.insert_one(invoice_dict)
    
    # Create cash flow entry
    # CashFlow model imported at top
    cash_flow = CashFlow(
        tenant_id=current_user.tenant_id,
        transaction_type='income',
        category='room_revenue' if request.booking_id else 'other_services',
        amount=total,
        description=f"Invoice {invoice_number}",
        reference_id=invoice.id,
        reference_type='invoice',
        date=datetime.now(timezone.utc),
        created_by=current_user.name
    )
    cf_dict = cash_flow.model_dump()
    cf_dict['date'] = cf_dict['date'].isoformat()
    cf_dict['created_at'] = cf_dict['created_at'].isoformat()
    await db.cash_flow.insert_one(cf_dict)
    
    return invoice



@router.get("/accounting/invoices")
async def get_accounting_invoices(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    invoice_type: Optional[str] = None,
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {'tenant_id': current_user.tenant_id}
    if start_date and end_date:
        query['issue_date'] = {'$gte': start_date, '$lte': end_date}
    if invoice_type:
        query['invoice_type'] = invoice_type
    if status:
        query['status'] = status
    
    invoices = await db.accounting_invoices.find(query, {'_id': 0}).sort('issue_date', -1).to_list(1000)
    return invoices



@router.put("/accounting/invoices/{invoice_id}")
async def update_accounting_invoice(invoice_id: str, updates: Dict[str, Any], current_user: User = Depends(get_current_user)):
    if 'status' in updates and updates['status'] == 'paid' and 'payment_date' not in updates:
        updates['payment_date'] = datetime.now(timezone.utc).isoformat()
    
    await db.accounting_invoices.update_one({'id': invoice_id, 'tenant_id': current_user.tenant_id}, {'$set': updates})
    invoice = await db.accounting_invoices.find_one({'id': invoice_id}, {'_id': 0})
    return invoice


@router.get("/accounting/cash-flow")
async def get_cash_flow(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    transaction_type: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {'tenant_id': current_user.tenant_id}
    if start_date and end_date:
        query['date'] = {'$gte': start_date, '$lte': end_date}
    if transaction_type:
        query['transaction_type'] = transaction_type
    
    flows = await db.cash_flow.find(query, {'_id': 0}).sort('date', -1).to_list(1000)
    
    total_income = sum(f['amount'] for f in flows if f['transaction_type'] == 'income')
    total_expense = sum(f['amount'] for f in flows if f['transaction_type'] == 'expense')
    net_cash_flow = total_income - total_expense
    
    return {
        'transactions': flows,
        'total_income': total_income,
        'total_expense': total_expense,
        'net_cash_flow': net_cash_flow
    }


@router.get("/accounting/reports/profit-loss")
@cached(ttl=900, key_prefix="report_profit_loss")  # Cache for 15 min
async def get_profit_loss_report(
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user)
):
    # Get all income
    invoices = await db.accounting_invoices.find({
        'tenant_id': current_user.tenant_id,
        'status': 'paid',
        'issue_date': {'$gte': start_date, '$lte': end_date}
    }, {'_id': 0}).to_list(1000)
    
    # Get all expenses
    expenses = await db.expenses.find({
        'tenant_id': current_user.tenant_id,
        'date': {'$gte': start_date, '$lte': end_date}
    }, {'_id': 0}).to_list(1000)
    
    total_revenue = sum(inv['total'] for inv in invoices)
    total_expenses = sum(exp['total_amount'] for exp in expenses)
    gross_profit = total_revenue - total_expenses
    profit_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0
    
    # Revenue breakdown
    revenue_by_category = {}
    for inv in invoices:
        for item in inv['items']:
            desc = item['description']
            revenue_by_category[desc] = revenue_by_category.get(desc, 0) + item['total']
    
    # Expense breakdown
    expense_by_category = {}
    for exp in expenses:
        cat = exp['category']
        expense_by_category[cat] = expense_by_category.get(cat, 0) + exp['total_amount']
    
    return {
        'period': {'start': start_date, 'end': end_date},
        'total_revenue': round(total_revenue, 2),
        'total_expenses': round(total_expenses, 2),
        'gross_profit': round(gross_profit, 2),
        'profit_margin': round(profit_margin, 2),
        'revenue_breakdown': revenue_by_category,
        'expense_breakdown': expense_by_category
    }



@router.get("/accounting/reports/vat-report")
async def get_vat_report(
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user)
):
    # Sales VAT (collected)
    invoices = await db.accounting_invoices.find({
        'tenant_id': current_user.tenant_id,
        'issue_date': {'$gte': start_date, '$lte': end_date}
    }, {'_id': 0}).to_list(1000)
    
    sales_vat = sum(inv['total_vat'] for inv in invoices)
    
    # Purchase VAT (paid)
    expenses = await db.expenses.find({
        'tenant_id': current_user.tenant_id,
        'date': {'$gte': start_date, '$lte': end_date}
    }, {'_id': 0}).to_list(1000)
    
    purchase_vat = sum(exp['vat_amount'] for exp in expenses)
    
    vat_payable = sales_vat - purchase_vat
    
    return {
        'period': {'start': start_date, 'end': end_date},
        'sales_vat': round(sales_vat, 2),
        'purchase_vat': round(purchase_vat, 2),
        'vat_payable': round(vat_payable, 2)
    }



@router.get("/accounting/reports/balance-sheet")
async def get_balance_sheet(current_user: User = Depends(get_current_user)):
    # Assets
    bank_accounts = await db.bank_accounts.find({'tenant_id': current_user.tenant_id}, {'_id': 0}).to_list(1000)
    total_cash = sum(acc['balance'] for acc in bank_accounts)
    
    inventory = await db.inventory_items.find({'tenant_id': current_user.tenant_id}, {'_id': 0}).to_list(1000)
    total_inventory = sum(item['quantity'] * item['unit_cost'] for item in inventory)
    
    # Receivables (unpaid invoices)
    receivables = await db.accounting_invoices.find({
        'tenant_id': current_user.tenant_id,
        'status': {'$in': ['pending', 'partial']}
    }, {'_id': 0}).to_list(1000)
    total_receivables = sum(inv['total'] for inv in receivables)
    
    total_assets = total_cash + total_inventory + total_receivables
    
    # Liabilities
    payables = await db.expenses.find({
        'tenant_id': current_user.tenant_id,
        'payment_status': 'pending'
    }, {'_id': 0}).to_list(1000)
    total_payables = sum(exp['total_amount'] for exp in payables)
    
    # Equity
    total_equity = total_assets - total_payables
    
    return {
        'assets': {
            'cash': round(total_cash, 2),
            'inventory': round(total_inventory, 2),
            'receivables': round(total_receivables, 2),
            'total': round(total_assets, 2)
        },
        'liabilities': {
            'payables': round(total_payables, 2),
            'total': round(total_payables, 2)
        },
        'equity': {
            'total': round(total_equity, 2)
        }
    }



@router.get("/accounting/dashboard")
@cached(ttl=600, key_prefix="accounting_dashboard")  # Cache for 10 minutes
async def get_accounting_dashboard(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    current_user = await get_current_user(credentials)
    
    # Get current month data
    today = datetime.now(timezone.utc)
    month_start = today.replace(day=1, hour=0, minute=0, second=0).isoformat()
    month_end = today.isoformat()
    
    invoices = await db.accounting_invoices.find({
        'tenant_id': current_user.tenant_id,
        'issue_date': {'$gte': month_start, '$lte': month_end}
    }, {'_id': 0}).to_list(1000)
    
    expenses = await db.expenses.find({
        'tenant_id': current_user.tenant_id,
        'date': {'$gte': month_start, '$lte': month_end}
    }, {'_id': 0}).to_list(1000)
    
    total_income = sum(inv.get('total', 0) for inv in invoices if inv.get('status') == 'paid')
    total_expenses = sum(exp.get('amount', 0) for exp in expenses)
    pending_invoices = len([inv for inv in invoices if inv.get('status') == 'pending'])
    overdue_invoices = len([inv for inv in invoices if inv.get('status') == 'overdue'])
    
    # Get bank balances
    bank_accounts = await db.bank_accounts.find({'tenant_id': current_user.tenant_id}, {'_id': 0}).to_list(1000)
    total_bank_balance = sum(acc['balance'] for acc in bank_accounts)
    
    return {
        'monthly_income': round(total_income, 2),
        'monthly_expenses': round(total_expenses, 2),
        'net_income': round(total_income - total_expenses, 2),
        'pending_invoices': pending_invoices,
        'overdue_invoices': overdue_invoices,
        'total_bank_balance': round(total_bank_balance, 2)
    }



@router.get("/accounting/currencies")
async def get_currencies(current_user: User = Depends(get_current_user)):
    """Get all supported currencies"""
    currencies = [
        {'code': 'TRY', 'name': 'Turkish Lira', 'symbol': '₺'},
        {'code': 'USD', 'name': 'US Dollar', 'symbol': '$'},
        {'code': 'EUR', 'name': 'Euro', 'symbol': '€'},
        {'code': 'GBP', 'name': 'British Pound', 'symbol': '£'}
    ]
    return {'currencies': currencies}


@router.post("/accounting/currency-rates")
async def create_currency_rate(
    request: CreateCurrencyRateRequest,
    current_user: User = Depends(get_current_user)
):
    """Create or update currency exchange rate"""
    rate = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'from_currency': request.from_currency,
        'to_currency': request.to_currency,
        'rate': request.rate,
        'effective_date': request.effective_date,
        'created_at': datetime.now(timezone.utc).isoformat(),
        'created_by': current_user.id
    }
    
    rate_copy = rate.copy()
    await db.currency_rates.insert_one(rate_copy)
    return rate


@router.get("/accounting/currency-rates")
async def get_currency_rates(
    from_currency: str = None,
    to_currency: str = None,
    date: str = None,
    current_user: User = Depends(get_current_user)
):
    """Get currency exchange rates"""
    query = {'tenant_id': current_user.tenant_id}
    
    if from_currency:
        query['from_currency'] = from_currency
    if to_currency:
        query['to_currency'] = to_currency
    if date:
        query['effective_date'] = {'$lte': date}
    
    rates = await db.currency_rates.find(
        query,
        {'_id': 0}
    ).sort('effective_date', -1).to_list(100)
    
    return {'rates': rates, 'count': len(rates)}


@router.post("/accounting/convert-currency")
async def convert_currency(
    request: ConvertCurrencyRequest,
    current_user: User = Depends(get_current_user)
):
    """Convert amount between currencies"""
    # If same currency, no conversion needed
    if request.from_currency == request.to_currency:
        return {
            'amount': request.amount,
            'from_currency': request.from_currency,
            'to_currency': request.to_currency,
            'rate': 1.0,
            'converted_amount': request.amount
        }
    
    # Get exchange rate
    query = {
        'tenant_id': current_user.tenant_id,
        'from_currency': request.from_currency,
        'to_currency': request.to_currency
    }
    
    if request.date:
        query['effective_date'] = {'$lte': request.date}
    
    rate_record = await db.currency_rates.find_one(
        query,
        {'_id': 0},
        sort=[('effective_date', -1)]
    )
    
    if not rate_record:
        # Try reverse rate
        reverse_query = {
            'tenant_id': current_user.tenant_id,
            'from_currency': request.to_currency,
            'to_currency': request.from_currency
        }
        if request.date:
            reverse_query['effective_date'] = {'$lte': request.date}
        
        reverse_rate = await db.currency_rates.find_one(
            reverse_query,
            {'_id': 0},
            sort=[('effective_date', -1)]
        )
        
        if reverse_rate:
            rate = 1.0 / reverse_rate['rate']
        else:
            # Default rates if not found
            default_rates = {
                ('TRY', 'USD'): 0.037,
                ('TRY', 'EUR'): 0.034,
                ('USD', 'TRY'): 27.0,
                ('EUR', 'TRY'): 29.5,
                ('USD', 'EUR'): 0.92,
                ('EUR', 'USD'): 1.09
            }
            rate = default_rates.get((request.from_currency, request.to_currency), 1.0)
    else:
        rate = rate_record['rate']
    
    converted_amount = request.amount * rate
    
    return {
        'amount': request.amount,
        'from_currency': request.from_currency,
        'to_currency': request.to_currency,
        'rate': round(rate, 4),
        'converted_amount': round(converted_amount, 2),
        'date': request.date or datetime.now(timezone.utc).date().isoformat()
    }


@router.post("/accounting/invoices/multi-currency")
async def create_multi_currency_invoice(
    request: CreateMultiCurrencyInvoiceRequest,
    current_user: User = Depends(get_current_user)
):
    """Create invoice in any currency with auto-conversion to TRY"""
    # Calculate totals in invoice currency
    subtotal = sum(item.get('quantity', 0) * item.get('unit_price', 0) for item in request.items)
    
    # Calculate VAT
    total_vat = 0
    for item in request.items:
        item_total = item.get('quantity', 0) * item.get('unit_price', 0)
        vat_rate = item.get('vat_rate', 18) / 100
        item['vat_amount'] = round(item_total * vat_rate, 2)
        total_vat += item['vat_amount']
    
    total = subtotal + total_vat
    
    # Convert to TRY if needed
    if request.currency != 'TRY':
        if request.exchange_rate:
            rate = request.exchange_rate
        else:
            # Get current rate
            conversion = await convert_currency(
                ConvertCurrencyRequest(
                    amount=1.0,
                    from_currency=request.currency,
                    to_currency='TRY'
                ),
                current_user
            )
            rate = conversion['rate']
        
        subtotal_try = subtotal * rate
        total_vat_try = total_vat * rate
        total_try = total * rate
    else:
        rate = 1.0
        subtotal_try = subtotal
        total_vat_try = total_vat
        total_try = total
    
    invoice_number = f"INV-{datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:8].upper()}"
    
    invoice = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'invoice_number': invoice_number,
        'customer_name': request.customer_name,
        'customer_email': request.customer_email,
        'customer_address': request.customer_address,
        'items': request.items,
        'currency': request.currency,
        'exchange_rate': rate,
        'subtotal': round(subtotal, 2),
        'total_vat': round(total_vat, 2),
        'total': round(total, 2),
        'subtotal_try': round(subtotal_try, 2),
        'total_vat_try': round(total_vat_try, 2),
        'total_try': round(total_try, 2),
        'payment_terms': request.payment_terms,
        'notes': request.notes,
        'issue_date': datetime.now(timezone.utc).date().isoformat(),
        'due_date': (datetime.now(timezone.utc) + timedelta(days=30)).date().isoformat(),
        'status': 'pending',
        'created_at': datetime.now(timezone.utc).isoformat(),
        'created_by': current_user.id
    }
    
    invoice_copy = invoice.copy()
    await db.accounting_invoices.insert_one(invoice_copy)
    
    return invoice



@router.post("/accounting/invoices/from-folio")
async def generate_invoice_from_folio(
    request: GenerateInvoiceFromFolioRequest,
    current_user: User = Depends(get_current_user)
):
    """Generate accounting invoice from PMS folio"""
    # Get folio
    folio = await db.folios.find_one({
        'id': request.folio_id,
        'tenant_id': current_user.tenant_id
    }, {'_id': 0})
    
    if not folio:
        raise HTTPException(status_code=404, detail="Folio not found")
    
    # Get folio charges
    charges = await db.folio_charges.find({
        'folio_id': request.folio_id,
        'tenant_id': current_user.tenant_id,
        'voided': False
    }, {'_id': 0}).to_list(1000)
    
    # Get booking info
    booking = await db.bookings.find_one({
        'folio_id': request.folio_id,
        'tenant_id': current_user.tenant_id
    }, {'_id': 0})
    
    # Convert charges to invoice items
    invoice_items = []
    for charge in charges:
        item = {
            'description': charge.get('description', 'Hotel Charge'),
            'quantity': 1,
            'unit_price': charge.get('amount', 0),
            'vat_rate': charge.get('vat_rate', 18),
            'total': charge.get('total', 0)
        }
        invoice_items.append(item)
    
    # Get customer info from booking or folio
    customer_name = booking.get('guest_name') if booking else folio.get('guest_name', 'Guest')
    customer_email = booking.get('guest_email') if booking else folio.get('guest_email', '')
    
    # Create invoice
    invoice_number = f"INV-{datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:8].upper()}"
    
    # Calculate totals
    subtotal = sum(item['unit_price'] * item['quantity'] for item in invoice_items)
    total_vat = sum(item['unit_price'] * item['quantity'] * (item['vat_rate'] / 100) for item in invoice_items)
    
    # Currency conversion if needed
    if request.invoice_currency != 'TRY':
        conversion = await convert_currency(
            ConvertCurrencyRequest(
                amount=subtotal + total_vat,
                from_currency='TRY',
                to_currency=request.invoice_currency
            ),
            current_user
        )
        exchange_rate = conversion['rate']
        total_foreign = conversion['converted_amount']
    else:
        exchange_rate = 1.0
        total_foreign = subtotal + total_vat
    
    invoice = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'invoice_number': invoice_number,
        'folio_id': request.folio_id,
        'booking_id': booking['id'] if booking else None,
        'customer_name': customer_name,
        'customer_email': customer_email,
        'customer_address': booking.get('guest_address', '') if booking else '',
        'items': invoice_items,
        'currency': request.invoice_currency,
        'exchange_rate': exchange_rate,
        'subtotal': round(subtotal, 2),
        'total_vat': round(total_vat, 2),
        'total': round(subtotal + total_vat, 2),
        'total_foreign_currency': round(total_foreign, 2),
        'payment_terms': 'Due on checkout',
        'issue_date': datetime.now(timezone.utc).date().isoformat(),
        'due_date': datetime.now(timezone.utc).date().isoformat(),
        'status': 'pending',
        'source': 'pms_folio',
        'created_at': datetime.now(timezone.utc).isoformat(),
        'created_by': current_user.id
    }
    
    invoice_copy = invoice.copy()
    await db.accounting_invoices.insert_one(invoice_copy)
    
    # Update folio with invoice reference
    await db.folios.update_one(
        {'id': request.folio_id},
        {'$set': {'invoice_id': invoice['id'], 'invoice_number': invoice_number}}
    )
    
    # Generate E-Fatura if requested
    if request.include_efatura:
        efatura_xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<Invoice xmlns="urn:oasis:names:specification:ubl:schema:xsd:Invoice-2">
    <ID>{invoice_number}</ID>
    <IssueDate>{invoice['issue_date']}</IssueDate>
    <InvoiceTypeCode>SATIS</InvoiceTypeCode>
    <DocumentCurrencyCode>{request.invoice_currency}</DocumentCurrencyCode>
    <LineCountNumeric>{len(invoice_items)}</LineCountNumeric>
    <LegalMonetaryTotal>
        <TaxExclusiveAmount currencyID="{request.invoice_currency}">{invoice['subtotal']}</TaxExclusiveAmount>
        <TaxInclusiveAmount currencyID="{request.invoice_currency}">{invoice['total']}</TaxInclusiveAmount>
    </LegalMonetaryTotal>
</Invoice>"""
        
        efatura_record = {
            'id': str(uuid.uuid4()),
            'tenant_id': current_user.tenant_id,
            'invoice_id': invoice['id'],
            'invoice_number': invoice_number,
            'efatura_uuid': str(uuid.uuid4()),
            'xml_content': efatura_xml,
            'status': 'generated',
            'generated_at': datetime.now(timezone.utc).isoformat()
        }
        
        efatura_copy = efatura_record.copy()
        await db.efatura_records.insert_one(efatura_copy)
        
        invoice['efatura_uuid'] = efatura_record['efatura_uuid']
        invoice['efatura_status'] = 'generated'
    
    return {
        'invoice': invoice,
        'message': 'Invoice generated from folio successfully',
        'efatura_generated': request.include_efatura
    }



@router.get("/accounting/invoices/{invoice_id}/efatura-status")
async def get_invoice_efatura_status(
    invoice_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get E-Fatura status for accounting invoice"""
    invoice = await db.accounting_invoices.find_one({
        'id': invoice_id,
        'tenant_id': current_user.tenant_id
    }, {'_id': 0})
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Get E-Fatura record
    efatura = await db.efatura_records.find_one({
        'invoice_id': invoice_id,
        'tenant_id': current_user.tenant_id
    }, {'_id': 0})
    
    if not efatura:
        return {
            'invoice_id': invoice_id,
            'invoice_number': invoice.get('invoice_number'),
            'efatura_status': 'not_generated',
            'message': 'E-Fatura has not been generated for this invoice'
        }
    
    return {
        'invoice_id': invoice_id,
        'invoice_number': invoice.get('invoice_number'),
        'efatura_uuid': efatura.get('efatura_uuid'),
        'efatura_status': efatura.get('status'),
        'generated_at': efatura.get('generated_at'),
        'sent_at': efatura.get('sent_at'),
        'gib_response': efatura.get('gib_response')
    }


@router.post("/accounting/invoices/{invoice_id}/generate-efatura")
async def generate_efatura_for_invoice(
    invoice_id: str,
    current_user: User = Depends(get_current_user)
):
    """Generate E-Fatura for existing accounting invoice"""
    invoice = await db.accounting_invoices.find_one({
        'id': invoice_id,
        'tenant_id': current_user.tenant_id
    }, {'_id': 0})
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Check if E-Fatura already exists
    existing_efatura = await db.efatura_records.find_one({
        'invoice_id': invoice_id,
        'tenant_id': current_user.tenant_id
    })
    
    if existing_efatura:
        return {
            'message': 'E-Fatura already exists for this invoice',
            'efatura_uuid': existing_efatura.get('efatura_uuid'),
            'status': existing_efatura.get('status')
        }
    
    # Generate E-Fatura XML
    currency = invoice.get('currency', 'TRY')
    efatura_xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<Invoice xmlns="urn:oasis:names:specification:ubl:schema:xsd:Invoice-2">
    <ID>{invoice.get('invoice_number')}</ID>
    <IssueDate>{invoice.get('issue_date')}</IssueDate>
    <InvoiceTypeCode>SATIS</InvoiceTypeCode>
    <DocumentCurrencyCode>{currency}</DocumentCurrencyCode>
    <LineCountNumeric>{len(invoice.get('items', []))}</LineCountNumeric>
    <AccountingSupplierParty>
        <Party>
            <PartyName>
                <Name>Hotel Name</Name>
            </PartyName>
        </Party>
    </AccountingSupplierParty>
    <AccountingCustomerParty>
        <Party>
            <PartyName>
                <Name>{invoice.get('customer_name', 'N/A')}</Name>
            </PartyName>
        </Party>
    </AccountingCustomerParty>
    <LegalMonetaryTotal>
        <TaxExclusiveAmount currencyID="{currency}">{invoice.get('subtotal', 0)}</TaxExclusiveAmount>
        <TaxInclusiveAmount currencyID="{currency}">{invoice.get('total', 0)}</TaxInclusiveAmount>
    </LegalMonetaryTotal>
</Invoice>"""
    
    efatura_record = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'invoice_id': invoice_id,
        'invoice_number': invoice.get('invoice_number'),
        'efatura_uuid': str(uuid.uuid4()),
        'xml_content': efatura_xml,
        'status': 'generated',
        'generated_at': datetime.now(timezone.utc).isoformat()
    }
    
    efatura_copy = efatura_record.copy()
    await db.efatura_records.insert_one(efatura_copy)
    
    # Update invoice with E-Fatura reference
    await db.accounting_invoices.update_one(
        {'id': invoice_id},
        {
            '$set': {
                'efatura_uuid': efatura_record['efatura_uuid'],
                'efatura_status': 'generated'
            }
        }
    )
    
    return {
        'message': 'E-Fatura generated successfully',
        'efatura_uuid': efatura_record['efatura_uuid'],
        'invoice_number': invoice.get('invoice_number')
    }



@router.get("/efatura/invoices")
async def get_efatura_invoices(current_user: User = Depends(get_current_user)):
    invoices = await db.invoices.find({
        'tenant_id': current_user.tenant_id
    }, {'_id': 0}).sort('created_at', -1).limit(50).to_list(50)
    
    # Add efatura status to each invoice
    for invoice in invoices:
        invoice['efatura_status'] = invoice.get('efatura_status', 'pending')
    
    return {'invoices': invoices}


@router.get("/efatura/settings")
async def get_efatura_settings(current_user: User = Depends(get_current_user)):
    settings = await db.efatura_settings.find_one({'tenant_id': current_user.tenant_id}, {'_id': 0})
    return settings or {'vkn': '1234567890', 'enabled': True, 'auto_send': False, 'last_sync': None}


@router.post("/efatura/send/{invoice_id}")
async def send_efatura(
    invoice_id: str,
    current_user: User = Depends(get_current_user)
):
    # Update invoice status
    await db.invoices.update_one(
        {'id': invoice_id},
        {'$set': {
            'efatura_status': 'sent',
            'efatura_sent_at': datetime.now(timezone.utc).isoformat()
        }}
    )
    return {'message': 'E-Fatura sent successfully'}


@router.post("/efatura/generate/{invoice_id}")
async def generate_efatura(
    invoice_id: str,
    current_user: User = Depends(get_current_user)
):
    """Generate E-Fatura XML for GIB"""
    invoice = await db.accounting_invoices.find_one(
        {'id': invoice_id, 'tenant_id': current_user.tenant_id},
        {'_id': 0}
    )
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Generate E-Fatura XML (simplified)
    efatura_xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<Invoice xmlns="urn:oasis:names:specification:ubl:schema:xsd:Invoice-2">
    <ID>{invoice['invoice_number']}</ID>
    <IssueDate>{invoice['invoice_date']}</IssueDate>
    <InvoiceTypeCode>SATIS</InvoiceTypeCode>
    <LineCountNumeric>{len(invoice.get('items', []))}</LineCountNumeric>
    <LegalMonetaryTotal>
        <TaxExclusiveAmount>{invoice.get('subtotal', 0)}</TaxExclusiveAmount>
        <TaxInclusiveAmount>{invoice.get('grand_total', 0)}</TaxInclusiveAmount>
    </LegalMonetaryTotal>
</Invoice>"""
    
    # Save E-Fatura record
    efatura_record = {
        'id': str(uuid.uuid4()),
        'tenant_id': current_user.tenant_id,
        'invoice_id': invoice_id,
        'invoice_number': invoice['invoice_number'],
        'efatura_uuid': str(uuid.uuid4()),
        'xml_content': efatura_xml,
        'status': 'generated',
        'generated_at': datetime.now(timezone.utc).isoformat()
    }
    
    efatura_copy = efatura_record.copy()
    await db.efatura_records.insert_one(efatura_copy)
    
    # Update invoice status
    await db.accounting_invoices.update_one(
        {'id': invoice_id},
        {'$set': {'efatura_status': 'generated', 'efatura_uuid': efatura_record['efatura_uuid']}}
    )
    
    return {
        'message': 'E-Fatura generated successfully',
        'efatura_uuid': efatura_record['efatura_uuid'],
        'xml_content': efatura_xml
    }


@router.post("/efatura/send-to-gib/{invoice_id}")
async def send_efatura_to_gib(
    invoice_id: str,
    current_user: User = Depends(get_current_user)
):
    """Send E-Fatura to GIB (Turkish Revenue Administration)"""
    efatura = await db.efatura_records.find_one(
        {'invoice_id': invoice_id, 'tenant_id': current_user.tenant_id},
        {'_id': 0}
    )
    
    if not efatura:
        raise HTTPException(status_code=404, detail="E-Fatura not found")
    
    # Mock GIB integration (in production, use actual GIB API)
    gib_response = {
        'status': 'success',
        'gib_id': str(uuid.uuid4()),
        'timestamp': datetime.now(timezone.utc).isoformat()
    }
    
    # Update E-Fatura status
    await db.efatura_records.update_one(
        {'id': efatura['id']},
        {
            '$set': {
                'status': 'sent_to_gib',
                'gib_response': gib_response,
                'sent_at': datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    await db.accounting_invoices.update_one(
        {'id': invoice_id},
        {'$set': {'efatura_status': 'sent', 'efatura_sent_at': datetime.now(timezone.utc).isoformat()}}
    )
    
    return {'message': 'E-Fatura sent to GIB successfully', 'gib_response': gib_response}


@router.post("/accounting/send-statement")
async def send_statement_email(
    company_id: str,
    email: Optional[str] = None,
    include_details: bool = True,
    current_user: User = Depends(get_current_user)
):
    """
    Send account statement to company with one click
    - Outstanding balance
    - Invoice details
    - Payment reminder
    """
    company = await db.companies.find_one({
        'id': company_id,
        'tenant_id': current_user.tenant_id
    })
    
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Get all open folios for company
    folios = []
    total_balance = 0
    async for folio in db.folios.find({
        'company_id': company_id,
        'tenant_id': current_user.tenant_id,
        'status': 'open'
    }):
        balance = folio.get('balance', 0)
        total_balance += balance
        folios.append({
            'folio_number': folio.get('folio_number'),
            'booking_id': folio.get('booking_id'),
            'balance': balance,
            'created_at': folio.get('created_at')
        })
    
    recipient_email = email or company.get('contact_email')
    
    if not recipient_email:
        raise HTTPException(status_code=400, detail="No email address provided")
    
    # Create statement document
    statement = {
        'company_name': company.get('name'),
        'statement_date': datetime.now(timezone.utc).isoformat(),
        'total_outstanding': round(total_balance, 2),
        'folios': folios,
        'payment_terms': company.get('payment_terms', 'Net 30'),
        'contact_person': company.get('contact_person')
    }
    
    # In production, send actual email via SMTP or email service
    # For now, simulate email sending
    
    return {
        'success': True,
        'message': f'Statement sent to {recipient_email}',
        'statement': statement,
        'note': 'In production, integrate with SendGrid, AWS SES, or SMTP server'
    }



@router.get("/accounting/smart-alerts")
async def get_smart_ar_alerts(
    current_user: User = Depends(get_current_user)
):
    """
    Smart AR/Collections alerts
    - Overdue invoices by company
    - Payment pattern analysis
    - Risk assessment
    """
    alerts = []
    
    # Get all companies with outstanding balances
    companies = []
    async for company in db.companies.find({
        'tenant_id': current_user.tenant_id,
        'status': 'active'
    }):
        # Get open folios
        total_balance = 0
        overdue_count = 0
        oldest_invoice_days = 0
        
        async for folio in db.folios.find({
            'company_id': company.get('id'),
            'tenant_id': current_user.tenant_id,
            'status': 'open'
        }):
            balance = folio.get('balance', 0)
            total_balance += balance
            
            # Check if overdue (based on payment terms)
            created_at = datetime.fromisoformat(folio.get('created_at'))
            days_old = (datetime.now(timezone.utc) - created_at).days
            
            # Default: Net 30 payment terms
            payment_terms_days = 30
            if company.get('payment_terms'):
                if 'Net 15' in company.get('payment_terms'): payment_terms_days = 15
                elif 'Net 45' in company.get('payment_terms'): payment_terms_days = 45
                elif 'Net 60' in company.get('payment_terms'): payment_terms_days = 60
            
            if days_old > payment_terms_days:
                overdue_count += 1
                oldest_invoice_days = max(oldest_invoice_days, days_old)
        
        if total_balance > 0:
            companies.append({
                'company_id': company.get('id'),
                'company_name': company.get('name'),
                'total_balance': total_balance,
                'overdue_invoices': overdue_count,
                'oldest_invoice_days': oldest_invoice_days
            })
    
    # Generate alerts
    for company in companies:
        if company['overdue_invoices'] >= 10:
            alerts.append({
                'type': 'critical',
                'priority': 'urgent',
                'icon': '🚨',
                'title': f"{company['company_name']} has {company['overdue_invoices']} overdue invoices",
                'description': f"Total outstanding: ${round(company['total_balance'], 2)}. Oldest invoice: {company['oldest_invoice_days']} days",
                'action': 'send_statement',
                'company_id': company['company_id']
            })
        elif company['overdue_invoices'] > 0:
            alerts.append({
                'type': 'warning',
                'priority': 'high',
                'icon': '⚠️',
                'title': f"{company['company_name']} - {company['overdue_invoices']} overdue invoices",
                'description': f"Outstanding: ${round(company['total_balance'], 2)}",
                'action': 'send_reminder',
                'company_id': company['company_id']
            })
        elif company['total_balance'] > 10000:
            alerts.append({
                'type': 'info',
                'priority': 'normal',
                'icon': 'ℹ️',
                'title': f"{company['company_name']} - High balance",
                'description': f"Outstanding: ${round(company['total_balance'], 2)}. Monitor payment",
                'action': 'monitor',
                'company_id': company['company_id']
            })
    
    # Sort by priority
    priority_order = {'urgent': 0, 'high': 1, 'normal': 2, 'low': 3}
    alerts.sort(key=lambda x: priority_order.get(x['priority'], 2))
    
    return {
        'total_alerts': len(alerts),
        'critical_count': sum(1 for a in alerts if a['type'] == 'critical'),
        'warning_count': sum(1 for a in alerts if a['type'] == 'warning'),
        'alerts': alerts
    }



class RecordPaymentRequest(BaseModel):
    folio_id: str
    amount: float
    payment_method: str
    notes: Optional[str] = None


@router.get("/finance/mobile/daily-collections")
async def get_daily_collections_mobile(
    date: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get daily collections for finance mobile dashboard"""
    current_user = await get_current_user(credentials)
    
    if date:
        target_date = datetime.fromisoformat(date)
    else:
        target_date = datetime.now(timezone.utc)
    
    start_of_day = target_date.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_day = target_date.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # Get payments for the day
    total_collected = 0.0
    payment_count = 0
    payment_methods = {}
    
    async for payment in db.payments.find({
        'tenant_id': current_user.tenant_id,
        'created_at': {
            '$gte': start_of_day,
            '$lte': end_of_day
        }
    }):
        amount = payment.get('amount', 0)
        total_collected += amount
        payment_count += 1
        
        method = payment.get('payment_method', 'unknown')
        payment_methods[method] = payment_methods.get(method, 0) + amount
    
    return {
        'date': target_date.date().isoformat(),
        'total_collected': total_collected,
        'payment_count': payment_count,
        'payment_methods': payment_methods,
        'average_transaction': total_collected / payment_count if payment_count > 0 else 0
    }



@router.get("/finance/mobile/monthly-collections")
async def get_monthly_collections_mobile(
    year: Optional[int] = None,
    month: Optional[int] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get monthly collections for finance mobile dashboard"""
    current_user = await get_current_user(credentials)
    
    today = datetime.now(timezone.utc)
    target_year = year or today.year
    target_month = month or today.month
    
    # First day of month
    start_of_month = datetime(target_year, target_month, 1, tzinfo=timezone.utc)
    
    # First day of next month
    if target_month == 12:
        end_of_month = datetime(target_year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        end_of_month = datetime(target_year, target_month + 1, 1, tzinfo=timezone.utc)
    
    # Get payments for the month
    total_collected = 0.0
    payments_by_method = {}
    
    async for payment in db.payments.find({
        'tenant_id': current_user.tenant_id,
        'created_at': {'$gte': start_of_month.isoformat(), '$lt': end_of_month.isoformat()}
    }):
        amount = payment.get('amount', 0)
        total_collected += amount
        
        method = payment.get('payment_method', 'unknown')
        payments_by_method[method] = payments_by_method.get(method, 0) + amount
    
    return {
        'total_collected': round(total_collected, 2),
        'month': target_month,
        'year': target_year,
        'payments_by_method': {k: round(v, 2) for k, v in payments_by_method.items()},
        'currency': 'TRY'
    }



@router.get("/finance/profit-loss")
async def get_profit_loss_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get Profit & Loss (P&L) report"""
    await get_current_user(credentials)
    
    if not start_date:
        # Default to current month
        start_date = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0)
    else:
        start_date = datetime.fromisoformat(start_date)
    
    if not end_date:
        # End of current month
        if start_date.month == 12:
            end_date = datetime(start_date.year + 1, 1, 1, tzinfo=timezone.utc) - timedelta(days=1)
        else:
            end_date = datetime(start_date.year, start_date.month + 1, 1, tzinfo=timezone.utc) - timedelta(days=1)
    else:
        end_date = datetime.fromisoformat(end_date)
    
    # REVENUE
    # Room Revenue



@router.get("/finance/cashier-shift-report")
async def get_cashier_shift_report(
    shift_date: Optional[str] = None,
    cashier_name: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get cashier shift report"""
    current_user = await get_current_user(credentials)
    
    if shift_date:
        target_date = datetime.fromisoformat(shift_date)
    else:
        target_date = datetime.now(timezone.utc)
    
    start_of_day = target_date.replace(hour=0, minute=0, second=0)
    end_of_day = target_date.replace(hour=23, minute=59, second=59)
    
    query = {
        'tenant_id': current_user.tenant_id,
        'created_at': {'$gte': start_of_day, '$lte': end_of_day}
    }
    
    if cashier_name:
        query['created_by'] = cashier_name
    
    # Get all payments/transactions
    total_cash = 0
    total_card = 0
    total_transfer = 0
    total_other = 0
    transaction_count = 0
    
    async for payment in db.payments.find(query):
        amount = payment.get('amount', 0)
        method = payment.get('payment_method', 'cash')
        
        if method == 'cash':
            total_cash += amount
        elif method == 'card':
            total_card += amount
        elif method == 'transfer':
            total_transfer += amount
        else:
            total_other += amount
        
        transaction_count += 1
    
    total_collected = total_cash + total_card + total_transfer + total_other
    
    # Get opening and closing balance (if tracked)
    opening_balance = 0  # Should be from shift start record
    expected_closing = opening_balance + total_cash
    
    # Calculate variances
    variance = 0  # Would be: actual_closing - expected_closing
    
    return {
        'shift_date': target_date.date().isoformat(),
        'cashier_name': cashier_name or 'All Cashiers',
        'opening_balance': opening_balance,
        'collections': {
            'cash': total_cash,
            'card': total_card,
            'transfer': total_transfer,
            'other': total_other,
            'total': total_collected
        },
        'expected_closing_balance': expected_closing,
        'variance': variance,
        'transaction_count': transaction_count,
        'average_transaction': total_collected / transaction_count if transaction_count > 0 else 0,
        'generated_at': datetime.now(timezone.utc).isoformat()
    }

    payment_count = 0
    
    async for payment in db.payments.find({
        'tenant_id': current_user.tenant_id,
        'created_at': {
            '$gte': start_of_month,
            '$lt': end_of_month
        }
    }):
        total_collected += payment.get('amount', 0)
        payment_count += 1
    
    # Calculate collection rate (collected vs expected)
    total_expected = 0.0
    async for booking in db.bookings.find({
        'tenant_id': current_user.tenant_id,
        'check_in': {
            '$gte': start_of_month,
            '$lt': end_of_month
        }
    }):
        total_expected += booking.get('total_amount', 0)
    
    collection_rate = (total_collected / total_expected * 100) if total_expected > 0 else 0
    
    return {
        'year': target_year,
        'month': target_month,
        'total_collected': total_collected,
        'payment_count': payment_count,
        'total_expected': total_expected,
        'collection_rate': collection_rate,
        'outstanding': total_expected - total_collected
    }



@router.get("/finance/mobile/pending-receivables")
async def get_pending_receivables_mobile(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get pending receivables for finance mobile dashboard"""
    current_user = await get_current_user(credentials)
    
    # Get all open folios with balance
    total_pending = 0.0
    overdue_amount = 0.0
    receivables = []
    
    today = datetime.now(timezone.utc)
    
    async for folio in db.folios.find({
        'tenant_id': current_user.tenant_id,
        'status': 'open',
        'balance': {'$gt': 0}
    }):
        balance = folio.get('balance', 0)
        total_pending += balance
        
        # Get booking info
        booking = await db.bookings.find_one({
            'id': folio.get('booking_id'),
            'tenant_id': current_user.tenant_id
        })
        
        is_overdue = False
        checkout_date_str = None
        
        if booking:
            checkout = booking.get('check_out')
            if checkout:
                try:
                    # Convert string to datetime for comparison
                    if isinstance(checkout, str):
                        checkout_dt = datetime.fromisoformat(checkout).replace(tzinfo=timezone.utc)
                        checkout_date_str = checkout
                    else:
                        checkout_dt = checkout if checkout.tzinfo else checkout.replace(tzinfo=timezone.utc)
                        checkout_date_str = checkout.isoformat()
                    
                    if checkout_dt < today:
                        is_overdue = True
                        overdue_amount += balance
                except (ValueError, AttributeError):
                    pass
        
        receivables.append({
            'folio_id': folio.get('id'),
            'folio_number': folio.get('folio_number'),
            'guest_name': booking.get('guest_name') if booking else 'Unknown',
            'balance': balance,
            'is_overdue': is_overdue,
            'checkout_date': checkout_date_str,
            'created_at': folio.get('created_at').isoformat() if isinstance(folio.get('created_at'), datetime) else folio.get('created_at')
        })
    
    # Sort by amount (highest first)
    receivables.sort(key=lambda x: x['balance'], reverse=True)
    
    return {
        'total_pending': total_pending,
        'overdue_amount': overdue_amount,
        'receivables_count': len(receivables),
        'receivables': receivables[:20]  # Top 20
    }



@router.get("/finance/mobile/monthly-costs")
async def get_monthly_costs_mobile(
    year: Optional[int] = None,
    month: Optional[int] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get monthly costs for finance mobile dashboard"""
    current_user = await get_current_user(credentials)
    
    today = datetime.now(timezone.utc)
    target_year = year or today.year
    target_month = month or today.month
    
    # First day of month
    start_of_month = datetime(target_year, target_month, 1, tzinfo=timezone.utc)
    
    # First day of next month
    if target_month == 12:
        end_of_month = datetime(target_year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        end_of_month = datetime(target_year, target_month + 1, 1, tzinfo=timezone.utc)
    
    # Get expenses for the month
    total_costs = 0.0
    costs_by_category = {}
    
    async for expense in db.expenses.find({
        'tenant_id': current_user.tenant_id,
        'expense_date': {
            '$gte': start_of_month,
            '$lt': end_of_month
        }
    }):
        amount = expense.get('amount', 0)
        total_costs += amount
        
        category = expense.get('category', 'other')
        costs_by_category[category] = costs_by_category.get(category, 0) + amount
    
    return {
        'year': target_year,
        'month': target_month,
        'total_costs': total_costs,
        'costs_by_category': costs_by_category
    }



@router.post("/finance/mobile/record-payment")
async def record_payment_mobile(
    request: RecordPaymentRequest,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Record a payment from finance mobile"""
    current_user = await get_current_user(credentials)
    folio_id = request.folio_id
    amount = request.amount
    payment_method = request.payment_method
    notes = request.notes
    
    # Validate folio
    folio = await db.folios.find_one({
        'id': folio_id,
        'tenant_id': current_user.tenant_id
    })
    
    if not folio:
        raise HTTPException(status_code=404, detail="Folio not found")
    
    # Create payment
    payment_id = str(uuid.uuid4())
    payment = {
        'id': payment_id,
        'tenant_id': current_user.tenant_id,
        'folio_id': folio_id,
        'booking_id': folio.get('booking_id'),
        'amount': amount,
        'payment_method': payment_method,
        'payment_type': 'final',
        'notes': notes,
        'created_at': datetime.now(timezone.utc),
        'created_by': current_user.username
    }
    
    await db.payments.insert_one(payment)
    
    # Update folio balance
    new_balance = folio.get('balance', 0) - amount
    await db.folios.update_one(
        {'id': folio_id, 'tenant_id': current_user.tenant_id},
        {'$set': {'balance': new_balance}}
    )
    
    # Close folio if balance is zero
    if abs(new_balance) < 0.01:
        await db.folios.update_one(
            {'id': folio_id, 'tenant_id': current_user.tenant_id},
            {
                '$set': {
                    'status': 'closed',
                    'closed_at': datetime.now(timezone.utc),
                    'closed_by': current_user.username
                }
            }
        )
    
    return {
        'message': 'Payment recorded successfully',
        'payment_id': payment_id,
        'folio_id': folio_id,
        'amount': amount,
        'new_balance': new_balance,
        'folio_closed': abs(new_balance) < 0.01
    }



@router.get("/finance/mobile/cash-flow-summary")
async def get_cash_flow_summary_mobile(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get cash flow summary for finance mobile dashboard
    - Today's cash inflow (tahsilat)
    - Today's cash outflow (gider)
    - Weekly collection/payment plan
    - Bank balance summaries
    """
    current_user = await get_current_user(credentials)
    today = datetime.now(timezone.utc).date()
    start_of_day = datetime.combine(today, datetime.min.time()).replace(tzinfo=timezone.utc)
    end_of_day = datetime.combine(today, datetime.max.time()).replace(tzinfo=timezone.utc)
    
    # Today's cash inflow (payments received)
    today_inflow = 0.0
    inflow_count = 0
    async for payment in db.payments.find({
        'tenant_id': current_user.tenant_id,
        'created_at': {'$gte': start_of_day, '$lte': end_of_day}
    }):
        today_inflow += payment.get('amount', 0)
        inflow_count += 1
    
    # Today's cash outflow (expenses)
    today_outflow = 0.0
    outflow_count = 0
    async for expense in db.expenses.find({
        'tenant_id': current_user.tenant_id,
        'date': {'$gte': start_of_day, '$lte': end_of_day},
        'paid': True
    }):
        today_outflow += expense.get('amount', 0)
        outflow_count += 1
    
    # Net cash flow today
    net_flow = today_inflow - today_outflow
    
    # Weekly collection plan (next 7 days expected collections)
    weekly_plan = []
    for days_ahead in range(7):
        target_date = today + timedelta(days=days_ahead)
        
        # Expected checkouts (potential collections)
        expected_collections = 0.0
        checkout_count = 0
        async for booking in db.bookings.find({
            'tenant_id': current_user.tenant_id,
            'check_out': target_date.isoformat(),
            'status': {'$in': ['confirmed', 'guaranteed', 'checked_in']}
        }):
            # Get folio balance for this booking
            folio = await db.folios.find_one({
                'tenant_id': current_user.tenant_id,
                'booking_id': booking.get('id'),
                'status': 'open'
            })
            if folio:
                expected_collections += folio.get('balance', 0)
                checkout_count += 1
        
        # Expected payments from companies (due date)
        expected_payments = 0.0
        async for invoice in db.accounting_invoices.find({
            'tenant_id': current_user.tenant_id,
            'due_date': target_date.isoformat(),
            'status': 'pending'
        }):
            expected_payments += invoice.get('total', 0)
        
        weekly_plan.append({
            'date': target_date.isoformat(),
            'day_name': target_date.strftime('%A'),
            'expected_collections': expected_collections,
            'expected_payments': expected_payments,
            'checkout_count': checkout_count
        })
    
    # Bank balance summaries
    bank_balances = []
    async for bank in db.bank_accounts.find({
        'tenant_id': current_user.tenant_id,
        'is_active': True
    }):
        bank_balances.append({
            'bank_name': bank.get('bank_name'),
            'account_number': bank.get('account_number')[-4:],  # Last 4 digits
            'currency': bank.get('currency', 'TRY'),
            'current_balance': bank.get('current_balance', 0),
            'available_balance': bank.get('available_balance', 0),
            'last_sync': bank.get('last_sync').isoformat() if bank.get('last_sync') else None
        })
    
    total_bank_balance = sum(b['current_balance'] for b in bank_balances if b['currency'] == 'TRY')
    
    return {
        'today': {
            'date': today.isoformat(),
            'cash_inflow': today_inflow,
            'cash_outflow': today_outflow,
            'net_flow': net_flow,
            'inflow_count': inflow_count,
            'outflow_count': outflow_count
        },
        'weekly_plan': weekly_plan,
        'bank_balances': bank_balances,
        'total_bank_balance_try': total_bank_balance
    }



@router.get("/finance/mobile/overdue-accounts")
async def get_overdue_accounts_mobile(
    min_days: int = 7,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get accounts overdue by more than specified days (default 7)
    Returns with risk level classification:
    - Normal: 0-7 days
    - Warning: 8-14 days (Yellow)
    - Critical: 15-30 days (Red)
    - Suspicious: 30+ days (Black)
    """
    current_user = await get_current_user(credentials)
    today = datetime.now(timezone.utc)
    
    overdue_accounts = []
    
    async for folio in db.folios.find({
        'tenant_id': current_user.tenant_id,
        'status': 'open',
        'balance': {'$gt': 0}
    }):
        booking = await db.bookings.find_one({
            'id': folio.get('booking_id'),
            'tenant_id': current_user.tenant_id
        })
        
        if booking:
            checkout = booking.get('check_out')
            if checkout:
                try:
                    if isinstance(checkout, str):
                        checkout_date = datetime.fromisoformat(checkout).replace(tzinfo=timezone.utc)
                    else:
                        checkout_date = checkout if checkout.tzinfo else checkout.replace(tzinfo=timezone.utc)
                    
                    days_overdue = (today - checkout_date).days
                    
                    if days_overdue >= min_days:
                        # Determine risk level
                        if days_overdue >= 30:
                            risk_level = RiskLevel.SUSPICIOUS
                            risk_color = "black"
                        elif days_overdue >= 15:
                            risk_level = RiskLevel.CRITICAL
                            risk_color = "red"
                        elif days_overdue >= 8:
                            risk_level = RiskLevel.WARNING
                            risk_color = "yellow"
                        else:
                            risk_level = RiskLevel.NORMAL
                            risk_color = "green"
                        
                        guest = await db.guests.find_one({
                            'id': booking.get('guest_id'),
                            'tenant_id': current_user.tenant_id
                        })
                        
                        overdue_accounts.append({
                            'folio_id': folio.get('id'),
                            'folio_number': folio.get('folio_number'),
                            'booking_id': booking.get('id'),
                            'guest_name': guest.get('name') if guest else 'Unknown',
                            'guest_email': guest.get('email') if guest else None,
                            'guest_phone': guest.get('phone') if guest else None,
                            'room_number': booking.get('room_number'),
                            'checkout_date': checkout_date.date().isoformat(),
                            'balance': folio.get('balance', 0),
                            'days_overdue': days_overdue,
                            'risk_level': risk_level.value,
                            'risk_color': risk_color
                        })
                except (ValueError, AttributeError):
                    pass
    
    # Sort by days overdue (most critical first)
    overdue_accounts.sort(key=lambda x: x['days_overdue'], reverse=True)
    
    # Summary statistics
    total_overdue = sum(acc['balance'] for acc in overdue_accounts)
    suspicious_count = len([a for a in overdue_accounts if a['risk_level'] == 'suspicious'])
    critical_count = len([a for a in overdue_accounts if a['risk_level'] == 'critical'])
    warning_count = len([a for a in overdue_accounts if a['risk_level'] == 'warning'])
    
    return {
        'overdue_accounts': overdue_accounts,
        'summary': {
            'total_count': len(overdue_accounts),
            'total_amount': total_overdue,
            'suspicious_count': suspicious_count,  # 30+ days
            'critical_count': critical_count,  # 15-30 days
            'warning_count': warning_count  # 8-14 days
        }
    }



@router.get("/finance/mobile/credit-limit-violations")
async def get_credit_limit_violations_mobile(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get companies exceeding their credit limits"""
    current_user = await get_current_user(credentials)
    
    violations = []
    
    async for credit_limit in db.credit_limits.find({
        'tenant_id': current_user.tenant_id
    }):
        company = await db.companies.find_one({
            'id': credit_limit.get('company_id'),
            'tenant_id': current_user.tenant_id
        })
        
        if not company:
            continue
        
        # Calculate current debt from open folios
        current_debt = 0.0
        async for folio in db.folios.find({
            'tenant_id': current_user.tenant_id,
            'company_id': credit_limit.get('company_id'),
            'status': 'open',
            'balance': {'$gt': 0}
        }):
            current_debt += folio.get('balance', 0)
        
        credit_limit_amount = credit_limit.get('credit_limit', 0)
        available_credit = credit_limit_amount - current_debt
        utilization_pct = (current_debt / credit_limit_amount * 100) if credit_limit_amount > 0 else 0
        
        # Check if exceeding limit
        if current_debt > credit_limit_amount:
            violations.append({
                'company_id': credit_limit.get('company_id'),
                'company_name': company.get('name'),
                'credit_limit': credit_limit_amount,
                'current_debt': current_debt,
                'over_limit_amount': current_debt - credit_limit_amount,
                'available_credit': available_credit,
                'utilization_percentage': utilization_pct,
                'payment_terms_days': credit_limit.get('payment_terms_days', 30),
                'contact_person': company.get('contact_person'),
                'contact_email': company.get('contact_email'),
                'contact_phone': company.get('contact_phone')
            })
        # Also include companies near limit (90%+)
        elif utilization_pct >= 90:
            violations.append({
                'company_id': credit_limit.get('company_id'),
                'company_name': company.get('name'),
                'credit_limit': credit_limit_amount,
                'current_debt': current_debt,
                'over_limit_amount': 0,
                'available_credit': available_credit,
                'utilization_percentage': utilization_pct,
                'payment_terms_days': credit_limit.get('payment_terms_days', 30),
                'contact_person': company.get('contact_person'),
                'contact_email': company.get('contact_email'),
                'contact_phone': company.get('contact_phone'),
                'warning': 'Near limit'
            })
    
    # Sort by over limit amount
    violations.sort(key=lambda x: x.get('over_limit_amount', 0), reverse=True)
    
    return {
        'violations': violations,
        'summary': {
            'total_count': len(violations),
            'over_limit_count': len([v for v in violations if v.get('over_limit_amount', 0) > 0]),
            'near_limit_count': len([v for v in violations if v.get('warning') == 'Near limit'])
        }
    }



@router.get("/finance/mobile/suspicious-receivables")
async def get_suspicious_receivables_mobile(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get suspicious receivables list (30+ days overdue + high amounts)"""
    current_user = await get_current_user(credentials)
    today = datetime.now(timezone.utc)
    
    suspicious_list = []
    
    async for folio in db.folios.find({
        'tenant_id': current_user.tenant_id,
        'status': 'open',
        'balance': {'$gt': 1000}  # Only significant amounts
    }):
        booking = await db.bookings.find_one({
            'id': folio.get('booking_id'),
            'tenant_id': current_user.tenant_id
        })
        
        if booking:
            checkout = booking.get('check_out')
            if checkout:
                try:
                    if isinstance(checkout, str):
                        checkout_date = datetime.fromisoformat(checkout).replace(tzinfo=timezone.utc)
                    else:
                        checkout_date = checkout if checkout.tzinfo else checkout.replace(tzinfo=timezone.utc)
                    
                    days_overdue = (today - checkout_date).days
                    
                    # Suspicious criteria: 30+ days OR high amount with 15+ days
                    balance = folio.get('balance', 0)
                    is_suspicious = (days_overdue >= 30) or (days_overdue >= 15 and balance > 5000)
                    
                    if is_suspicious:
                        guest = await db.guests.find_one({
                            'id': booking.get('guest_id'),
                            'tenant_id': current_user.tenant_id
                        })
                        
                        # Get payment history
                        payment_count = await db.payments.count_documents({
                            'tenant_id': current_user.tenant_id,
                            'folio_id': folio.get('id')
                        })
                        
                        suspicious_list.append({
                            'folio_id': folio.get('id'),
                            'folio_number': folio.get('folio_number'),
                            'guest_name': guest.get('name') if guest else 'Unknown',
                            'guest_email': guest.get('email') if guest else None,
                            'guest_phone': guest.get('phone') if guest else None,
                            'company_id': folio.get('company_id'),
                            'balance': balance,
                            'checkout_date': checkout_date.date().isoformat(),
                            'days_overdue': days_overdue,
                            'payment_history_count': payment_count,
                            'reason': '30+ days overdue' if days_overdue >= 30 else 'High amount + 15+ days overdue'
                        })
                except (ValueError, AttributeError):
                    pass
    
    # Sort by balance (highest first)
    suspicious_list.sort(key=lambda x: x['balance'], reverse=True)
    
    total_suspicious_amount = sum(s['balance'] for s in suspicious_list)
    
    return {
        'suspicious_receivables': suspicious_list,
        'summary': {
            'total_count': len(suspicious_list),
            'total_amount': total_suspicious_amount,
            'average_days_overdue': sum(s['days_overdue'] for s in suspicious_list) / len(suspicious_list) if suspicious_list else 0
        }
    }



@router.get("/finance/mobile/risk-alerts")
async def get_risk_alerts_mobile(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get comprehensive risk alerts for finance dashboard"""
    current_user = await get_current_user(credentials)
    
    alerts = []
    
    # Get overdue accounts (7+ days)
    overdue_response = await get_overdue_accounts_mobile(min_days=7, credentials=credentials)
    overdue_summary = overdue_response['summary']
    
    if overdue_summary['suspicious_count'] > 0:
        alerts.append({
            'id': str(uuid.uuid4()),
            'type': 'suspicious_receivables',
            'severity': 'critical',
            'title': 'Şüpheli Alacaklar',
            'message': f"{overdue_summary['suspicious_count']} adet 30+ gün gecikmiş alacak",
            'amount': sum(a['balance'] for a in overdue_response['overdue_accounts'] if a['risk_level'] == 'suspicious'),
            'action_required': True
        })
    
    if overdue_summary['critical_count'] > 0:
        alerts.append({
            'id': str(uuid.uuid4()),
            'type': 'critical_overdue',
            'severity': 'high',
            'title': 'Kritik Gecikmiş Ödemeler',
            'message': f"{overdue_summary['critical_count']} adet 15+ gün gecikmiş ödeme",
            'amount': sum(a['balance'] for a in overdue_response['overdue_accounts'] if a['risk_level'] == 'critical'),
            'action_required': True
        })
    
    # Get credit limit violations
    violations_response = await get_credit_limit_violations_mobile(credentials=credentials)
    violations_summary = violations_response['summary']
    
    if violations_summary['over_limit_count'] > 0:
        alerts.append({
            'id': str(uuid.uuid4()),
            'type': 'credit_limit_violation',
            'severity': 'critical',
            'title': 'Kredi Limiti Aşımı',
            'message': f"{violations_summary['over_limit_count']} firma limiti aştı",
            'action_required': True
        })
    
    if violations_summary['near_limit_count'] > 0:
        alerts.append({
            'id': str(uuid.uuid4()),
            'type': 'near_credit_limit',
            'severity': 'medium',
            'title': 'Limite Yaklaşan Firmalar',
            'message': f"{violations_summary['near_limit_count']} firma limitin %90'ına ulaştı",
            'action_required': False
        })
    
    # Check for large unpaid invoices
    large_unpaid = 0
    large_unpaid_amount = 0.0
    async for invoice in db.accounting_invoices.find({
        'tenant_id': current_user.tenant_id,
        'status': 'pending',
        'total': {'$gt': 10000}
    }):
        large_unpaid += 1
        large_unpaid_amount += invoice.get('total', 0)
    
    if large_unpaid > 0:
        alerts.append({
            'id': str(uuid.uuid4()),
            'type': 'large_unpaid_invoices',
            'severity': 'medium',
            'title': 'Büyük Ödenmemiş Faturalar',
            'message': f"{large_unpaid} adet büyük fatura ödenmedi (>₺10,000)",
            'amount': large_unpaid_amount,
            'action_required': False
        })
    
    # Sort by severity
    severity_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
    alerts.sort(key=lambda x: severity_order.get(x['severity'], 999))
    
    return {
        'alerts': alerts,
        'summary': {
            'total_alerts': len(alerts),
            'critical_count': len([a for a in alerts if a['severity'] == 'critical']),
            'high_count': len([a for a in alerts if a['severity'] == 'high']),
            'action_required_count': len([a for a in alerts if a.get('action_required')])
        }
    }



@router.get("/finance/mobile/daily-expenses")
async def get_daily_expenses_mobile(
    date: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get daily expense/cost summary"""
    current_user = await get_current_user(credentials)
    
    if date:
        target_date = datetime.fromisoformat(date).date()
    else:
        target_date = datetime.now(timezone.utc).date()
    
    start_of_day = datetime.combine(target_date, datetime.min.time()).replace(tzinfo=timezone.utc)
    end_of_day = datetime.combine(target_date, datetime.max.time()).replace(tzinfo=timezone.utc)
    
    # Get expenses by category
    expenses_by_category = {}
    expenses_by_department = {}
    total_expenses = 0.0
    expense_count = 0
    
    async for expense in db.expenses.find({
        'tenant_id': current_user.tenant_id,
        'date': {'$gte': start_of_day, '$lte': end_of_day}
    }):
        amount = expense.get('amount', 0)
        category = expense.get('category', 'Other')
        department = expense.get('department', 'other')
        
        total_expenses += amount
        expense_count += 1
        
        expenses_by_category[category] = expenses_by_category.get(category, 0) + amount
        expenses_by_department[department] = expenses_by_department.get(department, 0) + amount
    
    return {
        'date': target_date.isoformat(),
        'total_expenses': total_expenses,
        'expense_count': expense_count,
        'expenses_by_category': expenses_by_category,
        'expenses_by_department': expenses_by_department
    }



@router.get("/finance/mobile/folio-full-extract/{folio_id}")
async def get_folio_full_extract_mobile(
    folio_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get full folio extract with all charges and payments"""
    current_user = await get_current_user(credentials)
    
    folio = await db.folios.find_one({
        'id': folio_id,
        'tenant_id': current_user.tenant_id
    })
    
    if not folio:
        raise HTTPException(status_code=404, detail="Folio not found")
    
    # Get booking details
    booking = await db.bookings.find_one({
        'id': folio.get('booking_id'),
        'tenant_id': current_user.tenant_id
    })
    
    # Get guest details
    guest = None
    if booking:
        guest = await db.guests.find_one({
            'id': booking.get('guest_id'),
            'tenant_id': current_user.tenant_id
        })
    
    # Get all charges
    charges = []
    total_charges = 0.0
    async for charge in db.folio_charges.find({
        'folio_id': folio_id,
        'tenant_id': current_user.tenant_id,
        'voided': {'$ne': True}
    }).sort('created_at', 1):
        charge_amount = charge.get('total', 0)
        total_charges += charge_amount
        charges.append({
            'id': charge.get('id'),
            'date': charge.get('created_at').isoformat() if charge.get('created_at') else None,
            'category': charge.get('category'),
            'description': charge.get('description'),
            'quantity': charge.get('quantity', 1),
            'unit_price': charge.get('unit_price', 0),
            'amount': charge.get('amount', 0),
            'tax_amount': charge.get('tax_amount', 0),
            'total': charge_amount,
            'posted_by': charge.get('posted_by')
        })
    
    # Get all payments
    payments = []
    total_payments = 0.0
    async for payment in db.payments.find({
        'folio_id': folio_id,
        'tenant_id': current_user.tenant_id
    }).sort('created_at', 1):
        payment_amount = payment.get('amount', 0)
        total_payments += payment_amount
        payments.append({
            'id': payment.get('id'),
            'date': payment.get('created_at').isoformat() if payment.get('created_at') else None,
            'amount': payment_amount,
            'payment_method': payment.get('payment_method'),
            'payment_type': payment.get('payment_type'),
            'notes': payment.get('notes'),
            'posted_by': payment.get('created_by')
        })
    
    current_balance = total_charges - total_payments
    
    return {
        'folio': {
            'id': folio.get('id'),
            'folio_number': folio.get('folio_number'),
            'folio_type': folio.get('folio_type'),
            'status': folio.get('status'),
            'created_at': folio.get('created_at').isoformat() if folio.get('created_at') else None,
            'closed_at': folio.get('closed_at').isoformat() if folio.get('closed_at') else None
        },
        'guest': {
            'name': guest.get('name') if guest else 'Unknown',
            'email': guest.get('email') if guest else None,
            'phone': guest.get('phone') if guest else None,
            'id_number': guest.get('id_number') if guest else None
        } if guest else None,
        'booking': {
            'id': booking.get('id') if booking else None,
            'room_number': booking.get('room_number') if booking else None,
            'check_in': booking.get('check_in') if booking else None,
            'check_out': booking.get('check_out') if booking else None
        } if booking else None,
        'charges': charges,
        'payments': payments,
        'summary': {
            'total_charges': total_charges,
            'total_payments': total_payments,
            'current_balance': current_balance,
            'charge_count': len(charges),
            'payment_count': len(payments)
        }
    }



@router.get("/finance/mobile/invoices")
async def get_invoices_mobile(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    unpaid_only: bool = False,
    department: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get invoices with advanced filtering (date, unpaid, department)"""
    current_user = await get_current_user(credentials)
    
    query = {'tenant_id': current_user.tenant_id}
    
    # Date filter
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter['$gte'] = datetime.fromisoformat(start_date)
        if end_date:
            date_filter['$lte'] = datetime.fromisoformat(end_date)
        query['created_at'] = date_filter
    
    # Unpaid filter
    if unpaid_only:
        query['status'] = 'pending'
    
    # Department filter - will need to check items
    invoices = []
    total_amount = 0.0
    unpaid_amount = 0.0
    
    async for invoice in db.accounting_invoices.find(query).sort('created_at', -1).limit(100):
        # If department filter is specified, check invoice items
        if department:
            items = invoice.get('items', [])
            has_department = any(item.get('department') == department for item in items)
            if not has_department:
                continue
        
        invoice_total = invoice.get('total', 0)
        total_amount += invoice_total
        
        if invoice.get('status') == 'pending':
            unpaid_amount += invoice_total
        
        # Get company details if available
        company_name = None
        if invoice.get('company_id'):
            company = await db.companies.find_one({
                'id': invoice.get('company_id'),
                'tenant_id': current_user.tenant_id
            })
            if company:
                company_name = company.get('name')
        
        invoices.append({
            'id': invoice.get('id'),
            'invoice_number': invoice.get('invoice_number'),
            'invoice_type': invoice.get('invoice_type'),
            'status': invoice.get('status'),
            'customer_name': invoice.get('customer_name'),
            'company_name': company_name,
            'created_at': invoice.get('created_at').isoformat() if invoice.get('created_at') else None,
            'due_date': invoice.get('due_date'),
            'subtotal': invoice.get('subtotal', 0),
            'vat': invoice.get('vat', 0),
            'total': invoice_total,
            'currency': invoice.get('currency', 'TRY'),
            'has_efatura': invoice.get('efatura_uuid') is not None
        })
    
    return {
        'invoices': invoices,
        'summary': {
            'total_count': len(invoices),
            'total_amount': total_amount,
            'unpaid_amount': unpaid_amount,
            'paid_amount': total_amount - unpaid_amount
        }
    }



@router.get("/finance/mobile/invoice-pdf/{invoice_id}")
async def get_invoice_pdf_mobile(
    invoice_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Generate and return invoice PDF (dynamic generation)"""
    current_user = await get_current_user(credentials)
    
    invoice = await db.accounting_invoices.find_one({
        'id': invoice_id,
        'tenant_id': current_user.tenant_id
    })
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # For MVP, return invoice data for frontend PDF generation
    # In production, use libraries like WeasyPrint or ReportLab for server-side PDF generation
    
    # Get company details if available
    company = None
    if invoice.get('company_id'):
        company = await db.companies.find_one({
            'id': invoice.get('company_id'),
            'tenant_id': current_user.tenant_id
        })
    
    # Get tenant details
    tenant = await db.tenants.find_one({'id': current_user.tenant_id})
    
    pdf_data = {
        'invoice': {
            'invoice_number': invoice.get('invoice_number'),
            'invoice_type': invoice.get('invoice_type'),
            'invoice_date': invoice.get('created_at').isoformat() if invoice.get('created_at') else None,
            'due_date': invoice.get('due_date'),
            'status': invoice.get('status'),
            'customer_name': invoice.get('customer_name'),
            'customer_tax_number': invoice.get('customer_tax_number'),
            'customer_address': invoice.get('customer_address'),
            'items': invoice.get('items', []),
            'subtotal': invoice.get('subtotal', 0),
            'vat': invoice.get('vat', 0),
            'vat_rate': invoice.get('vat_rate', 20),
            'total': invoice.get('total', 0),
            'currency': invoice.get('currency', 'TRY'),
            'notes': invoice.get('notes'),
            'efatura_uuid': invoice.get('efatura_uuid')
        },
        'company': {
            'name': company.get('name') if company else None,
            'tax_number': company.get('tax_number') if company else None,
            'billing_address': company.get('billing_address') if company else None
        } if company else None,
        'hotel': {
            'name': tenant.get('hotel_name') if tenant else 'Hotel PMS',
            'address': tenant.get('address') if tenant else '',
            'tax_number': tenant.get('tax_number') if tenant else '',
            'phone': tenant.get('phone') if tenant else '',
            'email': tenant.get('email') if tenant else ''
        },
        'generated_at': datetime.now(timezone.utc).isoformat(),
        'pdf_ready': False,  # Flag for frontend to generate PDF
        'download_filename': f"Invoice_{invoice.get('invoice_number')}.pdf"
    }
    
    return pdf_data



@router.post("/finance/mobile/bank-balance-update")
async def update_bank_balance_mobile(
    bank_account_id: str,
    current_balance: float,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Manual bank balance update (for now, until API integration)"""
    current_user = await get_current_user(credentials)
    
    bank_account = await db.bank_accounts.find_one({
        'id': bank_account_id,
        'tenant_id': current_user.tenant_id
    })
    
    if not bank_account:
        raise HTTPException(status_code=404, detail="Bank account not found")
    
    await db.bank_accounts.update_one(
        {'id': bank_account_id, 'tenant_id': current_user.tenant_id},
        {
            '$set': {
                'current_balance': current_balance,
                'available_balance': current_balance,  # Simplified for now
                'last_sync': datetime.now(timezone.utc),
                'updated_at': datetime.now(timezone.utc)
            }
        }
    )
    
    return {
        'message': 'Bank balance updated successfully',
        'bank_account_id': bank_account_id,
        'current_balance': current_balance
    }



@router.get("/finance/mobile/bank-balances")
async def get_bank_balances_mobile(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get all bank account balances"""
    current_user = await get_current_user(credentials)
    
    bank_accounts = []
    total_balance_try = 0.0
    
    async for bank in db.bank_accounts.find({
        'tenant_id': current_user.tenant_id,
        'is_active': True
    }):
        balance = bank.get('current_balance', 0)
        bank_accounts.append({
            'id': bank.get('id'),
            'bank_name': bank.get('bank_name'),
            'account_number': bank.get('account_number'),
            'iban': bank.get('iban'),
            'currency': bank.get('currency', 'TRY'),
            'current_balance': balance,
            'available_balance': bank.get('available_balance', 0),
            'account_type': bank.get('account_type', 'checking'),
            'api_enabled': bank.get('api_enabled', False),
            'last_sync': bank.get('last_sync').isoformat() if bank.get('last_sync') else None
        })
        
        if bank.get('currency') == 'TRY':
            total_balance_try += balance
    
    return {
        'bank_accounts': bank_accounts,
        'total_balance_try': total_balance_try,
        'account_count': len(bank_accounts)
    }





@router.get("/finance/expense-summary")
async def get_expense_summary(
    date: Optional[str] = None,  # YYYY-MM-DD
    period: str = "today",  # today, week, month
    current_user: User = Depends(get_current_user)
):
    """
    Get expense summary with categories
    Categories: F&B costs, housekeeping, maintenance, staff, utilities, procurement
    """
    try:
        if date:
            target_date = datetime.fromisoformat(date).date()
        else:
            target_date = datetime.now(timezone.utc).date()
        
        # Calculate date range
        if period == "today":
            start_date = target_date
            end_date = target_date
        elif period == "week":
            start_date = target_date - timedelta(days=7)
            end_date = target_date
        else:  # month
            start_date = target_date.replace(day=1)
            end_date = target_date
        
        # Sample expense data (in production, fetch from expenses collection)
        expenses = {
            'fnb_costs': {
                'amount': 15420.50,
                'category': 'F&B Maliyetleri',
                'breakdown': {
                    'food_purchases': 8500.00,
                    'beverages': 4200.50,
                    'supplies': 2720.00
                }
            },
            'housekeeping_expenses': {
                'amount': 8750.00,
                'category': 'Temizlik Giderleri',
                'breakdown': {
                    'cleaning_supplies': 3200.00,
                    'laundry': 4050.00,
                    'equipment': 1500.00
                }
            },
            'maintenance_costs': {
                'amount': 5600.00,
                'category': 'Teknik Maliyetler',
                'breakdown': {
                    'repairs': 3200.00,
                    'parts': 1800.00,
                    'preventive': 600.00
                }
            },
            'staff_costs': {
                'amount': 45800.00,
                'category': 'Personel Maliyetleri',
                'breakdown': {
                    'hourly_wages': 28500.00,
                    'overtime': 8200.00,
                    'benefits': 9100.00
                },
                'hourly_rate_avg': 85.50
            },
            'utilities': {
                'amount': 12300.00,
                'category': 'Enerji & Utilities',
                'breakdown': {
                    'electricity': 7200.00,
                    'water': 2800.00,
                    'gas': 2300.00
                }
            },
            'procurement': {
                'amount': 9850.00,
                'category': 'Satın Alma',
                'breakdown': {
                    'supplies': 5200.00,
                    'equipment': 3150.00,
                    'other': 1500.00
                }
            }
        }
        
        # Calculate totals
        total_expenses = sum(cat['amount'] for cat in expenses.values())
        
        # Calculate daily average for the period
        days_in_period = (end_date - start_date).days + 1
        daily_avg = total_expenses / days_in_period if days_in_period > 0 else 0
        
        return {
            'period': period,
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'total_expenses': round(total_expenses, 2),
            'daily_average': round(daily_avg, 2),
            'categories': expenses,
            'top_expense': max(expenses.items(), key=lambda x: x[1]['amount'])[1]['category']
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get expense summary: {str(e)}")



@router.get("/finance/cash-flow-dashboard")
async def get_cash_flow_dashboard(
    current_user: User = Depends(get_current_user)
):
    """
    Comprehensive cash flow dashboard for Finance Manager
    Shows: today's cash in/out, weekly plan, bank balances, risk alerts
    """
    try:
        today = datetime.now(timezone.utc).date()
        today_str = today.isoformat()
        
        # Today's cash inflow (collections)
        today_checkins = await db.bookings.find({
            'check_in': today_str,
            'tenant_id': current_user.tenant_id
        }, {'_id': 0, 'total_amount': 1, 'paid_amount': 1}).to_list(200)
        
        today_collections = sum(b.get('paid_amount', 0) for b in today_checkins)
        
        # Today's cash outflow (expenses - sample data)
        today_expenses = {
            'staff_payments': 8500.00,
            'supplier_payments': 12300.00,
            'utility_bills': 3200.00,
            'other': 1500.00
        }
        today_outflow = sum(today_expenses.values())
        
        # Net cash flow
        net_cash_flow = today_collections - today_outflow
        
        # Weekly forecast (next 7 days)
        weekly_forecast = []
        for i in range(7):
            date = today + timedelta(days=i)
            date_str = date.isoformat()
            
            # Expected collections (check-ins + ongoing bookings)
            expected_checkins = await db.bookings.count_documents({
                'check_in': date_str,
                'tenant_id': current_user.tenant_id
            })
            
            expected_checkouts = await db.bookings.count_documents({
                'check_out': date_str,
                'tenant_id': current_user.tenant_id
            })
            
            # Simplified forecast
            expected_inflow = expected_checkins * 1500 + expected_checkouts * 500
            expected_outflow = 15000 if date.weekday() == 4 else 8000  # Friday = payroll
            
            weekly_forecast.append({
                'date': date_str,
                'day_name': date.strftime('%a'),
                'expected_inflow': round(expected_inflow, 2),
                'expected_outflow': round(expected_outflow, 2),
                'net': round(expected_inflow - expected_outflow, 2)
            })
        
        # Bank balances (sample - in production, integrate with bank API)
        bank_balances = [
            {'bank': 'Garanti BBVA', 'account': '****3421', 'balance': 285600.50, 'currency': 'TRY'},
            {'bank': 'İş Bankası', 'account': '****7832', 'balance': 142300.00, 'currency': 'TRY'},
            {'bank': 'Akbank', 'account': '****1259', 'balance': 95800.75, 'currency': 'TRY'}
        ]
        total_bank_balance = sum(b['balance'] for b in bank_balances)
        
        return {
            'today': {
                'date': today_str,
                'cash_inflow': round(today_collections, 2),
                'cash_outflow': round(today_outflow, 2),
                'net_cash_flow': round(net_cash_flow, 2),
                'outflow_breakdown': today_expenses
            },
            'weekly_forecast': weekly_forecast,
            'bank_balances': {
                'accounts': bank_balances,
                'total': round(total_bank_balance, 2)
            },
            'status': 'positive' if net_cash_flow > 0 else 'negative'
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get cash flow dashboard: {str(e)}")



@router.get("/finance/risk-alerts")
async def get_risk_alerts(
    current_user: User = Depends(get_current_user)
):
    """
    Financial risk monitoring: overdue accounts, credit limits, suspicious receivables
    """
    try:
        today = datetime.now(timezone.utc).date()
        
        # Get all folios with outstanding balance
        folios = await db.folios.find({
            'tenant_id': current_user.tenant_id,
            'status': {'$in': ['open', 'closed']}
        }, {'_id': 0}).to_list(1000)
        
        overdue_accounts = []
        over_limit_accounts = []
        suspicious_accounts = []
        
        for folio in folios:
            outstanding = folio.get('total_amount', 0) - folio.get('paid_amount', 0)
            
            if outstanding > 0:
                # Get booking info
                booking = await db.bookings.find_one({'id': folio.get('booking_id')}, {'_id': 0})
                
                if booking:
                    # Calculate days overdue
                    checkout_date = datetime.fromisoformat(booking['check_out']).date()
                    days_overdue = (today - checkout_date).days
                    
                    account_info = {
                        'folio_id': folio['id'],
                        'booking_id': booking['id'],
                        'guest_name': booking.get('guest_name'),
                        'room_number': booking.get('room_number'),
                        'outstanding_amount': round(outstanding, 2),
                        'checkout_date': booking['check_out'],
                        'days_overdue': days_overdue
                    }
                    
                    # Risk Categories
                    if days_overdue > 7:
                        account_info['risk_level'] = 'high' if days_overdue > 30 else 'medium'
                        overdue_accounts.append(account_info)
                    
                    # Credit limit check (sample: 10000 TL limit)
                    if outstanding > 10000:
                        account_info['risk_level'] = 'critical'
                        account_info['limit'] = 10000
                        over_limit_accounts.append(account_info)
                    
                    # Suspicious accounts (multiple unpaid bookings or high amount)
                    if outstanding > 20000 or days_overdue > 60:
                        account_info['risk_level'] = 'critical'
                        account_info['reason'] = 'High amount' if outstanding > 20000 else 'Long overdue'
                        suspicious_accounts.append(account_info)
        
        # Sort by risk
        overdue_accounts.sort(key=lambda x: x['days_overdue'], reverse=True)
        over_limit_accounts.sort(key=lambda x: x['outstanding_amount'], reverse=True)
        suspicious_accounts.sort(key=lambda x: x['outstanding_amount'], reverse=True)
        
        # Calculate totals
        total_overdue_amount = sum(acc['outstanding_amount'] for acc in overdue_accounts)
        total_at_risk = sum(acc['outstanding_amount'] for acc in suspicious_accounts)
        
        # Create notifications for critical cases
        for acc in suspicious_accounts[:5]:  # Top 5 critical
            existing_notif = await db.notifications.find_one({
                'related_id': acc['folio_id'],
                'type': 'financial_risk',
                'tenant_id': current_user.tenant_id
            }, {'_id': 0})
            
            if not existing_notif:
                await db.notifications.insert_one({
                    'id': str(uuid.uuid4()),
                    'tenant_id': current_user.tenant_id,
                    'user_role': 'finance',
                    'title': f'🚨 Yüksek Riskli Hesap - {acc["guest_name"]}',
                    'message': f'₺{acc["outstanding_amount"]:,.2f} ödenmemiş bakiye',
                    'type': 'financial_risk',
                    'priority': 'urgent',
                    'related_id': acc['folio_id'],
                    'read': False,
                    'created_at': datetime.now(timezone.utc).isoformat()
                })
        
        return {
            'overdue_accounts': overdue_accounts,
            'over_limit_accounts': over_limit_accounts,
            'suspicious_accounts': suspicious_accounts,
            'summary': {
                'total_overdue_count': len(overdue_accounts),
                'total_overdue_amount': round(total_overdue_amount, 2),
                'over_limit_count': len(over_limit_accounts),
                'suspicious_count': len(suspicious_accounts),
                'total_at_risk': round(total_at_risk, 2)
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get risk alerts: {str(e)}")



@router.get("/finance/folios-filtered")
@cached(ttl=300, key_prefix="finance_folios_filtered")  # Cache for 5 min
async def get_folios_filtered(
    customer_type: Optional[str] = None,  # vip, corporate, individual
    room_number: Optional[str] = None,
    status: Optional[str] = None,  # open, closed, cancelled
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Get filtered folios with customer type and room filters
    """
    try:
        filter_dict = {'tenant_id': current_user.tenant_id}
        
        if status:
            filter_dict['status'] = status
        
        folios = await db.folios.find(filter_dict, {'_id': 0}).sort('created_at', -1).limit(200).to_list(200)
        
        # Enrich with booking and guest data
        enriched_folios = []
        for folio in folios:
            if folio.get('booking_id'):
                booking = await db.bookings.find_one({'id': folio['booking_id']}, {'_id': 0})
                if booking:
                    folio['guest_name'] = booking.get('guest_name')
                    folio['room_number'] = booking.get('room_number')
                    folio['check_in'] = booking.get('check_in')
                    folio['check_out'] = booking.get('check_out')
                    
                    # Get guest info for customer type
                    if booking.get('guest_id'):
                        guest = await db.guests.find_one({'id': booking['guest_id']}, {'_id': 0})
                        if guest:
                            folio['customer_type'] = guest.get('customer_type', 'individual')
                            folio['guest_email'] = guest.get('email')
                            folio['guest_phone'] = guest.get('phone')
            
            # Apply filters
            if customer_type and folio.get('customer_type') != customer_type:
                continue
            
            if room_number and folio.get('room_number') != room_number:
                continue
            
            if date_from and folio.get('check_in', '') < date_from:
                continue
            
            if date_to and folio.get('check_out', '') > date_to:
                continue
            
            enriched_folios.append(folio)
        
        return {
            'folios': enriched_folios,
            'count': len(enriched_folios),
            'filters': {
                'customer_type': customer_type,
                'room_number': room_number,
                'status': status,
                'date_from': date_from,
                'date_to': date_to
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get filtered folios: {str(e)}")



@router.get("/finance/folio/{folio_id}/detail")
async def get_folio_detail(
    folio_id: str,
    current_user: User = Depends(get_current_user)
):
    """
    Get complete folio details with all charges and payments
    """
    try:
        folio = await db.folios.find_one({
            'id': folio_id,
            'tenant_id': current_user.tenant_id
        }, {'_id': 0})
        
        if not folio:
            raise HTTPException(status_code=404, detail="Folio not found")
        
        # Get booking details
        if folio.get('booking_id'):
            booking = await db.bookings.find_one({'id': folio['booking_id']}, {'_id': 0})
            folio['booking'] = booking
        
        # Calculate outstanding
        folio['outstanding'] = folio.get('total_amount', 0) - folio.get('paid_amount', 0)
        
        return folio
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get folio detail: {str(e)}")



@router.post("/finance/mobile-payment")
async def process_mobile_payment(
    folio_id: str,
    amount: float,
    payment_method: str,  # cash, card, bank_transfer
    notes: str = "",
    current_user: User = Depends(get_current_user)
):
    """
    Process payment from mobile device
    """
    try:
        folio = await db.folios.find_one({
            'id': folio_id,
            'tenant_id': current_user.tenant_id
        }, {'_id': 0})
        
        if not folio:
            raise HTTPException(status_code=404, detail="Folio not found")
        
        # Create payment record
        payment_id = str(uuid.uuid4())
        payment = {
            'id': payment_id,
            'folio_id': folio_id,
            'amount': amount,
            'payment_method': payment_method,
            'notes': notes,
            'processed_by': current_user.id,
            'processed_by_name': current_user.name,
            'processed_at': datetime.now(timezone.utc).isoformat(),
            'tenant_id': current_user.tenant_id
        }
        
        # Update folio
        new_paid_amount = folio.get('paid_amount', 0) + amount
        await db.folios.update_one(
            {'id': folio_id},
            {
                '$set': {
                    'paid_amount': new_paid_amount,
                    'updated_at': datetime.now(timezone.utc).isoformat()
                },
                '$push': {'payments': payment}
            }
        )
        
        # Create audit log
        await db.audit_logs.insert_one({
            'id': str(uuid.uuid4()),
            'tenant_id': current_user.tenant_id,
            'user_id': current_user.id,
            'user_name': current_user.name,
            'user_role': current_user.role,
            'action': 'MOBILE_PAYMENT',
            'entity_type': 'folio',
            'entity_id': folio_id,
            'changes': {'amount': amount, 'method': payment_method},
            'timestamp': datetime.now(timezone.utc).isoformat()
        })
        
        return {
            'message': 'Ödeme başarıyla işlendi',
            'payment_id': payment_id,
            'folio_id': folio_id,
            'amount': amount,
            'new_balance': round(folio.get('total_amount', 0) - new_paid_amount, 2)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process payment: {str(e)}")



@router.post("/cashiering/city-ledger")
async def create_city_ledger_account(
    account_data: dict,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Create a new city ledger account for direct billing"""
    current_user = await get_current_user(credentials)
    
    account = CityLedgerAccount(
        tenant_id=current_user.tenant_id,
        account_name=account_data['account_name'],
        company_name=account_data['company_name'],
        contact_person=account_data.get('contact_person'),
        email=account_data.get('email'),
        phone=account_data.get('phone'),
        address=account_data.get('address'),
        credit_limit=account_data.get('credit_limit', 0.0),
        payment_terms=account_data.get('payment_terms', 30)
    )
    
    await db.city_ledger_accounts.insert_one(account.model_dump())
    
    return {
        'success': True,
        'account_id': account.id,
        'account_name': account.account_name,
        'credit_limit': account.credit_limit
    }


@router.get("/cashiering/city-ledger")
async def get_city_ledger_accounts(
    is_active: bool = True,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get all city ledger accounts"""
    current_user = await get_current_user(credentials)
    
    query = {'tenant_id': current_user.tenant_id}
    if is_active is not None:
        query['is_active'] = is_active
    
    accounts = await db.city_ledger_accounts.find(query, {'_id': 0}).to_list(1000)
    
    return {
        'accounts': accounts,
        'total_count': len(accounts)
    }


@router.post("/cashiering/split-payment")
async def process_split_payment(
    booking_id: str,
    payments: List[dict],
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Process split payment (multiple payment methods for one bill)"""
    current_user = await get_current_user(credentials)
    
    # Get booking
    booking = await db.bookings.find_one({
        'id': booking_id,
        'tenant_id': current_user.tenant_id
    })
    
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    # Validate total matches booking amount
    total_payment = sum(p['amount'] for p in payments)
    
    if abs(total_payment - booking.get('total_amount', 0)) > 0.01:
        raise HTTPException(
            status_code=400,
            detail=f"Payment total ({total_payment}) doesn't match booking amount ({booking.get('total_amount', 0)})"
        )
    
    # Process each payment
    payment_records = []
    for payment in payments:
        payment_record = {
            'id': str(uuid.uuid4()),
            'tenant_id': current_user.tenant_id,
            'booking_id': booking_id,
            'payment_method': payment['payment_method'],
            'amount': payment['amount'],
            'reference': payment.get('reference'),
            'processed_at': datetime.now(timezone.utc).isoformat(),
            'processed_by': current_user.name
        }
        await db.payments.insert_one(payment_record)
        payment_records.append(payment_record)
    
    # Update booking status
    await db.bookings.update_one(
        {'id': booking_id},
        {'$set': {'payment_status': 'paid', 'paid_at': datetime.now(timezone.utc).isoformat()}}
    )
    
    return {
        'success': True,
        'booking_id': booking_id,
        'payments_processed': len(payment_records),
        'total_amount': total_payment,
        'payment_methods': [p['payment_method'] for p in payments]
    }


@router.get("/cashiering/ar-aging-report")
async def get_ar_aging_report(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get Accounts Receivable aging report (30/60/90 days)"""
    current_user = await get_current_user(credentials)
    
    today = datetime.now(timezone.utc)
    
    aging_buckets = {
        'current': [],
        '30_days': [],
        '60_days': [],
        '90_plus': []
    }
    
    # Get all city ledger accounts with balance
    accounts = await db.city_ledger_accounts.find({
        'tenant_id': current_user.tenant_id,
        'current_balance': {'$gt': 0}
    }, {'_id': 0}).to_list(1000)
    
    for account in accounts:
        # Get oldest transaction
        oldest_transaction = await db.city_ledger_transactions.find_one(
            {
                'account_id': account['id'],
                'transaction_type': 'charge'
            },
            {'_id': 0},
            sort=[('transaction_date', 1)]
        )
        
        if oldest_transaction:
            # Parse transaction_date safely
            transaction_date = oldest_transaction['transaction_date']
            if isinstance(transaction_date, str):
                transaction_date = datetime.fromisoformat(transaction_date.replace('Z', '+00:00'))
            elif not isinstance(transaction_date, datetime):
                continue  # Skip invalid data
            
            # Ensure timezone-aware
            if transaction_date.tzinfo is None:
                transaction_date = transaction_date.replace(tzinfo=timezone.utc)
            
            days_old = (today - transaction_date).days
            
            aging_entry = {
                'account_id': account['id'],
                'account_name': account['account_name'],
                'balance': account['current_balance'],
                'days_old': days_old
            }
            
            if days_old <= 30:
                aging_buckets['current'].append(aging_entry)
            elif days_old <= 60:
                aging_buckets['30_days'].append(aging_entry)
            elif days_old <= 90:
                aging_buckets['60_days'].append(aging_entry)
            else:
                aging_buckets['90_plus'].append(aging_entry)
    
    # Calculate totals
    totals = {
        'current': sum(a['balance'] for a in aging_buckets['current']),
        '30_days': sum(a['balance'] for a in aging_buckets['30_days']),
        '60_days': sum(a['balance'] for a in aging_buckets['60_days']),
        '90_plus': sum(a['balance'] for a in aging_buckets['90_plus'])
    }
    
    totals['total'] = sum(totals.values())
    
    return {
        'aging_buckets': aging_buckets,
        'totals': totals,
        'generated_at': today.isoformat()
    }


@router.post("/cashiering/credit-limit")
async def set_credit_limit(
    account_id: str,
    credit_limit: float,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Set credit limit for city ledger account"""
    current_user = await get_current_user(credentials)
    
    result = await db.city_ledger_accounts.update_one(
        {
            'id': account_id,
            'tenant_id': current_user.tenant_id
        },
        {'$set': {'credit_limit': credit_limit}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Account not found")
    
    return {
        'success': True,
        'account_id': account_id,
        'credit_limit': credit_limit
    }


@router.get("/cashiering/credit-limit/{account_id}")
async def get_credit_limit(
    account_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get credit limit and current balance for account"""
    current_user = await get_current_user(credentials)
    
    account = await db.city_ledger_accounts.find_one({
        'id': account_id,
        'tenant_id': current_user.tenant_id
    }, {'_id': 0})
    
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    available_credit = account['credit_limit'] - account['current_balance']
    
    return {
        'account_id': account_id,
        'account_name': account['account_name'],
        'credit_limit': account['credit_limit'],
        'current_balance': account['current_balance'],
        'available_credit': available_credit,
        'credit_status': 'ok' if available_credit > 0 else 'exceeded'
    }


@router.post("/cashiering/direct-bill")
async def post_to_city_ledger(
    booking_id: str,
    account_id: str,
    amount: float,
    description: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Post charge to city ledger (direct billing)"""
    current_user = await get_current_user(credentials)
    
    # Verify account
    account = await db.city_ledger_accounts.find_one({
        'id': account_id,
        'tenant_id': current_user.tenant_id
    })
    
    if not account:
        raise HTTPException(status_code=404, detail="City ledger account not found")
    
    # Check credit limit
    if account['current_balance'] + amount > account['credit_limit']:
        raise HTTPException(
            status_code=400,
            detail=f"Credit limit exceeded. Available: {account['credit_limit'] - account['current_balance']}"
        )
    
    # Create transaction
    transaction = CityLedgerTransaction(
        tenant_id=current_user.tenant_id,
        account_id=account_id,
        booking_id=booking_id,
        transaction_type='charge',
        amount=amount,
        description=description,
        posted_by=current_user.name
    )
    
    await db.city_ledger_transactions.insert_one(transaction.model_dump())
    
    # Update account balance
    await db.city_ledger_accounts.update_one(
        {'id': account_id},
        {'$inc': {'current_balance': amount}}
    )
    
    return {
        'success': True,
        'transaction_id': transaction.id,
        'account_name': account['account_name'],
        'amount_posted': amount,
        'new_balance': account['current_balance'] + amount
    }


@router.get("/cashiering/outstanding-balance")
async def get_outstanding_balances(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get all city ledger accounts with outstanding balances"""
    current_user = await get_current_user(credentials)
    
    accounts = await db.city_ledger_accounts.find({
        'tenant_id': current_user.tenant_id,
        'current_balance': {'$gt': 0}
    }, {'_id': 0}).sort('current_balance', -1).to_list(1000)
    
    total_outstanding = sum(a['current_balance'] for a in accounts)
    
    return {
        'accounts': accounts,
        'total_accounts': len(accounts),
        'total_outstanding': round(total_outstanding, 2)
    }


@router.post("/cashiering/city-ledger-payment")
async def post_city_ledger_payment(
    account_id: str,
    amount: float,
    payment_method: str,
    reference: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Post payment to city ledger account"""
    current_user = await get_current_user(credentials)
    
    account = await db.city_ledger_accounts.find_one({
        'id': account_id,
        'tenant_id': current_user.tenant_id
    })
    
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    transaction = CityLedgerTransaction(
        tenant_id=current_user.tenant_id,
        account_id=account_id,
        transaction_type='payment',
        amount=amount,
        description=f"Payment received via {payment_method}",
        reference_number=reference,
        posted_by=current_user.name
    )
    
    await db.city_ledger_transactions.insert_one(transaction.model_dump())
    
    new_balance = account['current_balance'] - amount
    await db.city_ledger_accounts.update_one(
        {'id': account_id},
        {'$set': {'current_balance': max(0, new_balance)}}
    )
    
    return {
        'success': True,
        'transaction_id': transaction.id,
        'account_name': account['account_name'],
        'amount_paid': amount,
        'new_balance': max(0, new_balance)
    }


@router.get("/cashiering/city-ledger/{account_id}/transactions")
async def get_city_ledger_transactions(
    account_id: str,
    limit: int = 100,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get transaction history for city ledger account"""
    current_user = await get_current_user(credentials)
    
    transactions = await db.city_ledger_transactions.find(
        {'account_id': account_id, 'tenant_id': current_user.tenant_id},
        {'_id': 0}
    ).sort('transaction_date', -1).limit(limit).to_list(limit)
    
    charges = sum(t['amount'] for t in transactions if t['transaction_type'] == 'charge')
    payments = sum(t['amount'] for t in transactions if t['transaction_type'] == 'payment')
    
    return {
        'account_id': account_id,
        'transactions': transactions,
        'summary': {
            'total_charges': round(charges, 2),
            'total_payments': round(payments, 2),
            'current_balance': round(charges - payments, 2),
            'transaction_count': len(transactions)
        }
    }





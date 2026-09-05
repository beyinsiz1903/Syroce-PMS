"""Reservation performance reports.

This module deliberately reports on *stay start dates*.  That makes the
numbers useful for front-office planning: arrivals, cancellations, channels,
and booked room nights all refer to the same requested stay period.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import UTC, date, datetime, timedelta
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query

from core.database import db
from core.helpers import require_module
from core.security import get_current_user
from core.utils import create_excel_workbook, excel_response
from models.schemas import User
from modules.pms_core.role_permission_service import require_op

try:
    from cache_manager import cached
except ImportError:

    def cached(ttl=300, key_prefix=""):
        def decorator(func):
            return func

        return decorator


sub_router = APIRouter()

_NON_COMMERCIAL_STATUSES = {"cancelled", "canceled", "no_show", "noshow"}
_STATUS_LABELS = {
    "confirmed": "Onaylandı",
    "guaranteed": "Garantili",
    "checked_in": "Giriş yaptı",
    "checked_out": "Çıkış yaptı",
    "cancelled": "İptal",
    "canceled": "İptal",
    "no_show": "No-show",
    "noshow": "No-show",
    "pending": "Bekliyor",
}
_LEAD_TIME_BUCKETS = (
    ("same_day", "Aynı gün", 0, 0),
    ("one_to_three", "1–3 gün", 1, 3),
    ("four_to_seven", "4–7 gün", 4, 7),
    ("eight_to_fourteen", "8–14 gün", 8, 14),
    ("fifteen_to_thirty", "15–30 gün", 15, 30),
    ("thirty_plus", "31+ gün", 31, None),
)


def _parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=value.tzinfo or UTC)
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except (TypeError, ValueError):
        return None
    return parsed.replace(tzinfo=parsed.tzinfo or UTC)


def _parse_date(value: str, field: str) -> date:
    try:
        return date.fromisoformat(value)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=422, detail=f"{field} YYYY-MM-DD formatında olmalı") from exc


def _amount(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _status(booking: dict[str, Any]) -> str:
    return str(booking.get("status") or "pending").strip().lower() or "pending"


def _channel(booking: dict[str, Any]) -> str:
    return str(
        booking.get("ota_channel")
        or booking.get("source_channel")
        or booking.get("channel")
        or booking.get("source")
        or "direct"
    ).strip() or "direct"


def _lead_time_bucket(days: int) -> str:
    for key, _label, lower, upper in _LEAD_TIME_BUCKETS:
        if days >= lower and (upper is None or days <= upper):
            return key
    return "same_day"


def aggregate_reservation_performance(
    bookings: list[dict[str, Any]], *, start_date: date, end_date: date
) -> dict[str, Any]:
    """Return a stable, display-ready reservation-report payload.

    Dates are parsed in Python as old import data can contain either ISO
    strings or BSON datetimes.  Bad date rows stay visible in the detailed
    list but never corrupt the report totals.
    """
    status_counts: dict[str, int] = defaultdict(int)
    channel_stats: dict[str, dict[str, Any]] = {}
    arrival_days: dict[str, dict[str, Any]] = {}
    lead_counts = {key: 0 for key, *_ in _LEAD_TIME_BUCKETS}
    rows: list[dict[str, Any]] = []

    commercial_count = cancelled_count = no_show_count = 0
    booked_revenue = cancelled_value = total_nights = lead_time_total = lead_time_count = 0.0

    for booking in bookings:
        status = _status(booking)
        channel = _channel(booking)
        check_in = _parse_datetime(booking.get("check_in"))
        check_out = _parse_datetime(booking.get("check_out"))
        created_at = _parse_datetime(booking.get("created_at") or booking.get("reservation_date"))
        amount = _amount(booking.get("total_amount"))
        nights = max((check_out.date() - check_in.date()).days, 0) if check_in and check_out else 0
        is_non_commercial = status in _NON_COMMERCIAL_STATUSES

        status_counts[status] += 1
        stat = channel_stats.setdefault(
            channel,
            {"channel": channel, "bookings": 0, "commercial_bookings": 0, "nights": 0, "revenue": 0.0, "cancelled": 0},
        )
        stat["bookings"] += 1

        if is_non_commercial:
            stat["cancelled"] += 1
            cancelled_value += amount
            if status in {"no_show", "noshow"}:
                no_show_count += 1
            else:
                cancelled_count += 1
        else:
            commercial_count += 1
            stat["commercial_bookings"] += 1
            stat["nights"] += nights
            stat["revenue"] += amount
            total_nights += nights
            booked_revenue += amount

        lead_days: int | None = None
        if check_in and created_at:
            lead_days = max((check_in.date() - created_at.date()).days, 0)
            lead_counts[_lead_time_bucket(lead_days)] += 1
            lead_time_total += lead_days
            lead_time_count += 1

        if check_in and start_date <= check_in.date() <= end_date:
            day = check_in.date().isoformat()
            daily = arrival_days.setdefault(day, {"date": day, "reservations": 0, "commercial_reservations": 0, "revenue": 0.0})
            daily["reservations"] += 1
            if not is_non_commercial:
                daily["commercial_reservations"] += 1
                daily["revenue"] += amount

        rows.append(
            {
                "booking_id": str(booking.get("id") or ""),
                "guest_name": str(booking.get("guest_name") or "Misafir"),
                "room_number": str(booking.get("room_number") or "-"),
                "check_in": check_in.date().isoformat() if check_in else str(booking.get("check_in") or ""),
                "check_out": check_out.date().isoformat() if check_out else str(booking.get("check_out") or ""),
                "status": status,
                "status_label": _STATUS_LABELS.get(status, status.replace("_", " ").title()),
                "channel": channel,
                "total_amount": round(amount, 2),
                "nights": nights,
                "lead_time_days": lead_days,
            }
        )

    total = len(bookings)
    status_breakdown = [
        {"status": key, "label": _STATUS_LABELS.get(key, key.replace("_", " ").title()), "count": count}
        for key, count in sorted(status_counts.items(), key=lambda item: (-item[1], item[0]))
    ]
    channels = sorted(channel_stats.values(), key=lambda item: (-item["bookings"], item["channel"].lower()))
    for item in channels:
        item["revenue"] = round(item["revenue"], 2)
    lead_time_breakdown = [
        {"bucket": key, "label": label, "count": lead_counts[key]}
        for key, label, *_ in _LEAD_TIME_BUCKETS
    ]
    rows.sort(key=lambda row: (row["check_in"], row["guest_name"], row["booking_id"]))

    return {
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "summary": {
            "total_bookings": total,
            "commercial_bookings": commercial_count,
            "booked_revenue": round(booked_revenue, 2),
            "total_room_nights": int(total_nights),
            "average_stay": round(total_nights / commercial_count, 2) if commercial_count else 0.0,
            "average_lead_time": round(lead_time_total / lead_time_count, 1) if lead_time_count else 0.0,
            "cancelled_count": cancelled_count,
            "no_show_count": no_show_count,
            "cancellation_rate": round((cancelled_count / total) * 100, 1) if total else 0.0,
            "cancelled_value": round(cancelled_value, 2),
        },
        "status_breakdown": status_breakdown,
        "channel_breakdown": channels,
        "lead_time_breakdown": lead_time_breakdown,
        "daily_arrivals": [arrival_days[key] for key in sorted(arrival_days)],
        "rows": rows,
    }


async def _reservation_performance_payload(
    *, tenant_id: str, start_date: date, end_date: date
) -> dict[str, Any]:
    # The report is explicitly arrival-date based.  This matches the calendars,
    # arrivals list, and operational expectations of front desk users.
    # ISO date strings compare correctly with both historic ``YYYY-MM-DD``
    # values and current full ISO datetimes.  Using a midnight timestamp as
    # the lower bound would silently exclude old date-only reservations.
    start_key = start_date.isoformat()
    end_exclusive_key = (end_date + timedelta(days=1)).isoformat()
    projection = {
        "_id": 0,
        "id": 1,
        "guest_name": 1,
        "room_number": 1,
        "check_in": 1,
        "check_out": 1,
        "status": 1,
        "channel": 1,
        "source_channel": 1,
        "source": 1,
        "ota_channel": 1,
        "total_amount": 1,
        "created_at": 1,
        "reservation_date": 1,
    }
    bookings = await db.bookings.find(
        {
            "tenant_id": tenant_id,
            "check_in": {"$gte": start_key, "$lt": end_exclusive_key},
        },
        projection,
    ).to_list(20000)
    return aggregate_reservation_performance(bookings, start_date=start_date, end_date=end_date)


@sub_router.get("/reports/reservation-performance")
@cached(ttl=180, key_prefix="report_reservation_performance")
async def get_reservation_performance_report(
    start_date: str | None = None,
    end_date: str | None = None,
    current_user: User = Depends(get_current_user),
    _: None = Depends(require_module("reports")),
    _perm=Depends(require_op("view_reports")),
    _nocache: bool = Query(False, alias="nocache"),
):
    end = _parse_date(end_date, "end_date") if end_date else datetime.now(UTC).date()
    start = _parse_date(start_date, "start_date") if start_date else end - timedelta(days=29)
    if start > end:
        raise HTTPException(status_code=422, detail="Başlangıç tarihi bitiş tarihinden sonra olamaz")
    if (end - start).days > 366:
        raise HTTPException(status_code=422, detail="Rezervasyon raporu en fazla 366 günlük aralıkta alınabilir")
    return await _reservation_performance_payload(tenant_id=current_user.tenant_id, start_date=start, end_date=end)


@sub_router.get("/reports/reservation-performance/excel")
async def export_reservation_performance_excel(
    start_date: str | None = None,
    end_date: str | None = None,
    current_user: User = Depends(get_current_user),
    _: None = Depends(require_module("reports")),
    _perm=Depends(require_op("view_reports")),
):
    end = _parse_date(end_date, "end_date") if end_date else datetime.now(UTC).date()
    start = _parse_date(start_date, "start_date") if start_date else end - timedelta(days=29)
    if start > end:
        raise HTTPException(status_code=422, detail="Başlangıç tarihi bitiş tarihinden sonra olamaz")
    payload = await _reservation_performance_payload(tenant_id=current_user.tenant_id, start_date=start, end_date=end)
    summary = payload["summary"]
    data = [
        ["Toplam rezervasyon", summary["total_bookings"]],
        ["Gelir getiren rezervasyon", summary["commercial_bookings"]],
        ["Rezerve gelir", summary["booked_revenue"]],
        ["Oda/gece", summary["total_room_nights"]],
        ["Ortalama konaklama", summary["average_stay"]],
        ["Ortalama rezervasyon süresi", summary["average_lead_time"]],
        ["İptal", summary["cancelled_count"]],
        ["No-show", summary["no_show_count"]],
        ["İptal oranı %", summary["cancellation_rate"]],
    ]
    workbook = create_excel_workbook(
        title=f"Rezervasyon Performansı ({start.isoformat()} - {end.isoformat()})",
        headers=["Gösterge", "Değer"],
        data=data,
        sheet_name="Özet",
    )
    from core.csv_safe import xlsx_safe
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.styles import Font, PatternFill

    def add_sheet(title: str, headers: list[str], values: list[list[Any]]):
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for row in values:
            sheet.append([
                xlsx_safe(ILLEGAL_CHARACTERS_RE.sub("", value)[:32767]) if isinstance(value, str) else value
                for value in row
            ])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(max((len(str(cell.value or "")) for cell in column), default=10) + 2, 36)

    add_sheet(
        "Kanallar",
        ["Kanal", "Rezervasyon", "Gelir getiren", "Gece", "Rezerve gelir", "İptal/No-show"],
        [[item["channel"], item["bookings"], item["commercial_bookings"], item["nights"], item["revenue"], item["cancelled"]] for item in payload["channel_breakdown"]],
    )
    add_sheet(
        "Durumlar",
        ["Durum", "Rezervasyon"],
        [[item["label"], item["count"]] for item in payload["status_breakdown"]],
    )
    add_sheet(
        "Rezervasyonlar",
        ["Misafir", "Oda", "Giriş", "Çıkış", "Durum", "Kanal", "Gece", "Tutar", "Rezervasyon süresi (gün)"],
        [[row["guest_name"], row["room_number"], row["check_in"], row["check_out"], row["status_label"], row["channel"], row["nights"], row["total_amount"], row["lead_time_days"]] for row in payload["rows"]],
    )
    return excel_response(workbook, f"rezervasyon_performansi_{start.isoformat()}_{end.isoformat()}.xlsx")

"""
Syroce PMS - Shared Utility Functions
Used across multiple routers. Extracted from server.py to avoid circular imports.
"""

import logging

logger = logging.getLogger(__name__)
import base64
import io
import secrets
from datetime import UTC, datetime, timedelta
from typing import Any

from core.database import db

# ── Folio Helpers ──


async def generate_folio_number(tenant_id: str) -> str:
    """Generate unique folio number"""
    year = datetime.now(UTC).year
    count = await db.folios.count_documents({"tenant_id": tenant_id}) + 1
    return f"F-{year}-{count:05d}"


async def calculate_folio_balance(folio_id: str, tenant_id: str) -> float:
    """Folio bakiyesi (charges − payments) — server-side $sum aggregation.

    - to_list cap yok: 500/1000 hard limit'lerin sessiz veri kaybını engeller.
    - charges.total yoksa amount fallback (schema esnek).
    - Fail-closed: Mongo hatası → exception bubble (caller karar versin).
      Önceki fail-open (return 0.0) yanlış pozitif "ödenmiş" durumu yaratıp
      ödenmemiş folyoların kapatılmasına yol açıyordu.
    """
    ch_pipe = [
        {"$match": {"folio_id": folio_id, "tenant_id": tenant_id, "voided": False}},
        {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$total", "$amount"]}}}},
    ]
    pay_pipe = [
        {"$match": {"folio_id": folio_id, "tenant_id": tenant_id, "voided": False}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}},
    ]
    ch_doc = await db.folio_charges.aggregate(ch_pipe).to_list(1)
    pay_doc = await db.payments.aggregate(pay_pipe).to_list(1)
    total_charges = float(ch_doc[0]["total"]) if ch_doc else 0.0
    total_payments = float(pay_doc[0]["total"]) if pay_doc else 0.0
    return round(total_charges - total_payments, 2)


# ── Excel Helpers ──

_XLSX_MAX_CELL_LEN = 32767


def _xlsx_sanitize_str(s: str) -> str:
    """Strip openpyxl-illegal control chars and cap cell length (Task #253).

    openpyxl raises `IllegalCharacterError` for C0 control chars
    (0x00-0x08, 0x0B-0x0C, 0x0E-0x1F) anywhere in a cell value, and a hard
    32767-char per-cell limit. Stress tenants accumulate residue (free-text
    descriptions, decoded byte arrays with errors='replace', user-typed
    notes) that can carry these characters; without this guard, a single
    bad row would 500 the entire export.
    """
    if not s:
        return s
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    cleaned = ILLEGAL_CHARACTERS_RE.sub("", s)
    if len(cleaned) > _XLSX_MAX_CELL_LEN:
        cleaned = cleaned[: _XLSX_MAX_CELL_LEN - 1] + "…"
    return cleaned


def create_excel_workbook(title: str, headers: list[str], data: list[list[Any]], sheet_name: str = "Report"):
    """Create a formatted Excel workbook with data"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # Bug AN: openpyxl interprets a leading '=' as a formula → spreadsheet
    # formula injection. xlsx_safe() prepends an apostrophe to neutralize.
    from core.csv_safe import xlsx_safe

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.merge_cells("A1:" + get_column_letter(len(headers)) + "1")
    title_cell = ws["A1"]
    title_cell.value = _xlsx_sanitize_str(title) if isinstance(title, str) else title
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        # Task #253: sanitize illegal chars + length cap before injection guard.
        cell.value = _xlsx_sanitize_str(xlsx_safe(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = 15

    for row_num, row_data in enumerate(data, 3):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            # Preserve numeric types (Excel needs them for sums/formatting);
            # only sanitize string-typed cells, which is where injection lives.
            # Task #253: also strip openpyxl-illegal control chars and cap at
            # 32767 chars so a single bad seed string can't 500 the export.
            if isinstance(value, str):
                cell.value = _xlsx_sanitize_str(xlsx_safe(value))
            else:
                cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    return wb


def excel_response(workbook, filename: str):
    """Convert workbook to StreamingResponse for download"""
    from fastapi.responses import StreamingResponse

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def apply_row_colors(ws, start_row=2):
    """Apply alternating colors to Excel worksheet rows and auto-adjust column widths."""
    from openpyxl.styles import PatternFill

    for row_num, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
        for cell in row:
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col in ws.columns:
        max_length = 0
        try:
            column = col[0].column_letter
        except AttributeError:
            continue
        for cell in col:
            try:
                if hasattr(cell, "value") and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    return ws


# ── QR Code Helpers ──


def generate_qr_code(data: str) -> str:
    """Generate QR code as base64 data URI"""
    try:
        import qrcode

        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")

        buffer = io.BytesIO()
        img.save(buffer, format="PNG")
        buffer.seek(0)
        img_base64 = base64.b64encode(buffer.getvalue()).decode()
        return f"data:image/png;base64,{img_base64}"
    except Exception:
        return ""


def generate_time_based_qr_token(booking_id: str, expiry_hours: int = 72) -> str:
    """Generate time-based QR token for booking"""
    import os

    import jwt as pyjwt

    JWT_SECRET = os.environ.get("JWT_SECRET")
    if not JWT_SECRET:
        # v107 (Bug DAG): hardcoded 'fallback-secret' was known-string — QR tokens
        # signed in fallback mode would NOT verify against core/security JWT_SECRET
        # (mismatch breaks decode); also enabled QR token forging if attacker knew
        # the fallback. Now fail-closed in production, random in dev.
        if os.environ.get("STRICT_JWT_SECRET") == "1" or os.environ.get("ENV", "").lower() == "production":
            raise RuntimeError("JWT_SECRET environment variable is required to sign QR tokens in production.")
        JWT_SECRET = secrets.token_urlsafe(64)
    JWT_ALGORITHM = "HS256"
    expiry = datetime.now(UTC) + timedelta(hours=expiry_hours)
    token = secrets.token_urlsafe(32)
    return pyjwt.encode(
        {
            "booking_id": booking_id,
            "token": token,
            "exp": expiry,
        },
        JWT_SECRET,
        algorithm=JWT_ALGORITHM,
    )


# ── Cancellation Policy ──


def get_cancellation_policy_details(policy: str):
    """Get cancellation policy details"""
    policies = {
        "flexible": {
            "name": "Flexible",
            "description": "Free cancellation up to 24 hours before check-in",
            "free_cancellation_hours": 24,
            "penalty_percentage": 0,
        },
        "moderate": {
            "name": "Moderate",
            "description": "Free cancellation up to 5 days before check-in",
            "free_cancellation_hours": 120,
            "penalty_percentage": 50,
        },
        "strict": {
            "name": "Strict",
            "description": "Free cancellation up to 14 days before check-in",
            "free_cancellation_hours": 336,
            "penalty_percentage": 100,
        },
        "non_refundable": {
            "name": "Non-Refundable",
            "description": "No refund upon cancellation",
            "free_cancellation_hours": 0,
            "penalty_percentage": 100,
        },
    }
    return policies.get(policy, policies.get("flexible"))


# ── Night Audit Helpers ──


async def night_audit_post_room_charges(tenant_id: str, date: str):
    """Post room charges for all occupied rooms"""
    posted_count = 0
    total_amount = 0

    async for booking in db.bookings.find(
        {
            "tenant_id": tenant_id,
            "status": "checked_in",
            "check_in": {"$lte": date},
            "check_out": {"$gte": date},
        }
    ):
        folio = await db.folios.find_one(
            {
                "booking_id": booking.get("id"),
                "folio_type": "guest",
                "status": "open",
            }
        )
        if folio:
            nights = (datetime.fromisoformat(booking.get("check_out")) - datetime.fromisoformat(booking.get("check_in"))).days
            room_rate = booking.get("total_amount", 0) / nights if nights > 0 else 0
            posted_count += 1
            total_amount += room_rate

    return {"charges_posted": posted_count, "total_amount": round(total_amount, 2)}


async def night_audit_calculate_revenue(tenant_id: str, date: str):
    """Calculate daily revenue breakdown"""
    revenue = {"room_revenue": 0, "fnb_revenue": 0, "other_revenue": 0, "total_revenue": 0}

    async for charge in db.folio_charges.find(
        {
            "tenant_id": tenant_id,
            "date": {"$gte": date, "$lt": (datetime.fromisoformat(date) + timedelta(days=1)).isoformat()},
        }
    ):
        category = charge.get("charge_category")
        amount = charge.get("total", 0)
        if category == "room":
            revenue["room_revenue"] += amount
        elif category in ["food", "beverage"]:
            revenue["fnb_revenue"] += amount
        else:
            revenue["other_revenue"] += amount
        revenue["total_revenue"] += amount

    return {k: round(v, 2) for k, v in revenue.items()}


async def night_audit_recalculate_ar(tenant_id: str):
    """Recalculate accounts receivable"""
    total_ar = 0
    open_folios = 0

    async for folio in db.folios.find(
        {
            "tenant_id": tenant_id,
            "status": "open",
            "folio_type": {"$in": ["company", "agency"]},
        }
    ):
        balance = folio.get("balance", 0)
        total_ar += balance
        open_folios += 1

    return {"total_ar": round(total_ar, 2), "open_folios": open_folios}


async def night_audit_housekeeping_rollup(tenant_id: str, date: str):
    """Housekeeping summary for the day"""
    tasks_completed = await db.housekeeping_tasks.count_documents(
        {
            "tenant_id": tenant_id,
            "status": "completed",
            "completed_at": {"$gte": date, "$lt": (datetime.fromisoformat(date) + timedelta(days=1)).isoformat()},
        }
    )
    return {"tasks_completed": tasks_completed, "date": date}


async def night_audit_ota_reconciliation(tenant_id: str, date: str):
    """OTA bookings reconciliation"""
    ota_bookings = 0
    ota_revenue = 0

    async for booking in db.bookings.find(
        {
            "tenant_id": tenant_id,
            "check_in": date,
            "ota_channel": {"$ne": None},
        }
    ):
        ota_bookings += 1
        ota_revenue += booking.get("total_amount", 0)

    return {"ota_bookings": ota_bookings, "ota_revenue": round(ota_revenue, 2)}
